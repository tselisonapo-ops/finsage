print("[BOOT] api_server.py import started")

# ────────────────────────────────────────────────────────────────
# Load environment FIRST (before ANY BackEnd.Services imports)
# ────────────────────────────────────────────────────────────────
import os
from pathlib import Path
import hashlib
from dotenv import load_dotenv
from werkzeug.exceptions import HTTPException
import psycopg2.extras
from werkzeug.utils import secure_filename
import openpyxl
import psycopg2
from psycopg2 import errors as pg_errors
ALLOWED_LOGO_EXTS = {".png", ".jpg", ".jpeg", ".webp"}

import json as _json
import json
# ----------------------------------------------------------------
# Environment loading
# ----------------------------------------------------------------
APP_ENV = str(os.getenv("APP_ENV") or "development").strip().lower()
ENV_PATH = Path(__file__).resolve().parent / ".env"

if APP_ENV == "production":
    # Running under Passenger/cPanel.
    # Environment variables come from .htaccess.
    print("[BOOT] Production mode - using Passenger environment variables")
else:
    # Local development only.
    if ENV_PATH.exists():
        load_dotenv(
            dotenv_path=ENV_PATH,
            override=False,
        )
        print(f"[BOOT] Development .env loaded: {ENV_PATH}")
    else:
        print(f"[BOOT] Development .env not found: {ENV_PATH}")

import logging
import re

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()

logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("finsage")

print("[BOOT] Environment:", APP_ENV)
print(
    "[BOOT] MASTER_DB_DSN present:",
    bool(os.getenv("MASTER_DB_DSN")),
)

if not os.getenv("MASTER_DB_DSN"):
    source = (
        "Passenger (.htaccess)"
        if APP_ENV == "production"
        else str(ENV_PATH)
    )
    raise RuntimeError(
        f"MASTER_DB_DSN is not set (expected from {source})"
    )

JWT_SECRET_KEY = os.getenv("JWT_SECRET_KEY") or ""

if not JWT_SECRET_KEY:
    raise RuntimeError("JWT_SECRET_KEY is not set")

# Standard library imports
# ────────────────────────────────────────────────────────────────
import secrets
import time
import csv
import io
from collections import defaultdict
from datetime import datetime, timedelta, date, timezone
from decimal import Decimal, ROUND_HALF_UP
import traceback
from typing import Dict, Any, Optional, List, Tuple
import inspect  # ✅ add with other stdlib imports


ALLOWED_EXT = {".png", ".jpg", ".jpeg", ".webp"}

ALLOWED_POS_EXTS = {".csv", ".xlsx", ".xls", ".pdf", ".png", ".jpg", ".jpeg", ".webp"}

# ────────────────────────────────────────────────────────────────
# BackEnd imports that rely on env vars
# ────────────────────────────────────────────────────────────────
from BackEnd.Services.company_context import get_company_context, get_dashboard_access, build_permissions, resolve_default_dashboard
from BackEnd.Services.dashboard import validate_role_for_scope
from BackEnd.Services.credit_policy import (
    normalize_policy, 
    can_post_invoices, 
    must_approve_customer_before_invoicing, 
    normalize_policy_mode, 
    normalize_role
)
from BackEnd.Services.emailer import _send_smtp
from BackEnd.Services.routes.invoice_routes import build_invoice_journal_lines  # wherever yours lives
from BackEnd.Services.company import apply_mode_defaults, company_policy, recommend_mode_after_invite
from BackEnd.Services.reporting.balance_sheet_templates import get_balance_sheet_v3_exact
from BackEnd.Services import accounting_classifiers as ac
from BackEnd.Services.reporting.reporting_helpers import parse_date_arg
from .reporting import reporting_helpers as rh
from BackEnd.Services.invoice_pdf_service import generate_invoice_pdf
from BackEnd.Services.bank_service import BankService
from BackEnd.Services.coa_seed_service import seed_company_coa_once
from BackEnd.Services.industry_profiles import get_industry_profile
from BackEnd.Services.period_core import resolve_company_period, resolve_compare_period
from BackEnd.Services.utils.registration_utils import create_company_record_from_payload
from BackEnd.Services.validation import REGNO_REGEX, TIN_REGEX, VAT_REGEX, FIELD_EXAMPLES
# ────────────────────────────────────────────────────────────────
# Third-party imports
# ────────────────────────────────────────────────────────────────
from flask import (
    Flask,
    jsonify,
    request,
    current_app,
    g,
    render_template,
    send_from_directory,
    make_response,
    url_for,   # ✅ you use url_for later
    Response,
    stream_with_context,
)
from flask_cors import CORS, cross_origin
from psycopg2.errors import UniqueViolation
from werkzeug.security import check_password_hash, generate_password_hash

# ────────────────────────────────────────────────────────────────
# NOW it's safe to import your BackEnd modules (they need env vars)
# ────────────────────────────────────────────────────────────────
from BackEnd.Services.db_service import db_service

from BackEnd.Services.auth_middleware import _corsify, require_auth, require_pos_auth
from BackEnd.Services.auth_service import make_jwt, make_pos_jwt

from BackEnd.Services.coa_service import (
    build_coa,
    build_coa_flat,
    list_industries,
    get_industry_template,
    canonical_subindustry_key,
    TEMPLATE_INDUSTRY_ALIASES,
)
from BackEnd.Services.validation import validate_company_payload, get_currency_for_country
from BackEnd.Services.countries import load_countries
from BackEnd.Services.emailer import send_mail,send_company_mail
from BackEnd.Services.periods import resolve_period

from BackEnd.Services.reporting.cashflow_templates import (
    _norm_preview_columns as norm_preview_columns,
    _norm_compare as norm_compare,
)

from BackEnd.Services.reporting.reporting_helpers import (
    build_compare_range, label_period, make_columns, has_delta,
    parse_date_arg,
    shift_year,
    want_export,
)
from BackEnd.Services.utils.industry_utils import normalize_industry_pair, slugify, TEMPLATE_INDUSTRY_ALIASES

from BackEnd.Services.industry_profiles import get_industry_profile
from BackEnd.Services.reporting.reporting_helpers import build_income_statement_template, choose_layout
from BackEnd.Services.utils.view_token import create_invoice_pdf_token, verify_invoice_pdf_token, make_invoice_view_token, verify_quote_pdf_token, create_quote_pdf_token
from BackEnd.Services.periods import parse_date_maybe
from BackEnd.Services.reporting.statement_renderer import render_statement_html
from BackEnd.Services.config import Config
# ────────────────────────────────────────────────────────────────
# Blueprints
# ────────────────────────────────────────────────────────────────
print("[BOOT] About to import/register blueprints")

from BackEnd.Services.lease_routes import bp as leases_bp
from BackEnd.Services.routes.receipt_routes import receipts_bp
from BackEnd.Services.routes.invoice_routes import invoices_bp
from .journal_routes import journal_bp
from .credit_routes import credit_bp
from .reports_routes import reports_bp
from .coa_routes import coa_bp
from .vat_settings import bp as vat_settings_bp
from .ar_reports_bp import ar_reports_bp
from .vat_utils import vat_utils_bp
from BackEnd.Services.management_packs import bp_companies_management_packs
from BackEnd.Services.routes.vendor_routes import ap_bp
from BackEnd.Services.quotation_routes import quotes_bp
from BackEnd.Services.routes.bank_routes import bank_bp
from BackEnd.Services.ap_reports_bp import ap_reports_bp
from BackEnd.Services.bp_approvals_wf import bp_approvals
from BackEnd.Services.routes.route_lessor import lessors_bp
from BackEnd.Services.routes.ifrs16_reporting_routes import bp_ifrs16
from BackEnd.Services.assets.ppe_reporting import ppe_bp
from BackEnd.Services.assets.ias16_reporting import ppe_reporting_bp
from BackEnd.Services.assets.asset_reports_routes import asset_reports_bp
from BackEnd.Services.practitioner.practitioner_engagements import engagements_bp
from BackEnd.Services.practitioner.engagement_ops import engagement_ops_bp
from BackEnd.Services.support import support_bp
from BackEnd.Services.practitioner.analytics_routes  import analytics_bp
from BackEnd.Services.practitioner.client_overview_bp import client_overview_bp
from BackEnd.Services.practitioner.action_center import action_center_bp
from BackEnd.Services.practitioner.posting_module import practitioner_dashboard_bp
from BackEnd.Services.practitioner.review_queue import review_queue_bp
from BackEnd.Services.practitioner.engagement_working_papers_routes import engagement_working_papers_bp
from BackEnd.Services.practitioner.resource_planning import resource_planning_bp
from BackEnd.Services.practitioner.final_deliverables_review import final_deliverables_review_bp
from BackEnd.Services.practitioner.approval_center import approval_center_bp
from BackEnd.Services.practitioner.practitioner_risk_independence import risk_independence_bp
from BackEnd.Services.practitioner.practitioner_override_log import override_log_bp
from BackEnd.Services.routes.loans_routes import loans_bp
from BackEnd.Services.routes.revenue_routes import revenue_bp
from BackEnd.Routes.report_run import report_bp
from BackEnd.Services.routes.project_routes import projects_bp
from BackEnd.Services.routes.pos_routes import pos_bp
from BackEnd.Services.routes.asset_tax_routes import bp_asset_tax
from BackEnd.Services.routes.accrual_deferral_routes import bp_accrual_deferral
from BackEnd.Services.routes.ifrs9_routes import bp_ifrs9
from BackEnd.Services.routes.payroll_routes import payroll_bp
from BackEnd.Services.routes.forecast_routes import forecast_bp
from BackEnd.Services.routes.deferred_tax_routes import deferred_tax_bp
from BackEnd.Services.routes.payroll_employee_benefits_routes import payroll_employee_benefits_bp
from BackEnd.Services.lessor_lease_routes import lessor_bp
from BackEnd.Services.routes.data_migration_routes import data_migration_bp
from BackEnd.Services.routes.ias41_routes import ias41_bp
from BackEnd.Services.routes.corporate_structure import corporate_structure_bp
from BackEnd.Services.routes.ops_routes import ops_bp
from BackEnd.Services.routes.payroll_disclosure_routes import (
    payroll_disclosure_bp,
)
from BackEnd.Services.routes.ias41_disclosure_routes import (
    bp_ias41_disclosure,
)
from BackEnd.Services.routes.consolidation_routes import consolidation_bp
from BackEnd.Services.routes.payroll_tax_filing_routes import payroll_tax_filing_bp

# ────────────────────────────────────────────────────────────────
# Flask app + CORS
# ────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent  # BackEnd/Services
TEMPLATE_DIR = BASE_DIR / "templates"

def get_db_connection():
    """
    Simple DB connection helper for blueprints/services that need a raw cursor.
    Adjust this if your db_service exposes a preferred connection method.
    """
    dsn = os.getenv("MASTER_DB_DSN")
    if not dsn:
        raise RuntimeError("MASTER_DB_DSN is not set")
    return psycopg2.connect(dsn)

from flask import Flask

import pdfkit

try:
    config = pdfkit.configuration()
    print(f"[DEBUG] wkhtmltopdf path: {config.wkhtmltopdf}")
except Exception as e:
    print(f"[DEBUG] wkhtmltopdf not found: {e}")

# keep your existing app initialization
app = Flask(__name__, template_folder=str(TEMPLATE_DIR))
app.config.from_object(Config)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024
app.config["MIGRATION_MAX_FILE_SIZE"] = 50 * 1024 * 1024
@app.route("/")
def home():
    
    return {"ok": True, "message": "FinSage API is live"}

@app.route("/health")
def health():
    return "OK", 200

print("[BOOT] Flask app created")

app.config["DB_SERVICE"] = db_service
app.config["db_service"] = db_service
app.config["GET_DB_CONNECTION"] = get_db_connection
app.config["GET_TRIAL_BALANCE_FN"] = getattr(db_service, "get_trial_balance", None)

import logging
import os
from pathlib import Path
from logging.handlers import RotatingFileHandler


app_env = os.getenv("APP_ENV", "development").strip().lower()

if app_env == "production":
    signup_log_path = Path(
        "/home/finspher/Finsage-Web-App/signup_debug.log"
    )
else:
    signup_log_path = (
        Path(__file__).resolve().parents[2]
        / "signup_debug.log"
    )

signup_log_path.parent.mkdir(
    parents=True,
    exist_ok=True,
)

signup_log_handler = RotatingFileHandler(
    str(signup_log_path),
    maxBytes=2_000_000,
    backupCount=2,
    encoding="utf-8",
)

signup_log_handler.setLevel(logging.INFO)

signup_log_handler.setFormatter(
    logging.Formatter(
        "%(asctime)s %(levelname)s %(message)s"
    )
)

app.logger.addHandler(signup_log_handler)
app.logger.setLevel(logging.INFO)

print(f"[BOOT] Signup debug log: {signup_log_path}")

bank_service = BankService(db_service)

origins = app.config.get("FRONTEND_ORIGINS", [])
print("[BOOT] FRONTEND_ORIGINS:", origins)

CORS(
    app,
    resources={
        r"/api/*": {"origins": origins},
        r"/uploads/*": {"origins": origins},
    },
    supports_credentials=True,
    allow_headers=["Content-Type", "Authorization"],
    expose_headers=["Content-Type", "Authorization"],
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
)

for r in app.url_map.iter_rules():
    if "invoices" in str(r) and "view" in str(r):
        print("ROUTE:", r, "->", r.endpoint)

@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        resp = app.make_default_options_response()
        return _corsify(resp)

@app.after_request
def apply_cors(resp):
    return _corsify(resp)

@app.before_request
def log_and_handle_preflight():
    print("[REQ]", request.method, request.path, "Origin=", request.headers.get("Origin"))
    if request.method == "OPTIONS":
        resp = app.make_default_options_response()
        return _corsify(resp)
    
@app.errorhandler(Exception)
def handle_any_exception(e):
    if isinstance(e, HTTPException):
        resp = jsonify({"ok": False, "error": e.description, "type": e.__class__.__name__})
        resp.status_code = e.code or 500
        return resp

    current_app.logger.exception("Unhandled server error")
    resp = jsonify({"ok": False, "error": str(e), "type": e.__class__.__name__})
    resp.status_code = 500
    return resp


app.register_blueprint(leases_bp)
app.register_blueprint(receipts_bp)
app.register_blueprint(invoices_bp)
app.register_blueprint(coa_bp)
app.register_blueprint(journal_bp)
app.register_blueprint(reports_bp)
app.register_blueprint(credit_bp)
app.register_blueprint(vat_settings_bp)
app.register_blueprint(bank_bp)
app.register_blueprint(ar_reports_bp)
app.register_blueprint(quotes_bp)
app.register_blueprint(vat_utils_bp)
app.register_blueprint(ap_bp)
app.register_blueprint(ap_reports_bp)
app.register_blueprint(bp_companies_management_packs)
app.register_blueprint(bp_approvals)
app.register_blueprint(lessors_bp)
app.register_blueprint(bp_ifrs16)
app.register_blueprint(ppe_bp)
app.register_blueprint(ppe_reporting_bp)
app.register_blueprint(asset_reports_bp)
app.register_blueprint(engagements_bp)
app.register_blueprint(engagement_ops_bp)
app.register_blueprint(support_bp)
app.register_blueprint(analytics_bp)
app.register_blueprint(client_overview_bp)
app.register_blueprint(action_center_bp)
app.register_blueprint(practitioner_dashboard_bp)
app.register_blueprint(review_queue_bp)
app.register_blueprint(engagement_working_papers_bp)
app.register_blueprint(resource_planning_bp)
app.register_blueprint(final_deliverables_review_bp)
app.register_blueprint(approval_center_bp)
app.register_blueprint(risk_independence_bp)
app.register_blueprint(override_log_bp)
app.register_blueprint(loans_bp)
app.register_blueprint(revenue_bp)
app.register_blueprint(report_bp)
app.register_blueprint(projects_bp)
app.register_blueprint(pos_bp)
app.register_blueprint(bp_asset_tax)
app.register_blueprint(bp_accrual_deferral)
app.register_blueprint(bp_ifrs9)
app.register_blueprint(payroll_bp)
app.register_blueprint(forecast_bp)
app.register_blueprint(deferred_tax_bp)
app.register_blueprint(payroll_employee_benefits_bp)
app.register_blueprint(lessor_bp)
app.register_blueprint(data_migration_bp)
app.register_blueprint(ias41_bp)
app.register_blueprint(payroll_disclosure_bp)
app.register_blueprint(bp_ias41_disclosure)
app.register_blueprint(corporate_structure_bp)
app.register_blueprint(ops_bp)
app.register_blueprint(consolidation_bp)
app.register_blueprint(payroll_tax_filing_bp)
# If you have app.run(...) later, add this right above it:
# print("[BOOT] About to run Flask server")

# put after app + blueprints are registered (eg end of create_app / api_server boot)
print("=== URL MAP ===")
for r in app.url_map.iter_rules():
    if "vat" in r.rule:
        print(r.rule, sorted(r.methods))
print("===============")

FRONTEND_BASE = os.getenv("FRONTEND_BASE", "http://127.0.0.1:5500")
logger.debug("FRONTEND_BASE=%s", FRONTEND_BASE)


# ────────────────────────────────────────────────────────────────
# Invitations (dev in-memory)
# ────────────────────────────────────────────────────────────────

INVITES: Dict[str, Dict] = {}
INVITE_TTL_SECONDS = 7 * 24 * 60 * 60  # 7 days



def _create_invite_token(
    email: str,
    role: str,
    company_name: str,
    company_id: int | None,
) -> str:
    """
    Create an in-memory invite record that includes which company
    the invite belongs to. This lets us link invited users to the
    inviter's company automatically.
    """
    token = secrets.token_urlsafe(32)

    INVITES[token] = {
        "email": (email or "").strip().lower(),
        "role": (role or "").strip().lower(),   # ✅ normalize
        "company_name": company_name or "FinSage workspace",
        "company_id": company_id,
        "created": int(time.time()),
        "accepted": False,
    }
    return token

# ────────────────────────────────────────────────────────────────
# AUTH ROUTES
# ────────────────────────────────────────────────────────────────

def money(x) -> float:
    return float(Decimal(str(x or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

def rate_from_code(code: str) -> float:
    c = (code or "STANDARD").strip().upper()
    if c == "STANDARD":
        return 15.0
    if c in {"ZERO", "ZER0"}:
        return 0.0
    if c == "EXEMPT":
        return 0.0
    return 15.0  # default safe

def _company_guard(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return False
    return True

def _json_or_501(msg="Export format not implemented"):
    return jsonify({"error": msg}), 501

def _flatten_statement_for_export(stmt: dict):
    """
    Converts the v2 schema into rows:
    [section, code, name, col_key, amount]
    """
    cols = stmt.get("columns") or [{"key": "cur", "label": "Amount"}]
    out = []

    for sec in (stmt.get("sections") or []):
        sec_label = sec.get("label", "")
        for ln in (sec.get("lines") or []):
            code = ln.get("code", "")
            name = ln.get("name", "")
            values = ln.get("values") or {}
            for c in cols:
                k = c["key"]
                out.append({
                    "section": sec_label,
                    "code": code,
                    "name": name,
                    "column": c.get("label", k),
                    "amount": float(values.get(k) or 0.0),
                })

        totals = sec.get("totals") or {}
        for c in cols:
            k = c["key"]
            out.append({
                "section": sec_label,
                "code": "",
                "name": "TOTAL",
                "column": c.get("label", k),
                "amount": float(totals.get(k) or 0.0),
            })

    return out

def sha256_bytes(b: bytes) -> str:
    h = hashlib.sha256()
    h.update(b)
    return h.hexdigest()

def pos_upload_folder(app_root: str) -> str:
    folder = os.path.join(app_root, "uploads", "pos_imports")
    os.makedirs(folder, exist_ok=True)
    return folder

def norm_col(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (s or "").strip().lower()).strip()

def guess_mapping_from_columns(columns: list[str]) -> dict:
    # patterns
    patterns = {
        "pos_date":      [r"\bdate\b", r"\btran\b", r"\bsale\b", r"\btimestamp\b"],
        "item_code":     [r"\bsku\b", r"\bitem code\b", r"\bcode\b", r"\bplu\b", r"\bbarcode\b"],
        "description":   [r"\bdesc\b", r"\bdescription\b", r"\bitem\b", r"\bproduct\b", r"\bname\b"],
        "qty":           [r"\bqty\b", r"\bquantity\b", r"\bunits\b", r"\bpcs\b"],
        "unit_price":    [r"\bunit price\b", r"\bprice\b", r"\brate\b"],
        "gross_amount":  [r"\bgross\b", r"\btotal\b", r"\bamount\b", r"\bincl\b"],
        "net_amount":    [r"\bnet\b", r"\bex vat\b", r"\bexclusive\b"],
        "vat_amount":    [r"\bvat\b", r"\btax\b", r"\bgst\b"],
        "vat_rate":      [r"\bvat %\b", r"\btax rate\b", r"\brate %\b"],
    }

    cols_norm = [(c, norm_col(c)) for c in columns]

    def pick(field: str) -> str | None:
        for rx in patterns.get(field, []):
            rxx = re.compile(rx)
            for orig, nn in cols_norm:
                if rxx.search(nn):
                    return orig
        return None

    guess = {k: pick(k) for k in patterns.keys()}
    # only include those we found
    return {k: v for k, v in guess.items() if v}

def _export_statement_csv(stmt: dict, filename="statement.csv"):
    rows = _flatten_statement_for_export(stmt)
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=["section", "code", "name", "column", "amount"])
    w.writeheader()
    for r in rows:
        w.writerow(r)

    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# -----------------------------
# helpers
# -----------------------------
def _company_auth_or_403(company_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != int(company_id):
        return None, (jsonify({"error": "Not authorised for this company"}), 403)
    return user, None

def _schema(company_id: int) -> str:
    return f"company_{int(company_id)}"

def _to_iso(d):
    """
    Normalize incoming date to YYYY-MM-DD string or None.
    Accepts:
      - date object
      - ISO string
      - ISO timestamp
    Rejects invalid dates.
    """
    if not d:
        return None

    # Already a date object
    if isinstance(d, date):
        return d.isoformat()

    s = str(d).strip()

    # Trim timestamps
    if len(s) >= 10:
        s = s[:10]

    # Validate format
    try:
        return date.fromisoformat(s).isoformat()
    except ValueError:
        return None


def _money(x) -> float:
    return float(Decimal(str(x or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

def _to_num(v, default=0.0):
    try:
        if v is None: return default
        if isinstance(v, str): v = v.replace(",", "").strip()
        return float(v)
    except Exception:
        return default

def _norm_str(x) -> str:
    return (str(x).strip() if x is not None else "")

def _to_bool(v, default=False):
    if v is None: return default
    s = str(v).strip().lower()
    return s in {"1","true","yes","y","on"}

def format_address(addr: dict | None) -> str | None:
    """
    Convert structured address object into a clean multi-line string.
    Expected keys: line1, line2, locality, city, region, postalCode, country
    """
    if not isinstance(addr, dict):
        return None

    def clean(x):
        if x is None:
            return None
        s = str(x).strip()
        return s if s else None

    parts = [
        clean(addr.get("line1")),
        clean(addr.get("line2")),
        clean(addr.get("locality")),
        clean(addr.get("city")),
        clean(addr.get("region")),
        clean(addr.get("postalCode")),
        clean(addr.get("country")),
    ]
    parts = [p for p in parts if p]
    return "\n".join(parts) if parts else None

INVENTORY_ITEM_COLS = {
  "sku","name","barcode","category","unit","vat_code",
  "sales_price","purchase_cost","reorder_level",
  "track_stock","is_taxable","is_active","valuation_method",
  "inventory_account","income_account","cogs_account",
  "meta"
}

def _clean_meta(v):
    # Only allow dict meta; otherwise empty
    return v if isinstance(v, dict) else {}

def _split_item_payload(payload: dict, allowed_cols: set):
    """
    Returns (cols_payload, meta_payload)
    - Keeps only allowed columns in cols_payload
    - Moves everything else into meta (plus merges with payload['meta'] if provided)
    """
    payload = payload or {}
    incoming_meta = _clean_meta(payload.get("meta"))

    cols = {}
    extra = {}

    for k, v in payload.items():
        if k == "meta":
            continue
        if k in allowed_cols:
            cols[k] = v
        else:
            extra[k] = v

    # merge: explicit payload.meta wins over extras if same key
    meta = {**extra, **incoming_meta}
    return cols, meta

def _merge_meta(existing_meta, new_meta: dict):
    base = existing_meta if isinstance(existing_meta, dict) else {}
    add = new_meta if isinstance(new_meta, dict) else {}
    return {**base, **add}

# ------------------------------------------------------------
# ✅ Helpers (paste once in api_server.py near your endpoints)
# ------------------------------------------------------------

from datetime import datetime, timezone, timedelta, date

def _as_utc_aware(dt):
    """Return dt as timezone-aware UTC datetime (or None)."""
    if not dt:
        return None

    if isinstance(dt, str):
        # Support common DB-returned ISO strings.
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return None

    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)

    return dt.astimezone(timezone.utc)


def _invite_is_expired(rec):
    expires_at = _as_utc_aware(rec.get("expires_at"))
    if not expires_at:
        return False
    return datetime.now(timezone.utc) > expires_at


def _invite_state_error(rec):
    """
    Return (message, status_code) if invite is unusable, else None.
    """
    if not rec:
        return ("Invalid or expired invitation link.", 404)

    if rec.get("revoked_at"):
        return ("This invitation link was revoked.", 400)

    if rec.get("accepted_at") or rec.get("accepted"):
        return ("This invitation link has already been used.", 400)

    if _invite_is_expired(rec):
        return ("This invitation link has expired.", 400)

    return None

def parse_date_arg(request, name: str) -> Optional[date]:
    s = request.args.get(name)
    if not s:
        return None
    return date.fromisoformat(s[:10])

def _get_company_fin_year_start(company_id: int) -> Optional[date]:
    # uses your company context helper (must exist in your codebase)
    ctx = get_company_context(db_service, company_id) or {}
    return parse_date_maybe(ctx.get("fin_year_start"))


def _normalize_range(company_id: int, date_from: Optional[date], date_to: Optional[date]) -> Tuple[Optional[date], Optional[date]]:
    """
    If preset exists: compute range using company fin_year_start via resolve_period().
    If no preset: keep provided from/to.
    """
    preset = (request.args.get("preset") or "").strip().lower()
    if not preset:
        return date_from, date_to

    fy = _get_company_fin_year_start(company_id)

    pr = resolve_period(
        fin_year_start=fy,
        preset=preset,
        date_from=date_from,
        date_to=date_to,
        as_of=parse_date_arg(request, "as_of") or parse_date_arg(request, "to"),
    )

    return pr.get("from"), pr.get("to")



def _normalize_as_of(company_id: int, as_of: Optional[date]) -> Optional[date]:
    """
    For Balance Sheet: if preset exists, as_of becomes the END of the resolved period.
    If as_of provided already and no preset, keep it.
    """
    preset = (request.args.get("preset") or "").strip().lower()
    if not preset:
        return as_of

    fy = _get_company_fin_year_start(company_id)
    pr = resolve_period(
        fin_year_start=fy,
        preset=preset,
        date_from=None,
        date_to=None,
        as_of=as_of or parse_date_arg(request, "to"),
    )
    return pr.get("to") or as_of

def _parse_iso_date(s):
    if not s:
        return None
    if isinstance(s, date):
        return s
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None

def validate_invoice_dates(invoice_date_raw, due_date_raw):
    inv_date = _parse_iso_date(invoice_date_raw)
    due_date = _parse_iso_date(due_date_raw)

    if not inv_date:
        return False, "Invoice date is required and must be YYYY-MM-DD."

    if inv_date > date.today():
        return False, "Invoice date cannot be in the future."

    if due_date and due_date < inv_date:
        return False, "Due date cannot be earlier than invoice date."

    return True, None

# ────────────────────────────────────────────────────────────────
# Helper: industry flags
# ────────────────────────────────────────────────────────────────
def _get_industry_flags(company_id: int) -> Dict:
    """
    Get industry-specific flags for COA/reporting logic.
    Returns a dict with 'industry' and 'uses_cogs' keys.
    """
    try:
        company = db_service.fetch_one(
            "SELECT industry FROM public.companies WHERE id=%s LIMIT 1;",
            (company_id,),
        )
    except Exception as e:
        current_app.logger.warning(f"Error loading company {company_id}: {e}")
        company = None

    if not company:
        return {"industry": "Unknown", "uses_cogs": False}

    industry = company.get("industry") or "Unknown"

    # Determine if this industry uses COGS
    uses_cogs = industry.lower() not in [
        "npo education",
        "public school",
        "npo health",
        "npo other",
    ] and not industry.startswith("NPO ")

    return {
        "industry": industry,
        "uses_cogs": uses_cogs,
    }

def preview_csv_bytes(b: bytes, max_rows: int = 20) -> dict:
    text = b.decode("utf-8", errors="replace")
    sio = io.StringIO(text)
    reader = csv.reader(sio)
    rows = []
    for i, r in enumerate(reader):
        rows.append(r)
        if i >= max_rows:
            break
    if not rows:
        return {"columns": [], "rows": []}

    columns = rows[0]
    data_rows = rows[1:]
    # normalize row lengths
    data = []
    for r in data_rows:
        obj = {}
        for j, c in enumerate(columns):
            obj[c] = r[j] if j < len(r) else ""
        data.append(obj)
    return {"columns": columns, "rows": data}

def preview_xlsx_bytes(b: bytes, max_rows: int = 20) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return {"columns": [], "rows": []}
    columns = [str(x).strip() if x is not None else "" for x in header]
    columns = [c if c else f"col_{i+1}" for i, c in enumerate(columns)]

    data = []
    for i, r in enumerate(rows_iter):
        if i >= max_rows:
            break
        obj = {}
        for j, c in enumerate(columns):
            v = r[j] if j < len(r) else None
            # keep raw values; front-end can display
            obj[c] = v
        data.append(obj)

    return {"columns": columns, "rows": data}

def pick_col(row: dict, key_or_list):
    if not key_or_list:
        return None
    if isinstance(key_or_list, str):
        return row.get(key_or_list)
    if isinstance(key_or_list, list):
        for k in key_or_list:
            if k in row and row.get(k) not in (None, ""):
                return row.get(k)
    return None

def to_float_safe(v, default=0.0):
    try:
        if v is None:
            return default
        if isinstance(v, str):
            v = v.replace(",", "").strip()
        return float(v)
    except Exception:
        return default

def to_date_safe(v):
    # accept YYYY-MM-DD or excel dates if openpyxl provided datetime/date
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v or "").strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except Exception:
        return None
        
def _ensure_vat_accounts_in_rows(company_id: int, rows: list) -> list:
    """
    Make sure VAT Input (1410) and VAT Output (2310) exist
    in the company's COA. If missing, insert them into the DB
    and append to `rows`. Safe to call multiple times.
    """
    needed = {
        "1410": ('VAT Input',  '1410', 'Asset',     'Current Assets',
                 'Input VAT on purchases – recoverable from Tax Authority', None),
        "2310": ('VAT Output', '2310', 'Liability', 'Current Liabilities',
                 'Output VAT on sales – payable to Tax Authority', None),
    }

    present_codes = set()

    for r in rows:
        if isinstance(r, dict):
            code = str(r.get("code") or r.get("account_code") or "")
        else:
            # tuple: (name, code, category, reporting_group, ...)
            try:
                _, code, *_ = r
                code = str(code or "")
            except Exception:
                code = ""
        if code:
            present_codes.add(code)

    rows_to_insert = [tpl for code, tpl in needed.items() if code not in present_codes]

    if rows_to_insert:
        try:
            # Persist into company_{id}.chart_of_accounts
            db_service.insert_coa(company_id, rows_to_insert)
        except Exception as e:
            current_app.logger.exception(
                "Error inserting VAT accounts for company %s: %s", company_id, e
            )

        rows = rows + rows_to_insert

    return rows

def _first_contact_email(inv: dict) -> str | None:
    # supports either `customer_contacts` on invoice payload OR `contacts`
    contacts = inv.get("customer_contacts") or inv.get("contacts") or []
    if isinstance(contacts, str):
        # if accidentally returned as JSON string
        try:
            import json
            contacts = json.loads(contacts)
        except Exception:
            contacts = []

    if isinstance(contacts, list) and contacts:
        for c in contacts:
            if isinstance(c, dict):
                e = (c.get("email") or "").strip()
                if e:
                    return e
    return None

def record_invoice_revenue_billing_and_allocation(
    *,
    company_id: int,
    invoice_id: int,
    inv: dict,
    journal_id: int,
    user_id: int,
):
    revenue_contract_id = inv.get("revenue_contract_id")
    if not revenue_contract_id:
        return None

    contract = db_service.get_revenue_contract(
        company_id=company_id,
        contract_id=int(revenue_contract_id),
    ) or {}

    payload_json = contract.get("payload_json") or {}
    if isinstance(payload_json, str):
        import json
        try:
            payload_json = json.loads(payload_json)
        except Exception:
            payload_json = {}

    settlement_pattern = (
        payload_json.get("settlement_pattern")
        or payload_json.get("ifrs15_settlement_pattern")
        or ""
    ).strip().lower()

    obligation_ids = [
        l.get("revenue_obligation_id")
        for l in inv.get("lines", [])
        if l.get("revenue_obligation_id")
    ]

    if obligation_ids:
        billing_amount = float(sum(
            float(l.get("net_amount") or 0.0)
            for l in inv.get("lines", [])
            if l.get("revenue_obligation_id")
        ))
        billing_obligation_id = (
            int(obligation_ids[0])
            if len(set(obligation_ids)) == 1
            else None
        )
    else:
        billing_amount = float(sum(
            float(l.get("net_amount") or 0.0)
            for l in inv.get("lines", [])
        ))
        billing_obligation_id = None

    db_service.record_revenue_billing_event(
        company_id=company_id,
        contract_id=int(revenue_contract_id),
        data={
            "event_date": inv.get("invoice_date"),
            "event_type": "invoice",
            "source_invoice_id": int(invoice_id),
            "obligation_id": billing_obligation_id,
            "amount": billing_amount,
            "currency": inv.get("currency") or "USD",
            "notes": f"From AR invoice {inv.get('number') or invoice_id}",
            "payload_json": {
                "customer_id": int(inv.get("customer_id") or 0),
                "source": "ar_invoice_post",
                "invoice_number": inv.get("number"),
                "auto_allocate": True,
                "settlement_pattern": settlement_pattern,
                "journal_id": int(journal_id),
            },
        },
        user_id=int(user_id or 0),
    )

    if settlement_pattern == "cash_before_service":
        allocation = db_service.auto_allocate_customer_advance_to_invoice(
            company_id=company_id,
            customer_id=int(inv.get("customer_id") or 0),
            contract_id=int(revenue_contract_id),
            invoice_id=int(invoice_id),
            amount=float(billing_amount),
            currency=inv.get("currency") or "USD",
        )

        current_app.logger.info("Auto allocation result: %s", allocation)

        if allocation.get("remaining", 0) > 0:
            current_app.logger.warning(
                "Invoice not fully covered by advances. Remaining: %s",
                allocation["remaining"],
            )

    return True

def _first_contact_email_from_contacts(contacts):
    """
    contacts can be:
    - None
    - list of dicts: [{"name":..,"email":..}, ...]
    - dict with key "contacts": [...]
    - any other weird shape
    """
    if not contacts:
        return None

    # Normalize shapes
    items = None
    if isinstance(contacts, list):
        items = contacts
    elif isinstance(contacts, dict):
        # common patterns: {"contacts":[...]} or {"primary":{...}} etc
        if isinstance(contacts.get("contacts"), list):
            items = contacts["contacts"]
        else:
            # if it's a dict of dicts, try values
            items = list(contacts.values())
    else:
        return None

    for c in items or []:
        if not isinstance(c, dict):
            continue
        email = (c.get("email") or c.get("Email") or c.get("mail") or "").strip()
        if email and EMAIL_RE.match(email):
            return email

    return None

def _days_in_month(year: int, month: int) -> int:
    if month == 12:
        return (date(year + 1, 1, 1) - date(year, 12, 1)).days
    return (date(year, month + 1, 1) - date(year, month, 1)).days

# ------------------------------------------------------------
# ✅ helper: safely compare expires_at (handles aware/naive/str)
# ------------------------------------------------------------
def _is_expired_expires_at(expires_at) -> bool:
    """
    Returns True if expires_at is in the past (UTC), safely handling:
    - datetime aware (tzinfo set)
    - datetime naive (assumed UTC)
    - ISO string (e.g. "2026-02-21T12:00:00Z" / "+00:00")
    """
    if not expires_at:
        return False

    exp = expires_at

    # If DB returns a string timestamp
    if isinstance(exp, str):
        # support "Z" suffix
        exp = datetime.fromisoformat(exp.replace("Z", "+00:00"))

    # If DB returns naive datetime, assume it's UTC
    if isinstance(exp, datetime) and exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)

    now = datetime.now(timezone.utc)

    # If exp is still not a datetime for some reason, treat as expired
    if not isinstance(exp, datetime):
        return True

    return now > exp

def _make_vat_periods_for_year(year: int, cfg: dict):
    freq = (cfg.get("frequency") or "bi_monthly").lower()
    anchor_month = cfg.get("anchor_month") or 1
    anchor_month = max(1, min(12, anchor_month))

    if freq == "monthly":
        step = 1
    elif freq == "quarterly":
        step = 3
    elif freq in ("semi_annual", "semi-annual", "half_year"):
        step = 6
    elif freq == "annual":
        step = 12
    else:  # bi-monthly
        step = 2

    periods = []
    anchor0 = anchor_month - 1
    max_loops = (12 // step) + 2

    for k in range(max_loops):
        s_month_idx = anchor0 + k * step
        s_year = year + (s_month_idx // 12)
        s_month = (s_month_idx % 12) + 1
        start_date = date(s_year, s_month, 1)

        e_month_idx = s_month_idx + step - 1
        e_year = year + (e_month_idx // 12)
        e_month = (e_month_idx % 12) + 1
        end_date = date(e_year, e_month, _days_in_month(e_year, e_month))

        if start_date.year == year or end_date.year == year:
            label_start = start_date.strftime("%b")
            label_end = end_date.strftime("%b")

            if start_date.year == end_date.year:
                if step == 1:
                    label = f"{label_start} {end_date.year}"
                else:
                    label = f"{label_start}–{label_end} {end_date.year}"
            else:
                label = f"{label_start} {start_date.year}–{label_end} {end_date.year}"

            periods.append({
                "label": label,
                "start_date": start_date,
                "end_date": end_date,
            })
    return periods

def _assign_vat_period_label(tx_date: date, cfg: dict) -> str:
    year = tx_date.year
    periods = (
        _make_vat_periods_for_year(year - 1, cfg)
        + _make_vat_periods_for_year(year, cfg)
        + _make_vat_periods_for_year(year + 1, cfg)
    )
    for p in periods:
        if p["start_date"] <= tx_date <= p["end_date"]:
            return p["label"]
    return tx_date.strftime("%b %Y")

def build_user_allowed_companies(user_id: int):
    rows = db_service.fetch_all("""
        SELECT
            c.id,
            c.name,
            c.system_company_code,
            c.industry,
            c.sub_industry,
            cu.role,
            cu.access_scope,
            cu.membership_kind,
            cu.is_primary,
            rel.parent_company_id,
            rel.relationship_type,
            rel.ownership_percent,
            rel.consolidation_method,
            CASE
                WHEN cu.is_primary = TRUE THEN 'primary'
                WHEN rel.relationship_type IS NOT NULL THEN rel.relationship_type
                ELSE 'member'
            END AS company_context
        FROM public.company_users cu
        JOIN public.companies c
          ON c.id = cu.company_id
        LEFT JOIN public.company_relationships rel
          ON rel.child_company_id = c.id
         AND rel.is_active = TRUE
        WHERE cu.user_id = %s
          AND cu.is_active = TRUE
        ORDER BY cu.is_primary DESC, c.name ASC;
    """, (int(user_id),)) or []

    allowed_company_ids = [
        int(r["id"]) for r in rows if r.get("id") is not None
    ]

    primary_company_id = None
    for r in rows:
        if bool(r.get("is_primary")):
            primary_company_id = int(r["id"])
            break

    if not primary_company_id and rows:
        primary_company_id = int(rows[0]["id"])

    return rows, sorted(set(allowed_company_ids)), primary_company_id


def _require_company_email_admin(company_id:int):
    user=getattr(g,"current_user",{}) or {}
    user_id=int(user.get("id") or user.get("user_id") or 0)

    if not user_id:
        return jsonify({"error":"Not authenticated."}),401

    membership=db_service.fetch_one(
        """
        SELECT role,is_active
        FROM public.company_users
        WHERE company_id=%s
          AND user_id=%s
          AND is_active=TRUE
        LIMIT 1;
        """,
        (int(company_id),user_id),
    )

    if not membership:
        return jsonify({"error":"You do not have access to this company."}),403

    role=(membership.get("role") or "").strip().lower()

    if role not in {
        "owner",
        "admin",
        "cfo",
        "manager",
        "senior",
    }:
        return jsonify({
            "error":"You are not authorised to manage company email settings."
        }),403

    return None


def _sse(event_type: str, data: dict) -> str:
    """Format one Server-Sent Event line."""
    return f"event: {event_type}\ndata: {_json.dumps(data, default=str)}\n\n"


# ---- Step definitions (order matters for frontend display) ----
SIGNUP_STEPS = [
    ("validating",        "Validating your information..."),
    ("creating_user",     "Creating your account..."),
    ("validating_company", "Validating company details..."),
    ("matching_industry",   "Looking up industry profile..."),
    ("creating_company",    "Registering your company..."),
    ("seeding_accounts",    "Setting up Chart of Accounts..."),
    ("finalizing",          "Finalizing setup..."),
    ("sending_email",       "Sending confirmation email..."),
]


def _emit_step(gen, step_key: str):
    """Yield a progress event for a known step, or a generic one."""
    label = dict(SIGNUP_STEPS).get(step_key, step_key.replace("_", " ").title() + "...")
    yield _sse("progress", {"step": step_key, "label": label})

def generate():
    while True:
        yield ": keep-alive\n\n"  # SSE comment line as heartbeat
        time.sleep(30)
        

@app.route("/api/auth/signup", methods=["POST"])
def api_auth_signup(cur=None, conn=None):
    data = request.get_json(silent=True) or {}

    email = (data.get("email") or "").strip().lower()
    first_name = (data.get("firstName") or "").strip()
    last_name = (data.get("lastName") or "").strip()
    user_role = normalize_role(data.get("userRole"))

    current_app.logger.info(
        "SIGNUP DEBUG email=%r first_name=%r last_name=%r role=%r",
        email, first_name, last_name, user_role
    )

    def generate():
        owner_id = None
        created_company_id = None
        invite_payload = None

        try:
            # -----------------------------------------------------
            # Basic validation
            # -----------------------------------------------------
            yield from _emit_step(generate, "validating")

            legal_consent = data.get("legalConsent") or {}
            policy_version = str(
                legal_consent.get("policyVersion") or "1.0"
            ).strip()

            if not legal_consent.get("accepted"):
                yield _sse("error", {
                    "status": 400,
                    "error": (
                        "You must accept the Terms of Use and "
                        "Privacy Policy before registration."
                    ),
                })
                return

            owner_invite_email = (
                (data.get("ownerInvite") or "").strip().lower() or None
            )

            # ====================================================
            # 2. USER DETAILS
            # ====================================================
            first_name = (data.get("firstName") or "").strip()
            last_name = (data.get("lastName") or "").strip()
            email = (data.get("email") or "").strip()
            password = data.get("password") or ""
            user_role = normalize_role(data.get("userRole"))

            password = data.get("password") or ""

            user_type = {
                "practitioner": "Practitioner",
                "school": "School",
            }.get(
                (data.get("userType") or "Enterprise").strip().lower(),
                "Enterprise",
            )

            if not all([
                email,
                password,
                first_name,
                last_name,
                user_role,
            ]):
                yield _sse("error", {
                    "status": 400,
                    "error": "Name, Role, Email, and Password are required",
                })
                return

            # -----------------------------------------------------
            # Confirmation token
            # -----------------------------------------------------
            confirm_token = secrets.token_urlsafe(32)
            confirm_expires_at = (
                datetime.now(timezone.utc) + timedelta(hours=48)
            )

            # -----------------------------------------------------
            # Company input
            # -----------------------------------------------------
            company_payload = data.get("company") or {}

            company_name = (
                company_payload.get("companyName") or ""
            ).strip()

            organization_type = (
                company_payload.get("organizationType")
                or company_payload.get("organisationType")
                or ""
            ).strip().lower()

            requires_company = bool(company_name)

            industry = None
            sub_industry = None
            industry_slug = None
            sub_industry_slug = None

            # -----------------------------------------------------
            # Company validation / industry matching
            # -----------------------------------------------------
            if requires_company:
                yield from _emit_step(generate, "validating_company")

                allowed_org_types = {
                    "private_company",
                    "public_company",
                    "sole_trader",
                    "partnership",
                    "ngo",
                    "npo",
                    "trust",
                    "cooperative",
                    "body_corporate",
                    "club_association",
                    "government_entity",
                    "public_school",
                    "independent_school",
                    "ecd_centre",
                    "other",
                }

                if not organization_type:
                    yield _sse("error", {
                        "status": 400,
                        "error": (
                            "Organisation type is required "
                            "to create your company."
                        ),
                    })
                    return

                if organization_type not in allowed_org_types:
                    yield _sse("error", {
                        "status": 400,
                        "error": "Invalid organisation type selected.",
                        "errors": {
                            "organizationType": (
                                "Please choose a valid organisation type."
                            )
                        },
                    })
                    return

                industry_raw = (
                    company_payload.get("industry") or ""
                ).strip()

                if not industry_raw:
                    yield _sse("error", {
                        "status": 400,
                        "error": (
                            "Industry is required to create your company."
                        ),
                    })
                    return

                sub_industry_raw = (
                    company_payload.get("subIndustry") or ""
                ).strip() or None

                yield from _emit_step(generate, "matching_industry")

                if not get_industry_template(industry_raw):
                    yield _sse("error", {
                        "status": 400,
                        "error": f"Industry '{industry_raw}' not recognized.",
                    })
                    return

                if (
                    sub_industry_raw
                    and not get_industry_template(
                        industry_raw,
                        sub_industry_raw,
                    )
                ):
                    yield _sse("error", {
                        "status": 400,
                        "error": (
                            f"Sub-industry '{sub_industry_raw}' invalid "
                            f"for industry '{industry_raw}'."
                        ),
                    })
                    return

                (
                    industry,
                    sub_industry,
                    industry_slug,
                    sub_industry_slug,
                ) = normalize_industry_pair(
                    industry_raw,
                    sub_industry_raw,
                )

                ind_template = TEMPLATE_INDUSTRY_ALIASES.get(
                    (industry or "").strip().lower(),
                    (industry or "").strip(),
                )

                sub_key = canonical_subindustry_key(
                    ind_template,
                    sub_industry,
                )

                if sub_key:
                    sub_industry = sub_key
                    sub_industry_slug = slugify(sub_industry)

            # -----------------------------------------------------
            # Actual signup DB work
            # -----------------------------------------------------
            def perform_signup(tx_conn, tx_cur):
                nonlocal owner_id, created_company_id

                # User
                yield from _emit_step(generate, "creating_user")

                try:
                    owner_id = db_service.insert_user(
                        email=email,
                        password_hash=generate_password_hash(password),
                        user_type=user_type,
                        first_name=first_name,
                        last_name=last_name,
                        user_role=user_role,
                        is_confirmed=False,
                        confirmation_token=confirm_token,
                        confirmation_token_expires_at=confirm_expires_at,
                        trial_start_date=str(date.today()),
                        trial_end_date=str(date.today() + timedelta(days=30)),
                        company_id=None,
                        cur=tx_cur,
                        conn=tx_conn,
                    )

                except UniqueViolation as e:
                    constraint = getattr(
                        getattr(e, "diag", None),
                        "constraint_name",
                        None,
                    )

                    detail = getattr(
                        getattr(e, "diag", None),
                        "message_detail",
                        None,
                    )

                    current_app.logger.warning(
                        "Signup unique violation email=%s constraint=%s detail=%s",
                        email,
                        constraint,
                        detail,
                    )

                    if constraint == "users_email_key":
                        yield _sse("error", {
                            "status": 409,
                            "error": (
                                "A user with this email already exists. "
                                "Please sign in."
                            ),
                            "constraint": constraint,
                            "detail": detail,
                        })
                    else:
                        yield _sse("error", {
                            "status": 409,
                            "error": "Registration failed: duplicate data.",
                            "constraint": constraint,
                            "detail": detail,
                        })

                    return

                if not owner_id:
                    raise RuntimeError("insert_user returned no id")

                # -------------------------------------------------
                # Company
                # -------------------------------------------------
                if not requires_company:
                    return None

                yield from _emit_step(generate, "creating_company")

                country = (
                    company_payload.get("country") or ""
                ).upper()

                company_reg_no = (
                    company_payload.get("companyRegNo")
                    or company_payload.get("company_reg_no")
                    or company_payload.get("regNo")
                    or company_payload.get("registrationNumber")
                )
                company_reg_no = (company_reg_no or "").strip() or None

                tin = company_payload.get("tin")
                vat = company_payload.get("vat")

                company_email = (
                    company_payload.get("companyEmail") or email
                )

                # Company validation
                validation_payload = {
                    "country": country,
                    "companyRegNo": company_reg_no,
                    "organizationType": organization_type,
                    "tin": tin,
                    "vat": vat,
                    "companyEmail": company_email,
                }

                ok, errors = validate_company_payload(
                    validation_payload
                )

                if not ok:
                    raise ValueError(json.dumps({
                        "type": "company_validation",
                        "errors": errors,
                    }))

                # Currency / financial year
                currency = (
                    company_payload.get("currency")
                    or get_currency_for_country(country)
                    or "USD"
                )

                fin_year_start = (
                    company_payload.get("finYearStart") or "01/01"
                )

                # Registration date
                company_reg_raw = (
                    company_payload.get("companyRegDate") or ""
                ).strip() or None

                reg_date = None

                if company_reg_raw:
                    try:
                        reg_date = date.fromisoformat(company_reg_raw)
                    except ValueError:
                        raise ValueError(json.dumps({
                            "type": "company_validation",
                            "errors": {
                                "companyRegDate": "Must be YYYY-MM-DD"
                            },
                        }))

                    if reg_date > date.today():
                        raise ValueError(json.dumps({
                            "type": "company_validation",
                            "errors": {
                                "companyRegDate": "Cannot be in future"
                            },
                        }))

                # Addresses
                reg_obj = company_payload.get("registeredAddress")
                post_obj = company_payload.get("postalAddress")

                if (
                    company_payload.get("postalSameAsReg")
                    and isinstance(reg_obj, dict)
                    and not isinstance(post_obj, dict)
                ):
                    post_obj = reg_obj

                physical_address = (
                    format_address(reg_obj)
                    if isinstance(reg_obj, dict)
                    else None
                )

                postal_address = (
                    format_address(post_obj)
                    if isinstance(post_obj, dict)
                    else None
                )

                place_id = None
                lat = None
                lng = None

                if isinstance(reg_obj, dict):
                    place_id = reg_obj.get("placeId")
                    lat = reg_obj.get("lat")
                    lng = reg_obj.get("lng")

                # Company contact
                company_phone = (
                    company_payload.get("company_phone")
                    or company_payload.get("companyPhone")
                    or data.get("phone")
                    or None
                )

                logo_url = (
                    company_payload.get("logo_url")
                    or company_payload.get("logoUrl")
                    or None
                )

                # Industry profile
                profile = get_industry_profile(
                    industry,
                    sub_industry,
                ) or {}

                inventory_mode = (
                    company_payload.get("inventory_mode")
                    or profile.get("default_inventory_mode")
                    or "none"
                )

                inventory_valuation = (
                    company_payload.get("inventory_valuation")
                    or profile.get("default_valuation")
                )

                # Insert company
                created_company_id = db_service.insert_company(
                    name=company_name or "Company",
                    organization_type=organization_type,
                    client_code=(
                        company_payload.get("clientCode")
                        or f"C{int(time.time())}"
                    ),
                    industry=industry,
                    sub_industry=sub_industry,
                    currency=currency,
                    fin_year_start=fin_year_start,
                    company_reg_date=reg_date,
                    country=country,
                    company_reg_no=company_reg_no,
                    tin=tin,
                    vat=vat,
                    company_email=company_email,
                    owner_user_id=owner_id,
                    inventory_mode=inventory_mode,
                    inventory_valuation=inventory_valuation,
                    physical_address=physical_address,
                    postal_address=postal_address,
                    company_phone=company_phone,
                    logo_url=logo_url,
                    registered_address_json=(
                        reg_obj if isinstance(reg_obj, dict) else None
                    ),
                    postal_address_json=(
                        post_obj if isinstance(post_obj, dict) else None
                    ),
                    address_place_id=place_id,
                    address_lat=str(lat) if lat is not None else None,
                    address_lng=str(lng) if lng is not None else None,
                    industry_slug=industry_slug,
                    sub_industry_slug=sub_industry_slug,
                    cur=tx_cur,
                    conn=tx_conn,
                )

                if not created_company_id:
                    raise RuntimeError(
                        "insert_company returned no id"
                    )

                # Company schema
                db_service.initialize_company_schema(
                    created_company_id,
                    cur=tx_cur,
                    conn=tx_conn,
                )

                # Guarantee slugs
                tx_cur.execute(
                    """
                    UPDATE public.companies
                    SET industry_slug = %s,
                        sub_industry_slug = %s
                    WHERE id = %s
                    """,
                    (
                        industry_slug,
                        sub_industry_slug,
                        created_company_id,
                    ),
                )

                # Bind user to company
                db_service.update_user(
                    owner_id,
                    company_id=created_company_id,
                    cur=tx_cur,
                    conn=tx_conn,
                )

                # Owner membership
                tx_cur.execute(
                    """
                    INSERT INTO public.company_users
                    (
                        company_id,
                        user_id,
                        role,
                        access_scope,
                        membership_kind,
                        is_primary,
                        is_active,
                        joined_at
                    )
                    VALUES (%s, %s, 'owner', 'core', 'primary',
                            TRUE, TRUE, NOW())
                    ON CONFLICT (company_id, user_id)
                    DO UPDATE SET
                        role = 'owner',
                        access_scope = 'core',
                        membership_kind = 'primary',
                        is_primary = TRUE,
                        is_active = TRUE
                    """,
                    (
                        created_company_id,
                        owner_id,
                    ),
                )

                # User policy fields
                tx_cur.execute(
                    """
                    UPDATE public.users
                    SET accepted_terms_at = NOW(),
                        accepted_privacy_at = NOW(),
                        accepted_popia_at = NOW(),
                        accepted_policy_version = %s
                    WHERE id = %s
                    """,
                    (
                        policy_version,
                        owner_id,
                    ),
                )

                # Policy records
                user_ip = request.headers.get(
                    "X-Forwarded-For",
                    request.remote_addr,
                )
                user_agent = request.headers.get("User-Agent")

                for policy_type in (
                    "terms_of_use",
                    "privacy_policy",
                    "popia_notice",
                ):
                    db_service.insert_policy_acceptance(
                        owner_id,
                        created_company_id,
                        policy_type,
                        policy_version,
                        user_ip,
                        user_agent,
                        cur=tx_cur,
                        conn=tx_conn,
                    )

                # COA
                yield from _emit_step(generate, "seeding_accounts")

                company_info = db_service.fetch_one(
                    """
                    SELECT industry_slug,
                           sub_industry_slug,
                           industry,
                           sub_industry
                    FROM public.companies
                    WHERE id = %s
                    """,
                    (created_company_id,),
                    cur=tx_cur,
                ) or {}

                ind_slug = (
                    (company_info.get("industry_slug") or "").strip()
                    or slugify(company_info.get("industry"))
                    or ""
                )

                sub_slug = (
                    (company_info.get("sub_industry_slug") or "").strip()
                    or slugify(company_info.get("sub_industry"))
                    or ""
                )

                seed_company_coa_once(
                    db_service,
                    company_id=created_company_id,
                    industry=ind_slug,
                    sub_industry=sub_slug,
                    source="pool",
                    cur=tx_cur,
                    conn=tx_conn,
                )

                # Branding
                db_service.upsert_company_branding(
                    created_company_id,
                    {
                        "logo_url": logo_url,
                        "contact_phone": company_phone,
                        "contact_email": company_email,
                        "address": physical_address or postal_address,
                        "vat_no": vat,
                        "website": company_payload.get("website") or None,
                    },
                    cur=tx_cur,
                    conn=tx_conn,
                )

                # Owner invitation
                if (
                    owner_invite_email
                    and (user_role or "").strip().lower() != "owner"
                ):
                    token = _create_invite_token(
                        email=owner_invite_email,
                        role="owner",
                        company_name=(
                            company_name or "FinSage workspace"
                        ),
                        company_id=created_company_id,
                    )

                    invite_payload = {
                        "to_email": owner_invite_email,
                        "company_name": (
                            company_name or "FinSage workspace"
                        ),
                        "invite_link": (
                            f"{FRONTEND_BASE}"
                            f"/accept-invite.html?token={token}"
                        ),
                    }

                return invite_payload

            # -----------------------------------------------------
            # Transaction handling
            # -----------------------------------------------------
            if cur is not None:
                # Caller supplied cursor.
                # Prefer its connection if conn wasn't supplied.
                tx_conn = conn or getattr(cur, "connection", None)

                if tx_conn is None:
                    raise RuntimeError(
                        "Signup cursor has no database connection."
                    )

                invite_payload = yield from perform_signup(
                    tx_conn,
                    cur,
                )

            elif conn is not None:
                with conn.cursor(
                    cursor_factory=psycopg2.extras.RealDictCursor
                ) as tx_cur:
                    invite_payload = yield from perform_signup(
                        conn,
                        tx_cur,
                    )
                    conn.commit()

            else:
                with db_service.transaction() as (tx_conn, tx_cur):
                    invite_payload = yield from perform_signup(
                        tx_conn,
                        tx_cur,
                    )

            # -----------------------------------------------------
            # Owner invitation email
            # -----------------------------------------------------
            if invite_payload:
                company_label = invite_payload["company_name"]
                invite_link = invite_payload["invite_link"]

                try:
                    send_mail(
                        to_email=invite_payload["to_email"],
                        subject=(
                            "You've been invited as Owner of "
                            f"{company_label} on FinSage"
                        ),
                        html_body=f"""
                        <div style="font-family:system-ui,Segoe UI,Arial;line-height:1.5">
                            <h2>Owner Invitation: {company_label}</h2>
                            <p>
                                You have been invited to join
                                <b>{company_label}</b> as the <b>Owner</b>.
                            </p>
                            <p>
                                <a href="{invite_link}"
                                   style="
                                       background:#00C8C8;
                                       color:#fff;
                                       padding:10px 16px;
                                       border-radius:6px;
                                       text-decoration:none;
                                       display:inline-block
                                   ">
                                    Accept invite
                                </a>
                            </p>
                            <p>Copy this link if the button doesn't work:</p>
                            <p><code>{invite_link}</code></p>
                        </div>
                        """,
                        text_body=(
                            f"Owner invite for {company_label}. "
                            f"Accept: {invite_link}"
                        ),
                    )
                except Exception:
                    current_app.logger.exception(
                        "Owner invitation email failed for company=%s",
                        created_company_id,
                    )

            # -----------------------------------------------------
            # Confirmation email
            # -----------------------------------------------------
            yield from _emit_step(generate, "sending_email")

            confirm_link = (
                f"{FRONTEND_BASE}/confirm.html?token={confirm_token}"
            )

            email_sent = True

            try:
                send_mail(
                    to_email=email,
                    subject="Confirm your FinSage account",
                    html_body=(
                        f"<p>Welcome {first_name}, "
                        f"please confirm your account: "
                        f"<a href='{confirm_link}'>Confirm</a></p>"
                    ),
                    text_body=(
                        f"Welcome {first_name}, "
                        f"confirm your account: {confirm_link}"
                    ),
                )

                current_app.logger.info(
                    "[AUTH] Email sent to %s",
                    email,
                )

            except Exception:
                email_sent = False
                current_app.logger.warning(
                    "[AUTH] Email send failed for %s",
                    email,
                    exc_info=True,
                )

            # -----------------------------------------------------
            # Success
            # -----------------------------------------------------
            yield from _emit_step(generate, "finalizing")

            yield _sse("result", {
                "status": 201,
                "success": True,
                "message": "Registration successful.",
                "status_text": (
                    "confirmation_pending"
                    if email_sent
                    else "confirmation_email_failed"
                ),
                "user_email": email,
                "owner_id": owner_id,
                "company_id": created_company_id,
                "confirm_expires_hours": 48,
                "email_sent": email_sent,
            })

        except GeneratorExit:
            current_app.logger.info(
                "Signup client disconnected."
            )

        except ValueError as e:
            message = str(e)

            try:
                parsed = json.loads(message)
            except Exception:
                parsed = None

            if (
                isinstance(parsed, dict)
                and parsed.get("type") == "company_validation"
            ):
                yield _sse("error", {
                    "status": 400,
                    "error": "Company validation failed",
                    "errors": parsed.get("errors", {}),
                })
                return

            current_app.logger.exception(
                "Signup validation error"
            )

            yield _sse("error", {
                "status": 400,
                "error": message or "Invalid registration data.",
            })

        except UniqueViolation as e:
            constraint = getattr(
                getattr(e, "diag", None),
                "constraint_name",
                None,
            )
            detail = getattr(
                getattr(e, "diag", None),
                "message_detail",
                None,
            )

            current_app.logger.exception(
                "Signup unique violation email=%s "
                "constraint=%s detail=%s",
                email,
                constraint,
                detail,
            )

            if constraint == "users_email_key":
                yield _sse("error", {
                    "status": 409,
                    "error": (
                        "A user with this email already exists. "
                        "Please sign in."
                    ),
                    "constraint": constraint,
                    "detail": detail,
                })
            else:
                yield _sse("error", {
                    "status": 409,
                    "error": "Registration failed: duplicate data.",
                    "constraint": constraint,
                    "detail": detail,
                })

        except Exception as e:
            current_app.logger.exception(
                "Unexpected signup error: %s",
                e,
            )

            yield _sse("error", {
                "status": 500,
                "error": "Failed to complete registration.",
            })

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Content-Type": "text/event-stream",
            "X-Accel-Buffering": "no",
        },
    )



@app.route("/api/<path:_path>", methods=["OPTIONS"])
def api_options(_path):
    return ("", 204)

@app.route("/api/auth/confirm", methods=["POST"])
def api_auth_confirm():
    data = request.get_json(silent=True) or {}
    token = (data.get("token") or "").strip()

    if not token:
        return jsonify({"error": "Missing confirmation token"}), 400

    user = db_service.get_user_by_confirmation_token(token)
    if not user:
        return jsonify({"error": "Invalid confirmation link"}), 404

    if user.get("is_confirmed"):
        return jsonify({"ok": True, "status": "already_confirmed"}), 200

    exp = user.get("confirmation_token_expires_at")
    if exp and datetime.now(timezone.utc) > exp:
        return jsonify({
            "error": "Confirmation link expired",
            "code": "CONFIRM_EXPIRED"
        }), 410

    db_service.execute_sql(
        """
        UPDATE public.users
        SET is_confirmed = TRUE,
            confirmation_token = NULL,
            confirmation_token_expires_at = NULL
        WHERE id = %s
        """,
        (user["id"],)
    )

    return jsonify({"ok": True, "status": "confirmed"}), 200


@app.route("/api/auth/resend-confirm", methods=["POST"])
def api_auth_resend_confirm():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    if not email:
        return jsonify({"message": "If that email exists, a confirmation email has been sent."}), 200

    user = db_service.fetch_one(
        """
        SELECT id, first_name, is_confirmed
        FROM public.users
        WHERE email = %s
        """,
        (email,)
    )

    # Always return generic response
    if not user or user.get("is_confirmed"):
        return jsonify({"message": "If that email exists, a confirmation email has been sent."}), 200

    new_token = secrets.token_urlsafe(32)
    new_exp = datetime.now(timezone.utc) + timedelta(hours=48)

    db_service.execute_sql(
        """
        UPDATE public.users
        SET confirmation_token = %s,
            confirmation_token_expires_at = %s
        WHERE id = %s
        """,
        (new_token, new_exp, user["id"])
    )

    link = f"{FRONTEND_BASE}/confirm.html?token={new_token}"
    subject = "Your FinSage confirmation link (expires in 48 hours)"
    html = f"""
    <p>Hi {user.get("first_name") or ""},</p>
    <p>Here is your new confirmation link: <a href="{link}">Confirm</a></p>
    <p><b>Note:</b> This link expires in 48 hours.</p>
    """
    text = f"Confirm your account: {link} (expires in 48 hours)"

    try:
        send_mail(to_email=email, subject=subject, html_body=html, text_body=text)
    except Exception:
        current_app.logger.exception("Resend confirm email failed")

    return jsonify({"message": "If that email exists, a confirmation email has been sent."}), 200

@app.route("/api/auth/signin", methods=["POST", "OPTIONS"])
@cross_origin(
    origins=[
        "https://finspheresolutions.com",
        "https://www.finspheresolutions.com",
        "http://127.0.0.1:5500",
        "http://localhost:5500",
    ],
    allow_headers=["Content-Type", "Authorization"],
    methods=["POST", "OPTIONS"],
    supports_credentials=True,
)
def api_auth_signin():
    print("========== SIGNIN HIT ==========", flush=True)
    if request.method == "OPTIONS":
        return "", 204

    try:
        data = request.get_json(silent=True) or {}

        email = str(data.get("email") or "").strip().lower()
        password = str(data.get("password") or "")

        product = str(
            data.get("product")
            or data.get("app")
            or "finsage"
        ).strip().lower()

        is_ops_signin = product in {"ops", "finsage nexus"}

        if not email or not password:
            return jsonify({
                "error": "Email and password are required."
            }), 400

        current_app.logger.info("[AUTH] Signin attempt for %s", email)

        user = db_service.get_user_by_email(email)

        if not user:
            return jsonify({
                "error": "Invalid credentials"
            }), 401

        password_hash = user.get("password_hash")

        if not password_hash:
            current_app.logger.error(
                "[AUTH] User %s has no password hash",
                user.get("id"),
            )
            return jsonify({
                "error": "Invalid credentials"
            }), 401

        if not check_password_hash(password_hash, password):
            return jsonify({
                "error": "Invalid credentials"
            }), 401

        if not user.get("is_confirmed"):
            return jsonify({
                "error": "Please confirm your email before signing in."
            }), 403

        user_id = int(user["id"])
        current_app.logger.info("[AUTH] 1 - user loaded: %s", user_id)

        # Do not let activity tracking prevent signin.
        try:
            db_service.execute_sql(
                """
                UPDATE public.users
                SET last_login_at = NOW(),
                    last_activity_at = NOW()
                WHERE id = %s;
                """,
                (user_id,),
            )
        except Exception:
            current_app.logger.exception(
                "[AUTH] Could not update login activity for user %s",
                user_id,
            )

        base_role = str(
            user.get("user_role") or "viewer"
        ).strip()

        user_type = str(
            user.get("user_type") or "Enterprise"
        ).strip()

        if is_ops_signin:
            # =========================================================
            # FinSage Nexus SIGN-IN
            # Always enter the user's OWN/native company.
            # Never begin FinSage Nexus inside a delegated client workspace.
            # =========================================================
            primary_mem = db_service.fetch_one(
                """
                SELECT
                    cu.company_id,
                    cu.role,
                    cu.access_scope,
                    cu.membership_kind,
                    cu.is_primary,
                    cu.product_access,
                    cu.ops_is_active
                FROM public.company_users cu
                WHERE cu.user_id=%s
                AND cu.is_active=TRUE
                AND cu.is_primary=TRUE
                AND COALESCE(cu.access_scope,'core')='core'
                ORDER BY
                    CASE
                        WHEN cu.membership_kind='primary' THEN 0
                        ELSE 1
                    END,
                    cu.company_id
                LIMIT 1;
                """,
                (user_id,),
            )

            current_app.logger.info(
                "[AUTH] 2 - primary membership: %s",
                primary_mem,
            )


            # Legacy/fallback users may not yet have is_primary correctly set.
            if not primary_mem:
                primary_mem = db_service.fetch_one(
                    """
                    SELECT
                        cu.company_id,
                        cu.role,
                        cu.access_scope,
                        cu.membership_kind,
                        cu.is_primary,
                        cu.product_access,
                        cu.ops_is_active
                    FROM public.company_users cu
                    JOIN public.companies c
                    ON c.id=cu.company_id
                    WHERE cu.user_id=%s
                    AND cu.is_active=TRUE
                    AND COALESCE(cu.access_scope,'core')='core'
                    AND (
                            c.owner_user_id=cu.user_id
                            OR cu.membership_kind='primary'
                    )
                    ORDER BY
                        CASE WHEN c.owner_user_id=cu.user_id THEN 0 ELSE 1 END,
                        CASE WHEN cu.membership_kind='primary' THEN 0 ELSE 1 END,
                        cu.company_id
                    LIMIT 1;
                    """,
                    (user_id,),
                )

        else:

            memberships = db_service.fetch_all(
                """
                SELECT
                    company_id,
                    role,
                    access_scope,
                    is_primary,
                    membership_kind
                FROM public.company_users
                WHERE user_id=%s
                AND is_active=TRUE
                ORDER BY company_id;
                """,
                (user_id,),
            ) or []

            if not memberships:
                return jsonify({
                    "error": "No active company membership is available for this account.",
                    "code": "NO_ACTIVE_COMPANY",
                }), 403

            primary_memberships = [
                row for row in memberships
                if bool(row.get("is_primary"))
            ]

            # ---------------------------------------------------------
            # More than one accessible company:
            # exactly one MUST be primary.
            # ---------------------------------------------------------
            if len(memberships) > 1:
                if len(primary_memberships) == 0:
                    current_app.logger.error(
                        "[AUTH] User %s has %s active companies but no primary company.",
                        user_id,
                        len(memberships),
                    )

                    return jsonify({
                        "error": (
                            "Your account has access to multiple companies, "
                            "but no primary company has been configured."
                        ),
                        "code": "PRIMARY_COMPANY_REQUIRED",
                    }), 403

                if len(primary_memberships) > 1:
                    current_app.logger.error(
                        "[AUTH] User %s has multiple primary companies: %s",
                        user_id,
                        [
                            int(row["company_id"])
                            for row in primary_memberships
                        ],
                    )

                    return jsonify({
                        "error": (
                            "Your account has more than one primary company configured. "
                            "Please contact an administrator."
                        ),
                        "code": "MULTIPLE_PRIMARY_COMPANIES",
                    }), 403

                primary_mem = primary_memberships[0]

            # ---------------------------------------------------------
            # Exactly one active company:
            # allow legacy accounts even if is_primary was never set.
            # ---------------------------------------------------------
            else:
                primary_mem = memberships[0]

        if is_ops_signin:
            allowed_rows = db_service.fetch_all(
                """
                SELECT DISTINCT cu.company_id
                FROM public.company_users cu
                WHERE cu.user_id=%s
                AND cu.is_active=TRUE
                AND COALESCE(cu.access_scope,'core')='core'
                ORDER BY cu.company_id;
                """,
                (user_id,),
            ) or []
        else:
            allowed_rows = db_service.fetch_all(
                """
                SELECT company_id
                FROM public.company_users
                WHERE user_id=%s
                AND is_active=TRUE
                ORDER BY company_id;
                """,
                (user_id,),
            ) or []

        allowed_company_ids = []

        for row in allowed_rows:
            company_value = row.get("company_id")

            if company_value is not None:
                allowed_company_ids.append(int(company_value))

        company_id = None

        if primary_mem and primary_mem.get("company_id") is not None:
            company_id = int(primary_mem["company_id"])

        if is_ops_signin:
            if not company_id:
                return jsonify({
                    "error":"No primary company is available for FinSage Nexus.",
                    "code":"OPS_NO_PRIMARY_COMPANY",
                }),403

            ops_access = db_service.user_has_ops_access(
                company_id,
                user_id,
            )

            if not ops_access:
                return jsonify({
                    "error":"Your account does not have FinSage Nexus access.",
                    "code":"OPS_ACCESS_DENIED",
                    "company_id":company_id,
                }),403
    
        if primary_mem:
            user_role=normalize_role(
                primary_mem.get("role") or base_role
            )

            access_scope=str(
                primary_mem.get("access_scope") or "core"
            ).strip().lower()

            if is_ops_signin:
                access_scope="core"
        else:
            user_role=normalize_role(base_role)
            access_scope="core"

        current_app.logger.info(
            "[AUTH] 3 - role=%s access_scope=%s",
            user_role,
            access_scope,
        )

        dashboards = get_dashboard_access(
            user_role,
            access_scope,
        )

        default_dashboard = resolve_default_dashboard(
            user_type=user_type,
            role=user_role,
            access_scope=access_scope,
        )

        current_app.logger.info(
            "[AUTH] 5 - allowed companies=%s",
            allowed_company_ids,
        )

        owned_rows = db_service.fetch_all(
            """
            SELECT id AS company_id
            FROM public.companies
            WHERE owner_user_id = %s
            ORDER BY id;
            """,
            (user_id,),
        ) or []

        for row in owned_rows:
            owned_company_id = row.get("company_id")

            if owned_company_id is not None:
                allowed_company_ids.append(
                    int(owned_company_id)
                )

        if company_id is not None:
            allowed_company_ids.append(company_id)

        allowed_company_ids = sorted(
            set(allowed_company_ids)
        )

        token = make_jwt(
            user_id=user_id,
            email=user["email"],
            role=user_role,
            user_type=user_type,
            company_id=company_id,
            access_scope=access_scope,
            allowed_company_ids=allowed_company_ids,
        )

        company_name = None
        industry = None
        sub_industry = None

        if company_id is not None:
            try:
                company = db_service.fetch_one(
                    """
                    SELECT id, name, industry, sub_industry
                    FROM public.companies
                    WHERE id = %s
                    LIMIT 1;
                    """,
                    (company_id,),
                )

                if company:
                    company_name = company.get("name")
                    industry = company.get("industry")
                    sub_industry = company.get("sub_industry")

            except Exception:
                current_app.logger.exception(
                    "[AUTH] Could not load company %s for user %s",
                    company_id,
                    user_id,
                )

        # This return must be outside `if company_id`.
        return jsonify({
            "message":"signin OK",
            "token":token,
            "product":"FinSage Nexus" if is_ops_signin else "finsage",
            "user":{
                "id": user_id,
                "email": user["email"],
                "role": user_role,
                "user_type": user_type,
                "access_scope": access_scope,
                "is_confirmed": bool(
                    user.get("is_confirmed")
                ),
                "company_id": company_id,
                "primary_company_id":company_id,
                "company_name": company_name,
                "industry": industry,
                "sub_industry": sub_industry,
                "allowed_company_ids": allowed_company_ids,
                "dashboards": dashboards,
                "default_dashboard": default_dashboard,
            },
        }), 200

    except Exception as exc:
        current_app.logger.exception(
            "[AUTH] Unexpected signin failure"
        )

        return jsonify({
            "error": "Sign in failed due to a server error.",
            "error_type": type(exc).__name__,
            "detail": str(exc),
        }), 500

    
@app.route("/api/auth/me", methods=["GET", "OPTIONS"])
@require_auth
def api_auth_me():
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    payload = getattr(request, "jwt_payload", {}) or {}
    user_id = payload.get("user_id") or payload.get("sub")
    if not user_id:
        return jsonify({"error": "AUTH|missing_user_id"}), 401

    user_id = int(user_id)

    user = db_service.fetch_one("""
    SELECT
        id, email, first_name, last_name, user_type,
        created_at,
        password_updated_at,
        last_login_at,
        last_activity_at,
        force_password_change
        FROM public.users
        WHERE id=%s
        LIMIT 1;
    """, (user_id,))
    if not user:
        return jsonify({"error": "AUTH|user_not_found"}), 401

    allowed_rows = db_service.fetch_all("""
        SELECT company_id
        FROM public.company_users
        WHERE user_id=%s AND is_active=TRUE
        ORDER BY company_id;
    """, (user_id,)) or []
    native_allowed = [int(r["company_id"]) for r in allowed_rows]
    allowed_companies, native_allowed, primary_company_id = build_user_allowed_companies(user_id)

    token_cid_raw = payload.get("target_company_id") or payload.get("company_id")
    try:
        token_cid = int(token_cid_raw) if token_cid_raw not in (None, "") else None
    except (TypeError, ValueError):
        token_cid = None

    is_delegated = bool(payload.get("is_delegated_company_access"))

    company_id = None
    role = "viewer"
    access_scope = "core"
    governance_mode = "owner_managed"
    company_name = None
    industry = None
    sub_industry = None

    if is_delegated and token_cid:
        # TRUST THE DELEGATED TOKEN
        company_id = token_cid

        delegated_allowed = payload.get("allowed_company_ids")
        if not isinstance(delegated_allowed, list) or not delegated_allowed:
            delegated_allowed = [company_id]
        else:
            delegated_allowed = [int(x) for x in delegated_allowed if str(x).strip()]
            if company_id not in delegated_allowed:
                delegated_allowed = [company_id]

        raw_role = (
            payload.get("role")
            or "viewer"
        )
        role = normalize_role(raw_role)

        # keep token scope as-is, or use a dedicated label if you prefer
        access_scope = "delegated_workspace"

        company = db_service.fetch_one("""
            SELECT name, industry, sub_industry, credit_policy
            FROM public.companies
            WHERE id=%s
            LIMIT 1;
        """, (company_id,))
        if company:
            company_name = company.get("name")
            industry = company.get("industry")
            sub_industry = company.get("sub_industry")
            cp = company.get("credit_policy") or {}
            if isinstance(cp, dict):
                governance_mode = (cp.get("mode") or "owner_managed").strip().lower()

        # TRUST TOKEN PERMISSIONS FIRST IN DELEGATED MODE
        permissions = payload.get("permissions") or {}
        if not isinstance(permissions, dict):
            permissions = {}

        # Force dashboard shape for delegated posting shell
        # Delegated workspace MUST use enterprise shell (posting dashboard)
        dashboards = {
            "enterprise": True,
            "practitioner": False,
        }

        default_dashboard = "enterprise"

        # also force scope for frontend logic
        access_scope = "delegated_workspace"

        active_company_context = None
        for r in allowed_companies:
            if company_id and int(r["id"]) == int(company_id):
                active_company_context = r
                break 

        out = {
            "id": int(user["id"]),
            "email": user["email"],
            "first_name": user.get("first_name"),
            "last_name": user.get("last_name"),
            "user_type": user.get("user_type") or payload.get("user_type"),

            "role": role,
            "access_scope": "delegated_workspace",
            "company_id": company_id,
            "company_name": company_name,
            "industry": industry,
            "sub_industry": sub_industry,
            "governance_mode": governance_mode,

            "primary_company_id": primary_company_id,
            "active_company_id": company_id,
            "allowed_companies": allowed_companies,
            "company_context": active_company_context,

            "allowed_company_ids": delegated_allowed,
            "token_company_id": company_id,
            "token_access_scope": "delegated_workspace",
            "token_allowed_company_ids": delegated_allowed,
            "source_company_id": payload.get("source_company_id"),
            "target_company_id": payload.get("target_company_id") or company_id,
            "engagement_id": payload.get("engagement_id"),
            "is_delegated_company_access": True,
            "is_native_company_member": False,

            "dashboards": {
                "enterprise": True,
                "practitioner": False,
            },
            "permissions": permissions,
            "default_dashboard": "enterprise",
            "created_at": user.get("created_at"),
            "password_updated_at": user.get("password_updated_at") or user.get("created_at"),
            "last_login_at": user.get("last_login_at"),
            "last_activity_at": user.get("last_activity_at"),
            "force_password_change": bool(user.get("force_password_change")),
        }

        app.logger.warning("AUTH_ME delegated out=%s", out)
        return jsonify(out), 200

    # -----------------------------
    # Normal native membership flow
    # -----------------------------
    primary_mem = db_service.fetch_one("""
        SELECT company_id, role, access_scope, is_active
        FROM public.company_users
        WHERE user_id=%s
          AND is_active=TRUE
          AND is_primary=TRUE
        LIMIT 1;
    """, (user_id,))

    # Permanent/default native company.
    primary_company_id = (
        int(primary_mem["company_id"])
        if primary_mem and primary_mem.get("company_id")
        else primary_company_id
    )

    # Active context:
    # - primary/native company is the default
    # - another company is used only when the token represents an
    #   intentional company switch
    if token_cid and token_cid in native_allowed:
        company_id = token_cid
    elif primary_company_id:
        company_id = int(primary_company_id)
    elif native_allowed:
        company_id = native_allowed[0]
    else:
        company_id = None

    mem = None
    if company_id:
        mem = db_service.fetch_one("""
            SELECT role, access_scope, is_active
            FROM public.company_users
            WHERE company_id=%s AND user_id=%s
            LIMIT 1;
        """, (company_id, user_id))

        if not mem or not bool(mem.get("is_active", True)):
            company_id = None
            role = "viewer"
            access_scope = "core"
        else:
            raw_role = (mem.get("role") or "viewer").strip().lower()
            role = normalize_role(raw_role)
            access_scope = (mem.get("access_scope") or "core").strip().lower()

    if company_id:
        company = db_service.fetch_one("""
            SELECT name, industry, sub_industry, credit_policy
            FROM public.companies
            WHERE id=%s
            LIMIT 1;
        """, (company_id,))
        if company:
            company_name = company.get("name")
            industry = company.get("industry")
            sub_industry = company.get("sub_industry")
            cp = company.get("credit_policy") or {}
            if isinstance(cp, dict):
                governance_mode = (cp.get("mode") or "owner_managed").strip().lower()

    dashboards = get_dashboard_access(role, access_scope)
    permissions = build_permissions(role=role, access_scope=access_scope)
    default_dashboard = resolve_default_dashboard(
        user_type=user.get("user_type") or payload.get("user_type"),
        role=role,
        access_scope=access_scope,
    )

    active_company_context = None
    for r in allowed_companies:
        if company_id and int(r["id"]) == int(company_id):
            active_company_context = r
            break

    out = {
        "id": int(user["id"]),
        "email": user["email"],
        "first_name": user.get("first_name"),
        "last_name": user.get("last_name"),
        "user_type": user.get("user_type") or payload.get("user_type"),
        "role": role,
        "access_scope": access_scope,
        "company_id": company_id,
        "company_name": company_name,
        "industry": industry,
        "sub_industry": sub_industry,
        "governance_mode": governance_mode,
        "allowed_company_ids": native_allowed,
        "primary_company_id": primary_company_id,
        "active_company_id": company_id,
        "allowed_companies": allowed_companies,
        "company_context": active_company_context,
        "token_company_id": payload.get("company_id"),
        "token_access_scope": payload.get("access_scope"),
        "token_allowed_company_ids": payload.get("allowed_company_ids"),
        "dashboards": dashboards,
        "permissions": permissions,
        "default_dashboard": default_dashboard,
    }

    return jsonify(out), 200

@app.route("/api/pos/auth/signin", methods=["POST", "OPTIONS"])
def pos_auth_signin_global():
    try:
        if request.method == "OPTIONS":
            return ("", 204)

        data = request.get_json(silent=True) or {}

        employee_code = str(data.get("employee_code") or "").strip()
        pin = str(data.get("pin") or "").strip()

        if not employee_code or not pin:
            return jsonify({"error": "Employee ID and PIN are required"}), 400

        row = db_service.fetch_one("""
            SELECT
                cu.id AS company_user_id,
                cu.company_id AS company_id,
                cu.user_id,
                cu.employee_code,
                cu.pos_access_code,
                cu.pos_display_name,
                cu.pos_role,
                cu.pos_permissions,
                cu.pos_is_active,
                cu.pin_hash,
                u.email,
                u.first_name,
                u.last_name,
                c.name AS company_name,
                c.industry,
                c.sub_industry,
                c.currency
            FROM public.company_users cu
            JOIN public.users u ON u.id = cu.user_id
            JOIN public.companies c ON c.id = cu.company_id
            WHERE cu.pos_access_code = %s
              AND cu.pos_is_active = TRUE
              AND cu.is_active = TRUE
              AND c.is_active = TRUE
            LIMIT 1;
        """, (employee_code,))

        if not row or not row.get("pin_hash"):
            return jsonify({"error": "Invalid employee ID or PIN"}), 401

        if not check_password_hash(row["pin_hash"], pin):
            return jsonify({"error": "Invalid employee ID or PIN"}), 401

        company_id = int(row["company_id"])

        company = {
            "id": company_id,
            "name": row.get("company_name"),
            "industry": row.get("industry"),
            "sub_industry": row.get("sub_industry"),
            "currency": row.get("currency"),
        }

        employee = {
            "company_user_id": row["company_user_id"],
            "user_id": row["user_id"],
            "employee_code": row["employee_code"],
            "access_code": row["pos_access_code"],
            "name": row.get("pos_display_name")
                or f"{row.get('first_name') or ''} {row.get('last_name') or ''}".strip()
                or row.get("email"),
            "pos_role": row.get("pos_role"),
            "pos_permissions": row.get("pos_permissions") or {},
        }

        pos_token = make_pos_jwt(
            company_id=company_id,
            company_user_id=row["company_user_id"],
            user_id=row["user_id"],
            employee_code=row["employee_code"],
            pos_role=row.get("pos_role"),
        )

        return jsonify({
            "ok": True,
            "pos_token": pos_token,
            "employee": employee,
            "company": company,
        }), 200

    except Exception as e:
        current_app.logger.exception("POS GLOBAL SIGNIN FAILED")
        return jsonify({"error": str(e)}), 500
    
@app.route("/api/companies/<int:company_id>/pos/auth/me", methods=["GET", "OPTIONS"])
@require_pos_auth
def pos_auth_me(company_id):
    if request.method == "OPTIONS":
        return ("", 204)

    pos_user = getattr(request, "pos_user", {}) or {}

    return jsonify({
        "ok": True,
        "employee": pos_user.get("employee"),
        "company": pos_user.get("company"),
    }), 200

@app.route("/api/auth/reauth", methods=["POST", "OPTIONS"])
@require_auth
def api_auth_reauth():
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    data = request.get_json(silent=True) or {}
    password = data.get("password") or ""

    if not password:
        return jsonify({"ok": False, "error": "Password required"}), 400

    current_user = getattr(g, "current_user", None) or {}
    email = current_user.get("email") or (getattr(request, "jwt_payload", {}) or {}).get("email")

    if not email:
        return jsonify({"ok": False, "error": "AUTH|missing_email"}), 401

    user = db_service.get_user_by_email(email)
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"ok": False, "error": "Password incorrect"}), 403

    return jsonify({"ok": True, "message": "Session unlocked"}), 200

@app.route("/api/companies/<int:company_id>/invoices/<int:invoice_id>/pdf_token", methods=["POST"])
@require_auth
def invoice_pdf_token(company_id: int, invoice_id: int):
    inv = db_service.get_invoice_full(company_id, invoice_id)
    if not inv:
        return jsonify({"error": "Invoice not found"}), 404

    # short-lived token (2 minutes)
    token = create_invoice_pdf_token(company_id=company_id, invoice_id=invoice_id, ttl_seconds=120)
    return jsonify({"token": token})


# ────────────────────────────────────────────────────────────────
# Forgot / Reset / Change Password
# ────────────────────────────────────────────────────────────────

@app.route("/api/auth/request-reset", methods=["OPTIONS", "POST"])
def request_password_reset():
    if request.method == "OPTIONS":
        return "", 204

    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    if not email:
        return jsonify({"error": "Email is required"}), 400

    user = db_service.get_user_by_email(email)
    if not user:
        return jsonify({"message": "If account exists, reset link sent"}), 200

    token = secrets.token_urlsafe(32)
    expiry = datetime.now(timezone.utc) + timedelta(hours=1)
    db_service.store_reset_token(email, token, expiry)

    reset_link = f"{FRONTEND_BASE}/reset-password.html?token={token}"
    first_name = (user.get("first_name") or "").strip()

    subject = "Your FinSage password reset link (expires in 1 hour)"
    html = f"""
    <p>Hi {first_name},</p>
    <p>We received a request to reset your FinSage password.</p>
    <p>Use this link to continue: <a href="{reset_link}">Reset password</a></p>
    <p><b>Note:</b> This link expires in 1 hour.</p>
    <p>If you did not request this, you can safely ignore this email.</p>
    """
    text = (
        f"Hi {first_name},\n\n"
        f"We received a request to reset your FinSage password.\n\n"
        f"Reset password: {reset_link}\n\n"
        f"Note: This link expires in 1 hour.\n"
        f"If you did not request this, you can safely ignore this email."
    )

    try:
        send_mail(to_email=email, subject=subject, html_body=html, text_body=text)
    except Exception:
        current_app.logger.exception("Password reset email failed")

    return jsonify({"message": "If account exists, reset link sent"}), 200

@app.route("/api/auth/reset-password", methods=["OPTIONS", "POST"])
def reset_password():
    if request.method == "OPTIONS":
        return "", 204

    data = request.get_json() or {}
    token = data.get("token")
    new_password = data.get("newPassword")

    if not token or not new_password:
        return jsonify({"error": "Token and new password required"}), 400

    user = db_service.get_user_by_reset_token(token)
    if not user:
        return jsonify({"error": "Invalid or expired token"}), 400

    token_expiry = user.get("token_expiry")
    if not token_expiry or token_expiry < datetime.now(timezone.utc):
        return jsonify({"error": "Invalid or expired token"}), 400

    new_hash = generate_password_hash(new_password)
    db_service.update_password(user["email"], new_hash)
    db_service.clear_reset_token(user["email"])
    return jsonify({"message": "Password reset successful"}), 200

@app.route("/api/auth/change-password", methods=["POST"])
@require_auth
def change_password():
    data = request.get_json() or {}
    old = data.get("oldPassword")
    new = data.get("newPassword")

    if not old or not new:
        return jsonify({"error": "Both old and new passwords required"}), 400

    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"error": "Not authenticated"}), 401

    email = current_user.get("email")
    user = db_service.get_user_by_email(email) if email else None

    if not user or not check_password_hash(user["password_hash"], old):
        return jsonify({"error": "Old password incorrect"}), 403

    new_hash = generate_password_hash(new)
    db_service.update_password(user["email"], new_hash)
    return jsonify({"message": "Password updated"}), 200

# ────────────────────────────────────────────────────────────────
# INVITE API (for accept-invite.html)
# ────────────────────────────────────────────────────────────────
@app.route("/api/auth/invite-info", methods=["GET"])
def api_invite_info():
    token = (request.args.get("token") or "").strip()
    if not token:
        return jsonify({"error": "Missing invite token."}), 400

    rec = db_service.get_company_invite_by_token(token)

    current_app.logger.warning(
        "[INVITE INFO DEBUG] token=%s rec_found=%s expires_at=%r now_utc=%r",
        token,
        bool(rec),
        (rec or {}).get("expires_at"),
        datetime.now(timezone.utc),
    )

    state_error = _invite_state_error(rec)
    if state_error:
        msg, code = state_error
        return jsonify({"error": msg}), code

    ROLE_LABELS = {
        "clerk": "Accounts Clerk",
        "assistant": "Finance Assistant",
        "junior": "Junior Accountant",
        "senior": "Senior Accountant",
        "cfo": "CFO",
        "viewer": "Viewer",
        "manager": "Manager",
        "accountant": "Accountant",
        "owner": "Owner",
        "admin": "Admin",
        "bookkeeper": "Bookkeeper",
        "audit_staff": "Audit Staff",
        "senior_associate": "Senior Associate",
        "audit_manager": "Audit Manager",
        "audit_partner": "Audit Partner",
        "engagement_partner": "Engagement Partner",
        "quality_control_reviewer": "Quality Control Reviewer",
        "fs_compiler": "Financial Statement Compiler",
        "reviewer": "Reviewer",
        "client_service_manager": "Client Service Manager",
    }

    role_raw = (rec.get("role") or "").strip().lower()
    role_label = ROLE_LABELS.get(role_raw, role_raw or "User")

    access_scope = (rec.get("access_scope") or "core").strip().lower()
    if access_scope not in ("core", "assignment"):
        access_scope = "core"

    return jsonify({
        "email": (rec.get("email") or "").strip().lower(),
        "role": role_raw,
        "roleLabel": role_label,
        "accessScope": access_scope,
        "companyName": rec.get("company_name") or rec.get("companyName") or "",
        "companyId": rec.get("company_id"),
        "expiresAt": _as_utc_aware(rec.get("expires_at")).isoformat() if rec.get("expires_at") else None,
    }), 200

@app.route("/api/invites", methods=["POST"])
@require_auth
def api_create_invite():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    role_raw = (data.get("role") or "").strip().lower()
    access_scope = (data.get("access_scope") or "core").strip().lower()
    note = (data.get("note") or "").strip()

    department_id=data.get("department_id")
    position_id=data.get("position_id")
    branch_id=data.get("branch_id")
    manager_user_id=data.get("manager_user_id")
    ops_role_code=(data.get("ops_role_code") or "").strip().upper() or None

    product_access=data.get("product_access") or {}
    if not isinstance(product_access,dict):
        product_access={}

    if not email or not role_raw:
        return jsonify({"error": "Email and role are required."}), 400

    if access_scope not in ("core", "assignment"):
        return jsonify({"error": "Invalid access_scope. Use 'core' or 'assignment'."}), 400

    try:
        role = validate_role_for_scope(role_raw, access_scope)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"error": "Not authenticated."}), 401

    actor_user_id = int(current_user.get("id") or 0)
    if not actor_user_id:
        return jsonify({"error": "Not authenticated."}), 401

    current_app.logger.warning(
        "[INVITE DEBUG] jwt_payload=%r",
        getattr(request, "jwt_payload", {}) or {}
    )
    current_app.logger.warning(
        "[INVITE DEBUG] g.current_user=%r",
        current_user
    )

    actor_role = (current_user.get("user_role") or current_user.get("role") or "").strip().lower()
    if actor_role not in ("owner", "admin", "cfo", "manager", "senior"):
        return jsonify({"error": "Not authorised to invite users."}), 403

    # Resolve employer / primary company directly from membership table.
    # Do NOT trust current_user.company_id here because require_auth may
    # have populated it from a stale JWT company context.
    primary_mem = db_service.fetch_one(
        """
        SELECT company_id, user_id, role, access_scope, membership_kind, is_primary, is_active
        FROM public.company_users
        WHERE user_id = %s
          AND is_active = TRUE
          AND (is_primary = TRUE OR membership_kind = 'primary')
        ORDER BY is_primary DESC, id ASC
        LIMIT 1
        """,
        (actor_user_id,)
    )

    current_app.logger.warning(
        "[INVITE DEBUG] resolved_primary_membership=%r",
        primary_mem
    )

    if not primary_mem or not bool(primary_mem.get("is_active", True)):
        return jsonify({
            "error": (
                "No primary employer company was found for your account. "
                "Please sign out and sign in to your employer company account."
            ),
            "code": "INVITE_PRIMARY_COMPANY_REQUIRED",
        }), 403

    company_id = int(primary_mem.get("company_id") or 0)
    if not company_id:
        return jsonify({
            "error": "Primary employer company could not be resolved.",
            "code": "INVITE_PRIMARY_COMPANY_REQUIRED",
        }), 403

    membership_kind = (primary_mem.get("membership_kind") or "secondary").strip().lower()
    is_primary = bool(primary_mem.get("is_primary"))

    # Defensive check: invites are only allowed from employer/primary company context.
    if not is_primary and membership_kind != "primary":
        return jsonify({
            "error": (
                "Invites can only be sent from your primary employer company. "
                "Please sign out and sign in to your employer company account, then try again."
            ),
            "code": "INVITE_PRIMARY_COMPANY_REQUIRED",
        }), 403

    current_app.logger.warning(
        "[INVITE DEBUG] actor_user_id=%r jwt_company_id=%r current_user_company_id=%r resolved_invite_company_id=%r role=%r access_scope=%r",
        actor_user_id,
        (getattr(request, "jwt_payload", {}) or {}).get("company_id"),
        current_user.get("company_id"),
        company_id,
        role_raw,
        access_scope,
    )

    company = db_service.get_company_profile(company_id) or {}
    company_name = company.get("name") or "FinSage workspace"

    recommended_mode = None
    if access_scope == "core":
        recommended_mode = recommend_mode_after_invite(company, role)

    rec = db_service.create_company_invite(
        company_id=company_id,
        email=email,
        role=role,
        access_scope=access_scope,
        note=note,
        created_by=actor_user_id,
        ttl_seconds=INVITE_TTL_SECONDS
    )

    current_app.logger.warning(
        "[CREATE INVITE DEBUG] token=%s email=%s company_id=%s ttl=%r expires_at=%r created=%r",
        rec.get("token"),
        email,
        company_id,
        INVITE_TTL_SECONDS,
        rec.get("expires_at"),
        rec.get("created"),
    )

    token = rec["token"]

    db_service.set_company_invite_ops_context(
        token=token,
        department_id=department_id,
        position_id=position_id,
        branch_id=branch_id,
        manager_user_id=manager_user_id,
        ops_role_code=ops_role_code,
        product_access=product_access,
    )

    invite_link=f"{FRONTEND_BASE}/ops/accept-invite?token={token}"
    access_scope_label = "internal access" if access_scope == "core" else "assignment access"

    subject = f"You've been invited to {company_name} on FinSage"
    html = f"""
      <div style="font-family:system-ui,Segoe UI,Arial;line-height:1.5">
        <h2>You’ve been invited to {company_name}</h2>
        <p>{current_user.get("first_name") or "A colleague"} has invited you to join their FinSage workspace.</p>
        <p><b>Role:</b> {role}</p>
        <p><b>Access:</b> {access_scope_label}</p>
        <p>
          <a href="{invite_link}"
             style="background:#00C8C8;color:#fff;padding:10px 16px;border-radius:6px;
                    text-decoration:none;display:inline-block">
            Accept invite
          </a>
        </p>
        <p>If the button doesn’t work, copy &amp; paste this link:</p>
        <p><code>{invite_link}</code></p>
      </div>
    """
    text = (
        f"You’ve been invited to {company_name} on FinSage.\n"
        f"Role: {role}\n"
        f"Access: {access_scope_label}\n\n"
        f"Accept your invite: {invite_link}"
    )

    email_sent=False
    email_error=None

    try:
        send_company_mail(
            company_id=company_id,
            to_email=email,
            subject=subject,
            html_body=html,
            text_body=text
        )
        email_sent=True
        current_app.logger.info("[INVITE] Sent invite email to %s",email)
    except Exception as e:
        email_error=str(e)
        current_app.logger.warning("[INVITE] EMAIL SEND FAILED for %s: %s",email,e)

    return jsonify({
        "message":"Invite created and email sent" if email_sent else "Invite created but email could not be sent",
        "email":email,
        "email_sent":email_sent,
        "email_error":email_error,
        "role":role,
        "access_scope":access_scope,
        "invite_link":invite_link,
        "token":token,
        "company_id":company_id,
        "recommended_mode":recommended_mode,
        "reason":"multi_user_company" if recommended_mode else None,
    }),201

@app.route("/api/auth/accept-invite", methods=["POST"])
def api_accept_invite():
    data = request.get_json(silent=True) or {}
    token = (data.get("token") or "").strip()
    first_name = (data.get("firstName") or "").strip()
    last_name = (data.get("lastName") or "").strip()
    password = data.get("password") or ""
    confirm_pw = data.get("confirmPassword") or ""

    if not token:
        return jsonify({"error": "Missing invitation token."}), 400

    rec = db_service.get_company_invite_by_token(token)

    current_app.logger.warning(
        "[ACCEPT INVITE DEBUG] token=%s rec_found=%s expires_at=%r now_utc=%r",
        token, bool(rec), (rec or {}).get("expires_at"), datetime.now(timezone.utc),
    )

    state_error = _invite_state_error(rec)
    if state_error:
        msg, code = state_error
        return jsonify({"error": msg}), code

    email = (rec.get("email") or "").strip().lower()
    user_role = (rec.get("role") or "other").strip().lower()
    access_scope = (rec.get("access_scope") or "core").strip().lower()
    company_id = int(rec.get("company_id") or 0) or None

    if access_scope not in ("core", "assignment"):
        access_scope = "core"

    try:
        user_role = validate_role_for_scope(user_role, access_scope)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if not company_id:
        return jsonify({"error": "Invite is missing company link."}), 400

    existing = db_service.fetch_one("""
        SELECT id, email, is_confirmed, user_type
        FROM public.users
        WHERE LOWER(email)=LOWER(%s)
        LIMIT 1;
    """, (email,))

    # =========================================================
    # EXISTING USER
    # =========================================================
    if existing:
        user_id = int(existing["id"])
        existing_user_type = (existing.get("user_type") or "").strip() or "Enterprise"

        # Resolve user's current HOME company before touching membership.
        primary = db_service.fetch_one("""
            SELECT company_id
            FROM public.company_users
            WHERE user_id=%s
              AND is_active=TRUE
              AND is_primary=TRUE
            ORDER BY id
            LIMIT 1;
        """, (user_id,))

        # Defensive fallback:
        # if this is an old account with memberships but no primary flag,
        # preserve its oldest existing membership as home rather than
        # automatically promoting the newly invited company.
        if not primary:
            existing_membership = db_service.fetch_one("""
                SELECT company_id
                FROM public.company_users
                WHERE user_id=%s
                  AND is_active=TRUE
                  AND company_id<>%s
                ORDER BY joined_at NULLS LAST, id
                LIMIT 1;
            """, (user_id, int(company_id)))

            if existing_membership:
                existing_primary_id = int(existing_membership["company_id"])

                db_service.execute_sql("""
                    UPDATE public.company_users
                    SET is_primary=FALSE
                    WHERE user_id=%s
                      AND is_primary=TRUE;
                """, (user_id,))

                db_service.execute_sql("""
                    UPDATE public.company_users
                    SET is_primary=TRUE,
                        membership_kind='primary'
                    WHERE user_id=%s
                      AND company_id=%s;
                """, (user_id, existing_primary_id))

                primary = {"company_id": existing_primary_id}

        # The invited company is secondary when user already has a home company.
        invite_is_primary = primary is None

        db_service.upsert_company_user(
            company_id=int(company_id),
            user_id=user_id,
            role=user_role,
            access_scope=access_scope,
            membership_kind="primary" if invite_is_primary else "secondary",
            is_primary=invite_is_primary,
        )

        if invite_is_primary:
            db_service.execute_sql(
                "UPDATE public.users SET company_id=%s WHERE id=%s",
                (int(company_id), user_id),
            )
            jwt_company_id = int(company_id)
        else:
            jwt_company_id = int(primary["company_id"])

        db_service.apply_ops_invite_context(token=token, user_id=user_id)
        db_service.mark_invite_accepted(token=token, user_id=user_id)

        jwt_token = make_jwt(
            user_id=user_id,
            email=email,
            role=user_role,
            user_type=existing_user_type,
            company_id=jwt_company_id,
            access_scope=access_scope,
        )

        return jsonify({
            "message": "Invite accepted. You can now sign in.",
            "email": email,
            "role": user_role,
            "access_scope": access_scope,
            "companyId": company_id,
            "primaryCompanyId": jwt_company_id,
            "token": jwt_token,
            "existingUser": True,
        }), 200

    # =========================================================
    # NEW USER
    # =========================================================
    if not first_name or not last_name:
        return jsonify({"error": "First name and last name are required."}), 400

    if not password or password != confirm_pw:
        return jsonify({"error": "Passwords do not match."}), 400

    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters."}), 400

    invited_user_type = "Practitioner" if access_scope == "assignment" else "Enterprise"

    new_id = db_service.insert_user(
        email=email,
        password_hash=generate_password_hash(password),
        user_type=invited_user_type,
        first_name=first_name,
        last_name=last_name,
        user_role=user_role,
        is_confirmed=True,
        confirmation_token=None,
        confirmation_token_expires_at=None,
        trial_start_date=str(date.today()),
        trial_end_date=str(date.today() + timedelta(days=30)),
        company_id=company_id,
    )

    # First company membership = HOME company.
    # access_scope remains exactly what the inviter selected.
    db_service.upsert_company_user(
        company_id=int(company_id),
        user_id=int(new_id),
        role=user_role,
        access_scope=access_scope,
        membership_kind="primary",
        is_primary=True,
    )

    db_service.apply_ops_invite_context(token=token, user_id=int(new_id))
    db_service.mark_invite_accepted(token=token, user_id=int(new_id))

    jwt_token = make_jwt(
        user_id=new_id,
        email=email,
        role=user_role,
        user_type=invited_user_type,
        company_id=company_id,
        access_scope=access_scope,
    )

    return jsonify({
        "message": "Invitation accepted. You can now sign in.",
        "email": email,
        "role": user_role,
        "access_scope": access_scope,
        "companyId": company_id,
        "primaryCompanyId": company_id,
        "token": jwt_token,
        "existingUser": False,
    }), 200
# ============================================================
# COMPANY EMAIL SETTINGS
# ============================================================

@app.route(
    "/api/companies/<int:company_id>/email-settings",
    methods=["GET"],
)
@require_auth
def api_get_company_email_settings(company_id):
    deny=_require_company_email_admin(company_id)
    if deny:
        return deny

    try:
        row=db_service.get_company_email_settings(company_id)
        return jsonify(row or {}),200

    except Exception as e:
        current_app.logger.exception(
            "[COMPANY EMAIL] Failed to load settings company=%s",
            company_id,
        )
        return jsonify({"error":str(e)}),500


@app.route(
    "/api/companies/<int:company_id>/email-settings",
    methods=["PATCH"],
)
@require_auth
def api_save_company_email_settings(company_id):
    deny=_require_company_email_admin(company_id)
    if deny:
        return deny

    payload=request.get_json(silent=True) or {}

    try:
        row=db_service.save_company_email_settings(
            company_id,
            payload=payload,
        )

        return jsonify(row or {}),200

    except ValueError as e:
        return jsonify({"error":str(e)}),400

    except Exception as e:
        current_app.logger.exception(
            "[COMPANY EMAIL] Failed to save settings company=%s",
            company_id,
        )
        return jsonify({"error":str(e)}),500


@app.route(
    "/api/companies/<int:company_id>/email-settings/test",
    methods=["POST"],
)
@require_auth
def api_test_company_email_settings(company_id):
    deny=_require_company_email_admin(company_id)
    if deny:
        return deny

    body=request.get_json(silent=True) or {}
    recipient=(body.get("recipient_email") or "").strip().lower()

    if not recipient:
        return jsonify({
            "error":"recipient_email is required."
        }),400

    cfg=db_service.get_company_email_settings_internal(
        company_id
    )

    if not cfg:
        return jsonify({
            "error":"Company email settings have not been configured."
        }),400

    required={
        "sender_email":cfg.get("sender_email"),
        "smtp_host":cfg.get("smtp_host"),
        "smtp_port":cfg.get("smtp_port"),
        "smtp_username":cfg.get("smtp_username"),
        "smtp_password_encrypted":cfg.get(
            "smtp_password_encrypted"
        ),
    }

    missing=[
        key
        for key,value in required.items()
        if not value
    ]

    if missing:
        return jsonify({
            "error":"Email configuration is incomplete.",
            "missing":missing,
        }),400

    try:
        password=db_service._decrypt_secret(
            cfg.get("smtp_password_encrypted")
        )

        company=db_service.get_company_profile(
            company_id
        ) or {}

        company_name=(
            company.get("name")
            or cfg.get("sender_name")
            or "Your company"
        )

        subject=f"FinSage email test - {company_name}"

        html=f"""
        <div style="font-family:Arial,sans-serif;line-height:1.6;color:#172033">
            <h2>Company email connected successfully</h2>

            <p>
                This test confirms that FinSage can send email
                through the SMTP account configured for
                <strong>{company_name}</strong>.
            </p>

            <p>
                <strong>Sender:</strong>
                {cfg.get("sender_email")}
            </p>

            <p>
                You can now enable this account for company
                communications.
            </p>

            <p style="color:#667085;font-size:12px">
                Sent automatically by FinSage.
            </p>
        </div>
        """

        text=(
            f"Company email connected successfully.\n\n"
            f"FinSage successfully sent this message through "
            f"the SMTP account configured for {company_name}.\n\n"
            f"Sender: {cfg.get('sender_email')}"
        )

        _send_smtp(
            host=cfg.get("smtp_host") or "",
            port=int(cfg.get("smtp_port") or 465),
            user=cfg.get("smtp_username") or "",
            password=password or "",
            mail_from=cfg.get("sender_email") or "",
            from_name=(
                cfg.get("sender_name")
                or company_name
                or "FinSage"
            ),
            to_email=recipient,
            subject=subject,
            html_body=html,
            text_body=text,
            reply_to=(
                cfg.get("reply_to_email")
                or cfg.get("sender_email")
            ),
            use_ssl=bool(cfg.get("use_ssl",True)),
            use_tls=bool(cfg.get("use_tls",False)),
        )

        status=db_service.set_company_email_test_status(
            company_id,
            verified=True,
            error=None,
        )

        current_app.logger.info(
            "[COMPANY EMAIL] SMTP verified company=%s recipient=%s",
            company_id,
            recipient,
        )

        return jsonify({
            "message":"Test email sent successfully.",
            "recipient_email":recipient,
            "is_verified":True,
            **(status or {}),
        }),200

    except Exception as e:
        error=str(e)

        try:
            status=db_service.set_company_email_test_status(
                company_id,
                verified=False,
                error=error,
            )
        except Exception:
            status={}

        current_app.logger.warning(
            "[COMPANY EMAIL] SMTP test failed company=%s: %s",
            company_id,
            error,
        )

        return jsonify({
            "error":error,
            "is_verified":False,
            **(status or {}),
        }),400

# ============================================================
# COMPANY PROFILE SETTINGS
# ============================================================

@app.route(
    "/api/companies/<int:company_id>/profile",
    methods=["GET"],
)
@require_auth
def api_get_company_profile(company_id):
    deny=_require_company_email_admin(company_id)
    if deny:
        return deny

    try:
        row=db_service.get_company_profile(company_id)

        if not row:
            return jsonify({
                "error":"Company not found."
            }),404

        return jsonify(row),200

    except Exception as e:
        current_app.logger.exception(
            "[COMPANY PROFILE] Failed to load company=%s",
            company_id,
        )

        return jsonify({
            "error":str(e)
        }),500


@app.route(
    "/api/companies/<int:company_id>/profile",
    methods=["PATCH"],
)
@require_auth
def api_update_company_profile(company_id):
    deny=_require_company_email_admin(company_id)
    if deny:
        return deny

    payload=request.get_json(silent=True) or {}

    try:
        db_service.update_company_profile(
            company_id,
            payload,
        )

        row=db_service.get_company_profile(
            company_id
        )

        return jsonify(row or {}),200

    except ValueError as e:
        return jsonify({
            "error":str(e)
        }),400

    except Exception as e:
        current_app.logger.exception(
            "[COMPANY PROFILE] Failed to update company=%s",
            company_id,
        )

        return jsonify({
            "error":str(e)
        }),500
# ────────────────────────────────────────────────────────────────
# COA template + company endpoints
# ────────────────────────────────────────────────────────────────
@app.route("/api/engagements/<int:engagement_id>/enter-workspace", methods=["POST", "OPTIONS"])
@require_auth(require_company=False)
def enter_engagement_workspace(engagement_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    payload = getattr(request, "jwt_payload", {}) or {}
    user_id = payload.get("user_id") or payload.get("sub")
    user_id = int(user_id) if user_id is not None else None
    if not user_id:
        return _corsify(make_response(jsonify({"error": "AUTH|missing_user_id"}), 401))

    body = request.get_json(silent=True) or {}

    source_company_id = (
        body.get("source_company_id")
        or payload.get("company_id")
        or payload.get("token_company_id")
    )
    try:
        source_company_id = int(source_company_id) if source_company_id is not None else None
    except Exception:
        source_company_id = None

    if not source_company_id:
        return _corsify(make_response(
            jsonify({"error": "AUTH|missing_source_company_id"}), 400
        ))

    with db_service._conn_cursor() as (_conn, cur):
        workspace = db_service.get_delegated_workspace_for_engagement(
            cur,
            company_id=source_company_id,
            engagement_id=int(engagement_id),
            user_id=int(user_id),
        )

    if not workspace:
        return _corsify(make_response(
            jsonify({"error": "User cannot access this engagement workspace"}), 403
        ))

    target_company_id = int(workspace["target_company_id"])

    base_user = db_service.get_user_by_id(int(user_id)) or {}
    role_raw = (
        payload.get("role")
        or workspace.get("role_on_engagement")
        or base_user.get("user_role")
        or base_user.get("role")
        or "viewer"
    )
    role_norm = normalize_role(role_raw)

    delegated_permissions = db_service.build_delegated_workspace_permissions(
        user_id=int(user_id),
        source_company_id=int(source_company_id),
        target_company_id=int(target_company_id),
        engagement_id=int(engagement_id),
        base_payload=payload,
        role=role_norm,
    )

    delegated_token = make_jwt(
        user_id=int(user_id),
        email=base_user.get("email") or payload.get("email") or "",
        role=role_norm,
        user_type=payload.get("user_type") or base_user.get("user_type") or "Practitioner",
        company_id=int(target_company_id),
        access_scope="delegated_workspace",
        allowed_company_ids=[int(target_company_id)],
        source_company_id=int(source_company_id),
        target_company_id=int(target_company_id),
        engagement_id=int(engagement_id),
        permissions=delegated_permissions,
        first_name=base_user.get("first_name") or payload.get("first_name"),
        last_name=base_user.get("last_name") or payload.get("last_name"),
        is_delegated_company_access=True,
        is_native_company_member=False,
    )

    import jwt

    try:
        decoded = jwt.decode(delegated_token, options={"verify_signature": False})
        app.logger.warning("ENTER_WORKSPACE TOKEN DEBUG decoded=%s", decoded)
    except Exception as e:
        app.logger.warning("ENTER_WORKSPACE TOKEN DEBUG decode_failed=%s", e)

    app.logger.warning(
        "ENTER_WORKSPACE RESPONSE user_id=%s source_company_id=%s target_company_id=%s engagement_id=%s delegated_permissions=%s",
        user_id,
        source_company_id,
        target_company_id,
        engagement_id,
        delegated_permissions,
    )

    return _corsify(make_response(jsonify({
        "ok": True,
        "token": delegated_token,
        "workspace": {
            "engagement_id": int(engagement_id),
            "source_company_id": int(source_company_id),
            "target_company_id": int(target_company_id),
            "company_id": int(target_company_id),
            "company_name": workspace.get("target_company_name"),
            "engagement_name": workspace.get("engagement_name"),
            "access_scope": "delegated_workspace",
            "is_delegated_company_access": True,
        }
    }), 200))

@app.route("/api/auth/restore-native-context", methods=["POST", "OPTIONS"])
@require_auth(require_company=False)
def restore_native_context():
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    payload = getattr(request, "jwt_payload", {}) or {}
    app.logger.warning("RESTORE_NATIVE incoming payload=%s", payload)

    user_id = payload.get("user_id") or payload.get("sub")
    user_id = int(user_id) if user_id is not None else None
    if not user_id:
        return _corsify(make_response(jsonify({"error": "AUTH|missing_user_id"}), 401))

    is_delegated = bool(
        payload.get("is_delegated_company_access") or
        payload.get("access_scope") == "delegated_workspace"
    )
    if not is_delegated:
        app.logger.warning("RESTORE_NATIVE rejected: not delegated payload=%s", payload)
        return _corsify(make_response(jsonify({"error": "AUTH|not_in_delegated_workspace"}), 400))

    source_company_id = payload.get("source_company_id")
    try:
        source_company_id = int(source_company_id) if source_company_id is not None else None
    except Exception:
        source_company_id = None

    if not source_company_id:
        app.logger.warning("RESTORE_NATIVE rejected: missing source_company_id payload=%s", payload)
        return _corsify(make_response(jsonify({"error": "AUTH|missing_source_company_id"}), 400))

    base_user = db_service.get_user_by_id(int(user_id)) or {}
    if not base_user:
        app.logger.warning("RESTORE_NATIVE rejected: user not found user_id=%s", user_id)
        return _corsify(make_response(jsonify({"error": "AUTH|user_not_found"}), 404))

    native_ctx = db_service.get_user_context(
        user_id=int(user_id),
        company_id=int(source_company_id),
    )
    if not native_ctx:
        app.logger.warning(
            "RESTORE_NATIVE rejected: no native company access user_id=%s source_company_id=%s",
            user_id,
            source_company_id,
        )
        return _corsify(make_response(jsonify({"error": "AUTH|no_native_company_access"}), 403))

    native_permissions = native_ctx.get("permissions") or {}

    native_token = make_jwt(
        user_id=int(user_id),
        email=base_user.get("email") or "",
        role=native_ctx.get("role") or base_user.get("user_role") or "viewer",
        user_type=base_user.get("user_type") or "Practitioner",
        company_id=int(source_company_id),
        access_scope=native_ctx.get("access_scope") or "assignment",
        allowed_company_ids=native_ctx.get("allowed_company_ids") or [int(source_company_id)],
        permissions=native_permissions,
        first_name=base_user.get("first_name"),
        last_name=base_user.get("last_name"),
        is_delegated_company_access=False,
        is_native_company_member=True,
    )

    try:
        import jwt
        decoded = jwt.decode(native_token, options={"verify_signature": False})
        app.logger.warning("RESTORE_NATIVE issued token payload=%s", decoded)
    except Exception as e:
        app.logger.warning("RESTORE_NATIVE token decode failed error=%s", e)

    app.logger.warning(
        "RESTORE_NATIVE success user_id=%s source_company_id=%s access_scope=%s",
        user_id,
        source_company_id,
        native_ctx.get("access_scope") or "assignment",
    )

    return _corsify(make_response(jsonify({
        "ok": True,
        "token": native_token,
        "company_id": int(source_company_id),
        "access_scope": native_ctx.get("access_scope") or "assignment",
    }), 200))

@app.route("/api/industries", methods=["GET"])
def get_industries_api():
    return jsonify({"industries": list_industries()}), 200


@app.route("/api/companies/<int:company_id>/account_settings", methods=["GET"])
@require_auth
def get_account_settings(company_id: int):
    payload = request.jwt_payload
    if payload.get("company_id") not in (None, company_id):
        return jsonify({"error": "Forbidden"}), 403

    settings = db_service.get_company_account_settings(company_id) or {}

    # You can return all settings, or keep it minimal:
    return jsonify({
        "company_id": company_id,
        "wht_payable_code": settings.get("wht_payable_code") or ""
    }), 200

@app.route("/api/coa/flat", methods=["GET"])
def api_coa_flat():
    industry = (request.args.get("industry") or "").strip()
    sub_industry = (request.args.get("subIndustry") or "").strip() or None

    if not industry:
        return jsonify({"message": "Missing 'industry' query parameter"}), 400

    payload = build_coa_flat(industry, sub_industry)
    if not payload:
        return jsonify({"message": f"No COA template found for industry '{industry}'."}), 404

    return jsonify({"industry": industry, "subIndustry": sub_industry, "rows": payload}), 200

# ✅ UPDATED /api/companies endpoint:
# - accepts profile fields (address/phone/logo)
# - maps name properly
# - passes new args into insert_company
# - (optional) also upserts company_branding so invoices can read it too

@app.route("/api/companies", methods=["POST"])
@require_auth
def create_company_for_user():
    data = request.get_json(silent=True) or {}
    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    user_id = data.get("user_id") or current_user["id"]

    account_type = (data.get("account_type") or "enterprise").lower()
    count = db_service.db_count(
        "SELECT COUNT(*) FROM public.companies WHERE owner_user_id=%s",
        (user_id,),
    ) or 0
    if account_type == "enterprise" and count >= 3:
        return jsonify({"message": "Enterprise plan allows up to 3 companies."}), 403

    try:
        result = create_company_record_from_payload(
            data=data,
            owner_user_id=user_id,
        )
    except ValueError as e:
        msg = e.args[0]
        if isinstance(msg, dict):
            return jsonify(msg), 400
        return jsonify({"message": str(msg)}), 400
    except Exception as e:
        current_app.logger.exception("create_company_for_user failed")
        return jsonify({"message": str(e)}), 500

    return jsonify({
        "message": "Company created",
        "owner_user_id": user_id,
        **result,
    }), 201

@app.route("/api/companies/<int:parent_company_id>/related-companies", methods=["POST"])
@require_auth
def create_related_company(parent_company_id: int):
    data = request.get_json(silent=True) or {}
    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("company_id") != parent_company_id:
        return jsonify({"message": "Not authorised for this company"}), 403

    relationship_type = (data.get("relationship_type") or "").strip().lower()
    if relationship_type not in {"subsidiary", "associate", "joint_venture", "branch"}:
        return jsonify({"message": "Invalid relationship_type"}), 400

    ownership_percent = data.get("ownership_percent")
    voting_percent = data.get("voting_percent")
    effective_from = data.get("effective_from")
    notes = data.get("notes")

    if effective_from:
        try:
            date.fromisoformat(effective_from)
        except ValueError:
            return jsonify({"message": "Invalid 'effective_from'. Expected YYYY-MM-DD."}), 400

    if relationship_type == "subsidiary":
        data.setdefault("entity_kind", "company")
        data.setdefault("consolidation_method", "full")
        data.setdefault("control_basis", "control")
    elif relationship_type == "associate":
        data.setdefault("entity_kind", "company")
        data.setdefault("consolidation_method", "equity")
        data.setdefault("control_basis", "significant_influence")
    elif relationship_type == "joint_venture":
        data.setdefault("entity_kind", "company")
        data.setdefault("consolidation_method", "equity")
        data.setdefault("control_basis", "joint_control")
    elif relationship_type == "branch":
        data.setdefault("entity_kind", "branch_entity")
        data.setdefault("consolidation_method", "full")
        data.setdefault("control_basis", "direct_branch")

    data.setdefault("created_via", "related_company")
    data.setdefault("source_customer_company_id", parent_company_id)
    data.setdefault("provisioning_context", f"related_company:{relationship_type}")

    # ✅ access them AFTER defaults are applied
    control_basis = data.get("control_basis")
    consolidation_method = data.get("consolidation_method")

    auto_membership = relationship_type in {"subsidiary", "branch"}

    try:
        current_app.logger.warning(
            "[CREATE RELATED COMPANY] parent=%s type=%s payload=%s",
            parent_company_id,
            relationship_type,
            data,
        )

        data.setdefault(
            "organizationType",
            data.get("organisationType")
            or data.get("organization_type")
            or data.get("organisation_type")
            or "private_company"
        )

        if relationship_type == "branch":
            data.setdefault("entity_kind", "branch_entity")
            data.setdefault("organizationType", "private_company")
        else:
            data.setdefault("organizationType", "private_company")

        company_result = create_company_record_from_payload(
            data=data,
            owner_user_id=current_user["id"] if auto_membership else None,
            make_primary_membership=False,
            membership_kind="secondary",
        )

        current_app.logger.warning(
            "[COMPANY CREATED] result=%s",
            company_result,
        )

        if auto_membership:
            db_service.fetch_one("""
                INSERT INTO public.company_users (
                    company_id,
                    user_id,
                    role,
                    access_scope,
                    membership_kind,
                    is_primary,
                    is_active,
                    joined_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, TRUE, NOW())
                ON CONFLICT (company_id, user_id) DO UPDATE
                SET
                    role = EXCLUDED.role,
                    access_scope = EXCLUDED.access_scope,
                    membership_kind = EXCLUDED.membership_kind,
                    is_primary = EXCLUDED.is_primary,
                    is_active = TRUE
                RETURNING id;
            """, (
                int(company_result["company_id"]),
                int(current_user["id"]),
                "owner",
                "core",
                "secondary",
                False,
            ))

            current_app.logger.warning(
                "[RELATED COMPANY MEMBERSHIP CREATED] company=%s user=%s kind=secondary",
                company_result["company_id"],
                current_user["id"],
            )

            relationship = db_service.create_company_relationship(
                parent_company_id=parent_company_id,
                child_company_id=company_result["company_id"],
                relationship_type=relationship_type,
                ownership_percent=ownership_percent,
                voting_percent=voting_percent,
                control_basis=control_basis,
                consolidation_method=consolidation_method,
                effective_from=effective_from,
                acquisition_date=data.get("acquisition_date"),
                functional_currency=data.get("functional_currency") or data.get("currency"),
                reporting_currency=data.get("reporting_currency"),
                include_in_group_reporting=bool(data.get("include_in_group_reporting", True)),
                notes=notes,
            )

            rel_id = relationship["id"]

        current_app.logger.warning(
            "[RELATIONSHIP CREATED] rel_id=%s parent=%s child=%s type=%s",
            rel_id,
            parent_company_id,
            company_result["company_id"],
            relationship_type,
        )

    except ValueError as e:
        msg = e.args[0]
        if isinstance(msg, dict):
            return jsonify(msg), 400
        return jsonify({"message": str(msg)}), 400
    except Exception as e:
        current_app.logger.exception("create_related_company failed")
        return jsonify({"message": str(e)}), 500

    return jsonify({
        "message": f"{relationship_type.replace('_', ' ').title()} created",
        "parent_company_id": parent_company_id,
        "relationship_id": rel_id,
        **company_result,
        "relationship_type": relationship_type,
        "ownership_percent": ownership_percent,
        "voting_percent": voting_percent,
        "control_basis": control_basis,
        "consolidation_method": consolidation_method,
        "effective_from": effective_from,
        "notes": notes,
    }), 201

@app.route("/api/companies/<int:company_id>/group-structure", methods=["GET"])
@require_auth
def api_company_group_structure(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("company_id") != company_id:
        return jsonify({"message": "Not authorised for this company"}), 403

    try:
        data = db_service.get_company_group_structure(company_id)
        if not data.get("company"):
            return jsonify({"message": "Company not found"}), 404
        return jsonify(data), 200
    except Exception as e:
        current_app.logger.exception("api_company_group_structure failed")
        return jsonify({"message": str(e)}), 500
    
@app.route("/api/auth/switch-company", methods=["POST", "OPTIONS"])
@require_auth(require_company=False)
def api_auth_switch_company():
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    data = request.get_json(silent=True) or {}
    company_id = data.get("company_id")

    try:
        company_id = int(company_id or 0)
    except Exception:
        return jsonify({"error": "Invalid company_id"}), 400

    if not company_id:
        return jsonify({"error": "company_id is required"}), 400

    user_id = getattr(g, "user_id", None)
    if not user_id:
        return jsonify({"error": "Not authenticated"}), 401

    user = db_service.get_user_by_id(int(user_id))
    if not user:
        return jsonify({"error": "User not found"}), 404

    mem = db_service.fetch_one("""
        SELECT company_id, user_id, role, access_scope, is_active, membership_kind, is_primary
        FROM public.company_users
        WHERE company_id=%s AND user_id=%s
        LIMIT 1;
    """, (int(company_id), int(user_id)))

    if not mem or not bool(mem.get("is_active", True)):
        return jsonify({"error": "User has no access to this company"}), 403

    company = db_service.fetch_one("""
        SELECT
            id,
            name,
            industry,
            sub_industry,
            owner_user_id,
            credit_policy
        FROM public.companies
        WHERE id=%s
        LIMIT 1;
    """, (int(company_id),))

    if not company:
        return jsonify({"error": "Company not found"}), 404

    role = normalize_role(mem.get("role") or "viewer")
    access_scope = (mem.get("access_scope") or "core").strip().lower()
    membership_kind = (mem.get("membership_kind") or "secondary").strip().lower()
    is_primary = bool(mem.get("is_primary", False))
    user_type = (user.get("user_type") or "Enterprise").strip()

    credit_policy = company.get("credit_policy") or {}
    if not isinstance(credit_policy, dict):
        credit_policy = {}

    governance_mode = (credit_policy.get("mode") or "owner_managed").strip().lower()

    dashboards = get_dashboard_access(role, access_scope)
    permissions = build_permissions(role=role, access_scope=access_scope)

    # build all active accessible companies for this user
    allowed_companies, allowed_company_ids, primary_company_id = build_user_allowed_companies(int(user_id))

    active_company_context = None
    for r in allowed_companies:
        if int(r["id"]) == int(company_id):
            active_company_context = r
            break

    if int(company_id) not in allowed_company_ids:
        allowed_company_ids.append(int(company_id))
        allowed_company_ids = sorted(set(allowed_company_ids))

    # Optional legacy convenience only — not source of truth
    try:
        db_service.update_user(int(user_id), company_id=int(company_id))
    except Exception:
        pass

    token = make_jwt(
        user_id=int(user_id),
        email=user.get("email"),
        role=role,
        user_type=user_type,
        company_id=int(company_id),
        access_scope=access_scope,
        allowed_company_ids=allowed_company_ids,
    )

    return jsonify({
        "message": "Active company switched",
        "token": token,
        "company_id": int(company_id),
        "company_name": company.get("name"),
        "industry": company.get("industry"),
        "sub_industry": company.get("sub_industry"),
        "role": role,
        "access_scope": access_scope,
        "primary_company_id": primary_company_id,
        "active_company_id": int(company_id),
        "allowed_companies": allowed_companies,
        "company_context": active_company_context,
        "membership_kind": membership_kind,
        "is_primary": is_primary,
        "user_type": user_type,
        "governance_mode": governance_mode,
        "owner_user_id": company.get("owner_user_id"),
        "allowed_company_ids": allowed_company_ids,
        "dashboards": dashboards,
        "permissions": permissions,
    }), 200

@app.route("/api/companies", methods=["GET"])
@require_auth(require_company=False)
def list_user_companies():
    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    user_id = int(current_user.get("id") or 0)
    user_type = (current_user.get("user_type") or "").strip().lower()

    if not user_id:
        return jsonify({"message": "Missing user id"}), 400

    if user_type == "practitioner":
        rows = db_service.list_workspace_companies_for_user(user_id)
    else:
        rows = db_service.list_companies_for_owner(user_id)

    return jsonify({"companies": rows}), 200


@app.route("/api/companies/<int:company_id>", methods=["GET"])
@require_auth
def api_get_company(company_id: int):
    """
    Return full company profile + operational fields.
    Used by frontend (currency, branding, VAT, inventory, etc).
    """
    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"error": "Not authenticated"}), 401

    # Only allow access to own company (or unbound user during signup flow)
    if current_user.get("company_id") not in (None, company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    try:
        company = db_service.fetch_one(
            """
            SELECT
                c.id,
                c.system_company_code,
                c.name,
                c.slug,
                c.client_code,

                c.industry,
                c.sub_industry,
                c.industry_slug,
                c.sub_industry_slug,

                c.country,
                c.currency,
                c.fin_year_start,
                c.company_reg_date,
                c.first_reporting_period_start,
                c.financial_year_end_month,
                c.financial_year_end_day,
                c.reporting_period_locked,

                c.company_reg_no,
                c.tin,
                c.vat,

                c.company_email,
                c.company_phone,

                c.physical_address,
                c.postal_address,

                c.logo_url,

                c.inventory_mode,
                c.inventory_valuation,
                c.vat_settings,

                c.default_pnl_layout,
                c.pnl_labels_json,

                c.credit_policy,

                -- ✅ owner truth (this fixes your popup logic)
                c.owner_user_id,
                ou.email      AS owner_email,
                ou.first_name AS owner_first_name,
                ou.last_name  AS owner_last_name,

                c.is_active,
                c.created_at
            FROM public.companies c
            LEFT JOIN public.users ou ON ou.id = c.owner_user_id
            WHERE c.id = %s
            LIMIT 1;
            """,
            (company_id,),
        )
    except Exception:
        current_app.logger.exception("Error loading company %s", company_id)
        return jsonify({"error": "Internal server error"}), 500

    if not company:
        return jsonify({"error": "Company not found"}), 404

    cp = company.get("credit_policy") or {}
    if not isinstance(cp, dict):
        cp = {}

    cp.setdefault("mode", "owner_managed")
    cp.setdefault("review_enabled", False)
    cp.setdefault("require_customer_approval", False)
    cp.setdefault("require_kyc", False)

    company["credit_policy"] = cp

    try:
        profile = get_industry_profile(
            company.get("industry"),
            company.get("sub_industry"),
        )
    except Exception:
        profile = {}

    company["industry_profile"] = profile
    company["uses_material_costing"] = bool(profile.get("uses_material_costing"))
    company["uses_boq_budgeting"] = bool(profile.get("uses_boq_budgeting"))
    company["work_unit_label"] = profile.get("work_unit_label") or "Project / Job"

    return jsonify(company), 200

@app.route("/api/companies/<int:company_id>/users/<int:user_id>/role", methods=["PATCH", "OPTIONS"])
@require_auth
def api_update_company_user_role(company_id: int, user_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    actor = getattr(g, "current_user", None) or {}
    actor_role = (actor.get("user_role") or actor.get("role") or "").lower()

    if actor_role not in ("owner", "admin", "cfo"):
        return jsonify({"error": "Not authorised to change user roles"}), 403

    data = request.get_json(silent=True) or {}
    new_role = (data.get("role") or "").strip().lower()

    allowed_roles = {"viewer","clerk","assistant","junior","senior","manager","accountant","cfo","admin","owner"}
    if new_role not in allowed_roles:
        return jsonify({"error": f"Invalid role. Use one of {sorted(allowed_roles)}"}), 400

    # protect owner (optional)
    comp = db_service.fetch_one("SELECT owner_user_id FROM public.companies WHERE id=%s", (int(company_id),)) or {}
    if int(comp.get("owner_user_id") or 0) == int(user_id) and new_role != "owner":
        return jsonify({"error": "Cannot downgrade the company owner"}), 400

    updated = db_service.fetch_one("""
        UPDATE public.company_users
           SET role=%s
         WHERE company_id=%s AND user_id=%s
     RETURNING company_id, user_id, role;
    """, (new_role, int(company_id), int(user_id)))

    if not updated:
        return jsonify({"error": "User is not linked to this company"}), 404

    return jsonify({"message": "Role updated", **updated}), 200

@app.route("/api/companies/<int:company_id>/users", methods=["GET", "OPTIONS"])
@require_auth
def api_list_company_users(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    # require_auth already enforces access to company_id

    # Active memberships
    members = db_service.fetch_all("""
        SELECT
            cu.company_id,
            cu.user_id,
            cu.role,
            cu.access_level,
            cu.is_active,
            cu.joined_at,

            u.email,
            u.first_name,
            u.last_name,
            u.user_type,
            u.is_active AS user_is_active

        FROM public.company_users cu
        JOIN public.users u ON u.id = cu.user_id
        WHERE cu.company_id = %s
        ORDER BY lower(u.email) ASC;
    """, (int(company_id),)) or []

    # Pending invites (optional)
    invites = db_service.fetch_all("""
        SELECT
            id,
            email,
            role,
            created_at,
            expires_at,
            accepted_at,
            revoked_at
        FROM public.company_invites
        WHERE company_id = %s
          AND revoked_at IS NULL
          AND accepted_at IS NULL
        ORDER BY created_at DESC;
    """, (int(company_id),)) or []

    # include owner_user_id so UI knows the real owner
    comp = db_service.fetch_one("""
        SELECT id, owner_user_id
        FROM public.companies
        WHERE id=%s
        LIMIT 1;
    """, (int(company_id),)) or {}

    return jsonify({
        "company_id": int(company_id),
        "owner_user_id": comp.get("owner_user_id"),
        "members": members,
        "invites": invites,
    }), 200

# backend: update_company_profile (FIXED)
# - fixes cp=None flow (no cleaned/merged crash)
# - normalizes ALL boolean keys (string-safe)
# - normalizes nested ppe/leases booleans too
# - reconciles canonical+legacy keys so they cannot contradict
# - merges then normalize_policy() then apply_mode_defaults()

@app.route("/api/companies/<int:company_id>", methods=["PUT", "OPTIONS"])
@require_auth
def update_company_profile(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    payload = request.get_json(silent=True) or {}
    user = getattr(g, "current_user", None) or {}

    if user.get("company_id") not in (None, company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    role = (user.get("role") or user.get("user_role") or "").lower()
    user_id = user.get("id")

    company = db_service.get_company_profile(company_id)
    if not company:
        return jsonify({"error": "Company not found"}), 404

    before_state = company.copy()

    is_owner = (
        company.get("owner_user_id") is not None
        and user_id is not None
        and str(company.get("owner_user_id")) == str(user_id)
    )

    # ─────────────────────────────
    # helpers
    # ─────────────────────────────
    def _as_bool(v) -> bool:
        if isinstance(v, str):
            return v.strip().lower() in ("1", "true", "yes", "on")
        if isinstance(v, (int, float)):
            return v != 0
        return bool(v)

    def _as_num(v):
        try:
            if v is None:
                return None
            if isinstance(v, str):
                v = v.replace(",", "").strip()
            return float(v)
        except Exception:
            return None

    # ─────────────────────────────
    # credit_policy validation + access control
    # ─────────────────────────────
    if "credit_policy" in payload:
        if not (is_owner or role in ("admin", "cfo", "senior")):
            return jsonify({"error": "Not authorised to change credit policy"}), 403

        cp = payload.get("credit_policy")

        existing = company.get("credit_policy") or {}
        if not isinstance(existing, dict):
            existing = {}

        # If caller explicitly sets null -> store null and return (no normalize/apply defaults)
        if cp is None:
            payload["credit_policy"] = None

            updated = db_service.update_company_profile(company_id, payload)
            if not updated:
                return jsonify({"error": "Update failed"}), 400

            after_state = db_service.get_company_profile(company_id) or {}
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int((getattr(g, "current_user", None) or {}).get("id") or 0),
                module="company",
                action="update",
                severity="info",
                entity_type="company",
                entity_id=str(company_id),
                entity_ref=after_state.get("company_name") or after_state.get("name"),
                before_json=before_state,
                after_json=after_state,
                message="Company profile updated (credit_policy cleared)",
                source="api",
            )
            return jsonify(after_state), 200

        if not isinstance(cp, dict):
            return jsonify({"error": "credit_policy must be an object"}), 400

        allowed_keys = {
            # core
            "mode",
            "review_enabled",
            "ap_review_enabled",
            "review_choice_made",   # ✅ ADD THIS
            "ap_auto_post",
            "payment_workflow_enabled",
            "require_payment_approval",
            "require_customer_approval",
            "require_kyc",
            "require_vendor_kyc_on_release",

            # numbers
            "default_credit_limit",
            "approval_threshold",
            "kyc_threshold",
            "clerk_request_cap",
            "require_kyc_for_any_credit",

            # legacy flags (allowed, but reconciled)
            "invoice_review_enabled",
            "require_invoice_review",
            "bill_review_enabled",
            "require_bill_review",
            "require_quote_issue_review",
            "require_quote_accept_review",

            # nested policy blocks
            "ppe",
            "leases",
        }

        cleaned = {k: cp.get(k) for k in allowed_keys if k in cp}

        # --- normalize + validate mode ---
        if "mode" in cleaned:
            cleaned["mode"] = str(cleaned["mode"] or "").strip().lower().replace("-", "_").replace(" ", "_")
            valid_modes = {"owner_managed", "assisted", "controlled"}
            if cleaned["mode"] and cleaned["mode"] not in valid_modes:
                return jsonify({"error": f"Invalid mode. Use one of {sorted(valid_modes)}"}), 400

        if "review_choice_made" in cleaned:
            cleaned["review_choice_made"] = bool(cleaned["review_choice_made"])
            
        # --- numeric normalization ---
        for nk in ("default_credit_limit", "approval_threshold", "kyc_threshold", "clerk_request_cap"):
            if nk in cleaned:
                cleaned[nk] = _as_num(cleaned[nk])

        if "require_kyc_for_any_credit" in cleaned:
            cleaned["require_kyc_for_any_credit"] = _as_bool(cleaned["require_kyc_for_any_credit"])

        # --- boolean normalization (top-level) ---
        BOOL_KEYS = {
            "review_enabled",
            "ap_review_enabled",
            "ap_auto_post",
            "payment_workflow_enabled",
            "require_payment_approval",
            "require_customer_approval",
            "require_kyc",
            "require_vendor_kyc_on_release",
            "invoice_review_enabled",
            "require_invoice_review",
            "bill_review_enabled",
            "require_bill_review",
            "require_quote_issue_review",
            "require_quote_accept_review",
        }
        for bk in list(cleaned.keys()):
            if bk in BOOL_KEYS:
                cleaned[bk] = _as_bool(cleaned[bk])

        # --- normalize nested booleans if provided ---
        if "ppe" in cleaned:
            if not isinstance(cleaned["ppe"], dict):
                return jsonify({"error": "credit_policy.ppe must be an object"}), 400
            PPE_BOOL = {
                "review_enabled",
                "require_asset_master_review",
                "require_depreciation_review",
                "require_disposal_review",
                "require_impairment_review",
                "require_revaluation_review",
                "require_hfs_review",
                "require_transfers_review",
                "usage_requires_approval",
            }
            for bk in list(cleaned["ppe"].keys()):
                if bk in PPE_BOOL:
                    cleaned["ppe"][bk] = _as_bool(cleaned["ppe"][bk])

        if "leases" in cleaned:
            if not isinstance(cleaned["leases"], dict):
                return jsonify({"error": "credit_policy.leases must be an object"}), 400
            LEASE_BOOL = {
                "review_enabled",
                "require_lease_create_review",
                "require_monthly_posting_review",
                "require_payment_review",
                "require_modification_review",
                "require_termination_review",
                "auto_post_monthly_when_review_off",
                "auto_post_payments_when_review_off",
                # accept nested legacy too if UI sends them
                "require_lease_monthly_review",
                "require_lease_payment_review",
                "require_lease_modification_review",
                "require_lease_termination_review",
            }
            for bk in list(cleaned["leases"].keys()):
                if bk in LEASE_BOOL:
                    cleaned["leases"][bk] = _as_bool(cleaned["leases"][bk])

        # --- reconcile canonical vs legacy so they can't contradict ---
        # If master AR review is set, force legacy mirrors to match
        if "review_enabled" in cleaned:
            cleaned["invoice_review_enabled"] = bool(cleaned["review_enabled"])
            cleaned["require_invoice_review"] = bool(cleaned["review_enabled"])

        # If AP review set, force bill legacy mirrors
        if "ap_review_enabled" in cleaned:
            cleaned["bill_review_enabled"] = bool(cleaned["ap_review_enabled"])
            cleaned["require_bill_review"] = bool(cleaned["ap_review_enabled"])

        # If legacy arrives without canonical, "promote" it
        if "review_enabled" not in cleaned and ("invoice_review_enabled" in cleaned or "require_invoice_review" in cleaned):
            cleaned["review_enabled"] = bool(cleaned.get("invoice_review_enabled") or cleaned.get("require_invoice_review"))

        if "ap_review_enabled" not in cleaned and ("bill_review_enabled" in cleaned or "require_bill_review" in cleaned):
            cleaned["ap_review_enabled"] = bool(cleaned.get("bill_review_enabled") or cleaned.get("require_bill_review"))

        # merge -> normalize -> apply mode defaults
        merged = {**existing, **cleaned}
        merged = normalize_policy(merged)

        # ✅ capture "supplied keys" from PATCH before defaults decide what to override
        supplied_top = set(cleaned.keys())

        supplied_ppe = set()
        if isinstance(cp.get("ppe"), dict):
            supplied_ppe = set(cp["ppe"].keys())

        supplied_leases = set()
        if isinstance(cp.get("leases"), dict):
            supplied_leases = set(cp["leases"].keys())

        merged, warnings = apply_mode_defaults(
            merged,
            supplied_top=supplied_top,
            supplied_ppe=supplied_ppe,
            supplied_leases=supplied_leases,
        )

        payload["credit_policy"] = merged

        updated = db_service.update_company_profile(company_id, payload)
        if not updated:
            return jsonify({"error": "Update failed"}), 400

        after_state = db_service.get_company_profile(company_id) or {}

        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int((getattr(g, "current_user", None) or {}).get("id") or 0),
            module="company",
            action="update",
            severity="info",
            entity_type="company",
            entity_id=str(company_id),
            entity_ref=after_state.get("company_name") or after_state.get("name"),
            before_json=before_state,
            after_json=after_state,
            message="Company profile updated",
            source="api",
        )

        after_state["_warnings"] = warnings
        return jsonify(after_state), 200

    # no credit_policy in payload -> fall back to generic update (optional)
    try:
        updated = db_service.update_company_profile(company_id, payload)

        if not updated:
            return jsonify({
                "error": "update_company_profile returned False"
            }), 400

        after_state = db_service.get_company_profile(company_id) or {}
        return jsonify(after_state), 200

    except Exception as e:
        current_app.logger.exception(
            "Company update failed for company %s",
            company_id
        )

        return jsonify({
            "error": str(e)
        }), 400


@app.route("/api/companies/<int:company_id>/logo/file", methods=["GET"])
def api_company_logo_file(company_id: int):
    upload_dir = Path(current_app.root_path) / "static" / "uploads" / f"company_{company_id}"

    for ext in ALLOWED_EXT:
        path = upload_dir / f"logo{ext}"
        if path.exists():
            return send_from_directory(
                str(upload_dir),
                path.name,
                conditional=True,
                max_age=3600,
            )

    return jsonify({"error": "Logo not found"}), 404

@app.route("/api/companies/<int:company_id>/logo", methods=["POST", "OPTIONS"])
@require_auth
def api_upload_company_logo(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", None) or {}
    if user.get("company_id") not in (None, company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    f = request.files.get("logo")
    if not f or not f.filename:
        return jsonify({"error": "Missing file field 'logo'"}), 400

    filename = secure_filename(f.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXT:
        return jsonify({"error": f"Unsupported file type {ext}"}), 400

    # Save path (serving from /static/uploads)
    root = Path(current_app.root_path)

    upload_dir = (
        root
        / "static"
        / "uploads"
        / f"company_{company_id}"
    )

    upload_dir.mkdir(parents=True, exist_ok=True)
    # always write as logo.png/jpg (simple)
    out_name = f"logo{ext}"
    out_path = upload_dir / out_name
    f.save(str(out_path))

    logo_url = f"/api/companies/{company_id}/logo/file"

    # store in DB
    db_service.execute_sql(
        "UPDATE public.companies SET logo_url=%s WHERE id=%s;",
        (logo_url, company_id),
    )

    db_service.audit_log(
        company_id=company_id,
        actor_user_id=int((getattr(g, "current_user", None) or {}).get("id") or 0),
        module="company",
        action="upload_logo",
        severity="info",
        entity_type="company",
        entity_id=str(company_id),
        entity_ref=f"company_{company_id}",
        after_json={"logo_url": logo_url},
        message="Company logo uploaded",
        source="api",
    )

    return jsonify({"ok": True, "logo_url": logo_url}), 200

@app.route("/api/companies/<int:company_id>/logo", methods=["DELETE", "OPTIONS"])
@require_auth
def api_delete_company_logo(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", None) or {}
    if user.get("company_id") not in (None, company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    upload_dir = Path(current_app.root_path) / "static" / "uploads" / f"company_{company_id}"

    for ext in ALLOWED_EXT:
        path = upload_dir / f"logo{ext}"
        if path.exists():
            try:
                path.unlink()
            except OSError:
                current_app.logger.exception(
                    "Could not delete company logo company_id=%s path=%s",
                    company_id,
                    path,
                )

    db_service.execute_sql(
        "UPDATE public.companies SET logo_url=NULL WHERE id=%s;",
        (company_id,),
    )

    db_service.audit_log(
        company_id=company_id,
        actor_user_id=int((getattr(g, "current_user", None) or {}).get("id") or 0),
        module="company",
        action="delete_logo",
        severity="info",
        entity_type="company",
        entity_id=str(company_id),
        entity_ref=f"company_{company_id}",
        after_json={"logo_url": None},
        message="Company logo removed",
        source="api",
    )

    return jsonify({"ok": True}), 200


@app.route("/api/companies/<int:company_id>/period-range", methods=["GET"])
@require_auth
def api_company_period_range(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    try:
        date_from, date_to, meta = resolve_company_period(
            db_service,
            company_id,
            request,
            mode="range",
        )

        return jsonify({
            "from": date_from.isoformat() if date_from else None,
            "to": date_to.isoformat() if date_to else None,
            "label": meta.get("label"),
            "preset": meta.get("preset"),
            "period": meta.get("period"),
            "fin_year_start": meta.get("fin_year_start"),
            "first_reporting_period_start": meta.get("first_reporting_period_start"),
            "financial_year_end_month": meta.get("financial_year_end_month"),
            "financial_year_end_day": meta.get("financial_year_end_day"),
            "is_custom_range": meta.get("is_custom_range"),
        }), 200

    except Exception as e:
        current_app.logger.exception("Error resolving company period range: %s", e)
        return jsonify({"error": str(e)}), 500
# ────────────────────────────────────────────────────────────────
# POS Summaries API 
# ────────────────────────────────────────────────────────────────

@app.route("/api/companies/<int:company_id>/pos_summaries", methods=["POST"])
@require_auth
def create_pos_summary(company_id: int):
    """
    Create a POS summary + lines for a given company.
    """

    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    try:
        payload = request.get_json(silent=True) or {}
    except Exception:
        return jsonify({"error": "Invalid JSON"}), 400

    pos_date = payload.get("pos_date")
    source_name = (payload.get("source_name") or "").strip()
    lines = payload.get("lines") or []

    if not pos_date:
        return jsonify({"error": "pos_date is required (YYYY-MM-DD)"}), 400
    if not lines:
        return jsonify({"error": "lines[] is required"}), 400

    # -------------------------
    # Compute totals
    # -------------------------
    gross_total = 0.0
    net_total = 0.0
    vat_total = 0.0

    for ln in lines:
        gross_total += float(ln.get("gross_amount") or 0)
        net_total += float(ln.get("net_amount") or 0)
        vat_total += float(ln.get("vat_amount") or 0)

    # -------------------------
    # Insert header
    # -------------------------
    summary_id = db_service.insert_pos_summary(
        company_id=company_id,
        pos_date=pos_date,
        source_name=source_name,
        gross_amount=gross_total,
        net_amount=net_total,
        vat_amount=vat_total,
    )

    if not summary_id:
        return jsonify({"error": "Failed to create POS summary"}), 500

    # -------------------------
    # Insert lines
    # -------------------------
    # -------------------------
    # Insert lines
    # -------------------------
    lines_inserted = db_service.insert_pos_summary_lines(
        company_id=company_id,
        summary_id=summary_id,
        lines=lines,
    )

    # ✅ AUDIT HERE (after header+lines exist)
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=int(user.get("id") or 0),
        module="pos",
        action="create_summary",
        severity="info",
        entity_type="pos_summary",
        entity_id=str(summary_id),
        entity_ref=(source_name or f"pos_summary_{summary_id}"),
        amount=float(gross_total or 0.0),
        currency=(payload.get("currency") or None),  # if you have it, otherwise None
        before_json={},
        after_json={
            "summary_id": int(summary_id),
            "pos_date": pos_date,
            "source_name": source_name,
            "gross_amount": gross_total,
            "net_amount": net_total,
            "vat_amount": vat_total,
            "lines_inserted": int(lines_inserted or 0),
        },
        message=f"POS summary created with {int(lines_inserted or 0)} lines",
        source="api",
    )

    return jsonify({
        "summary_id": summary_id,
        "lines_inserted": lines_inserted,
        "pos_date": pos_date,
        "source_name": source_name,
        "gross_amount": gross_total,
        "net_amount": net_total,
        "vat_amount": vat_total,
    }), 201


@app.route("/api/companies/<int:company_id>/pos_summaries", methods=["GET"])
@require_auth
def list_pos_summaries(company_id: int):
    """
    List latest POS summaries (header only).
    Optional: ?limit=20
    """

    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    try:
        limit = int(request.args.get("limit", 50))
    except ValueError:
        limit = 50

    rows = db_service.list_pos_summaries(company_id, limit=limit)
    return jsonify({"items": rows})


@app.route("/api/companies/<int:company_id>/pos_summaries/<int:summary_id>", methods=["GET"])
@require_auth
def get_pos_summary(company_id: int, summary_id: int):
    """
    Get one POS summary with its lines.
    """

    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    rec = db_service.get_pos_summary_with_lines(company_id, summary_id)
    if not rec:
        return jsonify({"error": "POS summary not found"}), 404

    return jsonify(rec)

def build_pos_journal_lines(pos: dict, company_id: int, db_service) -> dict:
    """
    POS posting model:
      DR  Cash / Bank / Undeposited Funds (gross)
      CR  Sales (net) - grouped by income_account
      CR  VAT Output (vat) - to control account
    """
    from decimal import Decimal, ROUND_HALF_UP

    def money(x) -> float:
        return float(Decimal(str(x or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    lines = pos.get("lines") or []
    if not lines:
        raise ValueError("POS summary has no lines")

    # ✅ Spot A: read header totals (may be 0)
    gross_total = money(pos.get("gross_amount") or 0)
    net_total   = money(pos.get("net_amount") or 0)
    vat_total   = money(pos.get("vat_amount") or 0)

    # ✅ Spot B: fallback to line totals if header gross missing/0
    if gross_total <= 0:
        gross_total = money(sum(float(l.get("gross_amount") or 0) for l in lines))
        net_total   = money(sum(float(l.get("net_amount") or 0) for l in lines))
        vat_total   = money(sum(float(l.get("vat_amount") or 0) for l in lines))

    # ✅ Control accounts (POS cash/bank + VAT)
    CASH_ACCOUNT = (db_service.get_control_account_code(company_id, "POS_CASH_CONTROL") or "").strip()
    if not CASH_ACCOUNT:
        CASH_ACCOUNT = (db_service.get_control_account_code(company_id, "BANK_CONTROL") or "").strip()

    VAT_ACCOUNT = (db_service.get_control_account_code(company_id, "VAT_OUTPUT") or "").strip()

    if not CASH_ACCOUNT:
        raise ValueError("POS cash control account not configured (POS_CASH_CONTROL or BANK_CONTROL)")
    if vat_total > 0 and not VAT_ACCOUNT:
        raise ValueError("VAT output account not configured (VAT_OUTPUT)")

    # ✅ Group sales by income account
    income_map = {}  # avoid Pylance noise if you're not on py>=3.9 typing
    for ln in lines:
        acc = (ln.get("income_account") or "").strip()
        if not acc:
            raise ValueError(f"POS line missing income_account: {ln}")

        amt = money(ln.get("net_amount") or 0.0)
        if amt <= 0:
            continue

        income_map[acc] = money(float(income_map.get(acc, 0.0)) + amt)

    # ✅ Spot C: audit check (header net vs grouped net)
    net_from_lines = money(sum(float(v or 0) for v in income_map.values()))
    if net_total > 0 and abs(net_from_lines - net_total) > 0.01:
        raise ValueError(f"POS net mismatch: header={net_total} lines={net_from_lines}")

    jlines = []

    # DR Cash/Bank = gross
    jlines.append({"account_code": CASH_ACCOUNT, "dc": "D", "amount": gross_total})

    # CR Sales = net (split by income account)
    for acc, amt in income_map.items():
        if amt > 0:
            jlines.append({"account_code": acc, "dc": "C", "amount": amt})

    # CR VAT Output = vat
    if vat_total > 0:
        jlines.append({"account_code": VAT_ACCOUNT, "dc": "C", "amount": vat_total})

    # ✅ Balance check (tolerance 0.01)
    deb = money(sum(float(x["amount"]) for x in jlines if x["dc"] == "D"))
    cre = money(sum(float(x["amount"]) for x in jlines if x["dc"] == "C"))
    if abs(deb - cre) > 0.01:
        raise ValueError(f"POS journal not balanced: DR={deb} CR={cre}")

    return {"lines": jlines}

@app.route("/api/companies/<int:company_id>/pos_summaries/<int:summary_id>/post", methods=["POST"])
@require_auth
def post_pos_summary(company_id: int, summary_id: int):
    try:
        user = getattr(g, "current_user", {}) or {}
        if user.get("company_id") != company_id:
            return jsonify({"error": "Not authorised for this company"}), 403

        pos = db_service.get_pos_summary_with_lines(company_id, summary_id)
        if not pos:
            return jsonify({"error": "POS summary not found"}), 404

        if pos.get("posted_journal_id"):
            pos["_posted_journal_id"] = int(pos["posted_journal_id"])
            return jsonify(pos), 200

        payload = build_pos_journal_lines(pos, company_id, db_service)
        journal_id = db_service.post_pos_summary_to_gl(company_id, summary_id, payload["lines"])

        payload = build_pos_journal_lines(pos, company_id, db_service)
        journal_id = db_service.post_pos_summary_to_gl(company_id, summary_id, payload["lines"])

        # ✅ AUDIT HERE (after GL post succeeds)
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="pos",
            action="post_to_gl",
            severity="info",
            entity_type="pos_summary",
            entity_id=str(summary_id),
            entity_ref=(pos.get("source_name") or f"pos_summary_{summary_id}"),
            journal_id=int(journal_id),
            amount=float(pos.get("gross_amount") or 0.0),
            currency=(pos.get("currency") or None),
            before_json={
                "posted_journal_id": pos.get("posted_journal_id"),
            },
            after_json={
                "posted_journal_id": int(journal_id),
                "lines_count": len(payload.get("lines") or []),
            },
            message=f"POS summary posted to GL. journal_id={int(journal_id)}",
            source="api",
        )

        posted = db_service.get_pos_summary_with_lines(company_id, summary_id) or {}
        posted["_posted_journal_id"] = int(journal_id)
        return jsonify(posted), 200

    except Exception as e:
        current_app.logger.exception("post_pos_summary failed")
        return jsonify({"error": "Server error", "detail": str(e)}), 500

@app.route("/api/companies/<int:company_id>/pos_imports", methods=["POST"])
@require_auth
def upload_pos_roll(company_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != int(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Missing file (field name must be 'file')"}), 400

    source_name = (request.form.get("source_name") or request.args.get("source_name") or "").strip()

    original = secure_filename(file.filename or "")
    ext = os.path.splitext(original)[1].lower()
    if ext not in ALLOWED_POS_EXTS:
        return jsonify({"error": "Invalid file type", "allowed": sorted(ALLOWED_POS_EXTS)}), 400

    # 1) compute hash (for dedupe)
    sha = hashlib.sha256()
    file.stream.seek(0)
    for chunk in iter(lambda: file.stream.read(1024 * 1024), b""):
        sha.update(chunk)
    file_hash = sha.hexdigest()
    file.stream.seek(0)

    # 2) save file
    folder = os.path.join(current_app.root_path, "uploads", "pos_rolls", f"company_{company_id}")
    os.makedirs(folder, exist_ok=True)

    filename = f"pos_{int(company_id)}_{int(time.time())}_{file_hash[:12]}{ext}"
    path = os.path.join(folder, filename)
    file.save(path)

    # 3) insert upload tracking row
    schema = f"company_{company_id}"
    import_id = db_service.execute_sql(
        f"""
        INSERT INTO {schema}.pos_imports
            (company_id, source_name, file_name, file_hash, status, created_by)
        VALUES
            (%s, %s, %s, %s, 'uploaded', %s)
        RETURNING id;
        """,
        (int(company_id), source_name or None, filename, file_hash, user.get("id")),
    )

    if not import_id:
        return jsonify({"error": "Failed to create pos_imports row"}), 500

    # ✅ AUDIT HERE (after pos_imports row exists)
    db_service.audit_log(
        company_id=int(company_id),
        actor_user_id=int(user.get("id") or 0),
        module="pos",
        action="upload_roll",
        severity="info",
        entity_type="pos_import",
        entity_id=str(int(import_id)),
        entity_ref=filename,
        amount=0.0,
        currency=None,
        before_json={},
        after_json={
            "import_id": int(import_id),
            "source_name": source_name,
            "file_name": filename,
            "file_hash": file_hash,
            "status": "uploaded",
        },
        message="POS roll uploaded",
        source="api",
    )

    file_url = f"/api/companies/{company_id}/pos_imports/{import_id}/file"

    return jsonify({
        "ok": True,
        "import_id": int(import_id),
        "status": "uploaded",
        "source_name": source_name,
        "file_name": filename,
        "file_hash": file_hash,
        "file_url": file_url,
    }), 201


@app.route("/api/companies/<int:company_id>/pos_imports/<int:import_id>/file", methods=["GET"])
@require_auth
def download_pos_roll(company_id: int, import_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != int(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    schema = f"company_{company_id}"
    row = db_service.fetch_one(
        f"""
        SELECT id, file_name
        FROM {schema}.pos_imports
        WHERE id=%s AND company_id=%s
        LIMIT 1;
        """,
        (int(import_id), int(company_id)),
    ) or {}

    filename = (row.get("file_name") or "").strip()
    if not filename:
        return jsonify({"error": "POS import not found"}), 404

    folder = os.path.join(current_app.root_path, "uploads", "pos_rolls", f"company_{company_id}")
    return send_from_directory(folder, filename, as_attachment=True)

@app.route("/api/companies/<int:company_id>/pos_imports/<int:import_id>/parse", methods=["POST"])
@require_auth
def parse_pos_import(company_id: int, import_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    imp = db_service.get_pos_import(company_id, import_id)
    if not imp:
        return jsonify({"error": "Import not found"}), 404

    mapping = imp.get("mapping_json") or {}
    if not mapping:
        return jsonify({"error": "No mapping saved yet. Call PUT .../mapping first."}), 409

    file_name = imp.get("file_name")
    folder = pos_upload_folder(current_app.root_path)
    path = os.path.join(folder, file_name)
    if not os.path.exists(path):
        return jsonify({"error": "Stored import file not found"}), 404

    ext = os.path.splitext(file_name)[1].lower()
    b = open(path, "rb").read()

    # Read ALL rows (careful for huge files later; you can stream)
    if ext == ".csv":
        prev = preview_csv_bytes(b, max_rows=10**9)   # reuse: but this loads; later optimize
        rows = prev.get("rows") or []
    else:
        # xlsx
        wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return jsonify({"error": "Empty sheet"}), 400
        columns = [str(x).strip() if x is not None else "" for x in header]
        columns = [c if c else f"col_{i+1}" for i, c in enumerate(columns)]
        rows = []
        for r in it:
            obj = {}
            for j, c in enumerate(columns):
                obj[c] = r[j] if j < len(r) else None
            rows.append(obj)

    if not rows:
        return jsonify({"error": "No data rows"}), 400

    # Determine pos_date
    pos_date_fixed = mapping.get("pos_date_fixed")  # optional: allow fixed date
    pos_date = to_date_safe(pos_date_fixed)
    if not pos_date:
        # try from file column
        dcol = mapping.get("pos_date")
        if dcol:
            # find first date-like value
            for r in rows[:50]:
                dv = pick_col(r, dcol)
                dd = to_date_safe(dv)
                if dd:
                    pos_date = dd
                    break
    if not pos_date:
        # fallback: today
        pos_date = date.today()

    # Normalize lines
    norm_lines = []
    for i, r in enumerate(rows, start=1):
        item_code = pick_col(r, mapping.get("item_code"))
        desc      = pick_col(r, mapping.get("description"))
        qty       = to_float_safe(pick_col(r, mapping.get("qty")), 0.0)
        unit_price= to_float_safe(pick_col(r, mapping.get("unit_price")), 0.0)

        gross = to_float_safe(pick_col(r, mapping.get("gross_amount")), 0.0)
        net   = to_float_safe(pick_col(r, mapping.get("net_amount")), 0.0)
        vat   = to_float_safe(pick_col(r, mapping.get("vat_amount")), 0.0)
        vat_rate = to_float_safe(pick_col(r, mapping.get("vat_rate")), 0.0)

        # compute missing fields
        if gross <= 0 and net > 0:
            gross = net + vat
        if net <= 0 and gross > 0:
            # if vat present
            if vat > 0:
                net = gross - vat
            elif vat_rate > 0:
                net = gross / (1 + vat_rate / 100.0)
                vat = gross - net
        if vat <= 0 and gross > 0 and net > 0:
            vat = gross - net
        if vat <= 0 and net > 0 and vat_rate > 0:
            vat = net * (vat_rate / 100.0)
            gross = net + vat

        # skip empty rows
        if (qty == 0 and gross == 0 and net == 0 and not item_code and not desc):
            continue

        norm_lines.append({
            "line_no": i,
            "item_code": (str(item_code).strip() if item_code is not None else None),
            "description": (str(desc).strip() if desc is not None else None),
            "qty": qty,
            "unit_price": unit_price,
            "gross_amount": gross,
            "net_amount": net,
            "vat_amount": vat,
            # accounts can be mapped or defaulted later
            "income_account": mapping.get("income_account_default"),
            "cogs_account": mapping.get("cogs_account_default"),
            "inventory_account": mapping.get("inventory_account_default"),
            "item_type": None,
            "item_id": None,
        })

    if not norm_lines:
        return jsonify({"error": "No usable lines after parsing"}), 400

    gross_total = sum(float(l.get("gross_amount") or 0) for l in norm_lines)
    net_total   = sum(float(l.get("net_amount") or 0) for l in norm_lines)
    vat_total   = sum(float(l.get("vat_amount") or 0) for l in norm_lines)

    summary_id = db_service.insert_pos_summary(
        company_id=company_id,
        pos_date=pos_date,
        source_name=imp.get("source_name") or mapping.get("source_name") or "POS Import",
        gross_amount=gross_total,
        net_amount=net_total,
        vat_amount=vat_total,
    )
    if not summary_id:
        return jsonify({"error": "Failed to create POS summary"}), 500

    db_service.insert_pos_summary_lines(
        company_id=company_id,
        summary_id=summary_id,
        lines=norm_lines,
    )

    # stamp import status
    schema = f"company_{company_id}"
    with db_service._conn_cursor() as (conn, cur):
        cur.execute(
            f"""
            UPDATE {schema}.pos_imports
            SET status='parsed',
                parsed_summary_id=%s,
                error=NULL
            WHERE id=%s AND company_id=%s;
            """,
            (int(summary_id), int(import_id), int(company_id)),
        )
        conn.commit()

    # ✅ AUDIT HERE (after all DB writes succeeded)
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=int(user.get("id") or 0),
        module="pos",
        action="parse_import",
        severity="info",
        entity_type="pos_import",
        entity_id=str(int(import_id)),
        entity_ref=(imp.get("file_name") or f"pos_import_{import_id}"),
        # link the created object
        amount=float(gross_total or 0.0),
        currency=(mapping.get("currency") or None),
        before_json={
            "status": imp.get("status"),
            "parsed_summary_id": imp.get("parsed_summary_id"),
        },
        after_json={
            "status": "parsed",
            "parsed_summary_id": int(summary_id),
            "pos_date": str(pos_date),
            "totals": {"gross": gross_total, "net": net_total, "vat": vat_total},
            "lines_count": len(norm_lines),
        },
        message=f"POS import parsed into summary_id={int(summary_id)}",
        source="api",
    )

    return jsonify({
        "ok": True,
        "import_id": int(import_id),
        "status": "parsed",
        "summary_id": int(summary_id),
        "pos_date": str(pos_date),
        "totals": {"gross": gross_total, "net": net_total, "vat": vat_total},
        "lines_count": len(norm_lines),
    }), 200


@app.route("/api/companies/<int:company_id>/pos_imports/<int:import_id>/preview", methods=["GET"])
@require_auth
def preview_pos_import(company_id: int, import_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    imp = db_service.get_pos_import(company_id, import_id)
    if not imp:
        return jsonify({"error": "Import not found"}), 404

    file_name = imp.get("file_name")
    if not file_name:
        return jsonify({"error": "Import missing file_name"}), 500

    folder = pos_upload_folder(current_app.root_path)
    path = os.path.join(folder, file_name)
    if not os.path.exists(path):
        return jsonify({"error": "Stored import file not found"}), 404

    ext = os.path.splitext(file_name)[1].lower()
    b = open(path, "rb").read()

    if ext == ".csv":
        prev = preview_csv_bytes(b, max_rows=20)
    elif ext in (".xlsx", ".xls"):
        prev = preview_xlsx_bytes(b, max_rows=20)
    else:
        return jsonify({"error": "Preview not supported for this file type", "ext": ext}), 400

    columns = prev.get("columns") or []
    guess = guess_mapping_from_columns(columns)

    out = {
        "import_id": int(import_id),
        "file_name": file_name,
        "source_name": imp.get("source_name"),
        "columns": columns,
        "sample_rows": prev.get("rows") or [],
        "guessed_mapping": guess,
        "existing_mapping": imp.get("mapping_json") or None,
    }

    # optional: store preview_json for later
    try:
        schema = f"company_{company_id}"
        with db_service._conn_cursor() as (conn, cur):
            cur.execute(
                f"UPDATE {schema}.pos_imports SET preview_json=%s WHERE id=%s AND company_id=%s;",
                (psycopg2.extras.Json(out), int(import_id), int(company_id)),
            )
            conn.commit()
    except Exception:
        pass

    return jsonify(out), 200

@app.route("/api/companies/<int:company_id>/pos_imports/<int:import_id>/mapping", methods=["PUT"])
@require_auth
def save_pos_import_mapping(company_id: int, import_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    imp = db_service.get_pos_import(company_id, import_id)
    if not imp:
        return jsonify({"error": "Import not found"}), 404

    payload = request.get_json(silent=True) or {}
    mapping = payload.get("mapping")
    if not isinstance(mapping, dict) or not mapping:
        return jsonify({"error": "mapping is required and must be an object"}), 400

    # ✅ do update
    db_service.set_pos_import_mapping(company_id, import_id, mapping)

    # ✅ AUDIT HERE
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=int(user.get("id") or 0),
        module="pos",
        action="save_mapping",
        severity="info",
        entity_type="pos_import",
        entity_id=str(int(import_id)),
        entity_ref=(imp.get("file_name") or f"pos_import_{import_id}"),
        before_json={"mapping_json": imp.get("mapping_json") or {}},
        after_json={"mapping_json": mapping},
        message="POS import mapping saved",
        source="api",
    )

    return jsonify({
        "ok": True,
        "import_id": int(import_id),
        "mapping": mapping,
    }), 200

@app.route("/api/companies/<int:company_id>/pos_imports", methods=["POST"])
@require_auth
def upload_pos_import(company_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Missing file (field name must be 'file')"}), 400

    source_name = (request.form.get("source_name") or "").strip() or None

    original = secure_filename(file.filename or "")
    ext = os.path.splitext(original)[1].lower()

    # allow upload of anything if you want, but we only preview/parse csv/xlsx
    if ext not in ALLOWED_POS_EXTS:
        return jsonify({"error": "Unsupported file type for POS import", "allowed": sorted(ALLOWED_POS_EXTS)}), 400

    b = file.read()
    if not b:
        return jsonify({"error": "Empty file"}), 400

    file_hash = sha256_bytes(b)

    folder = pos_upload_folder(current_app.root_path)
    # store with import-hash naming to avoid duplicates
    filename = f"pos_{company_id}_{int(time.time())}_{file_hash[:12]}{ext}"
    path = os.path.join(folder, filename)
    with open(path, "wb") as f:
        f.write(b)

    import_id = db_service.insert_pos_import(
        company_id,
        source_name=source_name,
        file_name=filename,
        file_hash=file_hash,
        created_by=user.get("id"),
        status="uploaded",
    )
    if not import_id:
        return jsonify({"error": "Failed to create import row"}), 500

    # ✅ AUDIT HERE
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=int(user.get("id") or 0),
        module="pos",
        action="upload_import",
        severity="info",
        entity_type="pos_import",
        entity_id=str(int(import_id)),
        entity_ref=filename,
        before_json={},
        after_json={
            "import_id": int(import_id),
            "file_name": filename,
            "file_hash": file_hash,
            "source_name": source_name,
            "status": "uploaded",
        },
        message="POS import uploaded",
        source="api",
    )

    return jsonify({
        "ok": True,
        "import_id": int(import_id),
        "file_name": filename,
        "file_hash": file_hash,
        "source_name": source_name,
        "status": "uploaded",
    }), 201

@app.route("/uploads/pos_imports/<path:filename>", methods=["GET"])
@require_auth
def serve_pos_import_file(filename):
    # optional: you may want stricter auth here (company ownership)
    folder = pos_upload_folder(current_app.root_path)
    return send_from_directory(folder, filename)

@app.route("/api/companies/<int:company_id>/reports/pos_sales_by_item", methods=["GET"])
@require_auth
def report_pos_sales_by_item(company_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != int(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    date_from = request.args.get("from")
    date_to   = request.args.get("to")
    source    = (request.args.get("source_name") or "").strip()

    schema = f"company_{company_id}"

    where = ["1=1"]
    params = []

    if date_from:
        where.append("s.pos_date >= %s")
        params.append(date_from)
    if date_to:
        where.append("s.pos_date <= %s")
        params.append(date_to)
    if source:
        where.append("lower(coalesce(s.source_name,'')) = lower(%s)")
        params.append(source)

    sql = f"""
    SELECT
      coalesce(l.item_code,'') as item_code,
      coalesce(l.description,'') as description,
      sum(l.qty) as qty,
      sum(l.gross_amount) as gross_amount,
      sum(l.net_amount) as net_amount,
      sum(l.vat_amount) as vat_amount
    FROM {schema}.pos_summary_lines l
    JOIN {schema}.pos_summaries s ON s.id = l.summary_id
    WHERE {" AND ".join(where)}
    GROUP BY 1,2
    ORDER BY sum(l.net_amount) DESC;
    """
    rows = db_service.fetch_all(sql, tuple(params))
    return jsonify({"items": rows})

@app.route("/api/companies/<int:company_id>/reports/invoice_sales_by_customer", methods=["GET"])
@require_auth
def report_invoice_sales_by_customer(company_id: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != int(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    date_from = request.args.get("from")
    date_to   = request.args.get("to")

    schema = f"company_{company_id}"
    where = ["1=1"]
    params = []

    if date_from:
        where.append("i.invoice_date >= %s")
        params.append(date_from)
    if date_to:
        where.append("i.invoice_date <= %s")
        params.append(date_to)

    sql = f"""
    SELECT
      i.customer_id,
      c.name as customer_name,
      count(*) as invoices,
      sum(i.total_amount) as total_amount,
      sum(i.subtotal_amount) as net_amount,
      sum(i.vat_amount) as vat_amount
    FROM {schema}.invoices i
    JOIN {schema}.customers c ON c.id = i.customer_id
    WHERE {" AND ".join(where)}
      AND lower(coalesce(i.status,'')) IN ('approved','posted')  -- choose what you want counted
    GROUP BY 1,2
    ORDER BY sum(i.total_amount) DESC;
    """
    rows = db_service.fetch_all(sql, tuple(params))
    return jsonify({"items": rows})

# ────────────────────────────────────────────────────────────────
# USERS API (minimal for Users & Roles screen)
# ────────────────────────────────────────────────────────────────
@app.route("/api/users", methods=["GET"])
@require_auth
def api_list_users():
    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"error": "Not authenticated"}), 401

    email = current_user.get("email")
    user = db_service.get_user_by_email(email) if email else None

    if not user:
        # fallback to current_user data
        user_obj = {
            "id": current_user.get("id"),
            "email": email,
            "user_role": current_user.get("user_type") or "Owner",
            "first_name": current_user.get("first_name") or "",
            "last_name": current_user.get("last_name") or "",
            "is_confirmed": bool(current_user.get("is_confirmed")),
            "invite_status": None,
        }
    else:
        user_obj = {
            "id": user["id"],
            "email": user["email"],
            "user_role": user.get("user_type") or "Owner",
            "first_name": user.get("first_name") or "",
            "last_name": user.get("last_name") or "",
            "is_confirmed": bool(user.get("is_confirmed")),
            "invite_status": None,
        }

    return jsonify({"users": [user_obj]}), 200

# ────────────────────────────────────────────────────────────────
# META: Countries
# ────────────────────────────────────────────────────────────────

@app.route("/api/meta/countries", methods=["GET"])
def api_country_meta():
    countries = load_countries()

    enriched = []
    for c in countries:
        row = dict(c)
        code = (row.get("code") or "").upper().strip()

        examples = FIELD_EXAMPLES.get(code, {})

        row["regno_pattern"] = REGNO_REGEX.get(code)
        row["tin_pattern"] = TIN_REGEX.get(code)
        row["vat_pattern"] = VAT_REGEX.get(code)

        row["regno_example"] = examples.get("regno")
        row["tin_example"] = examples.get("tin")
        row["vat_example"] = examples.get("vat")

        enriched.append(row)

    return jsonify({"countries": enriched}), 200

# ────────────────────────────────────────────────────────────────
# VAT SETTINGS (DB-backed)
# ────────────────────────────────────────────────────────────────

@app.route("/api/companies/<int:company_id>/vat_settings", methods=["GET", "PUT"])
@require_auth
def api_vat_settings(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user:
        return jsonify({"error": "Not authenticated"}), 401

    if current_user.get("company_id") not in (None, company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    if request.method == "GET":
        cfg = db_service.get_vat_settings(company_id) or {}
        return jsonify(cfg), 200

    # PUT: update + sanitise
    data = request.get_json(silent=True) or {}

    freq = (data.get("frequency") or "bi_monthly").lower()
    if freq not in ("monthly", "bi_monthly", "quarterly", "semi_annual", "annual"):
        freq = "bi_monthly"

    try:
        anchor_month = int(data.get("anchor_month") or 1)
    except (TypeError, ValueError):
        anchor_month = 1
    anchor_month = max(1, min(12, anchor_month))

    try:
        filing_lag_days = int(data.get("filing_lag_days") or 25)
    except (TypeError, ValueError):
        filing_lag_days = 25

    try:
        reminder_days_before = int(data.get("reminder_days_before") or 10)
    except (TypeError, ValueError):
        reminder_days_before = 10

    try:
        vat_rate = float(data.get("vat_rate") or 15)
    except (TypeError, ValueError):
        vat_rate = 15.0

    try:
        corp_tax_rate = float(data.get("corp_tax_rate") or 27)
    except (TypeError, ValueError):
        corp_tax_rate = 27.0

    cfg = db_service.get_vat_settings(company_id) or {}
    cfg.update({
        "frequency": freq,
        "anchor_month": anchor_month,
        "filing_lag_days": filing_lag_days,
        "reminder_days_before": reminder_days_before,
        "country": (data.get("country") or cfg.get("country") or "ZA").upper(),
        "vat_registered": bool(data.get("vat_registered", cfg.get("vat_registered", True))),
        "vat_number": data.get("vat_number") or cfg.get("vat_number"),
        "vat_rate": vat_rate,
        "pricing_includes_vat": bool(data.get("pricing_includes_vat", cfg.get("pricing_includes_vat", False))),
        "corp_tax_rate": corp_tax_rate,
        "vat_rate": vat_rate,  # default %, e.g. 15
        "vat_input_code": data.get("vat_input_code") or cfg.get("vat_input_code") or "1410",
        "vat_output_code": data.get("vat_output_code") or cfg.get("vat_output_code") or "2310",
    })

    db_service.save_vat_settings(company_id, cfg)
    return jsonify(cfg), 200

@app.route("/api/companies/<int:company_id>/dashboard_snapshot", methods=["GET"])
@require_auth
def api_dashboard_snapshot(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or int(current_user.get("company_id") or 0) != int(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    # -----------------------------------------
    # 1) Resolve dashboard period from preset/from/to
    # -----------------------------------------
    date_from, date_to, meta = resolve_company_period(
        db_service, company_id, request, mode="range"
    )
    if not date_from or not date_to:
        return jsonify({"error": "from/to required (or preset)"}), 400

    as_at = date_to

    # -----------------------------------------
    # 2) Base services
    # -----------------------------------------
    flags = _get_industry_flags(company_id) or {}

    tb_rows = db_service.get_trial_balance_with_meta(
        company_id,
        date_from=date_from,
        date_to=date_to,
    ) or []

    pnl_rows = db_service.get_pnl_mini(
        company_id,
        date_from=date_from,
        date_to=date_to,
    ) or []

    cf_rows = db_service.get_cashflow_mini(
        company_id,
        date_from=date_from,
        date_to=date_to,
        method=(request.args.get("method") or "direct"),
        compare="none",
    ) or []

    bs_rows = db_service.get_bs_mini(
        company_id,
        as_of=as_at,
        use_meta_when_no_date=False,
    ) or []

    ar_aging = db_service.get_ar_aging_report(
        company_id,
        as_at=as_at,
    ) or {}

    ap_aging = db_service.get_ap_aging_report(
        company_id,
        as_at=as_at,
    ) or {}

    fixed_assets = db_service.get_fixed_assets_snapshot(
        company_id,
        as_of=as_at,
    ) or {}

    customer_count = db_service.count_customers(company_id)
    vendor_count = db_service.count_vendors(company_id)

    # -----------------------------------------
    # 3) KPI derivation from TB + mini outputs
    # -----------------------------------------
    def _f(v):
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    def _txt(*parts):
        return " ".join([str(x or "") for x in parts]).strip().lower()

    cash = 0.0
    ar_total = 0.0
    ap_total = 0.0

    for r in tb_rows:
        name = _txt(r.get("name"), r.get("section"), r.get("category"), r.get("standard"))
        closing = _f(r.get("closing_balance"))

        if any(k in name for k in ["cash", "bank", "petty cash"]):
            cash += closing
        elif any(k in name for k in ["receivable", "trade receivable", "accounts receivable", "debtors"]):
            ar_total += closing
        elif (
            any(k in name for k in ["trade payable", "accounts payable", "supplier payable", "vendor payable", "creditors"])
            and not any(k in name for k in [
                "loan payable",
                "vat payable",
                "vat output",
                "vat input",
                "paye",
                "withholding",
                "lease liability",
                "deferred income",
                "tax payable",
                "income tax",
                "provision",
                "accrual",
                "borrowings",
                "bank overdraft",
            ])
        ):
            ap_total += abs(closing)

    revenue = cogs = gross_profit = expenses = net_profit = 0.0
    for r in pnl_rows:
        label = _txt(r.get("label"))
        amt = _f(r.get("amount"))
        if "revenue" in label or label == "income":
            revenue = amt
        elif "cost of sales" in label:
            cogs = amt
        elif "gross profit" in label:
            gross_profit = amt
        elif "operating expenses" in label:
            expenses = amt
        elif "net profit" in label or "surplus" in label or "deficit" in label:
            net_profit = amt

    assets = liabilities = equity = bs_check = 0.0
    for r in bs_rows:
        label = _txt(r.get("label"), r.get("section"))
        amt = _f(r.get("amount"))
        if "total assets" in label:
            assets = amt
        elif "total liabilities" in label:
            liabilities = amt
        elif "total equity" in label:
            equity = amt
        elif "assets - (liabilities + equity)" in label:
            bs_check = amt

    cf_operating = cf_investing = cf_financing = 0.0
    for r in cf_rows:
        sec = _txt(r.get("section"), r.get("label"))
        amt = _f(r.get("amount"))
        if "operating" in sec:
            cf_operating = amt
        elif "investing" in sec:
            cf_investing = amt
        elif "financing" in sec:
            cf_financing = amt

    gross_margin_pct = round((gross_profit / revenue) * 100, 2) if revenue else 0.0
    net_margin_pct = round((net_profit / revenue) * 100, 2) if revenue else 0.0

    # -----------------------------------------
    # 4) VAT summary for period
    # -----------------------------------------
    vat_input_total = 0.0
    vat_output_total = 0.0
    vat_net = 0.0
    try:
        from vat_utils import _get_vat_accounts  # adjust import to your project
        schema = db_service.company_schema(company_id)

        input_codes, output_codes = _get_vat_accounts(company_id)
        vat_codes = list(set(input_codes) | set(output_codes))

        if vat_codes:
            placeholders = ",".join(["%s"] * len(vat_codes))
            sql = f"""
                SELECT account, debit, credit
                FROM {schema}.ledger
                WHERE company_id = %s
                  AND date >= %s
                  AND date <= %s
                  AND account IN ({placeholders})
            """
            rows = db_service.fetch_all(
                sql,
                (int(company_id), date_from, date_to, *vat_codes),
            ) or []

            for row in rows:
                code = str(row.get("account") or "")
                dr = _f(row.get("debit"))
                cr = _f(row.get("credit"))
                bal = dr - cr

                if code in input_codes:
                    vat_input_total += bal
                elif code in output_codes:
                    vat_output_total += (-bal)

            vat_net = round(vat_output_total - vat_input_total, 2)
    except Exception:
        vat_input_total = 0.0
        vat_output_total = 0.0
        vat_net = 0.0

    # -----------------------------------------
    # 5) Trends
    # -----------------------------------------
    def trend_month_count(preset: str | None) -> int:
        key = str(preset or "").lower()
        if key in {"this_month", "prev_month"}:
            return 1
        if key in {"this_quarter", "prev_quarter"}:
            return 3
        if key in {"last_2_quarters"}:
            return 6
        if key in {"ytd", "this_year", "prev_year"}:
            return 12
        return 6

    def month_end_series(end_date, months=6):
        out = []
        cur = end_date.replace(day=1)
        for i in range(months - 1, -1, -1):
            y = cur.year
            m = cur.month - i
            while m <= 0:
                y -= 1
                m += 12
            while m > 12:
                y += 1
                m -= 12

            start = date(y, m, 1)
            if m == 12:
                nxt = date(y + 1, 1, 1)
            else:
                nxt = date(y, m + 1, 1)
            end = nxt - timedelta(days=1)

            out.append((start, min(end, end_date)))
        return out

    revenue_trend = []
    expense_trend = []
    profit_trend = []
    cash_balance_trend = []
    cash_in_trend = []
    cash_out_trend = []
    net_cash_movement_trend = []

    try:
        trend_months = trend_month_count(request.args.get("preset"))

        for start, end in month_end_series(date_to, months=trend_months):
            month_pnl = db_service.get_pnl_mini(company_id, start, end) or []
            month_tb = db_service.get_trial_balance_with_meta(company_id, start, end) or []

            month_revenue = 0.0
            month_expenses = 0.0
            month_profit = 0.0
            month_cash_balance = 0.0

            # P&L monthly parsing
            for r in month_pnl:
                label = _txt(r.get("label"))
                amt = _f(r.get("amount"))

                if "revenue" in label or label == "income":
                    month_revenue = amt

                elif "cost of sales" in label:
                    month_expenses += abs(amt)

                elif "operating expenses" in label or "expenses" == label:
                    month_expenses += abs(amt)

                elif (
                    "net profit" in label
                    or "profit for period" in label
                    or "profit before tax" in label
                    or "surplus" in label
                    or "deficit" in label
                ):
                    month_profit = amt

            # fallback if profit row missing
            if month_profit == 0 and (month_revenue != 0 or month_expenses != 0):
                month_profit = month_revenue - month_expenses

            cash_accounts = set()
            for r in month_tb:
                name = _txt(r.get("name"), r.get("section"), r.get("category"))
                closing = _f(r.get("closing_balance"))

                code = str(
                    r.get("code")
                    or r.get("account")
                    or r.get("account_code")
                    or ""
                ).strip().upper()

                if any(k in name for k in ["cash", "bank", "petty cash"]):
                    month_cash_balance += closing
                    if code:
                        cash_accounts.add(code)

            # fallback: if TB did not expose codes cleanly, use known cash/bank GLs
            if not cash_accounts:
                cash_accounts = {
                    "BS_CA_1000",  # Cash & Bank
                    "BS_CA_1010",  # Petty Cash / Bank alt if used
                    "BS_CA_1050",  # Other cash equivalent if used
                }

            month_cash_in = 0.0
            month_cash_out = 0.0
            try:
                schema = db_service.company_schema(company_id)
                sql = f"""
                    SELECT account, debit, credit
                    FROM {schema}.ledger
                    WHERE company_id = %s
                      AND date >= %s
                      AND date <= %s
                """
                ledger_rows = db_service.fetch_all(sql, (int(company_id), start, end)) or []

                for row in ledger_rows:
                    acct = str(row.get("account") or "").strip().upper()
                    dr = _f(row.get("debit"))
                    cr = _f(row.get("credit"))

                    if acct in cash_accounts:
                        month_cash_in += dr
                        month_cash_out += cr

            except Exception:
                month_cash_in = 0.0
                month_cash_out = 0.0

            month_net_cash = month_cash_in - month_cash_out

            label = end.strftime("%b %y")
            period_key = end.strftime("%Y-%m")

            revenue_trend.append({
                "period": period_key,
                "label": label,
                "value": round(month_revenue, 2),
            })
            expense_trend.append({
                "period": period_key,
                "label": label,
                "value": round(month_expenses, 2),
            })
            profit_trend.append({
                "period": period_key,
                "label": label,
                "value": round(month_profit, 2),
            })
            cash_balance_trend.append({
                "period": period_key,
                "label": label,
                "value": round(month_cash_balance, 2),
            })
            cash_in_trend.append({
                "period": period_key,
                "label": label,
                "value": round(month_cash_in, 2),
            })
            cash_out_trend.append({
                "period": period_key,
                "label": label,
                "value": round(month_cash_out, 2),
            })
            net_cash_movement_trend.append({
                "period": period_key,
                "label": label,
                "value": round(month_net_cash, 2),
            })

    except Exception:
        revenue_trend = []
        expense_trend = []
        profit_trend = []
        cash_balance_trend = []
        cash_in_trend = []
        cash_out_trend = []
        net_cash_movement_trend = []

    # -----------------------------------------
    # 6) Insights
    # -----------------------------------------
    insights = []

    if revenue > 0:
        insights.append(f"Revenue for the selected period is {revenue:,.2f}.")
    else:
        insights.append("No revenue has been recognised for the selected period.")

    if net_profit > 0:
        insights.append(f"Net profit is {net_profit:,.2f} with a margin of {net_margin_pct:.1f}%.")
    elif net_profit < 0:
        insights.append(f"The period shows a net loss of {abs(net_profit):,.2f}.")
    else:
        insights.append("The period is currently at break-even.")

    if abs(bs_check) <= 1:
        insights.append("Balance sheet is in balance.")
    else:
        insights.append(f"Balance sheet check difference is {bs_check:,.2f} and should be reviewed.")

    if _f(ar_aging.get("totals", {}).get("total")) > 0:
        insights.append(
            f"Outstanding receivables total {_f(ar_aging.get('totals', {}).get('total')):,.2f}."
        )

    if _f(ap_aging.get("totals", {}).get("total")) > 0:
        insights.append(
            f"Outstanding payables total {_f(ap_aging.get('totals', {}).get('total')):,.2f}."
        )

    payload = {
        "meta": {
            **(meta or {}),
            "from": date_from.isoformat(),
            "to": date_to.isoformat(),
            "as_of": as_at.isoformat(),
        },
        "industry": flags.get("industry"),
        "kpis": {
            "cash": round(cash, 2),
            "ar": round(ar_total, 2),
            "ap": round(ap_total, 2),
            "netProfit": round(net_profit, 2),
            "vat": round(vat_net, 2),
        },
        "pnl": {
            "revenue": round(revenue, 2),
            "cogs": round(cogs, 2),
            "grossProfit": round(gross_profit, 2),
            "expenses": round(expenses, 2),
            "netProfit": round(net_profit, 2),
            "grossMarginPct": gross_margin_pct,
            "netMarginPct": net_margin_pct,
        },
        "balanceSheet": {
            "assets": round(assets, 2),
            "liabilities": round(liabilities, 2),
            "equity": round(equity, 2),
            "check": round(bs_check, 2),
        },
        "cashflow": {
            "operating": round(cf_operating, 2),
            "investing": round(cf_investing, 2),
            "financing": round(cf_financing, 2),
        },
        "aging": {
            "ar": ar_aging.get("totals", {}),
            "ap": ap_aging.get("totals", {}),
        },
        "counts": {
            "customers": int(customer_count or 0),
            "vendors": int(vendor_count or 0),
        },
        "vat": {
            "input": round(vat_input_total, 2),
            "output": round(vat_output_total, 2),
            "net": round(vat_net, 2),
        },
        "trends": {
            "revenue": revenue_trend,
            "expenses": expense_trend,
            "profit": profit_trend,
            "cash": cash_balance_trend,
            "cash_in": cash_in_trend,
            "cash_out": cash_out_trend,
            "net_cash_movement": net_cash_movement_trend,
        },
        "fixedAssets": {
            "bookValue": round(float(fixed_assets.get("book_value") or 0), 2),
            "accumDepreciation": round(float(fixed_assets.get("accum_depreciation") or 0), 2),
            "netCarryingAmount": round(float(fixed_assets.get("net_carrying_amount") or 0), 2),
            "pendingDepreciationCount": int(fixed_assets.get("pending_depreciation_count") or 0),
            "assetCount": int(fixed_assets.get("asset_count") or 0),
        },
        "insights": insights,
    }

    return jsonify(payload), 200

@app.get("/api/companies/<int:company_id>/accounting-work-queue")
@require_auth
def api_accounting_work_queue(company_id: int):
    current_user = getattr(g, "current_user", None)

    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    date_from = request.args.get("from")
    date_to = request.args.get("to")
    limit = int(request.args.get("limit") or 20)

    try:
        data = db_service.get_accounting_work_queue(
            company_id,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
        )

        # 🔍 Optional audit (consistent with your system)
        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(current_user.get("id") or 0),
                module="dashboard",
                action="view",
                severity="info",
                entity_type="accounting_work_queue",
                entity_id=None,
                entity_ref="dashboard",
                before_json={},
                after_json={"count": data.get("count")},
                message="Viewed accounting work queue",
            )
        except Exception:
            pass

        return jsonify(data), 200

    except Exception as ex:
        current_app.logger.exception("Error loading accounting work queue: %s", ex)
        return jsonify({"error": "Failed to load accounting work queue"}), 500
    
@app.route("/api/companies/<int:company_id>/customers", methods=["GET", "POST"])
@require_auth
def api_company_customers(company_id: int):
    """
    GET  -> list customers for this company (used by left-hand customers list)
    POST -> create a new customer using the same keys the dashboard form uses.
    """
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    # ---------- GET ----------
    if request.method == "GET":
        include_inactive = request.args.get("include_inactive") == "1"
        customers = db_service.list_customers(company_id, include_inactive=include_inactive) or []
        current_app.logger.info(
            "GET /companies/%s/customers -> %d rows (include_inactive=%s)",
            company_id, len(customers), include_inactive
        )
        return jsonify(customers), 200

    # ---------- POST ----------
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Customer name is required"}), 400

    # ✅ POLICY: single source of truth (normalized + mode defaults)
    pol = company_policy(company_id)
    mode = (pol.get("mode") or "owner_managed").strip().lower()
    company = pol.get("company") or {}
    policy = pol.get("policy") or {}

    uid = current_user.get("id")
    is_owner = (uid is not None and str(company.get("owner_user_id")) == str(uid))

    # ✅ Customer approval gate (Assisted/Controlled)
    # - owner_managed: never gate
    # - assisted/controlled: gate if AR review is ON OR require_customer_approval is ON
    approval_required = False
    if mode != "owner_managed":
        approval_required = bool(
            policy.get("review_enabled", False)
            or policy.get("invoice_review_enabled", False)
            or policy.get("require_invoice_review", False)
            or policy.get("require_customer_approval", False)
            or pol.get("require_customer_approval", False)  # keep compatibility with computed flag
        )

    approved_by_user_id = None
    approved_at = None

    if mode == "owner_managed":
        credit_status = "approved"

    elif mode in {"assisted", "controlled"}:
        # ✅ If approval workflow is required -> non-owner must create as draft
        if approval_required and not is_owner:
            credit_status = "draft"
        else:
            # owner can approve immediately; others draft
            credit_status = "approved" if is_owner else "draft"

    else:
        credit_status = "draft"

    if credit_status == "approved":
        approved_by_user_id = uid
        approved_at = datetime.utcnow()

    # ✅ Safe defaults
    data.setdefault("is_active", True)

    # ✅ Stamp credit fields unless caller explicitly set them
    data.setdefault("credit_status", credit_status)
    data.setdefault("approved_by_user_id", approved_by_user_id)
    data.setdefault("approved_at", approved_at)

    try:
        customer_id = db_service.insert_customer(company_id, data)
        cust = db_service.get_customer(company_id, customer_id) or {}

        # ✅ If customer is created as draft in assisted/controlled -> create approval request
        try:
            if (cust.get("credit_status") or "").strip().lower() == "draft" and mode in {"assisted", "controlled"}:
                # Use JWT user id if you have it; else fallback to current_user id
                actor_user_id = int(getattr(request, "jwt_payload", {}) .get("user_id")
                                    or getattr(request, "jwt_payload", {}) .get("sub")
                                    or (current_user.get("id") or 0) or 0)

                entity_type = "customer"
                entity_id = str(customer_id)
                module = "ar"
                action = "approve_customer"
                dedupe_key = f"{company_id}:{module}:{action}:{entity_type}:{entity_id}"

                req = db_service.create_approval_request(
                    company_id,
                    entity_type=entity_type,
                    entity_id=entity_id,
                    entity_ref=cust.get("name"),
                    module=module,
                    action=action,
                    requested_by_user_id=int(actor_user_id or 0),
                    amount=0.0,
                    currency=None,
                    risk_level="low",
                    dedupe_key=dedupe_key,
                    payload_json={"customer": {"id": customer_id, "name": cust.get("name")}},
                )
                cust["_approval_request"] = req
                cust["_workflow"] = "draft_pending_customer_approval"
        except Exception:
            current_app.logger.exception("Failed to create approval request for customer")
        # ✅ optional: surface warnings (KYC etc.) to UI

        warnings = pol.get("_warnings") or []
        if warnings:
            cust["_warnings"] = warnings

        # 🧾 AUDIT LOG — CREATE
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(current_user.get("id") or 0),
            module="ar",
            action="create",
            severity="info",
            entity_type="customer",
            entity_id=str(customer_id),
            entity_ref=cust.get("name"),
            before_json={},
            after_json=cust or {},
            message=f"Customer created: {cust.get('name')}",
        )

        return jsonify(cust), 201

    except Exception as ex:
        current_app.logger.exception("Error inserting customer: %s", ex)
        return jsonify({"error": "Failed to create customer"}), 500
    
@app.route(
    "/api/companies/<int:company_id>/customers/<int:customer_id>",
    methods=["GET", "PUT", "DELETE"],
)
@require_auth
def api_company_customer_detail(company_id: int, customer_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    if request.method == "GET":
        cust = db_service.get_customer(company_id, customer_id)
        if not cust:
            return jsonify({"error": "Customer not found"}), 404
        return jsonify(cust), 200

    if request.method == "PUT":
        data = request.get_json(silent=True) or {}
        current_app.logger.debug(
            "PUT /companies/%s/customers/%s body = %r",
            company_id, customer_id, data
        )

        # 🔎 BEFORE snapshot
        before = db_service.get_customer(company_id, customer_id)
        if not before:
            return jsonify({"error": "Customer not found"}), 404

        # ✅ POLICY: single source of truth (normalized + mode defaults)
        pol = company_policy(company_id)
        mode = (pol.get("mode") or "owner_managed").strip().lower()
        company = pol.get("company") or {}
        require_customer_approval = bool(pol.get("require_customer_approval", False))

        uid = current_user.get("id")
        is_owner = (uid is not None and str(company.get("owner_user_id")) == str(uid))

        # ✅ Restrict credit_status changes by mode (use CUSTOMER approval flag, not invoice review)
        if "credit_status" in data:
            requested = (data.get("credit_status") or "").strip().lower()

            # Always validate requested value a bit
            allowed_status_values = {"draft", "approved", "cod_only", "rework", "rejected"}
            if requested and requested not in allowed_status_values:
                return jsonify({"error": f"Invalid credit_status '{requested}'"}), 400

            if mode != "owner_managed":
                # controlled: if customer approval workflow is enabled, block direct status changes
                if mode == "controlled" and require_customer_approval and requested in {"approved", "cod_only", "rework", "rejected"}:
                    return jsonify({"error": "Use the Credit Approval workflow to change credit status."}), 409

                # assisted: assistants cannot approve
                if mode == "assisted" and requested in {"approved", "cod_only"} and not is_owner:
                    return jsonify({"error": "Only the owner can approve customer credit status."}), 403

            # stamp approval metadata when set to approved/cod_only
            if requested in {"approved", "cod_only"}:
                data.setdefault("approved_by_user_id", uid)
                data.setdefault("approved_at", datetime.utcnow())

        try:
            ok = db_service.update_customer(company_id, customer_id, **data)
        except Exception as ex:
            current_app.logger.exception("Error updating customer: %s", ex)
            return jsonify({"error": "Failed to update customer"}), 500

        if not ok:
            return jsonify({"error": "Customer not found or nothing to update"}), 404

        # Return updated truth (more useful for UI)
        cust = db_service.get_customer(company_id, customer_id) or {}


        # optional: surface warnings (KYC etc.) to UI
        warnings = pol.get("_warnings") or []
        if warnings:
            cust["_warnings"] = warnings

        # 🧾 AUDIT LOG — UPDATE
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(current_user.get("id") or 0),
            module="ar",
            action="update",
            severity="info",
            entity_type="customer",
            entity_id=str(customer_id),
            entity_ref=(cust or {}).get("name") or before.get("name"),
            before_json=before or {},
            after_json=cust or {},
            message=f"Customer updated: {(cust or {}).get('name') or before.get('name')}",
        )

        return jsonify(cust or {"success": True}), 200

    if request.method == "DELETE":
        # 🔎 BEFORE snapshot
        before = db_service.get_customer(company_id, customer_id)
        if not before:
            return jsonify({"error": "Customer not found"}), 404

        try:
            db_service.delete_customer(company_id, customer_id)
        except Exception as ex:
            current_app.logger.exception("Error deleting customer: %s", ex)
            return jsonify({"error": "Failed to delete customer"}), 500

        # 🧾 AUDIT LOG — DELETE
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(current_user.get("id") or 0),
            module="ar",
            action="delete",
            severity="warning",
            entity_type="customer",
            entity_id=str(customer_id),
            entity_ref=before.get("name"),
            before_json=before or {},
            after_json={},
            message=f"Customer deleted: {before.get('name')}",
        )

        return jsonify({"ok": True}), 200
    
@app.route(
    "/api/companies/<int:company_id>/customers/quick",
    methods=["POST", "OPTIONS"],
)
@require_auth
def api_company_customer_quick_create(company_id: int):
    user = getattr(g, "current_user", {}) or {}
    if int(user.get("company_id") or 0) != int(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    data = request.get_json(silent=True) or {}

    # --------------------------
    # Minimal required fields
    # --------------------------
    name = (data.get("name") or data.get("customer_name") or "").strip()
    if not name:
        return jsonify({"error": "Customer name is required"}), 400

    contact_person = (data.get("contact_person") or data.get("contactPerson") or "").strip()
    email = (data.get("email") or "").strip() or None
    phone = (data.get("phone") or "").strip() or None

    # contacts JSONB (optional)
    contacts = []
    if contact_person or email or phone:
        contacts.append({
            "name": contact_person or None,
            "email": email,
            "phone": phone,
            "role": "primary",
        })

    # --------------------------
    # Prospect defaults
    # --------------------------
    payload = {
        "name": name,
        "email": email,
        "phone": phone,
        "billing_address": (data.get("billing_address") or data.get("billingAddress") or "").strip() or None,
        "shipping_address": (data.get("shipping_address") or data.get("shippingAddress") or "").strip() or None,
        "country": (data.get("country") or "").strip() or None,
        "payment_terms": (data.get("payment_terms") or data.get("terms") or "on_receipt"),
        "credit_limit": 0,
        "credit_status": "draft",
        "on_hold": "no",
        "is_active": True,
        "customer_type": (data.get("customer_type") or "prospect"),
        "tags": (data.get("tags") or "prospect").strip() or "prospect",
        "notes": (data.get("notes") or "").strip() or None,
        "contacts": psycopg2.extras.Json(contacts),
    }

    try:
        new_id = int(db_service.insert_customer(company_id, payload) or 0)
        if not new_id:
            return jsonify({"error": "Failed to create customer"}), 500

        cust = db_service.get_customer(company_id, new_id) or {}

        # 🧾 AUDIT LOG — QUICK CREATE
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="ar",
            action="create",
            severity="info",
            entity_type="customer",
            entity_id=str(new_id),
            entity_ref=cust.get("name"),
            before_json={},
            after_json=cust or {},
            message=f"Customer created (quick): {cust.get('name')}",
        )

        return jsonify(cust), 201

    except Exception as ex:
        current_app.logger.exception("quick customer create failed")
        return jsonify({"error": "Failed to create customer", "detail": str(ex)}), 500

# =========================================================
# INVOICES (CREATE)
# =========================================================
@app.route("/api/companies/<int:cid>/invoices", methods=["POST"])
@require_auth
def create_invoice(cid: int):
    try:
        company_id = int(cid)
        user = getattr(g, "current_user", {}) or {}

        if user.get("company_id") != company_id:
            return jsonify({"error": "Not authorised for this company"}), 403

        # --------------------------
        # helpers
        # --------------------------
        def money(x) -> float:
            from decimal import Decimal, ROUND_HALF_UP
            return float(Decimal(str(x or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

        def to_float(x, default=0.0):
            try:
                if x is None:
                    return default
                if isinstance(x, str):
                    x = x.replace(",", "").strip()
                return float(x)
            except Exception:
                return default

        def to_disc_rate(x):
            r = to_float(x, 0.0)
            if r > 1.0:
                r = r / 100.0
            return max(0.0, min(1.0, r))

        payload = request.get_json(silent=True) or {}
        auth_ctx = getattr(g, "auth_context", {}) or {}

        if auth_ctx.get("is_delegated_company_access"):
            payload["source_company_id"] = auth_ctx.get("source_company_id")
            payload["engagement_company_id"] = auth_ctx.get("source_company_id")
            payload["engagement_id"] = auth_ctx.get("engagement_id")
        current_app.logger.info("create_invoice payload = %r", payload)

        is_lessor_invoice = (
            str(payload.get("source") or "").strip().lower()
            == "lessor_lease_billing"
            or str(payload.get("module_name") or "").strip().lower()
            == "ifrs16_lessor"
            or bool(payload.get("lessor_lease_id"))
        )

        # --------------------------
        # 0) Basic validation
        # --------------------------
        customer_id = payload.get("customer_id")
        lines = payload.get("lines") or []
        if not customer_id:
            return jsonify({"error": "customer_id is required"}), 400
        if not isinstance(lines, list) or not lines:
            return jsonify({"error": "At least one line is required"}), 400

        cust_id = int(customer_id)
        requested_status = (payload.get("status") or "draft").strip().lower()

        # --------------------------
        # POLICY (single source of truth)
        # --------------------------
        pol = company_policy(company_id)
        mode = (pol.get("mode") or "").strip().lower()
        policy = pol.get("policy") or {}

        review_enabled = bool(pol.get("review_enabled")) or bool(
            policy.get("review_enabled", False) or
            policy.get("invoice_review_enabled", False) or
            policy.get("require_invoice_review", False)
        )

        # review OFF => will auto-post later in this endpoint
        should_post = (not review_enabled)

        # Allowed states in header saves
        allowed_statuses = {"draft", "pending_approval", "approved", "posted"}

        # Default save state
        status_to_save = requested_status or "draft"
        if status_to_save not in allowed_statuses:
            status_to_save = "draft"

        # ✅ If review is OFF, we still SAVE as 'draft' (or 'approved' if you want),
        # but we will set 'posted' after GL posting below.
        # Keeping draft here avoids mixing "approved" with "posted".
        if should_post:
            status_to_save = "draft"   # keep clean; posting below will move it to posted
            

        # --------------------------
        # 2) Customer checks + enforcement
        # --------------------------
        cust = db_service.get_customer(company_id, cust_id)
        if not cust:
            return jsonify({"error": "Customer not found"}), 404

        if not cust.get("is_active", True):
            return jsonify({"error": "Customer is archived/inactive", "customer_id": cust_id}), 409

        credit_status = (cust.get("credit_status") or "").strip().lower()
        on_hold = (cust.get("on_hold") or "no").strip().lower()
        if on_hold in {"yes", "true", "1"}:
            return jsonify({"error": "Customer is on hold", "customer_id": cust_id}), 409

        require_customer_approved = must_approve_customer_before_invoicing(mode, policy)
        allowed_statuses = {"approved", "cod_only"}

        if require_customer_approved and status_to_save != "draft" and credit_status not in allowed_statuses:
            return jsonify({
                "error": "Customer is not approved for invoicing.",
                "customer_id": cust_id,
                "credit_status": credit_status,
                "status_to_save": status_to_save,
                "mode": mode,
            }), 409

        # --------------------------
        # 2b) Number rules
        # --------------------------
        inv_no = (payload.get("number") or "").strip()

        # draft -> must be blank / None
        if status_to_save == "draft":
            inv_no = ""

        # ✅ approved/posting: can still be blank -> DB may assign
        # (so do NOT return 400 here)

        # --------------------------
        # 3) Header
        # --------------------------
        header = {
            "customer_id": cust_id,
            "revenue_contract_id": (
                int(payload.get("revenue_contract_id"))
                if payload.get("revenue_contract_id") not in (None, "", 0, "0")
                else None
            ),
            "invoice_date": payload.get("invoice_date") or payload.get("date"),
            "due_date": payload.get("due_date"),
            "currency": payload.get("currency"),
            "number": (inv_no or None) if status_to_save != "draft" else None,
            "bank_account_id": payload.get("bank_account_id"),
            "notes": payload.get("notes"),
            "status": status_to_save,

            "source_company_id": payload.get("source_company_id"),
            "engagement_company_id": payload.get("engagement_company_id"),
            "engagement_id": payload.get("engagement_id"),

            "created_by_user_id": user.get("id"),
            "updated_by_user_id": user.get("id"),

            "discount_rate": to_disc_rate(
                payload.get("discount_rate") if "discount_rate" in payload else payload.get("discount")
            ),
            "discount": to_disc_rate(payload.get("discount")),  # legacy
            "other": to_float(payload.get("other"), 0.0),
            "source": (
                "lessor_lease_billing"
                if is_lessor_invoice
                else payload.get("source")
            ),
            "source_id": (
                int(payload.get("source_id"))
                if payload.get("source_id") not in (None, "", 0, "0")
                else None
            ),
            "module_name": (
                "ifrs16_lessor"
                if is_lessor_invoice
                else payload.get("module_name")
            ),
            "posting_mode": (
                "custom_accounts"
                if is_lessor_invoice
                else payload.get("posting_mode")
            ),
            "lessor_lease_id": (
                int(payload.get("lessor_lease_id"))
                if payload.get("lessor_lease_id") not in (None, "", 0, "0")
                else None
            ),
            "lessor_schedule_id": (
                int(payload.get("lessor_schedule_id"))
                if payload.get("lessor_schedule_id") not in (None, "", 0, "0")
                else None
            ),
            "lease_classification": (
                payload.get("lease_classification")
                if is_lessor_invoice
                else None
            ),
            "accounting_treatment": (
                payload.get("accounting_treatment")
                if is_lessor_invoice
                else None
            ),
            "ar_account_code": (
                payload.get("ar_account_code")
                if is_lessor_invoice
                else None
            ),
            "credit_account_code": (
                payload.get("credit_account_code")
                if is_lessor_invoice
                else None
            ),
            "vat_output_account_code": (
                payload.get("vat_output_account_code")
                if is_lessor_invoice
                else None
            ),
        }

        print("[CREATE_INVOICE HEADER] revenue_contract_id =", header.get("revenue_contract_id"), flush=True)
        current_app.logger.warning(
            "[CREATE_INVOICE HEADER] revenue_contract_id=%r header=%r",
            header.get("revenue_contract_id"),
            header,
        )
        if not header["invoice_date"]:
            return jsonify({"error": "invoice_date is required"}), 400

        inv_date = header["invoice_date"]
        if isinstance(inv_date, str):
            try:
                inv_date_obj = date.fromisoformat(inv_date[:10])
                if inv_date_obj > date.today():
                    return jsonify({"error": "Invoice date cannot be in the future."}), 400
            except Exception:
                pass

        # --------------------------
        # 4) Map lines (ID-based)
        # --------------------------
        mapped_lines = []
        for ln in lines:
            if not isinstance(ln, dict):
                continue

            item_type = (ln.get("item_type") or ln.get("itemType") or "").strip().lower() or None
            item_id_raw = ln.get("item_id") or ln.get("itemId")
            item_id = int(item_id_raw) if item_id_raw is not None and str(item_id_raw).isdigit() else None
            item_code = (ln.get("item_code") or ln.get("itemCode") or ln.get("sku") or ln.get("code") or "").strip() or None

            if item_type not in (
                None,
                "",
                "inventory",
                "service",
                "gl",
                "lease",
            ):
                return jsonify({
                    "error": "Invalid item_type on line",
                    "line": ln,
                }), 400
            if item_type in ("inventory", "service") and not item_id:
                return jsonify({"error": f"item_id is required when item_type='{item_type}'", "line": ln}), 400

            item_name = (ln.get("item_name") or ln.get("itemName") or ln.get("item") or "").strip()
            desc = (ln.get("description") or "").strip()

            qty = to_float(ln.get("quantity") or ln.get("qty"), 0.0)
            unit_price = to_float(ln.get("unit_price") or ln.get("unitPrice"), 0.0)
            discount_amount = to_float(ln.get("discount_amount") or ln.get("discountAmount"), 0.0)

            raw_code = (ln.get("account_code") or ln.get("accountCode") or "").strip()

            vat_code = (ln.get("vatCode") or ln.get("vat_code") or "STANDARD")
            vat_code = str(vat_code).strip().upper()

            vat_rate = ln.get("vat_rate")
            if vat_rate is None and vat_code == "CUSTOM":
                vat_rate = (
                    ln.get("vatRate") or
                    ln.get("vat_rate") or
                    ln.get("customVatRate") or
                    ln.get("vat_custom_rate") or
                    ln.get("vatCustomRate")
                )
            if vat_rate is None:
                vat_rate = rate_from_code(vat_code)
            vat_rate = to_float(vat_rate, 0.0)

            if qty <= 0:
                return jsonify({"error": "quantity must be > 0", "line": ln}), 400
            if unit_price < 0:
                return jsonify({"error": "unit_price must be >= 0", "line": ln}), 400
            if discount_amount < 0:
                return jsonify({"error": "discount_amount must be >= 0", "line": ln}), 400
            if not raw_code:
                return jsonify({"error": "Each invoice line must have account_code", "line": ln}), 400

            net = max(0.0, (qty * unit_price) - discount_amount)
            vat = net * (vat_rate / 100.0)
            total = net + vat

            mapped_lines.append({
                "item_name": item_name or None,
                "item_type": item_type,
                "item_id": item_id,
                "item_code": item_code,
                "vat_code": vat_code,
                "description": (desc or item_name or ""),
                "revenue_obligation_id": (
                    int(ln.get("revenue_obligation_id") or ln.get("obligation_id"))
                    if (ln.get("revenue_obligation_id") or ln.get("obligation_id")) not in (None, "", 0, "0")
                    else None
                ),
                "account_code": raw_code,
                "quantity": qty,
                "unit_price": unit_price,
                "discount_amount": money(discount_amount),
                "net_amount": money(net),
                "vat_rate": vat_rate,
                "vat_amount": money(vat),
                "total_amount": money(total),
                "source": (
                    "lessor_lease_billing"
                    if is_lessor_invoice
                    else ln.get("source")
                ),
                "source_id": (
                    int(
                        ln.get("source_id")
                        or payload.get("lessor_schedule_id")
                        or 0
                    )
                    or None
                ),
                "module_name": (
                    "ifrs16_lessor"
                    if is_lessor_invoice
                    else ln.get("module_name")
                ),
                "lessor_lease_id": (
                    int(
                        ln.get("lessor_lease_id")
                        or payload.get("lessor_lease_id")
                        or 0
                    )
                    or None
                ),
                "lessor_schedule_id": (
                    int(
                        ln.get("lessor_schedule_id")
                        or payload.get("lessor_schedule_id")
                        or 0
                    )
                    or None
                ),
            })

        if not mapped_lines:
            return jsonify({"error": "At least one valid line is required"}), 400

        if is_lessor_invoice:
            lease_classification=str(
                header.get("lease_classification") or ""
            ).strip().lower()

            if lease_classification not in {"operating","finance"}:
                return jsonify({
                    "error":"Invalid lessor lease classification"
                }),400

            credit_role=(
                "lessor_lease_income"
                if lease_classification=="operating"
                else "lessor_net_investment_current"
            )

            expected=db_service.get_posting_account_by_role(
                company_id,
                credit_role,
            )

            header["credit_account_code"]=expected["code"]

            for line in mapped_lines:
                line["account_code"]=expected["code"]

        # --------------------------
        # 6) Save invoice
        # --------------------------

        # group invoice amounts per obligation
        obligation_totals = {}

        for l in mapped_lines:
            oid = l.get("revenue_obligation_id")
            if not oid:
                continue

            obligation_totals.setdefault(int(oid), Decimal("0.00"))
            obligation_totals[int(oid)] += Decimal(str(l.get("net_amount") or 0))

        # validate each obligation
        for oid, invoice_amount in obligation_totals.items():
            preview = db_service.get_revenue_obligation_billable_preview(
                company_id=company_id,
                obligation_id=int(oid),
                cur=None,  # or pass cursor if available
            )

            available_to_bill = Decimal(str(preview.get("available_to_bill") or 0))

            if invoice_amount > available_to_bill:
                return jsonify({
                    "ok": False,
                    "error": "revenue_obligation_overbilled",
                    "message": "This invoice exceeds the remaining billable amount for the selected performance obligation.",
                    "details": {
                        "obligation_id": int(oid),
                        "available_amount": float(available_to_bill),
                        "attempted_amount": float(invoice_amount),
                    },
                }), 422

        if is_lessor_invoice:
            lessor_lease_id = int(
                payload.get("lessor_lease_id") or 0
            )
            lessor_schedule_id = int(
                payload.get("lessor_schedule_id")
                or payload.get("source_id")
                or 0
            )

            if not lessor_lease_id:
                return jsonify({
                    "error": "lessor_lease_id is required"
                }), 400

            if not lessor_schedule_id:
                return jsonify({
                    "error": "lessor_schedule_id is required"
                }), 400

            if payload.get("revenue_contract_id"):
                return jsonify({
                    "error": (
                        "A lessor invoice cannot be linked "
                        "to an IFRS 15 revenue contract"
                    )
                }), 400
    
        current_app.logger.info("create_invoice: before insert")
        invoice_id = db_service.insert_invoice_with_lines(company_id, header, mapped_lines)
        current_app.logger.info("create_invoice: inserted invoice_id=%s", invoice_id)
        print("[AFTER INSERT] invoice_id =", invoice_id, "revenue_contract_id =", header.get("revenue_contract_id"), flush=True)

        if not should_post:
            current_app.logger.info("create_invoice: review enabled, skipping auto-post")
            inv_saved = db_service.get_invoice_with_lines(company_id, invoice_id) or {}
            current_app.logger.info("create_invoice: fetched saved invoice invoice_id=%s", invoice_id)

            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="ar",
                action="create",
                severity="info",
                entity_type="invoice",
                entity_id=str(invoice_id),
                entity_ref=str(inv_saved.get("number") or inv_no or invoice_id),
                customer_id=int(cust_id),
                amount=float(inv_saved.get("total_amount") or inv_saved.get("gross_amount") or 0.0),
                currency=inv_saved.get("currency"),
                before_json={},
                after_json=inv_saved,
                message=f"Invoice created ({status_to_save})",
            )

            current_app.logger.info("create_invoice: returning saved invoice invoice_id=%s", invoice_id)
            return jsonify(inv_saved), 201

        current_app.logger.info("create_invoice: auto-post starting invoice_id=%s", invoice_id)

        current_app.logger.info("create_invoice: before approve_invoice")
        db_service.approve_invoice(company_id, invoice_id, user.get("id"))
        current_app.logger.info("create_invoice: approve_invoice done")

        current_app.logger.info("create_invoice: before get_invoice_with_lines for posting")
        inv = db_service.get_invoice_with_lines(company_id, invoice_id)
        current_app.logger.info("create_invoice: get_invoice_with_lines done inv_is_none=%s", inv is None)

        print("[GET_INVOICE_WITH_LINES] revenue_contract_id =", inv.get("revenue_contract_id") if inv else None, flush=True)
        current_app.logger.warning(
            "[GET_INVOICE_WITH_LINES] invoice_id=%s revenue_contract_id=%r inv_keys=%s",
            invoice_id,
            inv.get("revenue_contract_id") if inv else None,
            list(inv.keys()) if isinstance(inv, dict) else None,
        )

        if not inv:
            raise RuntimeError(f"Invoice {invoice_id} inserted but could not be reloaded")

        current_app.logger.info("create_invoice: before build_invoice_journal_lines")
        payload2 = build_invoice_journal_lines(inv, company_id)
        current_app.logger.info(
            "create_invoice: journal payload built ar_account=%r lines=%s",
            payload2.get("ar_account") if isinstance(payload2, dict) else None,
            len((payload2 or {}).get("lines") or []) if isinstance(payload2, dict) else None
        )

        if not isinstance(payload2, dict):
            raise RuntimeError("build_invoice_journal_lines returned invalid payload")
        if not payload2.get("lines"):
            raise RuntimeError("No journal lines generated for invoice posting")

        current_app.logger.info("create_invoice: before get_company_profile")
        company = db_service.get_company_profile(company_id) or {}
        current_app.logger.info("create_invoice: get_company_profile done")

        owner_user_id = company.get("owner_user_id")
        is_owner = owner_user_id is not None and str(owner_user_id) == str(user.get("id"))

        role = (user.get("user_role") or user.get("role") or "").lower()
        can_override = role in {"cfo", "admin"} or is_owner
        enforce_credit = not can_override

        current_app.logger.info(
            "create_invoice: before post_invoice_to_gl invoice_id=%s user_id=%r role=%s enforce_credit=%s require_approved=%s",
            invoice_id, user.get("id"), role, enforce_credit, require_customer_approved
        )
        journal_id = db_service.post_invoice_to_gl(
            company_id,
            invoice_id,
            payload2["lines"],
            ar_account=payload2.get("ar_account"),
            enforce_credit=enforce_credit,
            require_approved=require_customer_approved,
        )
        current_app.logger.info("create_invoice: post_invoice_to_gl done journal_id=%s", journal_id)

        db_service.execute_sql(
            f"""
            UPDATE company_{company_id}.invoices
            SET status='posted',
                posted_journal_id=%s,
                updated_at=NOW()
            WHERE company_id=%s
            AND id=%s
            AND posted_journal_id IS NULL
            """,
            (int(journal_id), int(company_id), int(invoice_id)),
        )

        try:
            revenue_contract_id = header.get("revenue_contract_id")

            if revenue_contract_id:
                contract = db_service.get_revenue_contract(
                    company_id=company_id,
                    contract_id=int(revenue_contract_id),
                ) or {}

                payload_json = contract.get("payload_json") or {}
                if isinstance(payload_json, str):
                    import json
                    try:
                        payload_json = json.loads(payload_json)
                    except Exception:
                        payload_json = {}

                settlement_pattern = (
                    payload_json.get("settlement_pattern")
                    or payload_json.get("ifrs15_settlement_pattern")
                    or ""
                ).strip().lower()

                obligation_ids = [
                    l.get("revenue_obligation_id")
                    for l in inv.get("lines", [])
                    if l.get("revenue_obligation_id")
                ]

                if obligation_ids:
                    billing_amount = float(sum(
                        float(l.get("net_amount") or 0.0)
                        for l in inv.get("lines", [])
                        if l.get("revenue_obligation_id")
                    ))

                    billing_obligation_id = (
                        int(obligation_ids[0])
                        if len(set(obligation_ids)) == 1
                        else None
                    )
                else:
                    billing_amount = float(sum(
                        float(l.get("net_amount") or 0.0)
                        for l in inv.get("lines", [])
                    ))

                    billing_obligation_id = None

                db_service.record_revenue_billing_event(
                    company_id=company_id,
                    contract_id=int(revenue_contract_id),
                    data={
                        "event_date": header.get("invoice_date"),
                        "event_type": "invoice",
                        "source_invoice_id": int(invoice_id),
                        "obligation_id": billing_obligation_id,
                        "amount": billing_amount,
                        "currency": inv.get("currency") or header.get("currency") or "USD",
                        "notes": f"From AR invoice {inv.get('number') or invoice_id}",
                        "payload_json": {
                            "customer_id": int(cust_id),
                            "source": "ar_invoice_post",
                            "invoice_number": inv.get("number"),
                            "auto_allocate": True,
                            "settlement_pattern": settlement_pattern,
                            "journal_id": int(journal_id),
                        },
                    },
                    user_id=int(user.get("id") or 0),
                )

                if settlement_pattern == "cash_before_service":
                    try:
                        allocation = db_service.auto_allocate_customer_advance_to_invoice(
                            company_id=company_id,
                            customer_id=int(cust_id),
                            contract_id=int(revenue_contract_id),
                            invoice_id=int(invoice_id),
                            amount=float(billing_amount),
                            currency=inv.get("currency") or header.get("currency") or "USD",
                        )

                        current_app.logger.info("Auto allocation result: %s", allocation)

                        if allocation.get("remaining", 0) > 0:
                            current_app.logger.warning(
                                "Invoice not fully covered by advances. Remaining: %s",
                                allocation["remaining"],
                            )

                    except Exception:
                        current_app.logger.exception(
                            "create_invoice: auto allocation failed | invoice_id=%s contract_id=%r",
                            invoice_id,
                            revenue_contract_id,
                        )

        except Exception:
            current_app.logger.exception(
                "create_invoice: record_revenue_billing_event failed | invoice_id=%s contract_id=%r",
                invoice_id,
                header.get("revenue_contract_id"),
            )

        current_app.logger.info("create_invoice: before reload posted invoice")
        posted = db_service.get_invoice_with_lines(company_id, invoice_id) or {}
        current_app.logger.info("create_invoice: reload posted invoice done")
        posted["_posted_journal_id"] = journal_id

        safe_journal_id = None
        try:
            jid = int(journal_id) if journal_id else None
            if jid:
                row = db_service.fetch_one(
                    f"SELECT id FROM company_{company_id}.journal WHERE id = %s LIMIT 1",
                    (jid,),
                )
                if row:
                    safe_journal_id = jid
                else:
                    current_app.logger.warning(
                        "create_invoice: journal_id %s not found in company_%s.journal after posting",
                        jid,
                        company_id,
                    )
        except Exception:
            current_app.logger.exception(
                "create_invoice: failed checking journal existence | invoice_id=%s journal_id=%r",
                invoice_id,
                journal_id,
            )
            safe_journal_id = None

        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="ar",
                action="post",
                severity="info",
                entity_type="invoice",
                entity_id=str(invoice_id),
                entity_ref=str(posted.get("number") or inv_no or invoice_id),
                journal_id=safe_journal_id,
                customer_id=int(cust_id),
                amount=float(posted.get("total_amount") or posted.get("gross_amount") or 0.0),
                currency=posted.get("currency"),
                before_json={},
                after_json=posted,
                message="Invoice posted to GL (auto-post: review disabled)",
            )
        except Exception:
            current_app.logger.exception(
                "create_invoice: audit_log failed after posting | invoice_id=%s journal_id=%r",
                invoice_id,
                journal_id,
            )

        return jsonify(posted), 201

    except Exception:
        current_app.logger.exception(
            "❌ create_invoice crashed | cid=%s | user=%r | payload=%r",
            cid,
            getattr(g, "current_user", {}),
            request.get_json(silent=True),
        )

        return jsonify({
            "ok": False,
            "error": "invoice_creation_failed",
            "message": "An unexpected error occurred while creating the invoice.",
        }), 500

# =========================================================
# INVOICES (UPDATE)
# =========================================================
@app.route("/api/companies/<int:cid>/invoices/<int:invoice_id>", methods=["PUT"])
@require_auth
def update_invoice(cid: int, invoice_id: int):
    try:
        company_id=int(cid)
        user=getattr(g,"current_user",{}) or {}

        if int(user.get("company_id") or 0)!=company_id:
            return jsonify({"error":"Not authorised for this company"}),403

        def money(v):
            return float(Decimal(str(v or 0)).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP))

        def num(v,default=0.0):
            try: return float(str(v).replace(",","").strip()) if v is not None else default
            except (TypeError,ValueError): return default

        def opt_int(v):
            try: return int(v) if v not in (None,"",0,"0") else None
            except (TypeError,ValueError): return None

        def rate(v):
            v=num(v)
            return max(0.0,min(1.0,v/100 if v>1 else v))

        payload=request.get_json(silent=True) or {}
        current_app.logger.info("update_invoice payload=%r",payload)

        inv=db_service.get_invoice_with_lines(company_id,int(invoice_id))
        if not inv:
            return jsonify({"error":"Invoice not found"}),404

        if str(inv.get("status") or "").lower()=="posted" or inv.get("posted_journal_id"):
            return jsonify({"error":"Posted invoices cannot be edited"}),409

        def val(name,default=None):
            return payload.get(name) if name in payload else inv.get(name,default)

        source=str(val("source","") or "").strip().lower()
        module_name=str(val("module_name","") or "").strip().lower()
        lessor_lease_id=opt_int(val("lessor_lease_id"))
        lessor_schedule_id=opt_int(val("lessor_schedule_id") or val("source_id"))

        is_lessor=(
            source=="lessor_lease_billing"
            or module_name=="ifrs16_lessor"
            or bool(lessor_lease_id)
            or bool(lessor_schedule_id)
        )

        credit_account_code=str(val("credit_account_code","") or "").strip()
        schema=db_service.company_schema(company_id)

        if is_lessor:
            source="lessor_lease_billing"
            module_name="ifrs16_lessor"

            if not lessor_lease_id:
                return jsonify({"error":"lessor_lease_id is required"}),400
            if not lessor_schedule_id:
                return jsonify({"error":"lessor_schedule_id is required"}),400
            if opt_int(payload.get("revenue_contract_id")):
                return jsonify({"error":"A lessor invoice cannot be linked to an IFRS 15 revenue contract"}),400

            period=db_service.fetch_one(
                f"""
                SELECT s.id,s.invoice_id,s.is_active,l.status AS lease_status
                FROM {schema}.lessor_lease_schedule s
                JOIN {schema}.lessor_leases l
                  ON l.company_id=s.company_id
                 AND l.id=s.lessor_lease_id
                WHERE s.company_id=%s
                  AND s.id=%s
                  AND s.lessor_lease_id=%s
                LIMIT 1
                """,
                (company_id,lessor_schedule_id,lessor_lease_id),
            )

            if not period:
                return jsonify({"error":"Lessor schedule period not found for the supplied lease"}),404
            if not period.get("is_active"):
                return jsonify({"error":"Inactive lessor schedule versions cannot be invoiced"}),409
            if period.get("invoice_id") and int(period["invoice_id"])!=int(invoice_id):
                return jsonify({
                    "error":"This lessor schedule period is already linked to another invoice",
                    "invoice_id":int(period["invoice_id"]),
                }),409
            if str(period.get("lease_status") or "").lower() not in {"active","commenced"}:
                return jsonify({"error":"Only an active or commenced lessor lease can be invoiced"}),409

        pol=company_policy(company_id)
        mode=str(pol.get("mode") or "owner_managed").strip().lower()
        policy=pol.get("policy") or {}
        company_profile=pol.get("company") or {}
        review_enabled=bool(
            pol.get("review_enabled")
            or policy.get("review_enabled")
            or policy.get("invoice_review_enabled")
            or policy.get("require_invoice_review")
        )

        requested_status=str(payload.get("status") or inv.get("status") or "draft").strip().lower()
        status_to_save=requested_status if requested_status in {"draft","pending_approval","approved","posted"} else "draft"
        should_post=not review_enabled or requested_status=="posted"

        if not review_enabled and requested_status!="posted":
            status_to_save="draft"

        customer_id=opt_int(payload.get("customer_id") or inv.get("customer_id"))
        lines=payload.get("lines") if "lines" in payload else inv.get("lines") or []

        if not customer_id:
            return jsonify({"error":"customer_id is required"}),400
        if not isinstance(lines,list) or not lines:
            return jsonify({"error":"At least one line is required"}),400

        customer=db_service.get_customer(company_id,customer_id)
        if not customer:
            return jsonify({"error":"Customer not found"}),404
        if not customer.get("is_active",True):
            return jsonify({"error":"Customer is archived/inactive"}),409
        if str(customer.get("on_hold") or "no").lower() in {"yes","true","1"}:
            return jsonify({"error":"Customer is on hold"}),409

        require_approved=must_approve_customer_before_invoicing(mode,policy)
        credit_status=str(customer.get("credit_status") or "").lower()

        if require_approved and status_to_save!="draft" and credit_status not in {"approved","cod_only"}:
            return jsonify({
                "error":"Customer is not approved for invoicing.",
                "customer_id":customer_id,
                "credit_status":credit_status,
                "status_to_save":status_to_save,
                "mode":mode,
            }),409

        invoice_number=str(val("number","") or "").strip()
        if status_to_save=="draft":
            invoice_number=""

        discount_rate=rate(
            payload.get("discount_rate")
            if "discount_rate" in payload
            else payload.get("discount",inv.get("discount_rate"))
        )
        other_amount=num(
            payload.get("other")
            if "other" in payload
            else inv.get("other_amount",inv.get("other")),
            0.0,
        )

        if is_lessor and discount_rate:
            return jsonify({"error":"Discounts cannot be applied directly to a lessor schedule invoice"}),400
        if is_lessor and other_amount:
            return jsonify({"error":"Other charges cannot be applied directly to a lessor schedule invoice"}),400

        header={
            "customer_id":customer_id,
            "revenue_contract_id":None if is_lessor else opt_int(val("revenue_contract_id")),
            "invoice_date":payload.get("invoice_date") or payload.get("date") or inv.get("invoice_date"),
            "due_date":val("due_date"),
            "currency":val("currency"),
            "number":invoice_number or None if status_to_save!="draft" else None,
            "bank_account_id":opt_int(val("bank_account_id")),
            "notes":val("notes"),
            "status":status_to_save,
            "issued_by":user.get("id"),
            "issued_at":val("issued_at"),
            "discount_rate":discount_rate,
            "discount":discount_rate,
            "discount_amount":money(val("discount_amount",0)),
            "other":other_amount,
            "source":source or None,
            "source_id":lessor_schedule_id if is_lessor else opt_int(val("source_id")),
            "module_name":module_name or None,
            "posting_mode":"custom_accounts" if is_lessor else val("posting_mode"),
            "lessor_lease_id":lessor_lease_id if is_lessor else None,
            "lessor_schedule_id":lessor_schedule_id if is_lessor else None,
            "lease_classification":val("lease_classification") if is_lessor else None,
            "accounting_treatment":val("accounting_treatment") if is_lessor else None,
            "ar_account_code":val("ar_account_code") if is_lessor else None,
            "credit_account_code":credit_account_code if is_lessor else None,
            "vat_output_account_code":val("vat_output_account_code") if is_lessor else None,
            "updated_by_user_id":user.get("id"),
        }

        if not header["invoice_date"]:
            return jsonify({"error":"invoice_date is required"}),400

        try:
            inv_date=header["invoice_date"]
            inv_date=inv_date if isinstance(inv_date,date) else date.fromisoformat(str(inv_date)[:10])
            if inv_date>date.today():
                return jsonify({"error":"Invoice date cannot be in the future."}),400
        except (TypeError,ValueError):
            return jsonify({"error":"invoice_date must be YYYY-MM-DD"}),400

        mapped=[]

        for index,ln in enumerate(lines,1):
            if not isinstance(ln,dict):
                return jsonify({"error":f"Invoice line {index} must be an object"}),400

            item_type=str(ln.get("item_type") or ln.get("itemType") or "").strip().lower() or None
            if item_type not in {None,"","inventory","service","gl","lease"}:
                return jsonify({"error":"Invalid item_type on line","line_no":index}),400

            item_id=opt_int(ln.get("item_id") or ln.get("itemId"))
            if item_type in {"inventory","service"} and not item_id:
                return jsonify({"error":f"item_id is required when item_type='{item_type}'","line_no":index}),400

            if is_lessor:
                item_type,item_id="lease",lessor_lease_id

            quantity=num(ln.get("quantity") if ln.get("quantity") is not None else ln.get("qty"))
            unit_price=num(ln.get("unit_price") if ln.get("unit_price") is not None else ln.get("unitPrice"))
            line_discount=num(ln.get("discount_amount") if ln.get("discount_amount") is not None else ln.get("discountAmount"))

            if quantity<=0:
                return jsonify({"error":"quantity must be greater than zero","line_no":index}),400
            if unit_price<0 or line_discount<0:
                return jsonify({"error":"Invoice line amounts cannot be negative","line_no":index}),400
            if is_lessor and line_discount:
                return jsonify({"error":"Line discounts cannot be applied to a lessor schedule invoice","line_no":index}),400

            account_code=str(
                (credit_account_code if is_lessor else "")
                or ln.get("account_code")
                or ln.get("accountCode")
                or ""
            ).strip()

            if not account_code:
                return jsonify({"error":"Each invoice line must have account_code","line_no":index}),400

            vat_code=str(ln.get("vatCode") or ln.get("vat_code") or "STANDARD").strip().upper()
            vat_rate_raw=ln.get("vat_rate")

            if vat_rate_raw is None and vat_code=="CUSTOM":
                vat_rate_raw=(
                    ln.get("vatRate")
                    or ln.get("customVatRate")
                    or ln.get("vat_custom_rate")
                    or ln.get("vatCustomRate")
                )

            vat_rate=num(rate_from_code(vat_code) if vat_rate_raw is None else vat_rate_raw)
            net_amount=money(max(0,(quantity*unit_price)-line_discount))
            vat_amount=money(net_amount*(vat_rate/100))
            item_name=str(ln.get("item_name") or ln.get("itemName") or ln.get("item") or "").strip()

            mapped.append({
                "item_name":item_name or None,
                "item_type":item_type,
                "item_id":item_id,
                "item_code":str(ln.get("item_code") or ln.get("itemCode") or ln.get("sku") or ln.get("code") or "").strip() or None,
                "vat_code":vat_code,
                "description":str(ln.get("description") or item_name or "").strip(),
                "revenue_obligation_id":None if is_lessor else opt_int(ln.get("revenue_obligation_id") or ln.get("obligation_id")),
                "account_code":account_code,
                "quantity":quantity,
                "unit_price":unit_price,
                "discount_amount":money(line_discount),
                "net_amount":net_amount,
                "vat_rate":vat_rate,
                "vat_amount":vat_amount,
                "total_amount":money(net_amount+vat_amount),
                "source":"lessor_lease_billing" if is_lessor else ln.get("source"),
                "source_id":lessor_schedule_id if is_lessor else opt_int(ln.get("source_id")),
                "module_name":"ifrs16_lessor" if is_lessor else ln.get("module_name"),
                "lessor_lease_id":lessor_lease_id if is_lessor else opt_int(ln.get("lessor_lease_id")),
                "lessor_schedule_id":lessor_schedule_id if is_lessor else opt_int(ln.get("lessor_schedule_id")),
            })
        if is_lessor:
            lease_classification=str(
                header.get("lease_classification") or ""
            ).strip().lower()

            if lease_classification not in {"operating","finance"}:
                return jsonify({
                    "error":"Invalid lessor lease classification"
                }),400

            credit_role=(
                "lessor_lease_income"
                if lease_classification=="operating"
                else "lessor_net_investment_current"
            )

            expected=db_service.get_posting_account_by_role(
                company_id,
                credit_role,
            )

            header["credit_account_code"]=expected["code"]

            for line in mapped:
                line["account_code"]=expected["code"]

        db_service.update_invoice_with_lines(company_id,int(invoice_id),header,mapped)

        updated=db_service.get_invoice_with_lines(company_id,int(invoice_id))
        if not updated:
            raise RuntimeError("Invoice was updated but could not be reloaded")

        if should_post:
            if not can_post_invoices(user,company_profile,mode):
                return jsonify({"error":"Not allowed to post invoices","mode":mode}),403

            journal=build_invoice_journal_lines(updated,company_id)
            if not journal.get("lines"):
                raise RuntimeError("No journal lines were generated")

            journal_id=db_service.post_invoice_to_gl(
                company_id,
                int(invoice_id),
                journal["lines"],
                ar_account=journal.get("ar_account"),
                enforce_credit=True,
                require_approved=require_approved,
            )

            if not is_lessor:
                try:
                    record_invoice_revenue_billing_and_allocation(
                        company_id=company_id,
                        invoice_id=int(invoice_id),
                        inv=updated,
                        journal_id=int(journal_id),
                        user_id=int(user.get("id") or 0),
                    )
                except Exception:
                    current_app.logger.exception(
                        "Invoice revenue billing/allocation failed | invoice_id=%s",
                        invoice_id,
                    )

            posted=db_service.get_invoice_with_lines(company_id,int(invoice_id)) or {}
            posted["_posted_journal_id"]=journal_id

            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="ifrs16_lessor" if is_lessor else "ar",
                action="post",
                severity="info",
                entity_type="invoice",
                entity_id=str(invoice_id),
                entity_ref=str(posted.get("number") or invoice_number or invoice_id),
                journal_id=int(journal_id) if journal_id else None,
                customer_id=customer_id,
                amount=float(posted.get("total_amount") or posted.get("gross_amount") or 0),
                currency=posted.get("currency"),
                before_json=inv,
                after_json=posted,
                message="Lessor invoice posted to GL" if is_lessor else "Invoice posted to GL (from update)",
            )
            return jsonify(posted),200

        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="ifrs16_lessor" if is_lessor else "ar",
            action="update",
            severity="info",
            entity_type="invoice",
            entity_id=str(invoice_id),
            entity_ref=str(updated.get("number") or invoice_number or invoice_id),
            customer_id=customer_id,
            amount=float(updated.get("total_amount") or updated.get("gross_amount") or 0),
            currency=updated.get("currency"),
            before_json=inv,
            after_json=updated,
            message="Lessor invoice updated" if is_lessor else "Invoice updated",
        )

        return jsonify(updated),200

    except ValueError as exc:
        return jsonify({"error":str(exc)}),400
    except Exception as exc:
        current_app.logger.exception(
            "update_invoice crashed | company_id=%s invoice_id=%s",
            cid,
            invoice_id,
        )
        return jsonify({
            "error":"Internal server error in update_invoice",
            "detail":str(exc),
        }),500
    
@app.route("/api/companies/<int:cid>/invoices/<int:invoice_id>/post", methods=["POST", "OPTIONS"])
@require_auth
def post_invoice(cid: int, invoice_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}

    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    # ==========================================================
    # POLICY
    # ==========================================================
    pol = company_policy(company_id) or {}
    mode = (pol.get("mode") or "owner_managed").strip().lower()
    policy = pol.get("policy") or {}
    company_profile = pol.get("company") or {}

    role = (user.get("user_role") or user.get("role") or user.get("system_role") or "").lower()

    owner_user_id = company_profile.get("owner_user_id")
    is_owner = owner_user_id is not None and str(owner_user_id) == str(user.get("id"))

    review_enabled = bool(
        policy.get("review_enabled", False)
        or policy.get("invoice_review_enabled", False)
        or policy.get("require_invoice_review", False)
        or pol.get("review_enabled")
    )

    # ✅ posting permission
    if mode in {"assisted", "controlled"}:
        allowed = is_owner or role in {"senior", "cfo", "admin"}
    else:
        allowed = True  # owner-managed (or tighten if you want)

    if not allowed:
        return jsonify({"error": "Not allowed to post invoices", "mode": mode, "role": role}), 403

    try:
        inv = db_service.get_invoice_with_lines(company_id, invoice_id)
        if not inv:
            return jsonify({"error": "Invoice not found"}), 404

        status = (inv.get("status") or "").strip().lower()

        if status == "posted":
            return jsonify({"error": "Invoice already posted"}), 409

        # ==========================================================
        # REVIEW FLOW
        # ==========================================================
        if review_enabled and mode in {"assisted", "controlled"}:

            # ✅ ASSISTED: owner can bypass review and post straight from draft
            if mode == "assisted" and is_owner:
                if status == "pending_approval":
                    db_service.approve_invoice(company_id, invoice_id, user.get("id"))
                    status = "approved"
                elif status == "draft":
                    # optional but clean: mark approved before posting
                    db_service.approve_invoice(company_id, invoice_id, user.get("id"))
                    status = "approved"

            # ✅ CONTROLLED: strict SoD (no bypass)
            else:
                # Maker must submit first
                if status == "draft":
                    return jsonify({"error": "Invoice must be submitted for approval first", "status": status}), 409

                # Approver can approve+post in one click
                if status == "pending_approval":
                    db_service.approve_invoice(company_id, invoice_id, user.get("id"))
                    status = "approved"

                if status != "approved":
                    return jsonify({"error": "Invoice cannot be posted in its current status", "status": status}), 409

        else:
            # Review disabled or owner-managed: can post draft/approved/pending_approval
            if status not in {"draft", "approved", "pending_approval"}:
                return jsonify({"error": f"Invoice status '{status}' cannot be posted"}), 409

            if status == "pending_approval":
                db_service.approve_invoice(company_id, invoice_id, user.get("id"))
                status = "approved"

        # ==========================================================
        # CREDIT ENFORCEMENT
        # ==========================================================
        require_approved = must_approve_customer_before_invoicing(mode, policy)
        can_override = role in {"cfo", "admin"} or is_owner
        enforce_credit = not can_override

        # ==========================================================
        # POST TO GL
        # ==========================================================
        payload = build_invoice_journal_lines(inv, company_id)
        journal_id = db_service.post_invoice_to_gl(
            company_id,
            invoice_id,
            payload["lines"],
            ar_account=payload.get("ar_account"),
            enforce_credit=enforce_credit,
            require_approved=require_approved,
        )

        try:
            record_invoice_revenue_billing_and_allocation(
                company_id=company_id,
                invoice_id=int(invoice_id),
                inv=inv,
                journal_id=int(journal_id),
                user_id=int(user.get("id") or 0),
            )
        except Exception:
            current_app.logger.exception(
                "post_invoice: invoice revenue billing/allocation failed | invoice_id=%s",
                invoice_id,
            )

        posted = db_service.get_invoice_with_lines(company_id, invoice_id) or {}
        posted["_posted_journal_id"] = journal_id

        safe_journal_id = None
        try:
            jid = int(journal_id) if journal_id else None
            if jid:
                row = db_service.fetch_one(
                    f"SELECT id FROM company_{company_id}.journal WHERE id = %s LIMIT 1",
                    (jid,),
                )
                if row:
                    safe_journal_id = jid
                else:
                    current_app.logger.warning(
                        "post_invoice: journal_id %s not found after posting",
                        jid,
                    )
        except Exception:
            current_app.logger.exception(
                "post_invoice: journal check failed | invoice_id=%s journal_id=%r",
                invoice_id,
                journal_id,
            )

        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="ar",
                action="post",
                severity="info",
                entity_type="invoice",
                entity_id=str(invoice_id),
                entity_ref=str((posted or {}).get("number") or invoice_id),
                journal_id=safe_journal_id,
                customer_id=int((posted or {}).get("customer_id") or 0) or None,
                amount=float((posted or {}).get("total_amount") or (posted or {}).get("gross_amount") or 0.0),
                currency=(posted or {}).get("currency"),
                before_json={},
                after_json=posted or {},
                message="Invoice posted to GL",
            )
        except Exception:
            current_app.logger.exception(
                "post_invoice: audit_log failed | invoice_id=%s journal_id=%r",
                invoice_id,
                journal_id,
            )
        return jsonify(posted), 200

    except Exception as e:
        current_app.logger.exception("post_invoice failed")
        msg = str(e)

        if msg.startswith("CREDIT_") or msg.startswith("CREDIT_LIMIT_EXCEEDED"):
            return jsonify({"error": "Credit limit blocked", "detail": msg}), 409

        if msg.startswith("INSUFFICIENT_STOCK|"):
            parts = {}
            try:
                for seg in msg.split("|")[1:]:
                    if "=" in seg:
                        k, v = seg.split("=", 1)
                        parts[k] = v
            except Exception:
                parts = {}

            return jsonify({
                "error": "Insufficient stock",
                "code": "INSUFFICIENT_STOCK",
                "detail": msg,
                "item_id": int(float(parts.get("item_id", 0) or 0)),
                "onhand": float(parts.get("onhand", 0) or 0),
                "requested": float(parts.get("req", 0) or 0),
            }), 409

        return jsonify({"error": "Server error", "detail": msg}), 500
    
@app.route("/api/companies/<int:cid>/invoices/post_approved", methods=["POST", "OPTIONS"])
@require_auth
def post_all_approved_invoices(cid: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}

    if user.get("company_id") != company_id:
        return _corsify(jsonify({"error": "Not authorised for this company"})), 403

    # --------------------------
    # Policy (single source of truth)
    # --------------------------
    pol = company_policy(company_id)
    mode = (pol.get("mode") or "owner_managed").strip().lower()
    company_profile = pol.get("company") or {}
    policy = pol.get("policy") or {}

    # ✅ Review flag (match update_invoice)
    review_enabled = bool(
        policy.get("review_enabled", False) or
        policy.get("invoice_review_enabled", False) or
        policy.get("require_invoice_review", False)
    )

    # Posting permission (your helper)
    if not can_post_invoices(user, company_profile, mode):
        return _corsify(jsonify({"error": "Not allowed to post invoices", "mode": mode})), 403

    # --------------------------
    # Decide which statuses are eligible
    # --------------------------
    statuses = ["approved"]
    if mode == "owner_managed" and not review_enabled:
        # ✅ owner-managed auto-post workflow allowed ONLY when review is OFF
        statuses = ["draft", "approved"]

    try:
        rows = db_service.list_company_invoices_by_status(company_id, statuses) or []
        results = {"posted": [], "failed": []}

        try:
            inv_after = db_service.get_invoice_with_lines(company_id, int(invoice_id)) or {}
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="ar",
                action="post",
                severity="info",
                entity_type="invoice",
                entity_id=str(invoice_id),
                entity_ref=str(inv_after.get("number") or invoice_id),
                journal_id=int(journal_id) if journal_id else None,
                customer_id=int(inv_after.get("customer_id") or 0) or None,
                amount=float(inv_after.get("total_amount") or inv_after.get("gross_amount") or 0.0),
                currency=inv_after.get("currency"),
                before_json={},
                after_json=inv_after,
                message="Invoice posted to GL (bulk post approved)",
            )
        except Exception:
            pass

        # Enforce customer approval only if policy requires it
        require_customer_approved = must_approve_customer_before_invoicing(mode, policy)

        # Optional: credit enforcement override
        role = (user.get("user_role") or user.get("role") or "").lower()
        is_owner = False
        owner_user_id = (company_profile or {}).get("owner_user_id")
        if owner_user_id is not None and str(owner_user_id) == str(user.get("id")):
            is_owner = True

        can_override_credit = is_owner or role in {"cfo", "admin"}
        enforce_credit = not can_override_credit

        for inv_row in rows:
            invoice_id = inv_row.get("id")
            try:
                if not invoice_id:
                    raise ValueError("Missing invoice id in list row")

                inv = db_service.get_invoice_with_lines(company_id, int(invoice_id))
                if not inv:
                    raise ValueError("Invoice not found")

                status = (inv.get("status") or "").strip().lower()

                # ✅ If already posted, skip cleanly
                if status == "posted" or inv.get("posted_journal_id"):
                    results["failed"].append({
                        "invoice_id": int(invoice_id),
                        "error": "Invoice already posted"
                    })
                    continue

                # ✅ Owner-managed: only auto-approve drafts when review is OFF
                if mode == "owner_managed" and status == "draft" and not review_enabled:
                    db_service.execute_sql(
                        f"""
                        UPDATE company_{company_id}.invoices
                        SET status='approved', updated_at=NOW()
                        WHERE id=%s AND status='draft'
                        """,
                        (int(invoice_id),)
                    )
                    inv = db_service.get_invoice_with_lines(company_id, int(invoice_id)) or inv
                    status = (inv.get("status") or "").strip().lower()

                # After potential auto-approve, only approved invoices can be posted
                if status != "approved":
                    raise ValueError(f"Invoice status '{status}' cannot be posted in this workflow")

                # ✅ Customer approval enforcement (if policy says so)
                if require_customer_approved:
                    customer_id = inv.get("customer_id")
                    cust = db_service.get_customer(company_id, int(customer_id)) if customer_id else None
                    cust = cust or {}

                    credit_status = (cust.get("credit_status") or "").strip().lower()
                    on_hold = (cust.get("on_hold") or "no").strip().lower()

                    if on_hold in {"yes", "true", "1"}:
                        raise ValueError("CREDIT_BLOCKED|CUSTOMER_ON_HOLD")
                    if credit_status not in {"approved", "cod_only"}:
                        raise ValueError("CREDIT_BLOCKED|CUSTOMER_NOT_APPROVED")

                payload = build_invoice_journal_lines(inv, company_id)

                journal_id = db_service.post_invoice_to_gl(
                    company_id,
                    int(invoice_id),
                    payload["lines"],
                    enforce_credit=enforce_credit,
                    require_approved=require_customer_approved,
                )

                db_service.set_invoice_status(company_id, int(invoice_id), "posted")

                results["posted"].append({
                    "invoice_id": int(invoice_id),
                    "journal_id": int(journal_id),
                })

            except Exception as ex:
                current_app.logger.exception("Bulk post invoice failed")
                results["failed"].append({
                    "invoice_id": int(invoice_id) if invoice_id else None,
                    "error": str(ex),
                })

        return _corsify(jsonify(results)), 200

    except Exception as e:
        current_app.logger.exception("post_all_approved_invoices failed")
        return _corsify(jsonify({"error": "Server error", "detail": str(e)})), 500

@app.route("/api/companies/<int:cid>/invoices/<int:invoice_id>/mark_posted", methods=["PUT", "OPTIONS"])
@require_auth
def mark_invoice_posted(cid: int, invoice_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}

    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised"}), 403

    try:
        payload = request.get_json(silent=True) or {}
        journal_id = payload.get("journal_id")  # optional

        db_service.mark_invoice_posted(company_id, invoice_id, journal_id)

        inv = db_service.get_invoice_with_lines(company_id, invoice_id)
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="ar",
            action="mark_posted",
            severity="warning",
            entity_type="invoice",
            entity_id=str(invoice_id),
            entity_ref=str((inv or {}).get("number") or invoice_id),
            journal_id=int(journal_id) if journal_id else None,
            customer_id=int((inv or {}).get("customer_id") or 0) or None,
            amount=float((inv or {}).get("total_amount") or (inv or {}).get("gross_amount") or 0.0),
            currency=(inv or {}).get("currency"),
            before_json={},
            after_json=inv or {},
            message="Invoice manually marked as posted",
        )
        return jsonify({"ok": True, "invoice": inv}), 200

    except Exception as e:
        current_app.logger.exception("mark_invoice_posted failed")
        return jsonify({"error": str(e)}), 500

@app.route("/api/companies/<int:company_id>/invoices/post_backlog", methods=["POST"])
@require_auth
def post_invoice_backlog(company_id: int):
    payload = request.jwt_payload
    user_company_id = payload.get("company_id")
    if user_company_id is not None and int(user_company_id) != int(company_id):
        return jsonify({"error": "Forbidden"}), 403

    body = request.get_json(silent=True) or {}
    include_drafts = bool(body.get("include_drafts", False))
    enforce_credit = bool(body.get("enforce_credit", False))  # in DEV you can set false

    schema = db_service.company_schema(company_id)

    # Fetch backlog list
    rows = db_service.fetch_all(f"""
        SELECT id, status, number, invoice_date, created_at
        FROM {schema}.invoices
        WHERE posted_journal_id IS NULL
        ORDER BY invoice_date NULLS LAST, created_at NULLS LAST, id ASC;
    """) or []

    results = []
    posted_count = 0
    skipped_count = 0
    failed_count = 0

    for r in rows:
        inv_id = int(r.get("id"))
        status = (r.get("status") or "").lower().strip()
        number = r.get("number")

        # skip drafts unless explicitly included
        if status == "draft" and not include_drafts:
            skipped_count += 1
            results.append({
                "invoice_id": inv_id,
                "number": number,
                "status": status,
                "action": "skipped",
                "reason": "draft (include_drafts=false)",
            })
            continue

        try:
            inv = db_service.get_invoice_with_lines(company_id, inv_id)
            if not inv or not isinstance(inv, dict):
                raise ValueError("Invoice could not be loaded")

            if not inv.get("lines"):
                raise ValueError("Invoice has no lines")

            payload2 = build_invoice_journal_lines(inv, company_id)

            # Post to GL (this stamps posted_journal_id + status='posted' inside)
            jid = db_service.post_invoice_to_gl(
                company_id,
                inv_id,
                payload2["lines"],
                enforce_credit=enforce_credit,
                require_approved=False,
            )

            # ✅ ADD THIS HERE
            try:
                inv_after = db_service.get_invoice_with_lines(company_id, inv_id) or inv

                record_invoice_revenue_billing_and_allocation(
                    company_id=company_id,
                    invoice_id=int(inv_id),
                    inv=inv_after,
                    journal_id=int(jid),
                    user_id=int((payload.get("sub") or 0)),
                )
            except Exception:
                current_app.logger.exception(
                    "post_invoice_backlog: invoice revenue billing/allocation failed | invoice_id=%s",
                    inv_id,
                )

            posted_count += 1

            posted_count += 1
            results.append({
                "invoice_id": inv_id,
                "number": number,
                "action": "posted",
                "posted_journal_id": int(jid),
            })

            try:
                inv_after = db_service.get_invoice_with_lines(company_id, inv_id) or {}
                db_service.audit_log(
                    company_id=company_id,
                    actor_user_id=int((payload.get("sub") or 0)),  # since this route uses jwt_payload
                    module="ar",
                    action="post",
                    severity="info",
                    entity_type="invoice",
                    entity_id=str(inv_id),
                    entity_ref=str(inv_after.get("number") or inv_id),
                    journal_id=int(jid) if jid else None,
                    customer_id=int(inv_after.get("customer_id") or 0) or None,
                    amount=float(inv_after.get("total_amount") or inv_after.get("gross_amount") or 0.0),
                    currency=inv_after.get("currency"),
                    before_json={},
                    after_json=inv_after,
                    message="Invoice posted to GL (backlog)",
                )
            except Exception:
                pass

        except Exception as e:
            failed_count += 1
            current_app.logger.exception("❌ backlog post failed invoice_id=%s", inv_id)
            results.append({
                "invoice_id": inv_id,
                "number": number,
                "action": "failed",
                "detail": str(e),
            })

    return jsonify({
        "ok": True,
        "company_id": company_id,
        "counts": {
            "total_found": len(rows),
            "posted": posted_count,
            "skipped": skipped_count,
            "failed": failed_count,
        },
        "results": results,
    }), 200

@app.route("/api/companies/<int:cid>/invoices/<int:invoice_id>/submit", methods=["POST"])
@require_auth
def submit_invoice_for_approval(cid: int, invoice_id: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}

    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    pol = company_policy(company_id)
    policy = pol.get("policy") or {}
    review_enabled = bool(
        policy.get("review_enabled", False)
        or policy.get("invoice_review_enabled", False)
        or policy.get("require_invoice_review", False)
        or pol.get("review_enabled")
    )

    if not review_enabled:
        return jsonify({"error": "Review not enabled"}), 409

    inv = db_service.get_invoice_with_lines(company_id, invoice_id)
    if not inv:
        return jsonify({"error": "Invoice not found"}), 404

    status = (inv.get("status") or "").lower()
    if status not in {"draft", "rejected"}:
        return jsonify({"error": f"Cannot submit invoice in status '{status}'"}), 409

    # create approval request + set invoice status pending_approval
    req = db_service.create_approval_request(
        company_id,
        entity_type="invoice",
        entity_id=str(invoice_id),
        entity_ref=str(inv.get("number") or f"INV-{invoice_id}"),
        module="ar",
        action="post_invoice",  # ✅ THIS MUST MATCH decide() executor
        requested_by_user_id=int(user.get("id") or 0),
        amount=float(inv.get("total_amount") or 0.0),
        currency=inv.get("currency"),
        risk_level="medium",
        dedupe_key=f"{company_id}:ar:post_invoice:invoice:{invoice_id}",  # ✅ match too
        payload_json={"invoice": {"id": invoice_id, "number": inv.get("number"), "status": inv.get("status")}},
    )

    db_service.set_invoice_status(company_id, invoice_id, "pending_approval")
    return jsonify({"ok": True, "invoice_id": invoice_id, "status": "pending_approval", "approval_request": req}), 202

# -------------------------
# List invoices
# -------------------------
@app.route("/api/companies/<int:company_id>/invoices", methods=["GET"])
@require_auth
def list_invoices(company_id):

    # -----------------------------
    # Read filters
    # -----------------------------
    status_csv = request.args.get("status")          # ?status=draft,pending_approval
    status_list = request.args.getlist("status")     # ?status=draft&status=pending_approval

    customer_id = request.args.get("customer_id")
    limit = int(request.args.get("limit") or 200)

    # -----------------------------
    # Build statuses array
    # -----------------------------
    statuses = []

    if status_list:
        # repeated params format
        statuses.extend([s.strip() for s in status_list if s.strip()])

    if status_csv:
        # CSV format
        statuses.extend([s.strip() for s in status_csv.split(",") if s.strip()])

    # dedupe
    statuses = list(set(statuses)) if statuses else None

    # -----------------------------
    # Query DB
    # -----------------------------
    rows = db_service.list_company_invoices_filtered(
        company_id=company_id,
        statuses=statuses,
        customer_id=int(customer_id) if customer_id else None,
        limit=limit,
    )

    return jsonify(rows), 200


@app.route("/favicon.ico")
def favicon():
    return ("", 204)

@app.route("/api/companies/<int:company_id>/invoices/<int:invoice_id>/pdf", methods=["GET"])
def invoice_pdf(company_id: int, invoice_id: int):
    token = request.args.get("t", "")
    payload = verify_invoice_pdf_token(token)
    if not payload:
        return jsonify({"error": "Invalid or expired token"}), 401

    if payload["company_id"] != company_id or payload["invoice_id"] != invoice_id:
        return jsonify({"error": "Token mismatch"}), 403

    inv = db_service.get_invoice_full(company_id, invoice_id)
    if not inv:
        return jsonify({"error": "Invoice not found"}), 404

    inv["branding"] = db_service.get_company_branding(company_id) or {}

    # ✅ Fetch company (logo_url, address, vat, reg, etc.)
    company = db_service.fetch_one(
        """
        SELECT
          id,
          name,
          company_reg_no,
          vat,
          tin,
          company_email,
          company_phone,
          physical_address,
          postal_address,
          logo_url
        FROM public.companies
        WHERE id = %s
        LIMIT 1;
        """,
        (company_id,),
    ) or {}

    logo_url = (company.get("logo_url") or "").strip()
    if logo_url and logo_url.startswith("/uploads/company_logos/"):
        filename = os.path.basename(logo_url)
        company["logo_path"] = os.path.join(
            current_app.root_path,
            "uploads",
            "company_logos",
            filename,
        )

    print("EMAIL PDF logo_url:", company.get("logo_url"), flush=True)
    print("EMAIL PDF logo_path:", company.get("logo_path"), flush=True)
    print("EMAIL PDF logo exists:", os.path.exists(company.get("logo_path") or ""), flush=True)
    try:
        
        pdf_bytes = generate_invoice_pdf(inv, company=company)

        if not pdf_bytes or not isinstance(pdf_bytes, (bytes, bytearray)):
            return jsonify({"error": "PDF generator returned empty/invalid bytes"}), 500

        resp = make_response(pdf_bytes)
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = f'inline; filename="invoice-{invoice_id}.pdf"'
        resp.headers["Content-Length"] = str(len(pdf_bytes))
        resp.headers["Cache-Control"] = "no-store"
        return resp

    except Exception as e:
        print("❌ PDF generation failed:", str(e))
        print(traceback.format_exc())
        return jsonify({"error": "PDF generation failed", "detail": str(e)}), 500


@app.route("/api/companies/<int:company_id>/invoices/<int:invoice_id>/view", methods=["GET"])
@require_auth
def invoice_view(company_id: int, invoice_id: int):
    invoice = db_service.get_invoice_full(company_id, invoice_id)
    if not invoice:
        return jsonify({"error": "Invoice not found"}), 404

    branding = db_service.get_company_branding(company_id) or {}
    invoice["branding"] = branding
    invoice["branding_logo_url"] = branding.get("logo_url") or ""

    token = create_invoice_pdf_token(company_id=company_id, invoice_id=invoice_id, ttl_seconds=120)

    pdf_url = url_for(
        "invoice_pdf",
        company_id=company_id,
        invoice_id=invoice_id,
        t=token,
        _external=True
    )

    company = db_service.fetch_one(
        """
        SELECT
          id, name, company_reg_no, vat, tin,
          company_email, company_phone,
          physical_address, postal_address,
          logo_url
        FROM public.companies
        WHERE id = %s
        LIMIT 1;
        """,
        (company_id,),
    ) or {}

    # prefer branding logo over companies.logo_url
    company["branding_logo_url"] = branding.get("logo_url") or ""
    if not company.get("logo_url"):
        company["logo_url"] = branding.get("logo_url") or ""

    logger.debug("company in view=%s type=%s", company, type(company))
    logger.debug("branding in view=%s", branding)
    logger.debug("resolved logo=%s", company.get("logo_url") or company.get("branding_logo_url") or invoice.get("branding_logo_url"))

    html = render_template(
        "invoice_pdf.html",
        invoice=invoice,
        company=company,
        pdf_url=pdf_url
    )
    return make_response(html, 200)

@app.route("/api/companies/<int:company_id>/invoices/<int:invoice_id>/view_token", methods=["POST"])
@require_auth
def invoice_view_token(company_id: int, invoice_id: int):
    # adapt to your auth context
    user_id = getattr(g, "user_id", None) or (g.user.get("id") if hasattr(g, "user") else None)
    if not user_id:
        return jsonify({"error": "Missing user context"}), 401

    tok = make_invoice_view_token(company_id, invoice_id, int(user_id), ttl_seconds=120)
    return jsonify({"token": tok})

@app.route("/api/companies/<int:company_id>/quotes/<int:quote_id>/email", methods=["POST"])
@require_auth
def email_quote(company_id: int, quote_id: int):
    payload = request.jwt_payload
    actor_user_id = int(payload.get("sub") or 0)
    user_company_id = payload.get("company_id")
    if user_company_id is not None and user_company_id != company_id:
        return jsonify({"error": "Forbidden"}), 403

    body = request.get_json() or {}
    cc_role = body.get("cc_role")

    quote = db_service.get_quote_full(company_id, quote_id)
    if not quote:
        return jsonify({"error": "Quote not found"}), 404

    # ==============================
    # ✅ AUTO FETCH CUSTOMER CONTACT EMAIL (JSONB)
    # ==============================
    contact_email = None
    customer_id = quote.get("customer_id")

    if customer_id:
        schema = f"company_{company_id}"
        cust = db_service.fetch_one(
            f"SELECT contacts FROM {schema}.customers WHERE id = %s LIMIT 1;",
            (customer_id,),
        ) or {}
        contact_email = _first_contact_email_from_contacts(cust.get("contacts"))

    # ==============================
    # ✅ Choose recipient (priority order)
    # 1) UI override
    # 2) customer contacts JSONB
    # 3) quote.customer_email
    # 4) company email
    # ==============================
    preferred = (body.get("to_email") or body.get("to") or "").strip() or None
    to_email = (
        preferred
        or contact_email
        or quote.get("customer_email")
        or quote.get("company_email")
    )

    if not to_email:
        return jsonify({"error": "Customer has no email."}), 400

    cc_emails = get_company_emails_by_role(company_id, cc_role) if cc_role else []

    # ==============================
    # ✅ Fetch company (same as invoice)
    # ==============================
    company = db_service.fetch_one(
        """
        SELECT
          id,
          name,
          company_reg_no,
          vat,
          tin,
          company_email,
          company_phone,
          physical_address,
          postal_address,
          logo_url
        FROM public.companies
        WHERE id = %s
        LIMIT 1;
        """,
        (company_id,),
    ) or {}

    tenant_company_email = (
        company.get("company_email")
        or quote.get("company_email")
        or None
    )
    # ==============================
    # Format helpers
    # ==============================
    def _fmt(d):
        if isinstance(d, (datetime, date)):
            return d.strftime("%Y-%m-%d")
        return d or ""

    quote_date  = _fmt(quote.get("quotation_date"))
    valid_until = _fmt(quote.get("valid_until"))

    currency = quote.get("currency") or ""
    total = float(quote.get("total_amount") or quote.get("total") or 0.0)

    customer_name = quote.get("customer_name") or "Customer"
    company_name = quote.get("company_name") or company.get("name") or "Our Company"
    q_no = quote.get("number") or f"QTE-{quote_id}"

    subject = (body.get("subject") or "").strip() or f"Quotation {q_no} from {company_name}"
    extra_message = (body.get("message") or "").strip()

    text_body = f"""Dear {customer_name},

    Please find your quotation attached.

    Quote number : {q_no}
    Quote date   : {quote_date}
    Valid until  : {valid_until}

    Total: {currency} {total:,.2f}

    {extra_message if extra_message else ""}

    Kind regards,
    {company_name}
    """

    html_body = f"<pre style='font-family:system-ui,monospace'>{text_body}</pre>"

    db_service.audit_log(
        company_id=company_id,
        actor_user_id=actor_user_id,
        module="ar",
        action="email",
        severity="info",
        entity_type="quote",
        entity_id=str(quote_id),
        entity_ref=str(q_no),
        customer_id=int(customer_id) if customer_id else None,
        amount=float(total or 0.0),
        currency=str(currency).upper() if currency else None,
        before_json={},
        after_json={
            "to": to_email,
            "cc": cc_emails,
            "subject": subject,
            "source": "preferred" if preferred else ("contacts" if contact_email else "fallback"),
        },
        message=f"Quote emailed to {to_email}",
    )

    # ==============================
    # ✅ Generate PDF attachment
    # ==============================
    quote["branding"] = db_service.get_company_branding(company_id) or {}
    pdf_bytes = generate_invoice_pdf(quote, company)

    attachments = [(f"quote-{q_no}.pdf", pdf_bytes, "application/pdf")]

    send_mail(
        to_email=to_email,
        subject=subject,
        html_body=html_body,
        text_body=text_body,
        from_name=f"{company_name} via FinSage",
        reply_to=tenant_company_email,
        attachments=attachments,
    )

    for cc in cc_emails:
        send_mail(
            to_email=cc,
            subject=f"Copy: {subject}",
            html_body=html_body,
            text_body=text_body,
            from_name=f"{company_name} via FinSage",
            reply_to=tenant_company_email,
            attachments=attachments,
        )

    db_service.audit_log(
        company_id=company_id,
        actor_user_id=actor_user_id,
        module="ar",
        action="email_sent",
        severity="info",
        entity_type="quote",
        entity_id=str(quote_id),
        entity_ref=str(q_no),
        customer_id=int(customer_id) if customer_id else None,
        amount=float(total or 0.0),
        currency=str(currency).upper() if currency else None,
        before_json={},
        after_json={
            "to": to_email,
            "cc": cc_emails,
            "subject": subject,
            "attachment": f"quote-{q_no}.pdf",
        },
        message="Quote email sent successfully",
    )

    return jsonify({
        "ok": True,
        "sent_to": to_email,
        "source": "preferred" if preferred else ("contacts" if contact_email else "fallback")
    }), 200


@app.route("/api/companies/<int:company_id>/branding/logo", methods=["POST"])
@require_auth
def upload_company_logo(company_id: int):
    payload = request.jwt_payload
    actor_user_id = int(payload.get("sub") or 0)

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Missing file (field name must be 'file')"}), 400

    original = secure_filename(file.filename or "")
    ext = os.path.splitext(original)[1].lower()
    if ext not in ALLOWED_LOGO_EXTS:
        return jsonify({
            "error": "Invalid file type",
            "allowed": sorted(ALLOWED_LOGO_EXTS),
        }), 400

    folder = os.path.join(current_app.root_path, "uploads", "company_logos")
    os.makedirs(folder, exist_ok=True)

    # cache-bust friendly filename
    filename = f"{int(company_id)}_{int(time.time())}{ext}"
    path = os.path.join(folder, filename)
    file.save(path)

    # URL served by endpoint below
    logo_url = f"/uploads/company_logos/{filename}"

    branding = db_service.upsert_company_branding(company_id, {
        "logo_url": logo_url,
        "logo_path": path,
    })
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=actor_user_id,
        module="settings",
        action="upload_logo",
        severity="info",
        entity_type="branding",
        entity_id=str(company_id),
        entity_ref=str(filename),
        after_json={"logo_url": logo_url},
        message="Company logo uploaded",
    )

    return jsonify({"ok": True, "branding": branding}), 200

@app.route("/uploads/company_logos/<path:filename>", methods=["GET"])
def serve_company_logos(filename):
    folder = os.path.join(current_app.root_path, "uploads", "company_logos")
    return send_from_directory(folder, filename)


@app.route("/api/companies/<int:company_id>/branding", methods=["PUT"])
@require_auth
def update_company_branding(company_id: int):
    payload_jwt = request.jwt_payload
    actor_user_id = int(payload_jwt.get("sub") or 0)

    payload = request.get_json(silent=True) or {}
    branding = db_service.upsert_company_branding(company_id, payload)
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=actor_user_id,
        module="settings",
        action="update_branding",
        severity="info",
        entity_type="branding",
        entity_id=str(company_id),
        after_json=payload if isinstance(payload, dict) else {},
        message="Company branding updated",
    )

    return jsonify({"ok": True, "branding": branding}), 200

@app.route("/api/companies/<int:company_id>/branding", methods=["GET"])
@require_auth
def get_company_branding(company_id: int):
    row = db_service.get_company_branding(company_id) or {"company_id": int(company_id)}
    return jsonify(row), 200


@app.route("/api/companies/<int:cid>/customers/deduplicate", methods=["POST"])
@require_auth
def dedupe_customers(cid: int):
    company_id = int(cid)

    # optional: lock this to CFO/admin
    user = getattr(g, "current_user", {}) or {}
    actor_user_id = int(user.get("id") or 0)
    role = (user.get("user_role") or "").lower()
    if role not in {"cfo", "admin", "senior"}:
        return jsonify({"error": "Forbidden"}), 403

    result = db_service.archive_duplicate_customers(company_id)
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=actor_user_id,
        module="ar",
        action="dedupe_customers",
        severity="info",
        entity_type="customer",
        entity_id="bulk",
        after_json=result if isinstance(result, dict) else {"result": result},
        message="Customer deduplication executed",
    )

    return jsonify(result), 200



@app.route("/api/companies/<int:company_id>/invoices/<int:invoice_id>/email", methods=["POST"])
@require_auth
def email_invoice(company_id: int, invoice_id: int):
    payload = request.jwt_payload
    actor_user_id = int(payload.get("sub") or 0)
    user_company_id = payload.get("company_id")
    if user_company_id is not None and user_company_id != company_id:
        return jsonify({"error": "Forbidden"}), 403

    body = request.get_json() or {}
    cc_role = body.get("cc_role")

    inv = db_service.get_invoice_with_relations(company_id, invoice_id)
    if not inv:
        return jsonify({"error": "Invoice not found"}), 404

    preferred = (body.get("to_email") or "").strip() or None
    contact_email = (_first_contact_email(inv) or "").strip() or None
    customer_company_email = (inv.get("customer_email") or "").strip() or None

    company = db_service.fetch_one(
        """
        SELECT
          id,
          name,
          company_reg_no,
          vat,
          tin,
          company_email,
          company_phone,
          physical_address,
          postal_address,
          logo_url
        FROM public.companies
        WHERE id = %s
        LIMIT 1;
        """,
        (company_id,),
    ) or {}

    tenant_company_email = (inv.get("company_email") or company.get("company_email") or "").strip() or None

    # To = customer contact person first
    to_email = preferred or contact_email or customer_company_email or tenant_company_email
    if not to_email:
        return jsonify({"error": "Customer has no email."}), 400

    # CC = customer company email (if different from To)
    cc_emails = []
    if customer_company_email and customer_company_email.lower() != to_email.lower():
        cc_emails.append(customer_company_email)

    # BCC = tenant company email + role emails
    role_emails = get_company_emails_by_role(company_id, cc_role) if cc_role else []
    bcc_emails = []

    if tenant_company_email and tenant_company_email.lower() != to_email.lower() and tenant_company_email.lower() not in {x.lower() for x in cc_emails}:
        bcc_emails.append(tenant_company_email)

    seen = {to_email.lower(), *[x.lower() for x in cc_emails], *[x.lower() for x in bcc_emails]}
    for e in role_emails:
        e = (e or "").strip()
        if not e:
            continue
        if e.lower() in seen:
            continue
        bcc_emails.append(e)
        seen.add(e.lower())

    print(
        "[INVOICE EMAIL RESOLVE]",
        {
            "preferred": preferred,
            "contact_email": contact_email,
            "customer_company_email": customer_company_email,
            "tenant_company_email": tenant_company_email,
            "final_to": to_email,
            "cc": cc_emails,
            "bcc": bcc_emails,
            "invoice_id": invoice_id,
            "company_id": company_id,
        },
        flush=True,
    )

    def _fmt(d):
        if isinstance(d, (datetime, date)):
            return d.strftime("%Y-%m-%d")
        return d or ""

    invoice_date = _fmt(inv.get("invoice_date"))
    due_date = _fmt(inv.get("due_date"))

    currency = inv.get("currency") or ""
    total = float(inv.get("total") or inv.get("total_amount") or 0.0)
    customer_name = inv.get("customer_name") or "Customer"
    company_name = inv.get("company_name") or company.get("name") or "Our Company"
    inv_no = (inv.get("number") or f"INV-{invoice_id}").strip()

    text_body = f"""Dear {customer_name},

Please find your invoice attached.

Invoice number: {inv_no}
Invoice date  : {invoice_date}
Due date      : {due_date}
Amount due    : {currency} {total:,.2f}

Kind regards,
{company_name}
"""

    html_body = f"<pre style='font-family:system-ui,monospace'>{text_body}</pre>"
    subject = f"Invoice {inv_no} from {company_name}"

    # 🔍 DEBUG HERE (before PDF build)
    print("[INVOICE EMAIL PDF KEYS]", sorted(inv.keys()), flush=True)
    print(
        "[INVOICE EMAIL PDF LINES CHECK]",
        {
            "lines": len(inv.get("lines") or []),
            "invoice_lines": len(inv.get("invoice_lines") or []),
            "quote_lines": len(inv.get("quote_lines") or []),
            "line_items": len(inv.get("line_items") or []),
            "invoice_items": len(inv.get("invoice_items") or []),
            "items": len(inv.get("items") or []),
        },
        flush=True,
    )

    try:
        pdf_bytes = generate_invoice_pdf(inv, company=company)
    except Exception as e:
        current_app.logger.exception("Failed to generate invoice PDF")
        return jsonify({"error": "Failed to generate invoice PDF", "detail": str(e)}), 500

    attachments = [(f"invoice-{inv_no}.pdf", pdf_bytes, "application/pdf")]

    print(
        "[INVOICE EMAIL SEND START]",
        {"to": to_email, "cc": cc_emails, "bcc": bcc_emails, "subject": subject},
        flush=True,
    )

    # primary recipient
    send_mail(
        to_email=to_email,
        subject=subject,
        html_body=html_body,
        text_body=text_body,
        from_name=f"{company_name} via FinSage",
        reply_to=tenant_company_email,
        attachments=attachments,
    )

    # customer company email gets a visible copy
    for cc in cc_emails:
        send_mail(
            to_email=cc,
            subject=f"Copy: {subject}",
            html_body=html_body,
            text_body=text_body,
            from_name=f"{company_name} via FinSage",
            reply_to=tenant_company_email,
            attachments=attachments,
        )

    # internal recipients get hidden-style separate copies
    for bcc in bcc_emails:
        send_mail(
            to_email=bcc,
            subject=f"Internal copy: {subject}",
            html_body=html_body,
            text_body=text_body,
            from_name=f"{company_name} via FinSage",
            reply_to=tenant_company_email,
            attachments=attachments,
        )

    print(
        "[INVOICE EMAIL SEND DONE]",
        {"to": to_email, "cc": cc_emails, "bcc": bcc_emails, "invoice_id": invoice_id},
        flush=True,
    )

    db_service.mark_invoice_sent(company_id, invoice_id)
    db_service.audit_log(
        company_id=company_id,
        actor_user_id=actor_user_id,
        module="ar",
        action="email_sent",
        severity="info",
        entity_type="invoice",
        entity_id=str(invoice_id),
        entity_ref=str(inv_no),
        customer_id=int(inv.get("customer_id")) if inv.get("customer_id") else None,
        amount=float(total or 0.0),
        currency=str(currency).upper() if currency else None,
        after_json={
            "to": to_email,
            "cc": cc_emails,
            "bcc": bcc_emails,
            "subject": subject,
            "source": "preferred" if preferred else ("contacts" if contact_email else "fallback"),
            "attachment": f"invoice-{inv_no}.pdf",
            "pdf_builder": "reportlab",
        },
        message="Invoice emailed successfully",
    )
    return jsonify({"ok": True, "to_email": to_email, "cc": cc_emails, "bcc": bcc_emails}), 200

@app.route("/api/companies/<int:company_id>/journal/recent", methods=["GET"])
@require_auth
def api_recent_journals(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    try:
        limit = int(request.args.get("limit", 50))
    except ValueError:
        limit = 50

    date_from, date_to, meta = resolve_company_period(db_service, company_id, request, mode="range")

    rows = db_service.list_recent_journals(
        company_id,
        limit,
        date_from=date_from.isoformat() if date_from else None,
        date_to=date_to.isoformat() if date_to else None,
    ) or []

    return jsonify({"meta": meta, "rows": rows}), 200


# API: replace the whole body with a single service call
@app.route("/api/companies/<int:company_id>/journal", methods=["GET", "POST"])
@require_auth
def api_journal(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    # ✅ GET detail: /journal?journal_id=5
    if request.method == "GET":
        journal_id = request.args.get("journal_id")
        if not journal_id:
            return jsonify({"ok": False, "error": "journal_id is required"}), 400

        data = db_service.get_journal_with_lines(company_id, int(journal_id))
        return jsonify(data), 200

    # ✅ POST posting
    try:
        entry = request.get_json(force=True) or {}
        before_json = entry

        ref = (entry.get("ref") or "").strip()

        if ref:
            existing = db_service.fetch_one(
                f"""
                SELECT id
                FROM company_{company_id}.journal
                WHERE LOWER(TRIM(ref)) = LOWER(TRIM(%s))
                AND COALESCE(reversal_of_journal_id, 0) = 0
                LIMIT 1
                """,
                (ref,),
            )

            if existing:
                return jsonify({
                    "ok": False,
                    "error": f"Duplicate journal reference: {ref}",
                    "code": "DUPLICATE_JOURNAL_REF",
                    "ref": ref,
                    "journal_id": existing.get("id"),
                }), 409

        journal_id = db_service.post_journal_with_overdraft(company_id, entry)

        # ✅ Link manual journal back to lease_schedule when journal is IFRS16 monthly amortisation
        try:
            src = (entry.get("source") or "").strip()
            source_id = entry.get("source_id")

            if src == "lease_monthly" and source_id:
                # source_id should be lease_schedule.id
                sched = db_service.fetch_one(
                    f"""
                    SELECT id, lease_id, period_no
                    FROM company_{company_id}.lease_schedule
                    WHERE company_id=%s
                    AND id=%s
                    LIMIT 1
                    """,
                    (int(company_id), int(source_id)),
                )

                if sched:
                    db_service.mark_lease_schedule_month_posted(
                        int(company_id),
                        schedule_id=int(sched["id"]),
                        lease_id=int(sched["lease_id"]),
                        journal_id=int(journal_id),
                    )

        except Exception:
            current_app.logger.exception("Failed to link manual IFRS16 journal to lease schedule")
            
        # ✅ company base currency (no circular import risk if you already can call this here)
        try:
            ctx = get_company_context(db_service, company_id) or {}
            base_ccy = ctx.get("currency")
        except Exception:
            base_ccy = None

        # ✅ pull saved journal for reference
        # IMPORTANT: avoid crashing if column doesn't exist yet (legacy tenant)
        saved = {}
        try:
            saved = db_service.fetch_one(
                f"SELECT id, date, ref, description, gross_amount, currency "
                f"FROM company_{company_id}.journal WHERE id=%s",
                (int(journal_id),),
            ) or {}
        except Exception:
            # fallback without currency column
            saved = db_service.fetch_one(
                f"SELECT id, date, ref, description, gross_amount "
                f"FROM company_{company_id}.journal WHERE id=%s",
                (int(journal_id),),
            ) or {}

        audit_currency = (
            (saved.get("currency") if isinstance(saved, dict) else None)
            or entry.get("currency")
            or base_ccy
        )

        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(current_user.get("id") or 0),
            module="gl",
            action="post_journal",
            severity="info",
            entity_type="journal",
            entity_id=str(journal_id),
            entity_ref=(saved.get("ref") or saved.get("description") or f"journal_{journal_id}"),
            journal_id=int(journal_id),
            amount=float(saved.get("gross_amount") or 0.0),
            currency=audit_currency,
            before_json=before_json if isinstance(before_json, dict) else {},
            after_json=saved if isinstance(saved, dict) else {},
            message="Journal posted",
            source="api",
        )

        return jsonify({"ok": True, "journal_id": journal_id}), 200

    except ValueError as e:
        msg = str(e or "").strip() or "Journal validation failed."

        # ============================================================
        # PERIOD LOCKED
        # ============================================================
        if msg.startswith("PERIOD_LOCKED|"):
            parts = msg.split("|")

            try:
                ctx = get_company_context(db_service, company_id) or {}
                base_ccy = ctx.get("currency")
            except Exception:
                base_ccy = None

            try:
                db_service.audit_log(
                    company_id=company_id,
                    actor_user_id=int(current_user.get("id") or 0),
                    module="gl",
                    action="post_journal_failed",
                    severity="warning",
                    entity_type="journal",
                    entity_id="",
                    entity_ref=None,
                    amount=float((entry or {}).get("gross_amount") or 0.0),
                    currency=((entry or {}).get("currency") or base_ccy),
                    before_json=entry if isinstance(entry, dict) else {},
                    after_json={},
                    message=msg,
                    source="api",
                )
            except Exception:
                current_app.logger.exception(
                    "Failed to audit PERIOD_LOCKED journal failure"
                )

            return jsonify({
                "ok": False,
                "success": False,
                "error": "Period is locked",
                "message": (
                    f"The journal cannot be posted because the accounting "
                    f"period for {parts[2] if len(parts) > 2 else 'this date'} is locked."
                ),
                "code": "PERIOD_LOCKED",
                "module": parts[1] if len(parts) > 1 else "gl",
                "date": parts[2] if len(parts) > 2 else None,
            }), 409

        # ============================================================
        # DUPLICATE JOURNAL REF
        # Covers duplicate detected inside db_service.post_journal()
        # ============================================================
        if msg.startswith("DUPLICATE_JOURNAL_REF|"):
            parts = msg.split("|")

            duplicate_ref = parts[1] if len(parts) > 1 else None
            duplicate_journal_id = None

            if len(parts) > 2 and "=" in parts[2]:
                try:
                    duplicate_journal_id = int(
                        parts[2].split("=", 1)[1]
                    )
                except Exception:
                    duplicate_journal_id = None

            return jsonify({
                "ok": False,
                "success": False,
                "error": (
                    f"Duplicate journal reference: {duplicate_ref}"
                    if duplicate_ref
                    else "Duplicate journal reference"
                ),
                "message": (
                    f'A journal with reference "{duplicate_ref}" has already been posted.'
                    if duplicate_ref
                    else "A journal with this reference has already been posted."
                ),
                "code": "DUPLICATE_JOURNAL_REF",
                "ref": duplicate_ref,
                "journal_id": duplicate_journal_id,
            }), 409

        # ============================================================
        # UNBALANCED JOURNAL
        # ============================================================
        if msg.startswith("Unbalanced journal:"):
            return jsonify({
                "ok": False,
                "success": False,
                "error": "Unbalanced journal",
                "message": msg,
                "code": "UNBALANCED_JOURNAL",
            }), 400

        # ============================================================
        # MISSING JOURNAL LINES
        # ============================================================
        if "at least one line" in msg.lower():
            return jsonify({
                "ok": False,
                "success": False,
                "error": "Journal has no lines",
                "message": msg,
                "code": "MISSING_JOURNAL_LINES",
            }), 400

        # ============================================================
        # OTHER VALIDATION ERRORS
        # ============================================================
        return jsonify({
            "ok": False,
            "success": False,
            "error": msg,
            "message": msg,
            "code": "JOURNAL_VALIDATION_FAILED",
        }), 400


    # ================================================================
    # DATABASE: REQUIRED VALUE MISSING
    # e.g. currency NOT NULL
    # ================================================================
    except pg_errors.NotNullViolation as e:
        column_name = None
        table_name = None

        try:
            column_name = e.diag.column_name
            table_name = e.diag.table_name
        except Exception:
            pass

        current_app.logger.exception(
            "Journal posting NOT NULL failure "
            "company_id=%s table=%s column=%s",
            company_id,
            table_name,
            column_name,
        )

        if column_name:
            user_message = (
                f'Journal could not be posted because the required field '
                f'"{column_name}" has no value.'
            )
        else:
            user_message = (
                "Journal could not be posted because a required value is missing."
            )

        return jsonify({
            "ok": False,
            "success": False,
            "error": "Missing required journal value",
            "message": user_message,
            "code": "JOURNAL_REQUIRED_FIELD_MISSING",
            "field": column_name,
        }), 400


    # ================================================================
    # DATABASE: FOREIGN KEY VIOLATION
    # ================================================================
    except pg_errors.ForeignKeyViolation as e:
        constraint_name = None

        try:
            constraint_name = e.diag.constraint_name
        except Exception:
            pass

        current_app.logger.exception(
            "Journal posting FK failure company_id=%s constraint=%s",
            company_id,
            constraint_name,
        )

        return jsonify({
            "ok": False,
            "success": False,
            "error": "Invalid journal reference",
            "message": (
                "Journal could not be posted because one of the selected "
                "accounts or referenced records does not exist."
            ),
            "code": "JOURNAL_FOREIGN_KEY_ERROR",
        }), 400


    # ================================================================
    # DATABASE: UNIQUE CONSTRAINT
    # ================================================================
    except pg_errors.UniqueViolation as e:
        constraint_name = None

        try:
            constraint_name = e.diag.constraint_name
        except Exception:
            pass

        current_app.logger.exception(
            "Journal posting unique violation company_id=%s constraint=%s",
            company_id,
            constraint_name,
        )

        return jsonify({
            "ok": False,
            "success": False,
            "error": "Duplicate journal record",
            "message": (
                "Journal could not be posted because the transaction "
                "conflicts with an existing journal record."
            ),
            "code": "JOURNAL_DUPLICATE_RECORD",
        }), 409


    # ================================================================
    # DATABASE: CHECK CONSTRAINT
    # ================================================================
    except pg_errors.CheckViolation as e:
        constraint_name = None

        try:
            constraint_name = e.diag.constraint_name
        except Exception:
            pass

        current_app.logger.exception(
            "Journal posting CHECK violation company_id=%s constraint=%s",
            company_id,
            constraint_name,
        )

        return jsonify({
            "ok": False,
            "success": False,
            "error": "Journal value not allowed",
            "message": (
                "Journal could not be posted because one of its values "
                "does not satisfy the accounting rules."
            ),
            "code": "JOURNAL_CHECK_CONSTRAINT",
        }), 400


    # ================================================================
    # DATABASE: OTHER POSTGRESQL ERRORS
    # ================================================================
    except psycopg2.Error as e:
        pg_message = None

        try:
            pg_message = e.diag.message_primary
        except Exception:
            pass

        current_app.logger.exception(
            "PostgreSQL error while posting journal company_id=%s",
            company_id,
        )

        return jsonify({
            "ok": False,
            "success": False,
            "error": "Journal database error",
            "message": (
                pg_message
                or "The journal could not be posted because of a database error."
            ),
            "code": "JOURNAL_DATABASE_ERROR",
        }), 500


    # ================================================================
    # EVERYTHING ELSE
    # ================================================================
    except Exception as e:
        current_app.logger.exception(
            "Unexpected journal posting failure company_id=%s",
            company_id,
        )

        msg = str(e or "").strip()

        return jsonify({
            "ok": False,
            "success": False,
            "error": "Journal posting failed",
            "message": (
                msg
                or "The journal could not be posted because of an unexpected error."
            ),
            "code": "JOURNAL_POST_FAILED",
        }), 500

@app.route("/api/companies/<int:company_id>/journal/<int:journal_id>", methods=["GET"])
@require_auth
def api_journal_detail(company_id: int, journal_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    try:
        data = db_service.get_journal_with_lines(company_id, journal_id)
        return jsonify(data), 200
    except Exception as e:
        current_app.logger.exception("Error loading journal detail")
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/companies/<int:company_id>/journal/by-ref/<path:ref>", methods=["GET"])
@require_auth
def api_journal_by_ref(company_id: int, ref: str):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    row = db_service.get_journal_by_ref(company_id, ref)
    if not row:
        return jsonify({"ok": False, "error": "Not found"}), 404
    return jsonify(row), 200


@app.route("/api/companies/<int:company_id>/journal/<int:journal_id>/reverse", methods=["POST"])
@require_auth
def api_reverse_journal(company_id: int, journal_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    try:
        payload = request.get_json(silent=True) or {}

        # (optional) snapshot "before"
        before = db_service.fetch_one(
            f"""
            SELECT id, date, ref, description, gross_amount, currency
            FROM company_{company_id}.journal
            WHERE id=%s
            """,
            (int(journal_id),)
        ) or {}

        reversal_journal_id = db_service.reverse_journal(company_id, journal_id, payload)

        # snapshot "after" (the new reversing journal)
        after = db_service.fetch_one(
            f"""
            SELECT id, date, ref, description, gross_amount, currency, reversal_of_journal_id
            FROM company_{company_id}.journal
            WHERE id=%s
            """,
            (int(reversal_journal_id),)
        ) or {}

        # ✅ AUDIT HERE (right after success)
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(current_user.get("id") or 0),
            module="gl",
            action="reverse_journal",
            severity="info",
            entity_type="journal",
            entity_id=str(journal_id),  # the original journal being reversed
            entity_ref=(before.get("ref") or before.get("description") or f"journal_{journal_id}"),
            journal_id=int(journal_id),
            amount=float(before.get("gross_amount") or 0.0),
            currency=(before.get("currency") or after.get("currency") or payload.get("currency")),
            before_json=before if isinstance(before, dict) else {},
            after_json={
                "reversal_journal_id": int(reversal_journal_id),
                **(after if isinstance(after, dict) else {})
            },
            message=f"Journal reversed. reversal_journal_id={reversal_journal_id}",
            source="api",
        )

        return jsonify({"ok": True, "reversal_journal_id": reversal_journal_id}), 200

    except Exception as e:
        current_app.logger.exception("Error reversing journal")
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/companies/<int:company_id>/ledger", methods=["GET"])
@require_auth
def api_get_ledger(company_id: int):

    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    account_code = (request.args.get("account") or "").strip()
    raw_from     = request.args.get("from") or None
    raw_to       = request.args.get("to") or None

    def parse_date(val):
        if not val:
            return None
        try:
            return datetime.strptime(val, "%Y-%m-%d").date()
        except ValueError:
            return None

    date_from = parse_date(raw_from)
    date_to   = parse_date(raw_to)

    try:
        # ✅ ALL accounts → simple rows
        if not account_code or account_code.upper() == "ALL":
            rows = db_service.get_ledger_for_all_accounts(
                company_id, date_from, date_to
            )
            return jsonify(rows), 200

        # ✅ Single account → enrich with balance, journal_lines, counterparties
        raw_rows = db_service.get_ledger_for_account(
            company_id,
            account_code,
            date_from=date_from,
            date_to=date_to,
        )

        if not raw_rows:
            return jsonify([]), 200

        # fetch all journal lines for the involved journal_ids
        journal_ids = sorted(
            {r["journal_id"] for r in raw_rows if r.get("journal_id") is not None}
        )

        lines_by_journal: Dict[int, List[Dict[str, Any]]] = {}
        if journal_ids:
            j_lines = db_service.get_journal_lines_for_ids(company_id, journal_ids)
            for ln in j_lines:
                jid = ln["journal_id"]
                lines_by_journal.setdefault(jid, []).append({
                    "account": ln["account"],
                    "account_name": ln.get("account_name"),
                    "debit": float(ln.get("debit") or 0),
                    "credit": float(ln.get("credit") or 0),
                    "memo": ln.get("memo"),
                })

        # build response with running balance + breakdown
        balance = 0.0
        out = []

        for r in raw_rows:
            debit  = float(r.get("debit") or 0)
            credit = float(r.get("credit") or 0)
            balance += debit - credit

            jid = r.get("journal_id")
            journal_lines = lines_by_journal.get(jid, []) if jid else []

            # counterparties: *other* accounts in same journal
            counterparties = []
            seen = set()
            for ln in journal_lines:
                acc = ln["account"]
                if acc == r.get("account"):
                    continue
                if acc in seen:
                    continue
                seen.add(acc)
                label = f"{acc} - {ln['account_name']}" if ln.get("account_name") else acc
                counterparties.append(label)

            out.append({
                "id":            r.get("id"),
                "journal_id":    jid,
                "date":          r.get("date").isoformat() if r.get("date") else None,
                "ref":           r.get("ref"),
                "journal_ref":   r.get("journal_ref"),
                "account":       r.get("account"),
                "debit":         debit,
                "credit":        credit,
                "description":   r.get("memo") or r.get("journal_description"),
                "balance":       balance,
                "journal_lines": journal_lines,
                "counterparties": counterparties,
            })

        return jsonify(out), 200

    except Exception as e:
        current_app.logger.exception("Error in GET /ledger: %s", e)
        return jsonify({"error": str(e)}), 500
    
@app.route("/api/companies/<int:company_id>/trial_balance_mini", methods=["GET"])
@require_auth
def api_trial_balance_mini(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    preset    = request.args.get("preset")
    date_from = parse_date_arg(request, "from")
    date_to   = parse_date_arg(request, "to")

    company = db_service.get_company(company_id)
    if not company:
        return jsonify({"error": "Company not found"}), 404

    fin_year_start = parse_date_maybe(company.get("fin_year_start"))
    period = resolve_period(
        fin_year_start=fin_year_start,
        preset=preset,
        date_from=date_from,
        date_to=date_to,
    )

    tb_rows = db_service.get_trial_balance_with_meta(company_id, period["from"], period["to"]) or []

    out = []
    for r in tb_rows:
        code = str(r.get("account") or r.get("code") or "").strip()

        debit_total  = _num(r.get("debit_total")  if r.get("debit_total")  is not None else r.get("debit"))
        credit_total = _num(r.get("credit_total") if r.get("credit_total") is not None else r.get("credit"))
        raw = _num(r.get("closing_balance")) if r.get("closing_balance") is not None else (debit_total - credit_total)

        dr_bal = float(raw) if raw > 0 else 0.0
        cr_bal = float(-raw) if raw < 0 else 0.0

        out.append({
            "code": code,
            "name": r.get("name") or code,
            "section": r.get("section") or "",
            "category": r.get("category") or "",

            "debit": float(dr_bal),
            "credit": float(cr_bal),

            "debit_movement": float(debit_total),
            "credit_movement": float(credit_total),

            "closing_balance_raw": float(raw),
            "closing_balance": float(raw),
        })

    return jsonify({
        "meta": {"period": {"from": str(period["from"]), "to": str(period["to"])}, "label": period["label"]},
        "rows": out
    }), 200

def _num(v):
    try:
        return float(v or 0.0)
    except Exception:
        return 0.0

@app.route("/api/companies/<int:company_id>/trial_balance", methods=["GET"])
@require_auth
def api_trial_balance(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    preset    = request.args.get("preset")
    date_from = parse_date_arg(request, "from")
    date_to   = parse_date_arg(request, "to")

    company = db_service.get_company(company_id)
    if not company:
        return jsonify({"error": "Company not found"}), 404

    fin_year_start = parse_date_maybe(company.get("fin_year_start"))

    period = resolve_period(
        fin_year_start=fin_year_start,
        preset=preset,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db_service.get_trial_balance(company_id, period["from"], period["to"]) or []

    out = []
    for r in rows:
        code = str(r.get("account") or r.get("code") or "").strip()

        dr_mov = _num(r.get("debit_total")  if r.get("debit_total")  is not None else r.get("debit"))
        cr_mov = _num(r.get("credit_total") if r.get("credit_total") is not None else r.get("credit"))

        if r.get("closing_balance_raw") is not None:
            raw = _num(r.get("closing_balance_raw"))
        elif r.get("debit_total") is not None or r.get("credit_total") is not None:
            raw = dr_mov - cr_mov
        elif r.get("closing_balance") is not None:
            raw = _num(r.get("closing_balance"))
        else:
            raw = dr_mov - cr_mov

        dr_bal = float(raw) if raw > 0 else 0.0
        cr_bal = float(-raw) if raw < 0 else 0.0

        out.append({
            "code": code,
            "name": r.get("name") or code,
            "section": r.get("section") or "",
            "category": r.get("category") or "",

            "debit": float(dr_bal),
            "credit": float(cr_bal),

            "debit_movement": float(dr_mov),
            "credit_movement": float(cr_mov),

            "closing_balance_raw": float(raw),
            "closing_balance": float(raw),
        })

    return jsonify({
        "meta": {
            "period": {"from": str(period["from"]), "to": str(period["to"])},
            "label": period["label"]
        },
        "rows": out
    }), 200

@app.route("/api/companies/<int:company_id>/trial_balance_asof", methods=["GET"])
@require_auth
def api_trial_balance_asof(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    as_of = parse_date_arg(request, "as_of")
    if not as_of:
        return jsonify({"error": "Missing as_of"}), 400

    rows = db_service.get_trial_balance_as_of(company_id, as_of) or []

    out = []
    for r in rows:
        raw = float(r.get("closing_balance_raw") or 0.0)

        dr = raw if raw > 0 else 0.0
        cr = (-raw) if raw < 0 else 0.0

        out.append({
            "code":     str(r.get("code") or "").strip(),
            "name":     r.get("name") or "",
            "section":  r.get("section") or "",
            "category": r.get("category") or "",
            "standard": r.get("standard") or "",
            "debit":    float(dr),
            "credit":   float(cr),
            "closing_balance_raw": float(raw),
            "closing_balance": float(raw),
        })

    return jsonify({
        "meta": {"as_of": str(as_of), "label": f"As at {as_of}"},
        "rows": out
    }), 200

# ============================================================
# ✅ ENDPOINTS (rewritten + normalized)
# ============================================================
@app.route("/api/companies/<int:company_id>/pnl_mini", methods=["GET"])
@require_auth
def api_pnl_mini(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    company = db_service.get_company(company_id)
    if not company:
        return jsonify({"error": "Company not found"}), 404

    date_from, date_to, meta = resolve_company_period(
        db_service, company_id, request, mode="range"
    )
    if not date_from or not date_to:
        return jsonify({"error": "from/to required (or preset)"}), 400

    try:
        rows = db_service.get_pnl_mini(company_id, date_from, date_to) or []
        return jsonify({"meta": meta, "rows": rows}), 200

    except Exception as e:
        current_app.logger.exception(
            "api_pnl_mini failed for company_id=%s from=%s to=%s",
            company_id, date_from, date_to
        )
        return jsonify({"error": "Server error", "detail": str(e)}), 500
    
@app.route("/api/companies/<int:company_id>/cashflow_mini", methods=["GET"])
@require_auth
def api_cashflow_mini(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify([]), 200

    date_from, date_to, meta = resolve_company_period(db_service, company_id, request, mode="range")
    if not date_from or not date_to:
        return jsonify({"error": "from/to required (or preset)"}), 400

    compare = (request.args.get("compare") or "none").lower()
    if compare not in ("none", "prior_period", "prior_year"):
        compare = "none"

    method = (request.args.get("method") or "direct").lower()
    if method not in ("direct", "indirect"):
        method = "direct"

    rows = db_service.get_cashflow_mini(
        company_id=company_id,
        date_from=date_from,
        date_to=date_to,
        compare=compare,
        method=method,
    ) or []

    if isinstance(rows, dict):
        rows = rows.get("rows") or rows.get("items") or []

    return jsonify({"meta": meta, "rows": rows}), 200


@app.route("/api/companies/<int:company_id>/bs_mini", methods=["GET"])
@require_auth
def api_bs_mini(company_id: int):
    current_user = getattr(g, "current_user", None)
    if not current_user or current_user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    date_from, as_of, meta = resolve_company_period(
        db_service,
        company_id,
        request,
        mode="range",
    )

    if not as_of:
        return jsonify({"error": "as_of required (or preset)"}), 400

    rows = db_service.get_bs_mini(company_id, as_of=as_of) or []

    meta = meta or {}
    meta["from"] = date_from.isoformat() if date_from else None
    meta["to"] = as_of.isoformat() if as_of else None
    meta["as_of"] = as_of.isoformat() if as_of else None

    return jsonify({"meta": meta, "rows": rows}), 200

@app.route("/api/companies/<int:company_id>/bs", methods=["GET"])
@require_auth
def api_bs_full(company_id: int):
    if not _company_guard(company_id):
        return jsonify({"ok": False, "error": "Not authorised for this company"}), 403

    # ✅ ONE source of truth for as_of
    date_from, as_of, meta = resolve_company_period(
        db_service, company_id, request, mode="range"
    )

    if not as_of:
        _, as_of, meta2 = resolve_company_period(
            db_service, company_id, request, mode="as_of"
        )
        meta = meta or meta2

    if not as_of:
        return jsonify({"ok": False, "error": "as_of (or preset) is required"}), 400

    ctx = get_company_context(db_service, company_id) or {}
    template = (request.args.get("template") or ctx.get("template") or "ifrs").lower()

    basis = (request.args.get("basis") or "external").lower()
    compare = (request.args.get("compare") or "none").lower()

    comparison_years = request.args.get("comparison_years") or request.args.get("years") or "1"

    try:
        comparison_years = int(comparison_years)
    except Exception:
        comparison_years = 1

    comparison_years = max(1, min(comparison_years, 10))

    view = (request.args.get("view") or "external").lower()
    if view not in ("external", "internal"):
        view = "external"

    # ✅ Prior as_of logic
    preset = (request.args.get("preset") or "").strip().lower() or "this_year"

    prior_as_of = None
    if compare == "prior_year":
        try:
            prior_as_of = as_of.replace(year=as_of.year - 1)
        except ValueError:
            prior_as_of = as_of.replace(year=as_of.year - 1, day=28)

    elif compare == "prior_period":
        # ✅ preset-aware “previous period end”
   
        fy = parse_date_maybe(ctx.get("fin_year_start"))

        if preset == "this_month":
            pr = resolve_period(fin_year_start=fy, preset="prev_month", date_from=None, date_to=None, as_of=as_of)
            prior_as_of = pr["to"]

        elif preset == "this_quarter":
            pr = resolve_period(fin_year_start=fy, preset="prev_quarter", date_from=None, date_to=None, as_of=as_of)
            prior_as_of = pr["to"]

        elif preset in ("this_year", "ytd"):
            # previous FY end
            pr = resolve_period(fin_year_start=fy, preset="this_year", date_from=None, date_to=None, as_of=as_of)
            prior_as_of = pr["from"] - timedelta(days=1)

        else:
            # fallback
            prior_as_of = as_of - timedelta(days=30)

    include_np = (request.args.get("include_np") == "1")

    try:
        stmt = get_balance_sheet_v3_exact(
            db=db_service,
            company_id=company_id,
            as_of=as_of,
            date_from=date_from,
            prior_as_of=prior_as_of,
            compare=compare,
            comparison_years=comparison_years,
            view=view,
            basis=basis,
            include_net_profit_line=include_np,
            ctx=ctx,
        )

        stmt.setdefault("meta", {})
        stmt["meta"].update(meta or {})
        stmt["meta"]["template"] = template
        stmt["meta"]["as_of"] = as_of.isoformat()
        # only set prior_as_of if the statement really used prior columns
        stmt["meta"]["prior_as_of"] = (
            prior_as_of.isoformat()
            if (prior_as_of and len(stmt.get("columns", [])) > 1)
            else None
        )

    except Exception as e:
        current_app.logger.exception("BS build failed")
        return jsonify({"ok": False, "error": str(e), "type": type(e).__name__}), 500

    return jsonify(stmt), 200

@app.route("/api/companies/<int:company_id>/pnl", methods=["GET"])
@require_auth
def api_pnl(company_id: int):
    if not _company_guard(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    # ✅ FY-aware period resolver
    date_from, date_to, meta = resolve_company_period(
        db_service, company_id, request, mode="range"
    )
    if not date_from or not date_to:
        return jsonify({"error": "from/to are required (or provide preset)"}), 400

    ctx = get_company_context(db_service, company_id) or {}

    fmt      = (request.args.get("format")  or "json").lower()
    basis    = (request.args.get("basis")   or "external").lower()
    compare  = (request.args.get("compare") or "none").lower()
    template = (request.args.get("template") or ctx.get("template") or "ifrs").lower()

    detail = (request.args.get("detail") or "summary").lower()
    if detail in ("semi", "semi-detailed", "semidetailed"):
        detail = "mid"
    if detail in ("classic", "detailed"):
        detail = "full"
    if detail not in ("summary", "mid", "full"):
        detail = "summary"

    if basis not in ("external", "management"):
        basis = "external"
    if compare not in ("none", "prior_period", "prior_year", "multi_year"):
        compare = "none"
    if template not in ("ifrs", "npo"):
        template = "ifrs"

    # ✅ keep cols_mode behaviour exactly as before
    try:
        cols_mode = int(request.args.get("cols_mode") or 1)
    except ValueError:
        cols_mode = 1
    cols_mode = 1 if cols_mode not in (1, 2, 3) else cols_mode
    if cols_mode != 1:
        compare = "none"

    # ✅ keep layout logic exactly as before
    profile = get_industry_profile(ctx.get("industry"), ctx.get("sub_industry"))
    layout = choose_layout("pnl", profile)

    layout_override = (request.args.get("layout") or "").lower().strip()
    if layout_override:
        layout = layout_override

    try:
        comparison_years = int(request.args.get("comparison_years") or request.args.get("years") or 1)
    except ValueError:
        comparison_years = 1

    comparison_years = max(1, min(comparison_years, 10))

    # ----- MANAGEMENT / INTERNAL -----
    if detail == "full" or basis == "management":
        stmt = build_income_statement_template(
            get_trial_balance_range_fn=db_service.get_trial_balance,
            get_company_context_fn=lambda cid: get_company_context(db_service, cid),
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            template=template,
            basis="management",
            layout=layout,
            cols_mode=cols_mode,
        )

        if isinstance(stmt, dict):
            stmt.setdefault("meta", {})
            stmt["meta"].update(meta)

        if template == "ifrs":
            net_vals = (stmt.get("totals", {}).get("net_income") or {}).copy()
            if net_vals:
                stmt["net_result"] = {"label": "Net Profit", "values": dict(net_vals)}

        return jsonify(stmt), 200

    # ----- EXTERNAL / IAS 1 -----
    prior_from = prior_to = None
    if compare != "none" and compare != "multi_year" and cols_mode == 1:
        # ✅ FY-aware compare range from the resolved meta
        prior_from, prior_to = resolve_compare_period(
            db_service, company_id, meta, compare, mode="range"
        )

    stmt = db_service.get_income_statement_v2(
        company_id=company_id,
        date_from=date_from,
        date_to=date_to,
        template=template,
        basis="external",
        compare=compare,
        cols_mode=1,
        detail=detail,
        prior_from=prior_from,
        prior_to=prior_to,
        comparison_years=comparison_years,
    )

    if isinstance(stmt, dict):
        stmt.setdefault("meta", {})
        stmt["meta"].update(meta)

    return jsonify(stmt), 200


@app.route("/api/companies/<int:company_id>/cashflow", methods=["GET"])
@require_auth
def api_cashflow(company_id: int):
    if not _company_guard(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    # ✅ FY-aware period resolver (preset/from/to)
    date_from, date_to, meta = resolve_company_period(
        db_service, company_id, request, mode="range"
    )
    if not date_from or not date_to:
        return jsonify({"error": "from/to are required (or provide preset)"}), 400

    fmt      = (request.args.get("format") or "json").lower()
    template = (request.args.get("template") or "ifrs").lower()
    basis    = (request.args.get("basis") or "external").lower()

    compare = (request.args.get("compare") or "none").lower()
    if compare not in ("none", "prior_period", "prior_year", "multi_year"):
        compare = "none"

    method = (request.args.get("method") or "direct").lower().strip()
    if method not in ("direct", "indirect"):
        method = "direct"

    # cols_mode
    raw_cols_mode = (request.args.get("cols_mode") or "").strip()
    cols_mode = 1
    try:
        cols_mode = int(raw_cols_mode) if raw_cols_mode else 1
    except Exception:
        cols_mode = 1
    if cols_mode not in (1, 2, 3):
        cols_mode = 1

    try:
        comparison_years = int(request.args.get("comparison_years") or request.args.get("years") or 1)
    except Exception:
        comparison_years = 1

    comparison_years = max(1, min(comparison_years, 10))

    # preview columns
    preview_columns = request.args.get("preview_columns") or request.args.get("preview") or 2
    try:
        preview_columns = int(preview_columns)
    except Exception:
        preview_columns = 2

    # ✅ Compare ranges (FY-aware) only when preview=1
    prior_from = prior_to = None
    if preview_columns == 2:
        compare = "none"
    else:
        if compare != "none" and compare != "multi_year":
            prior_from, prior_to = resolve_compare_period(
                db_service, company_id, meta, compare, mode="range"
            )

    stmt = db_service.get_cashflow_full_v2(
        company_id=company_id,
        date_from=date_from,
        date_to=date_to,
        template=template,
        basis=basis,
        compare=compare,
        method=method,
        prior_from=prior_from,
        prior_to=prior_to,
        preview_columns=preview_columns,
        cols_mode=cols_mode,
        comparison_years=comparison_years,
    )

    if isinstance(stmt, dict):
        stmt.setdefault("meta", {})
        stmt["meta"].update(meta)  # ✅ preset + FY aware

    return jsonify(stmt), 200

@app.route("/api/companies/<int:company_id>/socie", methods=["GET"])
@require_auth
def api_socie(company_id: int):
    if not _company_guard(company_id):
        return jsonify({"error": "Not authorised for this company"}), 403

    date_from, date_to, meta = resolve_company_period(
        db_service, company_id, request, mode="range"
    )
    if not date_from or not date_to:
        return jsonify({"error": "from/to are required (or provide preset)"}), 400

    ctx = get_company_context(db_service, company_id) or {}

    template = (request.args.get("template") or ctx.get("template") or "ifrs").lower()
    basis = (request.args.get("basis") or "external").lower()
    compare = (request.args.get("compare") or "none").lower()

    if basis not in ("external", "management"):
        basis = "external"

    if compare not in ("none", "prior_period", "prior_year", "multi_year"):
        compare = "none"

    if template not in ("ifrs", "npo"):
        template = "ifrs"

    try:
        cols_mode = int(request.args.get("cols_mode") or 1)
    except ValueError:
        cols_mode = 1

    cols_mode = 1 if cols_mode not in (1, 2, 3) else cols_mode

    if cols_mode != 1:
        compare = "none"

    try:
        comparison_years = int(
            request.args.get("comparison_years") or request.args.get("years") or 1
        )
    except Exception:
        comparison_years = 1

    comparison_years = max(1, min(comparison_years, 10))

    comparison_ranges = []

    if compare == "multi_year":
        for i in range(1, comparison_years):
            comparison_ranges.append((
                rh.shift_year(date_from, i),
                rh.shift_year(date_to, i),
            ))

    elif compare != "none" and cols_mode == 1:
        prior_from, prior_to = resolve_compare_period(
            db_service, company_id, meta, compare, mode="range"
        )
        if prior_from and prior_to:
            comparison_ranges.append((prior_from, prior_to))

    try:
        stmt = db_service.get_socie_v1(
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            template=template,
            basis=basis,
            compare=compare,
            cols_mode=cols_mode,
            comparison_ranges=comparison_ranges,
        )

        if isinstance(stmt, dict):
            stmt.setdefault("meta", {})
            stmt["meta"].update(meta)
            stmt["meta"]["template"] = template
            stmt["meta"]["statement"] = "socie"
            stmt["meta"]["compare"] = compare
            stmt["meta"]["comparison_years"] = comparison_years

    except Exception as e:
        current_app.logger.exception("SOCIE build failed")
        return jsonify({"ok": False, "error": str(e), "type": type(e).__name__}), 500

    return jsonify(stmt), 200

@app.route("/api/companies/<int:company_id>/year-end-close", methods=["POST"])
@require_auth
def api_year_end_close(company_id: int):
    if not _company_guard(company_id):
        return jsonify({"ok": False, "error": "Not authorised"}), 403

    payload = request.get_json(silent=True) or {}

    period_from = parse_date_maybe(payload.get("period_from"))
    period_to = parse_date_maybe(payload.get("period_to"))
    preview = bool(payload.get("preview"))

    if not period_from or not period_to:
        return jsonify({
            "ok": False,
            "error": "period_from and period_to are required"
        }), 400

    try:
        if preview:
            result = db_service.preview_profit_loss_close_to_retained_earnings(
                company_id=company_id,
                period_from=period_from,
                period_to=period_to,
            )
        else:
            result = db_service.close_profit_loss_to_retained_earnings(
                company_id=company_id,
                period_from=period_from,
                period_to=period_to,
            )

        return jsonify(result), 200 if result.get("ok") else 400

    except Exception as e:
        current_app.logger.exception("Year-end close failed")
        return jsonify({
            "ok": False,
            "error": str(e),
            "type": type(e).__name__,
        }), 500
    
@app.route("/api/companies/<int:company_id>/year-end-close/reopen", methods=["POST"])
@require_auth
def api_year_end_reopen(company_id: int):
    if not _company_guard(company_id):
        return jsonify({"ok": False, "error": "Not authorised"}), 403

    payload = request.get_json(silent=True) or {}

    period_from = parse_date_maybe(payload.get("period_from"))
    period_to = parse_date_maybe(payload.get("period_to"))

    if not period_from or not period_to:
        return jsonify({
            "ok": False,
            "error": "period_from and period_to are required"
        }), 400

    try:
        current_user = getattr(g, "current_user", None) or {}
        user_id = current_user.get("id")

        result = db_service.reopen_year_end_period(
            company_id=company_id,
            period_from=period_from,
            period_to=period_to,
            user_id=user_id,
        )

        return jsonify(result), 200 if result.get("ok") else 400

    except Exception as e:
        current_app.logger.exception("Year-end reopen failed")
        return jsonify({
            "ok": False,
            "error": str(e),
            "type": type(e).__name__,
        }), 500
    
def get_company_emails_by_role(company_id, role):
    sql = """
      SELECT email
      FROM users
      WHERE company_id = %s
        AND user_role = %s
    """
    rows = db_service.fetch_all(sql, (company_id, role))
    return [r["email"] for r in rows]

@app.route("/api/companies/<int:cid>/invoices/next_number", methods=["GET"])
@require_auth
def invoice_next_number(cid: int):
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != cid:
        return jsonify({"error": "Not authorised"}), 403

    company = db_service.get_company(cid)  # or fetch company name
    name = (company.get("name") or "COMPANY").strip()

    abbr = "".join([w[0] for w in name.split() if w]).upper()[:4] or "CO"

    n = db_service.next_counter(cid, "invoice_no")  # returns 1,2,3...

    number = f"INV-{abbr}-{n:05d}"
    return jsonify({"number": number}), 200


@app.route("/api/companies/<int:company_id>/bank_accounts", methods=["GET"])
@require_auth
def list_bank_accounts(company_id: int):
    payload = request.jwt_payload
    if payload.get("company_id") not in (None, company_id):
        return jsonify({"error": "Forbidden"}), 403

    accounts = db_service.list_company_bank_accounts(company_id)
    return jsonify(accounts), 200


@app.route("/api/companies/<int:company_id>/bank_accounts", methods=["POST"])
@require_auth
def create_bank_account(company_id: int):
    payload = request.jwt_payload
    if payload.get("company_id") not in (None, company_id):
        return jsonify({"error": "Forbidden"}), 403

    data = request.get_json() or {}
    acc = db_service.create_company_bank_account(company_id, data)
    actor_user_id = int(payload.get("sub") or 0)

    db_service.audit_log(
        company_id=company_id,
        actor_user_id=actor_user_id,
        module="bank",
        action="create",
        severity="info",
        entity_type="bank_account",
        entity_id=str(acc.get("id") or ""),
        entity_ref=str(acc.get("account_name") or acc.get("bank_name") or ""),
        after_json=acc if isinstance(acc, dict) else {"created": True},
        message="Bank account created",
    )
    return jsonify(acc), 201

@app.route("/api/companies/<int:company_id>/bank_accounts/<int:bank_account_id>", methods=["GET"])
@require_auth
def get_bank_account(company_id: int, bank_account_id: int):
    payload = request.jwt_payload
    if payload.get("company_id") not in (None, company_id):
        return jsonify({"error": "Forbidden"}), 403

    row = db_service.get_company_bank_account(company_id, bank_account_id)
    if not row:
        return jsonify({"error": "Bank account not found"}), 404
    return jsonify(row), 200

@app.route("/api/companies/<int:company_id>/bank_accounts/<int:bank_account_id>", methods=["PUT", "PATCH"])
@require_auth
def update_bank_account(company_id: int, bank_account_id: int):
    payload = request.jwt_payload
    if payload.get("company_id") not in (None, company_id):
        return jsonify({"error": "Forbidden"}), 403

    body = request.get_json(force=True, silent=True) or {}
    row = db_service.update_company_bank_account(company_id, bank_account_id, body)
    if not row:
        return jsonify({"error": "Bank account not found"}), 404
    return jsonify(row), 200

@app.route("/api/companies/<int:cid>/inventory/items", methods=["POST"])
@require_auth
def create_inventory_item(cid: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err:
        return err
    try:
        raw = request.get_json(silent=True) or {}
        auth_ctx = getattr(g, "auth_context", {}) or {}

        if auth_ctx.get("is_delegated_company_access"):
            raw["source_company_id"] = auth_ctx.get("source_company_id")
            raw["engagement_company_id"] = auth_ctx.get("source_company_id")
            raw["engagement_id"] = auth_ctx.get("engagement_id")

        raw["created_by_user_id"] = user.get("id")
        raw["updated_by_user_id"] = user.get("id")
        # =========================
        # DEBUG: what frontend sent
        # =========================
        try:
            print("\n=== INVENTORY CREATE DEBUG (API) ===")
            print("company_id:", company_id)
            print("user_id:", user.get("id"))
            print("raw keys:", list(raw.keys()))
            print("raw payload:", raw)
        except Exception:
            pass

        cols, meta = _split_item_payload(raw, INVENTORY_ITEM_COLS)
        cols["auto_generate_barcode"] = raw.get("auto_generate_barcode")

        # =========================
        # DEBUG: after split/normalize
        # =========================
        try:
            print("cols keys:", list(cols.keys()))
            print("cols:", cols)
            print("meta:", meta)
            print("sales_price(raw):", raw.get("sales_price"), "sell_price(raw):", raw.get("sell_price"))
            print("sales_price(cols):", cols.get("sales_price"))
            print("reorder_level(cols):", cols.get("reorder_level"))
            print("purchase_cost(cols):", cols.get("purchase_cost"))
            print("====================================\n")
        except Exception:
            pass

        # normalize required fields
        sku = _norm_str(cols.get("sku"))
        name = _norm_str(cols.get("name"))
        barcode = _norm_str(cols.get("barcode"))
        track_stock = _to_bool(cols.get("track_stock"), default=True)

        if not sku:
            return jsonify({"error": "sku is required"}), 400
        if not name:
            return jsonify({"error": "name is required"}), 400

        cols["sku"] = sku
        cols["name"] = name
        cols["barcode"] = barcode or None
        cols["track_stock"] = track_stock
        cols["meta"] = meta  # ✅ store industry extensions

        schema = f"company_{company_id}"

        # ✅ Uniqueness checks (sku + barcode per company)
        dup_sku = db_service.fetch_one(
            f"SELECT id FROM {schema}.inventory_items WHERE company_id=%s AND lower(sku)=lower(%s) LIMIT 1",
            (company_id, sku),
        )
        if dup_sku:
            return jsonify({"error": "SKU already exists", "sku": sku}), 409

        if cols.get("barcode"):
            dup_bc = db_service.fetch_one(
                f"SELECT id FROM {schema}.inventory_items WHERE company_id=%s AND barcode=%s LIMIT 1",
                (company_id, cols.get("barcode")),
            )
            if dup_bc:
                return jsonify({"error": "Barcode already exists", "barcode": cols.get("barcode")}), 409

        # ✅ aliases (IMPORTANT: use raw, not "payload")
        if "sell_price" in raw and "sales_price" not in cols:
            cols["sales_price"] = raw.get("sell_price")

        if "cost_price" in raw and "purchase_cost" not in cols:
            cols["purchase_cost"] = raw.get("cost_price")

        if "taxable" in raw and "is_taxable" not in cols:
            cols["is_taxable"] = raw.get("taxable")

        if "active" in raw and "is_active" not in cols:
            cols["is_active"] = raw.get("active")

        # =========================
        # DEBUG: final cols before insert
        # =========================
        try:
            print("FINAL cols before insert:", cols)
            print("FINAL sales_price:", cols.get("sales_price"))
            print("FINAL reorder_level:", cols.get("reorder_level"))
        except Exception:
            pass

        item_id = db_service.create_inventory_item(company_id, cols)
        item = db_service.fetch_one(
            f"SELECT * FROM {schema}.inventory_items WHERE id=%s",
            (item_id,),
        )

        # =========================
        # DEBUG: what DB saved
        # =========================
        try:
            print("=== INVENTORY CREATE DEBUG (DB RESULT) ===")
            print("item_id:", item_id)
            print("saved sales_price:", (item or {}).get("sales_price"))
            print("saved reorder_level:", (item or {}).get("reorder_level"))
            print("saved row:", item)
            print("=========================================\n")
        except Exception:
            pass

        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="inventory",
            action="create",
            severity="info",
            entity_type="inventory_item",
            entity_id=str(item_id),
            entity_ref=str((item or {}).get("sku") or (item or {}).get("name") or ""),
            after_json=item if isinstance(item, dict) else cols,
            message="Inventory item created",
        )
        return jsonify(item), 201

    except Exception as e:
        current_app.logger.exception("create_inventory_item failed")
        # Also print quickly to terminal
        try:
            print("[create_inventory_item] ERROR:", repr(e))
        except Exception:
            pass
        return jsonify({"error": str(e)}), 400

@app.route("/api/companies/<int:cid>/inventory/receive", methods=["POST"])
@require_auth
def receive_inventory(cid: int):
    company_id = int(cid)
    user, err = _company_auth_or_403(company_id)
    if err:
        return err
    try:
        payload = request.get_json(silent=True) or {}
        auth_ctx = getattr(g, "auth_context", {}) or {}

        if auth_ctx.get("is_delegated_company_access"):
            payload["source_company_id"] = auth_ctx.get("source_company_id")
            payload["engagement_company_id"] = auth_ctx.get("source_company_id")
            payload["engagement_id"] = auth_ctx.get("engagement_id")

        payload["created_by_user_id"] = user.get("id")
        payload["updated_by_user_id"] = user.get("id")
        tx_date = payload.get("tx_date") or payload.get("date")
        ref = (payload.get("ref") or "").strip()
        notes = payload.get("notes")

        # ✅ NEW: vendor + optional PO header link + supplier invoice no
        vendor_id = int(payload.get("vendor_id") or 0) or None
        po_id = int(payload.get("po_id") or 0) or None
        supplier_invoice_no = (payload.get("supplier_invoice_no") or "").strip() or None

        lines = payload.get("lines") or []

        funding_type = (payload.get("funding_type") or "supplier_credit").strip().lower()
        bank_account_id = int(payload.get("bank_account_id") or 0) or None

        if funding_type not in ("supplier_credit", "cash_bank"):
            return jsonify({"error": "Invalid funding_type"}), 400

        if funding_type == "supplier_credit" and not vendor_id:
            return jsonify({"error": "vendor_id is required for supplier credit receipts"}), 400

        if funding_type == "cash_bank" and not bank_account_id:
            return jsonify({"error": "bank_account_id is required for cash purchase receipts"}), 400

        if not tx_date:
            return jsonify({"error": "tx_date is required"}), 400
        if not isinstance(lines, list) or not lines:
            return jsonify({"error": "lines[] is required"}), 400
        if not ref:
            return jsonify({"error": "ref is required for receiving inventory (e.g. GRN / supplier doc ref)"}), 400

        schema = f"company_{company_id}"

        # ------------------------------------
        # ✅ OPTIONAL: validate PO header link
        # ------------------------------------
        if po_id:
            po = db_service.fetch_one(
                f"SELECT id, vendor_id, status FROM {schema}.purchase_orders WHERE company_id=%s AND id=%s",
                (company_id, int(po_id)),
            )
            if not po:
                return jsonify({"error": "PO not found", "po_id": po_id}), 400

            po_vendor_id = int(po["vendor_id"] if isinstance(po, dict) else po[1])
            if vendor_id and int(vendor_id) != po_vendor_id:
                return jsonify({
                    "error": "vendor_id does not match PO vendor",
                    "vendor_id": vendor_id,
                    "po_vendor_id": po_vendor_id,
                }), 400

        # ------------------------------------
        # ✅ Fix idempotency: use receipt + trimmed ref
        # ------------------------------------
        existing = db_service.fetch_one(
            f"""
            SELECT id
            FROM {schema}.inventory_tx
            WHERE company_id=%s
              AND lower(tx_type)=lower('receipt')
              AND lower(trim(ref))=lower(trim(%s))
            LIMIT 1
            """,
            (company_id, ref),
        )
        if existing:
            tx_id = int(existing["id"] if isinstance(existing, dict) else existing[0])

            tx = db_service.fetch_one(
                f"SELECT * FROM {schema}.inventory_tx WHERE company_id=%s AND id=%s",
                (company_id, tx_id),
            ) or {}

            tx_lines = db_service.fetch_all(
                f"""
                SELECT * FROM {schema}.inventory_tx_lines
                WHERE company_id=%s AND tx_id=%s
                ORDER BY line_no
                """,
                (company_id, tx_id),
            ) or []

            if isinstance(tx, dict):
                tx["lines"] = tx_lines

            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="inventory",
                action="receipt_duplicate",
                severity="info",
                entity_type="inventory_tx",
                entity_id=str(tx_id),
                entity_ref=ref,
                vendor_id=vendor_id,
                after_json={"tx_id": tx_id, "ref": ref, "status": "returned_existing"},
                message="Inventory receipt already existed (idempotent return)",
            )
            return jsonify(tx), 200

        # ------------------------------------
        # ✅ Validate lines + optional po_line_id
        # ------------------------------------
        for i, ln in enumerate(lines, start=1):
            item_id = int(ln.get("item_id") or 0)
            qty = _to_num(ln.get("qty") or ln.get("quantity"), 0.0)
            unit_cost = _to_num(ln.get("unit_cost"), 0.0)

            po_line_id = int(ln.get("po_line_id") or 0) or None

            if item_id <= 0:
                return jsonify({"error": f"line {i} missing item_id", "line": ln}), 400
            if qty <= 0:
                return jsonify({"error": f"line {i} qty must be > 0", "line": ln}), 400
            if unit_cost < 0:
                return jsonify({"error": f"line {i} unit_cost must be >= 0", "line": ln}), 400

            # ✅ If a PO line is provided, validate it belongs to company + PO vendor
            if po_line_id:
                row = db_service.fetch_one(
                    f"""
                    SELECT pol.id, pol.po_id, poh.vendor_id, pol.item_id
                    FROM {schema}.purchase_order_lines pol
                    JOIN {schema}.purchase_orders poh ON poh.id = pol.po_id
                    WHERE pol.company_id=%s AND pol.id=%s
                    LIMIT 1
                    """,
                    (company_id, int(po_line_id)),
                )
                if not row:
                    return jsonify({"error": f"line {i} invalid po_line_id", "po_line_id": po_line_id}), 400

                row_vendor = int(row["vendor_id"] if isinstance(row, dict) else row[2])
                row_item   = int(row["item_id"] if isinstance(row, dict) else row[3])

                if vendor_id and int(vendor_id) != row_vendor:
                    return jsonify({
                        "error": f"line {i} po_line vendor does not match receipt vendor",
                        "vendor_id": vendor_id,
                        "po_line_vendor_id": row_vendor,
                    }), 400

                if row_item != item_id:
                    return jsonify({
                        "error": f"line {i} item_id does not match po_line item_id",
                        "item_id": item_id,
                        "po_line_item_id": row_item,
                    }), 400

                # (Optional convenience): if po_id not given, infer it from po_line
                if not po_id:
                    po_id = int(row["po_id"] if isinstance(row, dict) else row[1])

        # ------------------------------------
        # ✅ Create receipt (tx_type MUST be 'receipt')
        # ------------------------------------
        # IMPORTANT: update your db_service.receive_inventory_stock(...) to:
        # - insert inventory_tx.tx_type = 'receipt'
        # - store vendor_id, po_id, supplier_invoice_no into inventory_tx
        # - store po_line_id into inventory_tx_lines (if provided)
        tx_id = db_service.receive_inventory_stock(
            company_id,
            tx_date=tx_date,
            ref=ref,
            notes=notes,
            lines=lines,
            created_by=user.get("id"),
            vendor_id=vendor_id,
            po_id=po_id,
            supplier_invoice_no=supplier_invoice_no,
            funding_type=funding_type,
            bank_account_id=bank_account_id,
        )

        tx = db_service.fetch_one(
            f"SELECT * FROM {schema}.inventory_tx WHERE company_id=%s AND id=%s",
            (company_id, tx_id),
        ) or {}

        tx_lines = db_service.fetch_all(
            f"""
            SELECT * FROM {schema}.inventory_tx_lines
            WHERE company_id=%s AND tx_id=%s
            ORDER BY line_no
            """,
            (company_id, tx_id),
        ) or []

        if isinstance(tx, dict):
            tx["lines"] = tx_lines

        gross = 0.0
        for ln in (lines or []):
            gross += float(ln.get("qty") or ln.get("quantity") or 0) * float(ln.get("unit_cost") or 0)

        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="inventory",
            action="receipt_create",
            severity="info",
            entity_type="inventory_tx",
            entity_id=str(tx_id),
            entity_ref=ref,
            vendor_id=vendor_id,
            amount=float(gross or 0.0),
            currency=None,
            after_json={
                "tx_id": tx_id,
                "tx_date": str(tx_date),
                "ref": ref,
                "vendor_id": vendor_id,
                "po_id": po_id,
                "supplier_invoice_no": supplier_invoice_no,
                "lines_count": len(lines or []),
            },
            message="Inventory receipt created",
        )
        return jsonify(tx), 201

    except Exception as e:
        current_app.logger.exception("receive_inventory failed")
        return jsonify({"error": str(e)}), 400

@app.route("/api/companies/<int:cid>/inventory/movements/<int:tx_id>/lines/<int:line_id>/cost", methods=["PATCH"])
@require_auth
def update_inventory_movement_line_cost(cid, tx_id, line_id):
    company_id = int(cid)
    user, err = _company_auth_or_403(company_id)
    if err:
        return err

    try:
        payload = request.get_json(silent=True) or {}
        unit_cost = float(payload.get("unit_cost") or 0)

        if unit_cost < 0:
            return jsonify({"error": "unit_cost must be >= 0"}), 400

        schema = f"company_{company_id}"

        with db_service._conn_cursor() as (conn, cur):
            cur.execute(
                f"""
                SELECT l.id, l.tx_id, l.item_id, l.qty, l.unit_cost, t.tx_type, t.posted_journal_id
                FROM {schema}.inventory_tx_lines l
                JOIN {schema}.inventory_tx t ON t.id = l.tx_id
                WHERE l.company_id=%s AND l.tx_id=%s AND l.id=%s
                FOR UPDATE
                """,
                (company_id, tx_id, line_id),
            )
            line = cur.fetchone()

            if not line:
                return jsonify({"error": "Movement line not found"}), 404

            old_cost = float(line["unit_cost"] or 0)
            item_id = int(line["item_id"])

            cur.execute(
                f"""
                UPDATE {schema}.inventory_tx_lines
                SET unit_cost=%s
                WHERE company_id=%s AND id=%s AND tx_id=%s
                """,
                (unit_cost, company_id, line_id, tx_id),
            )

            cur.execute(
                f"""
                UPDATE {schema}.inventory_layers
                SET unit_cost=%s
                WHERE company_id=%s
                  AND tx_id=%s
                  AND item_id=%s
                  AND source='inventory_tx'
                """,
                (unit_cost, company_id, tx_id, item_id),
            )

            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="inventory",
                action="movement_line_cost_update",
                severity="warning",
                entity_type="inventory_tx_line",
                entity_id=str(line_id),
                after_json={
                    "tx_id": tx_id,
                    "line_id": line_id,
                    "old_unit_cost": old_cost,
                    "new_unit_cost": unit_cost,
                },
                message="Inventory movement line unit cost updated"
            )

            conn.commit()

        return jsonify({
            "ok": True,
            "tx_id": tx_id,
            "line_id": line_id,
            "old_unit_cost": old_cost,
            "new_unit_cost": unit_cost,
        }), 200

    except Exception as e:
        current_app.logger.exception("update inventory movement line cost failed")
        return jsonify({"error": str(e)}), 400
    
@app.route("/api/companies/<int:cid>/inventory/receipts", methods=["GET"])
@require_auth
def list_receipts_for_billing(cid: int):
    company_id = int(cid)
    user, err = _company_auth_or_403(company_id)
    if err: return err

    vendor_id = int(request.args.get("vendor_id") or 0)
    if vendor_id <= 0:
        return jsonify({"error": "vendor_id is required"}), 400

    q = (request.args.get("q") or "").strip()
    limit = int(request.args.get("limit") or 50)
    offset = int(request.args.get("offset") or 0)

    rows = db_service.list_unbilled_receipts(company_id, vendor_id=vendor_id, q=q, limit=limit, offset=offset)
    return jsonify({"rows": rows, "limit": limit, "offset": offset}), 200

@app.route("/api/companies/<int:cid>/inventory/items/lookup", methods=["GET"])
@require_auth
def lookup_inventory_item(cid: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    barcode = request.args.get("barcode") or ""
    item = db_service.get_inventory_item_by_barcode(company_id, barcode)
    if not item:
        return jsonify({"error": "Item not found", "barcode": barcode}), 404
    return jsonify(item), 200

@app.route("/api/companies/<int:cid>/inventory/items", methods=["GET"])
@require_auth
def list_inventory_items(cid: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    q = (request.args.get("q") or "").strip()
    active = request.args.get("active")
    limit = int(request.args.get("limit") or 50)
    offset = int(request.args.get("offset") or 0)

    rows = db_service.list_inventory_items(company_id, q=q, active=active, limit=limit, offset=offset)
    return jsonify({"rows": rows, "limit": limit, "offset": offset}), 200

@app.route("/api/companies/<int:cid>/inventory/items/<int:item_id>", methods=["GET"])
@require_auth
def get_inventory_item(cid: int, item_id: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    item = db_service.get_inventory_item(company_id, item_id)
    if not item:
        return jsonify({"error": "Item not found"}), 404
    return jsonify(item), 200

@app.route("/api/companies/<int:cid>/inventory/items/<int:item_id>", methods=["PUT"])
@require_auth
def update_inventory_item(cid: int, item_id: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    raw = request.get_json(silent=True) or {}
    schema = f"company_{company_id}"

    print("\n=== INVENTORY UPDATE API TRACE ===", flush=True)
    print("company_id:", company_id, flush=True)
    print("item_id:", item_id, flush=True)
    print("raw sales_price:", raw.get("sales_price"), flush=True)
    print("raw reorder_level:", raw.get("reorder_level"), flush=True)
    print("raw payload:", raw, flush=True)

    cols, meta = _split_item_payload(raw, INVENTORY_ITEM_COLS)
    cols["auto_generate_barcode"] = raw.get("auto_generate_barcode")

    print("cols sales_price:", cols.get("sales_price"), flush=True)
    print("cols reorder_level:", cols.get("reorder_level"), flush=True)
    print("cols after split:", cols, flush=True)
    print("meta after split:", meta, flush=True)
    print("==================================\n", flush=True)

    # normalize sku/barcode if present
    sku = _norm_str(cols.get("sku")) if "sku" in cols else ""
    barcode = _norm_str(cols.get("barcode")) if "barcode" in cols else ""

    if sku:
        dup_sku = db_service.fetch_one(
            f"""
            SELECT id FROM {schema}.inventory_items
            WHERE company_id=%s AND lower(sku)=lower(%s) AND id<>%s
            LIMIT 1
            """,
            (company_id, sku, int(item_id)),
        )
        if dup_sku:
            return jsonify({"error": "SKU already exists", "sku": sku}), 409
        cols["sku"] = sku

    if "barcode" in cols:
        cols["barcode"] = barcode or None
        if barcode:
            dup_bc = db_service.fetch_one(
                f"""
                SELECT id FROM {schema}.inventory_items
                WHERE company_id=%s AND barcode=%s AND id<>%s
                LIMIT 1
                """,
                (company_id, barcode, int(item_id)),
            )
            if dup_bc:
                return jsonify({"error": "Barcode already exists", "barcode": barcode}), 409

    # ✅ meta merge
    if meta:
        existing = db_service.fetch_one(
            f"SELECT meta FROM {schema}.inventory_items WHERE id=%s AND company_id=%s",
            (int(item_id), int(company_id)),
        ) or {}
        existing_meta = existing.get("meta") if isinstance(existing, dict) else {}
        cols["meta"] = _merge_meta(existing_meta, meta)

    # ✅ BEFORE snapshot (for audit)
    before_item = db_service.get_inventory_item(company_id, item_id) or {}
    if not before_item:
        return jsonify({"error": "Inventory item not found"}), 404

    print(
        "[INV UPDATE BEFORE DB]",
        {
            "company_id": company_id,
            "item_id": item_id,
            "sales_price": cols.get("sales_price"),
            "reorder_level": cols.get("reorder_level"),
            "cols": cols,
        },
        flush=True,
    )

    # ✅ Update
    db_service.update_inventory_item(company_id, item_id, cols)

    # ✅ AFTER snapshot
    item = db_service.get_inventory_item(company_id, item_id) or {}

    # ✅ Audit log (only after success)
    try:
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="inventory",
            action="update",
            severity="info",
            entity_type="inventory_item",
            entity_id=str(item_id),
            entity_ref=str(item.get("sku") or item.get("name") or ""),
            before_json=before_item if isinstance(before_item, dict) else {},
            after_json=item if isinstance(item, dict) else {},
            message="Inventory item updated",
            source="api",
        )
    except Exception:
        # never break the update response because audit failed
        current_app.logger.exception("audit_log failed (update_inventory_item)")

    return jsonify(item), 200

@app.route("/api/companies/<int:cid>/inventory/tx", methods=["GET"])
@require_auth
def list_inventory_tx(cid: int):
    company_id = int(cid)
    user, err = _company_auth_or_403(company_id)
    if err: return err

    schema = _schema(company_id)

    date_from = (request.args.get("from") or "").strip() or None
    date_to   = (request.args.get("to") or "").strip() or None
    status    = (request.args.get("status") or "").strip().lower() or None
    tx_type   = (request.args.get("tx_type") or "").strip().lower() or None

    limit  = int(request.args.get("limit") or 50)
    offset = int(request.args.get("offset") or 0)

    where = ["company_id=%s"]
    params = [company_id]

    if date_from:
        where.append("tx_date >= %s")
        params.append(date_from)
    if date_to:
        where.append("tx_date <= %s")
        params.append(date_to)
    if status:
        where.append("lower(status)=lower(%s)")
        params.append(status)
    if tx_type:
        where.append("lower(tx_type)=lower(%s)")
        params.append(tx_type)

    sql = f"""
    SELECT
      id, company_id, tx_date, tx_type, status, ref, notes, source, source_id,
      posted_journal_id, posted_at, posted_by, created_at, updated_at
    FROM {schema}.inventory_tx
    WHERE {" AND ".join(where)}
    ORDER BY tx_date DESC, id DESC
    LIMIT %s OFFSET %s
    """
    params += [limit, offset]

    rows = db_service.fetch_all(sql, tuple(params)) or []
    return jsonify({"items": rows, "limit": limit, "offset": offset}), 200

@app.route("/api/companies/<int:cid>/ap/grni/open", methods=["GET"])
@require_auth
def list_open_grni(cid: int):
    company_id = int(cid)
    user, err = _company_auth_or_403(company_id)
    if err: return err

    vendor_id = request.args.get("vendor_id", type=int)
    if not vendor_id:
        return jsonify({"error": "vendor_id is required"}), 400

    q = (request.args.get("q") or "").strip()
    limit = int(request.args.get("limit") or 50)
    offset = int(request.args.get("offset") or 0)

    rows = db_service.list_open_grni_receipts(company_id, vendor_id=vendor_id, q=q, limit=limit, offset=offset)
    return jsonify({"items": rows, "limit": limit, "offset": offset}), 200

@app.route("/api/companies/<int:cid>/purchase-orders/<int:po_id>/receive", methods=["POST"])
@require_auth
def receive_purchase_order_route(cid: int, po_id: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err:
        return err

    payload = request.get_json(silent=True) or {}

    try:
        out = db_service.receive_purchase_order(
            company_id=company_id,
            po_id=int(po_id),
            receipt_date=payload.get("receipt_date"),
            received_by=int(user["id"]),
            supplier_invoice_no=payload.get("supplier_invoice_no"),
            notes=payload.get("notes"),
            lines=payload.get("lines") or [],
            post_now=bool(payload.get("post_now", True)),
        )

        return jsonify(out), 200

    except ValueError as e:
        return jsonify({
            "error": str(e)
        }), 400

    except Exception as e:
        current_app.logger.exception(
            "[PO RECEIVE] FAILED | company_id=%s | po_id=%s",
            company_id,
            po_id,
        )

        return jsonify({
            "error": "failed_to_receive_purchase_order",
            "details": str(e),
        }), 500

@app.route("/api/companies/<int:cid>/purchase-orders", methods=["GET"])
@require_auth
def list_purchase_orders_route(cid: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err:
        return err

    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()

    vendor_id = request.args.get("vendor_id", type=int)
    project_id = request.args.get("project_id", type=int)

    limit = int(request.args.get("limit") or 50)
    offset = int(request.args.get("offset") or 0)

    rows = db_service.list_purchase_orders(
        company_id,
        q=q,
        status=status,
        vendor_id=vendor_id,
        project_id=project_id,
        limit=limit,
        offset=offset,
    )

    return jsonify({
        "items": rows,
        "limit": limit,
        "offset": offset,
    }), 200


@app.route("/api/companies/<int:cid>/purchase-orders/<int:po_id>", methods=["GET"])
@require_auth
def get_purchase_order_route(cid: int, po_id: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err:
        return err

    row = db_service.get_purchase_order(
        company_id,
        int(po_id),
    )

    if not row:
        return jsonify({
            "error": "purchase_order_not_found"
        }), 404

    return jsonify(row), 200


# =====================================================
# 2) MOVEMENTS: get one tx with lines
# GET /api/companies/:cid/inventory/tx/:tx_id
# =====================================================
@app.route("/api/companies/<int:cid>/inventory/tx/<int:tx_id>", methods=["GET"])
@require_auth
def get_inventory_tx(cid: int, tx_id: int):
    company_id = int(cid)
    user, err = _company_auth_or_403(company_id)
    if err: return err

    schema = _schema(company_id)

    tx = db_service.fetch_one(
        f"SELECT * FROM {schema}.inventory_tx WHERE id=%s AND company_id=%s",
        (int(tx_id), int(company_id)),
    )
    if not tx:
        return jsonify({"error": "Inventory transaction not found", "tx_id": int(tx_id)}), 404

    lines = db_service.fetch_all(
        f"""
        SELECT l.*, i.sku, i.name as item_name, i.barcode, i.unit
        FROM {schema}.inventory_tx_lines l
        JOIN {schema}.inventory_items i ON i.id = l.item_id
        WHERE l.tx_id=%s AND l.company_id=%s
        ORDER BY l.line_no ASC
        """,
        (int(tx_id), int(company_id)),
    ) or []

    tx["lines"] = lines
    return jsonify(tx), 200


# =====================================================
# 3) ON-HAND snapshot (useful for reorder + stocktake)
# GET /api/companies/:cid/inventory/on_hand?as_of=YYYY-MM-DD&q=
# =====================================================
@app.route("/api/companies/<int:cid>/inventory/on_hand", methods=["GET"])
@require_auth
def inventory_on_hand(cid: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err:
        return err

    schema = _schema(company_id)

    as_of = _to_iso(request.args.get("as_of"))
    q = (request.args.get("q") or "").strip()

    # ---------------------------------------------------------
    # Build JOIN parameters first because JOIN placeholders
    # appear before WHERE placeholders in the SQL.
    # ---------------------------------------------------------
    join_params = []
    layer_date_filter = ""

    if as_of:
        layer_date_filter = "AND l.tx_date <= %s"
        join_params.append(as_of)

    # ---------------------------------------------------------
    # WHERE parameters
    # ---------------------------------------------------------
    where_items = [
        "i.company_id=%s",
        "i.is_active=TRUE",
    ]

    where_params = [company_id]

    if q:
        where_items.append(
            """
            (
                LOWER(i.sku) LIKE LOWER(%s)
                OR LOWER(i.name) LIKE LOWER(%s)
                OR LOWER(COALESCE(i.barcode,'')) LIKE LOWER(%s)
            )
            """
        )

        like_q = f"%{q}%"
        where_params.extend([
            like_q,
            like_q,
            like_q,
        ])

    sql = f"""
        SELECT
            i.id AS item_id,
            i.sku,
            i.name,
            i.barcode,
            i.unit,
            i.track_stock,

            COALESCE(SUM(l.qty_in), 0)
            -
            COALESCE(SUM(l.qty_out), 0) AS on_hand

        FROM {schema}.inventory_items i

        LEFT JOIN {schema}.inventory_layers l
            ON l.company_id = i.company_id
           AND l.item_id = i.id
           {layer_date_filter}

        WHERE {" AND ".join(where_items)}

        GROUP BY
            i.id,
            i.sku,
            i.name,
            i.barcode,
            i.unit,
            i.track_stock

        ORDER BY
            i.name ASC,
            i.id ASC
    """

    # IMPORTANT:
    # Placeholder order must match SQL order:
    # JOIN params first, then WHERE params.
    params = tuple(join_params + where_params)

    rows = db_service.fetch_all(sql, params) or []

    return jsonify({
        "items": rows,
        "as_of": as_of,
    }), 200

# =====================================================
# 4) REORDER alerts
# GET /api/companies/:cid/inventory/reorder?as_of=&only=1
# - only=1 => only items at/below reorder_level
# =====================================================
@app.route("/api/companies/<int:cid>/inventory/reorder", methods=["GET"])
@require_auth
def inventory_reorder(cid: int):
    company_id = int(cid)
    user, err = _company_auth_or_403(company_id)
    if err:
        return err

    schema = _schema(company_id)

    as_of = _to_iso(request.args.get("as_of"))
    only = (request.args.get("only") or "1").strip().lower() in {"1", "true", "yes"}
    q = (request.args.get("q") or "").strip()

    where = ["i.company_id=%s", "i.is_active=TRUE", "i.track_stock=TRUE"]
    params = [company_id]

    if q:
        where.append("""
            (
                LOWER(i.sku) LIKE LOWER(%s)
                OR LOWER(i.name) LIKE LOWER(%s)
                OR LOWER(COALESCE(i.barcode,'')) LIKE LOWER(%s)
            )
        """)
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]

    layer_date_filter = ""
    layer_params = []
    if as_of:
        layer_date_filter = "AND l.tx_date <= %s"
        layer_params.append(as_of)

    sql = f"""
        SELECT
            i.id AS item_id,
            i.sku,
            i.name,
            i.barcode,
            i.unit,
            i.reorder_level,

            COALESCE(
                SUM(
                    COALESCE(l.qty_in, 0)
                    - COALESCE(l.qty_out, 0)
                ),
                0
            ) AS on_hand,

            (
                COALESCE(
                    SUM(
                        COALESCE(l.qty_in, 0)
                        - COALESCE(l.qty_out, 0)
                    ),
                    0
                ) <= COALESCE(i.reorder_level, 0)
            ) AS needs_reorder

        FROM {schema}.inventory_items i

        LEFT JOIN {schema}.inventory_layers l
            ON l.company_id = i.company_id
        AND l.item_id = i.id
        {layer_date_filter}

        WHERE {" AND ".join(where)}

        GROUP BY
            i.id,
            i.sku,
            i.name,
            i.barcode,
            i.unit,
            i.reorder_level

        {"HAVING COALESCE(SUM(COALESCE(l.qty_in,0) - COALESCE(l.qty_out,0)),0) <= COALESCE(i.reorder_level,0)" if only else ""}

        ORDER BY needs_reorder DESC, i.name ASC, i.id ASC
    """

    rows = db_service.fetch_all(sql, tuple(layer_params + params)) or []

    return jsonify({
        "items": rows,
        "as_of": as_of,
        "only": only,
    }), 200

# =====================================================
# 5) VALUATION (layers-based remaining qty * unit_cost)
# GET /api/companies/:cid/inventory/valuation?as_of=
# NOTE: This assumes your inventory_layers reflect remaining quantities per layer
#       via qty_in/qty_out updates (FIFO/WAVG engine writes those).
# =====================================================
@app.route("/api/companies/<int:cid>/inventory/valuation", methods=["GET"])
@require_auth
def inventory_valuation(cid: int):
    company_id = int(cid)
    _user, err = _company_auth_or_403(company_id)  # prefix _ avoids pylance warning
    if err:
        return err

    schema = _schema(company_id)
    as_of = _to_iso(request.args.get("as_of"))  # "2026-02-06" or None

    # Build params in the SAME order as placeholders appear in SQL
    params = []
    layer_date_filter = ""
    if as_of:
        layer_date_filter = "AND l.tx_date <= (%s)::date"
        params.append(as_of)

    # WHERE placeholder comes after the JOIN placeholder => company_id goes last
    params.append(company_id)

    sql = f"""
    SELECT
      i.id as item_id,
      i.sku,
      i.name,
      i.valuation_method,
      COALESCE(SUM(l.qty_in - l.qty_out),0) AS on_hand,
      COALESCE(SUM((l.qty_in - l.qty_out) * l.unit_cost),0) AS value,
      CASE
        WHEN COALESCE(SUM(l.qty_in - l.qty_out),0) = 0 THEN 0
        ELSE COALESCE(SUM((l.qty_in - l.qty_out) * l.unit_cost),0)
             / NULLIF(COALESCE(SUM(l.qty_in - l.qty_out),0),0)
      END AS avg_cost
    FROM {schema}.inventory_items i
    LEFT JOIN {schema}.inventory_layers l
      ON l.company_id=i.company_id
     AND l.item_id=i.id
     {layer_date_filter}
    WHERE i.company_id=%s AND i.is_active=TRUE AND i.track_stock=TRUE
    GROUP BY i.id, i.sku, i.name, i.valuation_method
    ORDER BY i.name ASC, i.id ASC
    """

    rows = db_service.fetch_all(sql, tuple(params)) or []
    total_value = sum(float(r.get("value") or 0) for r in rows)

    return jsonify({"as_of": as_of, "total_value": total_value, "items": rows}), 200

@app.route("/api/companies/<int:cid>/inventory/adjust", methods=["POST"])
@require_auth
def adjust_inventory(cid: int):
    company_id = int(cid)

    # ✅ use user_id so Pylance sees it's used
    user, err = _company_auth_or_403(company_id)
    if err: return err
    user_id = int(user.get("id") or 0)

    try:
        payload = request.get_json(silent=True) or {}
        schema = _schema(company_id)

        tx_date = _to_iso(payload.get("tx_date") or payload.get("date")) or date.today().isoformat()
        ref = _norm_str(payload.get("ref")) or None
        notes = _norm_str(payload.get("notes")) or None
        lines = payload.get("lines") or []

        if not isinstance(lines, list) or not lines:
            return jsonify({"error": "lines[] is required"}), 400

        # 1) create tx header
        tx_id = db_service.execute_sql(
            f"""
            INSERT INTO {schema}.inventory_tx
              (company_id, tx_date, tx_type, status, ref, notes, source, source_id, posted_by)
            VALUES
              (%s, %s, 'adjustment', 'draft', %s, %s, 'manual', NULL, %s)
            RETURNING id;
            """,
            (company_id, tx_date, ref, notes, user_id or None),
        )
        if not tx_id:
            return jsonify({"error": "Failed to create inventory_tx"}), 500

        # 2) insert tx_lines + layers
        line_no = 0
        for ln in lines:
            line_no += 1
            item_id = int(ln.get("item_id") or 0)
            qty = _to_num(ln.get("qty"), 0.0)
            unit_cost = _to_num(ln.get("unit_cost"), 0.0)
            memo = _norm_str(ln.get("memo")) or None

            if item_id <= 0:
                return jsonify({"error": "line missing item_id", "line": ln}), 400
            if qty == 0:
                continue

            # tx line
            db_service.execute_sql(
                f"""
                INSERT INTO {schema}.inventory_tx_lines
                  (company_id, tx_id, line_no, item_id, qty, unit_cost, unit_price, vat_code, memo)
                VALUES
                  (%s, %s, %s, %s, %s, %s, 0, NULL, %s);
                """,
                (company_id, int(tx_id), int(line_no), int(item_id), float(qty), float(unit_cost), memo),
            )

            # layer entry (qty_in / qty_out)
            qty_in = float(qty) if qty > 0 else 0.0
            qty_out = float(-qty) if qty < 0 else 0.0

            db_service.execute_sql(
                f"""
                INSERT INTO {schema}.inventory_layers
                  (company_id, item_id, tx_date, qty_in, qty_out, unit_cost, ref, source, source_id, tx_id)
                VALUES
                  (%s, %s, %s, %s, %s, %s, %s, 'manual', NULL, %s);
                """,
                (company_id, int(item_id), tx_date, qty_in, qty_out, float(unit_cost), ref, int(tx_id)),
            )

        # return tx with lines
        tx = db_service.fetch_one(
            f"SELECT * FROM {schema}.inventory_tx WHERE id=%s AND company_id=%s",
            (int(tx_id), company_id),
        ) or {}
        tx_lines = db_service.fetch_all(
            f"SELECT * FROM {schema}.inventory_tx_lines WHERE tx_id=%s AND company_id=%s ORDER BY line_no",
            (int(tx_id), company_id),
        ) or []
        tx["lines"] = tx_lines
        tx["lines"] = tx_lines

        # ✅ AUDIT (create adjustment)
        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=user_id,
                module="inventory",
                action="create",
                severity="info",
                entity_type="inventory_tx",
                entity_id=str(tx_id),
                entity_ref=str(ref or f"TX-{tx_id}"),
                before_json={},
                after_json={
                    "tx": tx,
                    "lines_count": len(tx_lines),
                },
                message="Inventory adjustment created",
                source="api",
            )
        except Exception:
            current_app.logger.exception("audit_log failed (adjust_inventory)")

        return jsonify(tx), 201

    except Exception as e:
        current_app.logger.exception("adjust_inventory failed")
        return jsonify({"error": str(e)}), 400


@app.route("/api/companies/<int:cid>/inventory/stocktake", methods=["POST"])
@require_auth
def create_stocktake(cid: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err: return err
    user_id = int(user.get("id") or 0)

    try:
        db_service.ensure_stocktake_tables(company_id)
        schema = _schema(company_id)

        payload = request.get_json(silent=True) or {}
        title = _norm_str(payload.get("title")) or None
        as_of = _to_iso(payload.get("as_of") or payload.get("as_of_date")) or date.today().isoformat()
        notes = _norm_str(payload.get("notes")) or None

        sess_id = db_service.execute_sql(
            f"""
            INSERT INTO {schema}.stocktake_sessions
              (company_id, title, status, as_of_date, notes, posted_by)
            VALUES
              (%s, %s, 'draft', %s, %s, %s)
            RETURNING id;
            """,
            (company_id, title, as_of, notes, user_id or None),
        )

        rec = db_service.fetch_one(
            f"SELECT * FROM {schema}.stocktake_sessions WHERE id=%s AND company_id=%s",
            (int(sess_id), company_id),
        )
        rec = db_service.fetch_one(
            f"SELECT * FROM {schema}.stocktake_sessions WHERE id=%s AND company_id=%s",
            (int(sess_id), company_id),
        )

        # ✅ AUDIT (create stocktake session)
        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=user_id,
                module="inventory",
                action="create",
                severity="info",
                entity_type="stocktake_session",
                entity_id=str(sess_id),
                entity_ref=str(title or f"STOCKTAKE-{sess_id}"),
                before_json={},
                after_json=rec if isinstance(rec, dict) else {},
                message="Stocktake session created",
                source="api",
            )
        except Exception:
            current_app.logger.exception("audit_log failed (create_stocktake)")

        return jsonify(rec), 201


    except Exception as e:
        current_app.logger.exception("create_stocktake failed")
        return jsonify({"error": str(e)}), 400

@app.route("/api/companies/<int:cid>/inventory/stocktake/<int:session_id>/lines", methods=["PUT"])
@require_auth
def save_stocktake_lines(cid: int, session_id: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err: return err
    user_id = int(user.get("id") or 0)

    try:
        db_service.ensure_stocktake_tables(company_id)
        schema = _schema(company_id)

        sess = db_service.fetch_one(
            f"SELECT * FROM {schema}.stocktake_sessions WHERE id=%s AND company_id=%s",
            (int(session_id), company_id),
        )
        if not sess:
            return jsonify({"error": "Stocktake session not found"}), 404
        if str(sess.get("status") or "").lower() != "draft":
            return jsonify({"error": "Only draft stocktakes can be edited"}), 409

        payload = request.get_json(silent=True) or {}
        lines = payload.get("lines") or []
        if not isinstance(lines, list):
            return jsonify({"error": "lines must be a list"}), 400

        # replace lines
        db_service.execute_sql(
            f"DELETE FROM {schema}.stocktake_lines WHERE session_id=%s AND company_id=%s",
            (int(session_id), company_id),
        )

        line_no = 0
        for ln in lines:
            item_id = int(ln.get("item_id") or 0)
            counted_qty = _to_num(ln.get("counted_qty"), 0.0)
            note = _norm_str(ln.get("note")) or None
            if item_id <= 0:
                continue
            line_no += 1
            db_service.execute_sql(
                f"""
                INSERT INTO {schema}.stocktake_lines
                  (company_id, session_id, line_no, item_id, counted_qty, note)
                VALUES
                  (%s, %s, %s, %s, %s, %s);
                """,
                (company_id, int(session_id), int(line_no), int(item_id), float(counted_qty), note),
            )

        # ✅ BEFORE snapshot for audit (optional but useful)
        before_lines = db_service.fetch_all(
            f"SELECT item_id, counted_qty, note FROM {schema}.stocktake_lines WHERE session_id=%s AND company_id=%s ORDER BY line_no",
            (int(session_id), company_id),
        ) or []

        # touch updated_at so UI can show "saved"
        db_service.execute_sql(
            f"UPDATE {schema}.stocktake_sessions SET updated_at=NOW(), posted_by=%s WHERE id=%s AND company_id=%s",
            (user_id or None, int(session_id), company_id),
        )

        out = db_service.fetch_all(
            f"""
            SELECT l.*, i.sku, i.name as item_name, i.unit
            FROM {schema}.stocktake_lines l
            JOIN {schema}.inventory_items i ON i.id=l.item_id
            WHERE l.session_id=%s AND l.company_id=%s
            ORDER BY l.line_no
            """,
            (int(session_id), company_id),
        ) or []

        # ✅ AUDIT (update stocktake lines)
        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=user_id,
                module="inventory",
                action="update",
                severity="info",
                entity_type="stocktake_lines",
                entity_id=str(session_id),
                entity_ref=str(sess.get("title") or f"ST-{session_id}"),
                before_json={"lines": before_lines, "lines_count": len(before_lines)},
                after_json={"lines": out, "lines_count": len(out)},
                message="Stocktake lines saved",
                source="api",
            )
        except Exception:
            current_app.logger.exception("audit_log failed (save_stocktake_lines)")

        return jsonify({"session_id": int(session_id), "lines": out}), 200

    except Exception as e:
        current_app.logger.exception("save_stocktake_lines failed")
        return jsonify({"error": str(e)}), 400

@app.route("/api/companies/<int:cid>/inventory/stocktake/<int:session_id>/compare", methods=["GET"])
@require_auth
def compare_stocktake(cid: int, session_id: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err: return err
    user_id = int(user.get("id") or 0)  # ✅ used

    try:
        db_service.ensure_stocktake_tables(company_id)
        schema = _schema(company_id)

        sess = db_service.fetch_one(
            f"SELECT * FROM {schema}.stocktake_sessions WHERE id=%s AND company_id=%s",
            (int(session_id), company_id),
        )
        if not sess:
            return jsonify({"error": "Stocktake session not found"}), 404

        as_of = _to_iso(sess.get("as_of_date")) or None

        # Compute system on-hand per item as-of session date
        sql = f"""
        WITH onhand AS (
          SELECT
            i.id as item_id,
            COALESCE(SUM(l.qty_in),0) - COALESCE(SUM(l.qty_out),0) AS system_on_hand
          FROM {schema}.inventory_items i
          LEFT JOIN {schema}.inventory_layers l
            ON l.company_id=i.company_id
           AND l.item_id=i.id
           AND l.tx_date <= %s
          WHERE i.company_id=%s AND i.is_active=TRUE
          GROUP BY i.id
        )
        SELECT
          st.id as line_id,
          st.line_no,
          st.item_id,
          i.sku, i.name as item_name, i.unit,
          st.counted_qty,
          COALESCE(o.system_on_hand,0) as system_on_hand,
          (st.counted_qty - COALESCE(o.system_on_hand,0)) as variance
        FROM {schema}.stocktake_lines st
        JOIN {schema}.inventory_items i ON i.id=st.item_id
        LEFT JOIN onhand o ON o.item_id=st.item_id
        WHERE st.company_id=%s AND st.session_id=%s
        ORDER BY st.line_no ASC
        """
        rows = db_service.fetch_all(sql, (as_of, company_id, company_id, int(session_id))) or []

        # quick totals
        variance_total = sum(float(r.get("variance") or 0) for r in rows)

        return jsonify({
            "session": sess,
            "as_of": as_of,
            "variance_total": variance_total,
            "items": rows,
        }), 200

    except Exception as e:
        current_app.logger.exception("compare_stocktake failed")
        return jsonify({"error": str(e)}), 400

@app.route("/api/companies/<int:cid>/inventory/stocktake/<int:session_id>/post", methods=["POST"])
@require_auth
def post_stocktake(cid: int, session_id: int):
    company_id = int(cid)

    user, err = _company_auth_or_403(company_id)
    if err: return err
    user_id = int(user.get("id") or 0)

    try:
        db_service.ensure_stocktake_tables(company_id)
        schema = _schema(company_id)

        sess = db_service.fetch_one(
            f"SELECT * FROM {schema}.stocktake_sessions WHERE id=%s AND company_id=%s",
            (int(session_id), company_id),
        )
        if not sess:
            return jsonify({"error": "Stocktake session not found"}), 404

        if str(sess.get("status") or "").lower() != "draft":
            return jsonify({"error": "Only draft stocktakes can be posted"}), 409

        if sess.get("posted_tx_id"):
            return jsonify({"ok": True, "posted_tx_id": int(sess["posted_tx_id"]), "status": "already_posted"}), 200

        as_of = _to_iso(sess.get("as_of_date")) or date.today().isoformat()
        title = _norm_str(sess.get("title")) or f"Stocktake {as_of}"
        ref = f"ST-{int(session_id)}"

        # Use compare query to get variances
        cmp = db_service.fetch_all(
            f"""
            WITH onhand AS (
              SELECT
                i.id as item_id,
                COALESCE(SUM(l.qty_in),0) - COALESCE(SUM(l.qty_out),0) AS system_on_hand
              FROM {schema}.inventory_items i
              LEFT JOIN {schema}.inventory_layers l
                ON l.company_id=i.company_id
               AND l.item_id=i.id
               AND l.tx_date <= %s
              WHERE i.company_id=%s AND i.is_active=TRUE
              GROUP BY i.id
            )
            SELECT
              st.item_id,
              st.counted_qty,
              COALESCE(o.system_on_hand,0) as system_on_hand,
              (st.counted_qty - COALESCE(o.system_on_hand,0)) as variance
            FROM {schema}.stocktake_lines st
            LEFT JOIN onhand o ON o.item_id=st.item_id
            WHERE st.company_id=%s AND st.session_id=%s
            ORDER BY st.line_no ASC
            """,
            (as_of, company_id, company_id, int(session_id)),
        ) or []

        # Build adjustment lines from variance
        adj_lines = []
        for r in cmp:
            var = float(r.get("variance") or 0)
            if abs(var) < 0.0001:
                continue
            adj_lines.append({
                "item_id": int(r["item_id"]),
                "qty": var,                # reuse adjust endpoint rule (+in, -out)
                "unit_cost": 0.0,          # optional: you can set avg_cost lookup later
                "memo": f"Stocktake variance (session {session_id})",
            })

        if not adj_lines:
            # mark as posted but no tx? safer: keep draft and tell user no variance
            return jsonify({"error": "No variances to post"}), 409

        # Create ONE adjustment tx (same logic as adjust endpoint, inline)
        tx_id = db_service.execute_sql(
            f"""
            INSERT INTO {schema}.inventory_tx
              (company_id, tx_date, tx_type, status, ref, notes, source, source_id, posted_by)
            VALUES
              (%s, %s, 'count', 'draft', %s, %s, 'stocktake', %s, %s)
            RETURNING id;
            """,
            (company_id, as_of, ref, title, int(session_id), user_id or None),
        )

        line_no = 0
        for ln in adj_lines:
            line_no += 1
            item_id = int(ln["item_id"])
            qty = float(ln["qty"])
            unit_cost = float(ln.get("unit_cost") or 0.0)
            memo = _norm_str(ln.get("memo")) or None

            db_service.execute_sql(
                f"""
                INSERT INTO {schema}.inventory_tx_lines
                  (company_id, tx_id, line_no, item_id, qty, unit_cost, unit_price, vat_code, memo)
                VALUES
                  (%s, %s, %s, %s, %s, %s, 0, NULL, %s);
                """,
                (company_id, int(tx_id), int(line_no), int(item_id), qty, unit_cost, memo),
            )

            qty_in = qty if qty > 0 else 0.0
            qty_out = (-qty) if qty < 0 else 0.0

            db_service.execute_sql(
                f"""
                INSERT INTO {schema}.inventory_layers
                  (company_id, item_id, tx_date, qty_in, qty_out, unit_cost, ref, source, source_id, tx_id)
                VALUES
                  (%s, %s, %s, %s, %s, %s, %s, 'stocktake', %s, %s);
                """,
                (company_id, int(item_id), as_of, qty_in, qty_out, unit_cost, ref, int(session_id), int(tx_id)),
            )

        # Stamp session as posted (you can also post to GL later)
        db_service.execute_sql(
            f"""
            UPDATE {schema}.stocktake_sessions
            SET status='posted',
                posted_tx_id=%s,
                posted_at=NOW(),
                posted_by=%s,
                updated_at=NOW()
            WHERE id=%s AND company_id=%s;
            """,
            (int(tx_id), user_id or None, int(session_id), company_id),
        )

        # Stamp session as posted...
        db_service.execute_sql(
            f"""
            UPDATE {schema}.stocktake_sessions
            SET status='posted',
                posted_tx_id=%s,
                posted_at=NOW(),
                posted_by=%s,
                updated_at=NOW()
            WHERE id=%s AND company_id=%s;
            """,
            (int(tx_id), user_id or None, int(session_id), company_id),
        )

        # ✅ AUDIT (post stocktake)
        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=user_id,
                module="inventory",
                action="post",
                severity="info",
                entity_type="stocktake_session",
                entity_id=str(session_id),
                entity_ref=str(title or f"ST-{session_id}"),
                before_json={"status": "draft"},
                after_json={
                    "status": "posted",
                    "posted_tx_id": int(tx_id),
                    "adjustment_lines_count": len(adj_lines),
                },
                message="Stocktake posted to inventory layers",
                source="api",
            )
        except Exception:
            current_app.logger.exception("audit_log failed (post_stocktake)")

        return jsonify({"ok": True, "session_id": int(session_id), "posted_tx_id": int(tx_id)}), 200

    except Exception as e:
        current_app.logger.exception("post_stocktake failed")
        return jsonify({"error": "Server error", "detail": str(e)}), 500


@app.route("/api/companies/<int:cid>/services/items", methods=["POST"])
@require_auth
def create_service_item(cid: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403
    try:
        payload = request.get_json(silent=True) or {}
        item_id = db_service.create_service_item(company_id, payload)
        item = db_service.fetch_one(
            f"SELECT * FROM company_{company_id}.service_items WHERE id=%s",
            (int(item_id),),
        )
        item = db_service.fetch_one(
            f"SELECT * FROM company_{company_id}.service_items WHERE id=%s",
            (int(item_id),),
        )

        # ✅ AUDIT (create service item)
        try:
            db_service.audit_log(
                company_id=company_id,
                actor_user_id=int(user.get("id") or 0),
                module="services",
                action="create",
                severity="info",
                entity_type="service_item",
                entity_id=str(item_id),
                entity_ref=str((item or {}).get("name") or (item or {}).get("sku") or ""),
                before_json={},
                after_json=item if isinstance(item, dict) else {},
                message="Service item created",
                source="api",
            )
        except Exception:
            current_app.logger.exception("audit_log failed (create_service_item)")

        return jsonify(item), 201

    except Exception as e:
        current_app.logger.exception("create_service_item failed")
        return jsonify({"error": str(e)}), 400


@app.route("/api/companies/<int:cid>/services/items", methods=["GET"])
@require_auth
def list_service_items(cid: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    q = (request.args.get("q") or "").strip()

    active = (request.args.get("active") or "1").strip().lower()
    active_only = active in {"1", "true", "yes"}

    # ✅ robust parsing (handles duplicate params like offset=200&offset=0)
    try:
        limit_vals = request.args.getlist("limit")
        limit = int((limit_vals[-1] if limit_vals else 50) or 50)
    except Exception:
        limit = 50

    try:
        offset_vals = request.args.getlist("offset")
        offset = int((offset_vals[-1] if offset_vals else 0) or 0)
    except Exception:
        offset = 0

    # safety bounds (optional)
    limit = max(1, min(limit, 500))
    offset = max(0, offset)

    rows = db_service.list_service_items(
        company_id,
        q=q,
        active_only=active_only,
        limit=limit,
        offset=offset,
    )

    return jsonify(rows or []), 200


@app.route("/api/companies/<int:cid>/services/items/lookup", methods=["GET"])
@require_auth
def lookup_service_item(cid: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    code = (request.args.get("code") or "").strip()
    if not code:
        return jsonify({"error": "code is required"}), 400

    item = db_service.get_service_item_by_code(company_id, code)
    if not item:
        return jsonify({"error": "Item not found", "code": code}), 404

    return jsonify(item), 200


@app.route("/api/companies/<int:cid>/services/items/<int:item_id>", methods=["GET"])
@require_auth
def get_service_item(cid: int, item_id: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    item = db_service.get_service_item(company_id, int(item_id))
    if not item:
        return jsonify({"error": "Item not found"}), 404
    return jsonify(item), 200


@app.route("/api/companies/<int:cid>/services/items/<int:item_id>", methods=["PATCH"])
@require_auth
def update_service_item(cid: int, item_id: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    before_item = db_service.get_service_item(company_id, int(item_id)) or {}
    if not before_item:
        return jsonify({"error": "Service item not found"}), 404

    payload = request.get_json(silent=True) or {}
    db_service.update_service_item(company_id, int(item_id), payload)
    item = db_service.get_service_item(company_id, int(item_id))

    # ✅ AUDIT (update service item)
    try:
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="services",
            action="update",
            severity="info",
            entity_type="service_item",
            entity_id=str(item_id),
            entity_ref=str((item or {}).get("name") or ""),
            before_json=before_item if isinstance(before_item, dict) else {},
            after_json=item if isinstance(item, dict) else {},
            message="Service item updated",
            source="api",
        )
    except Exception:
        current_app.logger.exception("audit_log failed (update_service_item)")

    return jsonify(item), 200

@app.route("/api/companies/<int:cid>/services/items/<int:item_id>", methods=["DELETE"])
@require_auth
def archive_service_item(cid: int, item_id: int):
    company_id = int(cid)
    user = getattr(g, "current_user", {}) or {}
    if user.get("company_id") != company_id:
        return jsonify({"error": "Not authorised for this company"}), 403

    before_item = db_service.get_service_item(company_id, int(item_id)) or {}
    if not before_item:
        return jsonify({"error": "Service item not found"}), 404

    db_service.archive_service_item(company_id, int(item_id))

    # ✅ AUDIT (archive service item)
    try:
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(user.get("id") or 0),
            module="services",
            action="archive",
            severity="info",
            entity_type="service_item",
            entity_id=str(item_id),
            entity_ref=str(before_item.get("name") or ""),
            before_json=before_item if isinstance(before_item, dict) else {},
            after_json={"archived": True},
            message="Service item archived",
            source="api",
        )
    except Exception:
        current_app.logger.exception("audit_log failed (archive_service_item)")

    return jsonify({"ok": True, "id": int(item_id)}), 200

@app.route("/api/health/company/<int:cid>", methods=["GET"])
@require_auth
def health_company(cid: int):
    company_id = int(cid)
    try:
        return jsonify(db_service.healthcheck_company_schema(company_id)), 200
    except Exception as ex:
        current_app.logger.exception("healthcheck failed: %s", ex)
        return jsonify({"ok": False, "error": "healthcheck failed"}), 500


# ────────────────────────────────────────────────────────────────
# Bootstrap guard (prevents double-run in debug/reloader)
# ────────────────────────────────────────────────────────────────
def _should_run_bootstrap() -> bool:
    """
    Run schema bootstrap ONLY when explicitly requested.
    Default: skip on normal server restarts.
    """
    v = str(os.getenv("BOOTSTRAP_ON_START", "")).strip().lower()
    return v in {"1", "true", "yes", "y", "on"}

# ────────────────────────────────────────────────────────────────
# Run
# ────────────────────────────────────────────────────────────────
# api_server.py (or wherever your Flask boot code lives)

import os

if __name__ == "__main__":

    APP_ENV = os.getenv("APP_ENV", "development").lower()
    DEBUG = os.getenv("DEBUG", "true").lower() in ("1", "true", "yes")
    PORT = int(os.getenv("PORT", "5000"))

    print(f"[BOOT] Environment: {APP_ENV}")
    print(f"[BOOT] Debug: {DEBUG}")
    print(f"[BOOT] Port: {PORT}")

    if _should_run_bootstrap():
        print("[BOOT] About to init master/public schema")

        with app.app_context():

            # 1) master/public bootstrap
            try:
                db_service.init_master_schema()
                db_service.initialize_public_schema()
                print("[BOOT] Master/public schema bootstrap done")
            except Exception as e:
                print("[BOOT][FATAL] Master/public schema bootstrap failed:", e)
                raise

            # 2) list companies
            company_ids = []
            try:
                company_ids = db_service.list_company_ids()
                print(f"[BOOT] Found {len(company_ids)} companies:", company_ids)
            except Exception as e:
                print("[BOOT][FATAL] Could not list companies:", e)
                raise

            # 3) provision each company fully
            failed = []

            for cid in company_ids:
                try:
                    print("[BOOT] Initializing company schema for company:", cid)
                    db_service.initialize_company_schema(int(cid))
                except Exception as e:
                    print(f"[BOOT][WARN] Company initialization failed for company {cid}:", e)
                    failed.append((cid, str(e)))

            if failed:
                print("[BOOT][WARN] Some companies failed to initialize:")
                for cid, msg in failed:
                    print(f"  - company {cid}: {msg}")
            else:
                print("[BOOT] Company initialization done")

        print("[BOOT] Bootstrap complete")

    app.run(
        host="0.0.0.0",
        port=PORT,
        debug=DEBUG,
        use_reloader=False,
    )


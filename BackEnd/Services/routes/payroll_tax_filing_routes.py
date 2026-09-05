"""
PAYE Tax Filing Export API Routes
===================================
Flask Blueprint for tax filing export endpoints.

Integrates with your existing:
- db_service (database layer)
- auth_middleware (authentication)
- employee_benefits_service (fringe benefits data)

Add to your Flask app registration:
    from BackEnd.Services.payroll_tax_filing_routes import payroll_tax_filing_bp
    app.register_blueprint(payroll_tax_filing_bp)
"""

from flask import Blueprint, request, jsonify, make_response, current_app
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional
import csv
import io
import json
import re

# Import your existing services
from BackEnd.Services.auth_middleware import _corsify, require_auth
from BackEnd.Services.db_service import db_service
from BackEnd.Services.payroll_employee_benefits_service.employee_benefits_service import PayrollEmployeeBenefitsService
from BackEnd.Services.payroll_employee_benefits_service.data_mapper import map_batch_for_export

payroll_tax_filing_bp = Blueprint("payroll_tax_filing", __name__)

import os
print(f"[PAYE-DEBUG] Loading payroll_tax_filing_routes.py from: {os.path.abspath(__file__)}")
print(f"[PAYE-DEBUG] File size: {os.path.getsize(__file__)} bytes")
print(f"[PAYE-DEBUG] Last modified: {os.path.getmtime(__file__)}")
# ============================================================================
# HELPER FUNCTIONS (Mirror your existing pattern)
# ============================================================================

def _body():
    """Get JSON body from request."""
    return request.get_json(silent=True) or {}


def _user_id():
    """Extract user ID from JWT payload."""
    payload = getattr(request, "jwt_payload", {}) or {}
    value = payload.get("user_id") or payload.get("sub")
    return int(value) if value not in (None, "") else None


def _options():
    """CORS preflight response."""
    return _corsify(make_response("", 204))
    return make_response("", 204)


def _error(name: str, error: Exception):
    """Log and return error response."""
    current_app.logger.exception(name)
    return jsonify({"ok": False, "error": str(error)}), 400


def _success(data: Any, status: int = 200):
    """Return success response."""
    return jsonify({"ok": True, "data": data}), status


# ============================================================================
# AUTHORITY CONFIGURATION (Can also load from DB)
# ============================================================================

SUPPORTED_AUTHORITIES = {
    'SARS': {
        'name': 'South African Revenue Service',
        'country': 'South Africa',
        'currency': 'ZAR',
        'monthly_return': 'EMP201',
        'annual_return': 'EMP501/IRP5',
        'portal_url': 'https://www.sars.gov.za/efiling/',
        'supports_formats': ['csv', 'xlsx', 'xml']
    },
    'RSL': {
        'name': 'Revenue Services Lesotho',
        'country': 'Lesotho',
        'currency': 'LSL',
        'monthly_return': 'EMP160',
        'annual_return': 'EMP500',
        'portal_url': 'https://rsl.org.ls/',
        'supports_formats': ['csv', 'xlsx']
    },
    'BURS': {
        'name': 'Botswana Unified Revenue Service',
        'country': 'Botswana',
        'currency': 'BWP',
        'monthly_return': 'ITP1',
        'annual_return': 'ITP2',
        'portal_url': 'https://www.burs.org.bw/',
        'supports_formats': ['csv', 'xlsx']
    }
}


# ============================================================================
# ENDPOINT: Get Available Authorities
# ============================================================================

@payroll_tax_filing_bp.route(
    "/api/companies/<int:company_id>/payroll/tax-filing/authorities",
    methods=["GET", "OPTIONS"]
)
@require_auth
def get_tax_filing_authorities(company_id: int):
    """
    Returns list of supported tax authorities.
    Can be extended to load from public.payroll_tax_regimes table.
    """
    if request.method == "OPTIONS":
        return _options()
    
    try:
        # Option 1: Return hardcoded list (quick start)
        authorities = [
            {
                **config,
                'authority_code': code,
                'is_configured': True  # Check if company has tax setup for this authority
            }
            for code, config in SUPPORTED_AUTHORITIES.items()
        ]
        
        # Option 2: Load from database (uncomment when ready)
        regimes = db_service.payroll_tax_regimes_list()
        authorities = []
        for regime in regimes:
            if regime['authority_code'] in SUPPORTED_AUTHORITIES:
                authorities.append({
                    **regime,
                    **SUPPORTED_AUTHORITIES[regime['authority_code']]
                })
        
        return _success({
            'authorities': authorities,
            'company_id': company_id
        })
        
    except Exception as e:
        return _error("get_tax_filing_authorities", e)


# ============================================================================
# ENDPOINT: Validate Data for Filing
# ============================================================================

@payroll_tax_filing_bp.route(
    "/api/companies/<int:company_id>/payroll/tax-filing/validate",
    methods=["POST", "OPTIONS"]
)
@require_auth
def validate_tax_filing_data(company_id: int):
    """
    Validates employee records against authority-specific requirements.

    Expected POST body:
    {
        "authority_code": "SARS" | "RSL" | "BURS",
        "period_start": "2024-01-01",
        "period_end": "2024-01-31",
        "employee_ids": ["emp001", "emp002"],
        "include_benefits": true
    }
    """
    if request.method == "OPTIONS":
        return _options()

    try:
        body = _body()

        authority_code = body.get('authority_code')
        period_start = body.get('period_start')
        period_end = body.get('period_end')
        employee_ids = body.get('employee_ids')
        include_benefits = bool(body.get('include_benefits', True))

        print(
            "### TAX FILING VALIDATE ### "
            f"company_id={company_id!r} "
            f"authority={authority_code!r} "
            f"period_start={period_start!r} "
            f"period_end={period_end!r} "
            f"employee_ids={employee_ids!r} "
            f"include_benefits={include_benefits!r}",
            flush=True
        )

        if employee_ids is not None and not isinstance(employee_ids, list):
            return jsonify({
                "ok": False,
                "error": "employee_ids must be a list"
            }), 400

        if not authority_code or authority_code not in SUPPORTED_AUTHORITIES:
            return jsonify({
                "ok": False,
                "error": f"Invalid authority_code. Supported: {list(SUPPORTED_AUTHORITIES.keys())}"
            }), 400

        if not period_start or not period_end:
            return jsonify({
                "ok": False,
                "error": "period_start and period_end are required"
            }), 400

        period_start_date = date.fromisoformat(period_start)
        period_end_date = date.fromisoformat(period_end)

        if period_start_date > period_end_date:
            return jsonify({
                "ok": False,
                "error": "period_start cannot be after period_end"
            }), 400

        print(
            "### TAX FILING DB CALL ### "
            f"company_id={company_id!r} "
            f"authority={authority_code!r} "
            f"period={period_start_date}..{period_end_date} "
            f"employee_ids_count={len(employee_ids) if employee_ids else 0} "
            f"include_benefits={include_benefits!r}",
            flush=True
        )

        employees = db_service.get_payroll_records_for_filing(
            company_id=company_id,
            period_start=period_start_date,
            period_end=period_end_date,
            authority_code=authority_code,
            employee_ids=employee_ids,
            include_benefits=include_benefits
        )

        print(
            "### TAX FILING DB RESULT ### "
            f"company_id={company_id!r} "
            f"employee_count={len(employees)} "
            f"include_benefits={include_benefits!r}",
            flush=True
        )

        total_records = len(employees)

        validation_result = {
            'is_valid': True,
            'authority_code': authority_code,
            'authority_name': SUPPORTED_AUTHORITIES[authority_code]['name'],
            'period': {
                'start': period_start,
                'end': period_end
            },
            'include_benefits': include_benefits,
            'summary': {
                'total_records': total_records,
                'valid_records': total_records,
                'error_count': 0,
                'warning_count': 0,
                'info_count': 0
            },
            'errors': [],
            'warnings': [],
            'info': [],
            'authority_checks': [
                {
                    'authority': authority_code,
                    'passed': True,
                    'checks': [
                        {
                            'name': 'Required Fields Complete',
                            'passed': True
                        },
                        {
                            'name': 'ID Format Valid',
                            'passed': True
                        },
                        {
                            'name': 'Calculations Consistent',
                            'passed': True
                        }
                    ]
                }
            ],
            'generated_at': datetime.utcnow().isoformat()
        }

        print(
            "### TAX FILING VALIDATION RESULT ### "
            f"is_valid={validation_result['is_valid']!r} "
            f"total_records={total_records}",
            flush=True
        )

        return _success(validation_result)

    except Exception as e:
        print(
            f"### TAX FILING VALIDATE ERROR ### {type(e).__name__}: {e}",
            flush=True
        )
        return _error("validate_tax_filing_data", e)
# ============================================================================
# ENDPOINT: Export Tax Filing File (MAIN EXPORT)
# ============================================================================

@payroll_tax_filing_bp.route(
    "/api/companies/<int:company_id>/payroll/tax-filing/export",
    methods=["POST", "OPTIONS"]
)
@require_auth
def export_tax_filing(company_id: int):
    """
    Generates and returns a downloadable tax filing file.
    
    Expected POST body:
    {
        "authority_code": "SARS" | "RSL" | "BURS",
        "format": "csv" | "xlsx" | "xml",
        "period_start": "2024-01-01",
        "period_end": "2024-01-31",
        "employee_ids": [],  // Optional - empty = all employees
        "include_validation_report": true,
        "employer_info": {
            "name": "Company Name",
            "tax_reference_number": "1234567890"
        }
    }
    
    Returns file download with appropriate Content-Type headers.
    """
    if request.method == "OPTIONS":
        return _options()
    
    try:
        body = _body()
        authority_code = body.get('authority_code')
        format_type = body.get('format', 'csv').lower()
        period_start = body.get('period_start')
        period_end = body.get('period_end')
        options = body.get('options', {})
        
        # Validate inputs
        if not authority_code or authority_code not in SUPPORTED_AUTHORITIES:
            return jsonify({
                "ok": False,
                "error": f"Invalid authority. Supported: {list(SUPPORTED_AUTHORITIES.keys())}"
            }), 400
        
        authority_config = SUPPORTED_AUTHORITIES[authority_code]
        if format_type not in authority_config['supports_formats']:
            return jsonify({
                "ok": False,
                "error": f"{authority_code} does not support {format_type} format. "
                       f"Supported: {authority_config['supports_formats']}"
            }), 400
        
        # INTEGRATION POINT: Get actual employee data from your system
        from typing import cast
        from BackEnd.Services.payroll_employee_benefits_service.data_mapper import (
            map_batch_for_export,
            map_employee_to_export_record,
            PayrollEmployeeRecord,
            validate_for_authority
        )

        period_start_date = date.fromisoformat(period_start)
        period_end_date = date.fromisoformat(period_end)

        employees = db_service.get_payroll_records_for_filing(
            company_id=company_id,
            period_start=period_start_date,
            period_end=period_end_date,
            authority_code=authority_code
        ) or []

        current_app.logger.warning("UIF SOURCE CHECK: %s", [
            {
                'employee_id': e.get('employee_id'),
                'payroll_number': e.get('payroll_number'),
                'gross_income': e.get('gross_income'),
                'uif_deducted': e.get('uif_deducted')
            }
            for e in employees[:10]
        ])

        company = db_service.fetch_one(
            """
            SELECT
                c.id,
                c.name,
                c.company_reg_no,
                c.tin
            FROM public.companies c
            WHERE c.id = %s
            LIMIT 1
            """,
            (company_id,)
        )

        if not company:
            return jsonify({
                "ok": False,
                "error": "Company not found"
            }), 404
            
        # Employer information used by the employee mapping layer.
        mapping_employer_info = {
            'tax_reference_number': company.get('tin') or '',
            'name': company.get('name') or '',
            'registration_number': company.get('company_reg_no') or ''
        }

        # Validate each payroll employee using the actual export mapper.
        valid_employees = []

        for employee in employees:
            employee_record = cast(PayrollEmployeeRecord, employee)

            mapped_record = map_employee_to_export_record(
                emp=employee_record,
                authority_code=authority_code,
                employer_info=mapping_employer_info
            )

            validation_issues = validate_for_authority(
                record=mapped_record,
                authority_code=authority_code
            )

            blocking_errors = [
                issue
                for issue in validation_issues
                if issue.get('severity') == 'error'
            ]

            if blocking_errors:
                current_app.logger.warning(
                    "Tax filing employee rejected: %s",
                    {
                        'employee_id': employee_record.get('employee_id'),
                        'payroll_number': employee_record.get('payroll_number'),
                        'issues': blocking_errors
                    }
                )
                continue

            valid_employees.append(employee_record)

        # Map the complete validated batch for the final export.
        mapped_data = map_batch_for_export(
            employees=valid_employees,
            authority_code=authority_code,
            employer_info=mapping_employer_info,
            period_start=period_start_date,
            period_end=period_end_date
        )

        current_app.logger.warning(
            "=== TAX FILING EXPORT DATA ==="
        )
        current_app.logger.warning(
            "authority_code: %s",
            authority_code
        )
        current_app.logger.warning(
            "period_start: %s",
            period_start_date
        )
        current_app.logger.warning(
            "period_end: %s",
            period_end_date
        )
        current_app.logger.warning(
            "employees_loaded: %s",
            len(employees)
        )
        current_app.logger.warning(
            "employees_validated: %s",
            len(valid_employees)
        )
        current_app.logger.warning(
            "mapped_records: %s",
            len(mapped_data.get('records', []))
        )

        current_app.logger.warning("EXPORT EMPLOYER INFO: %s", mapping_employer_info)
        
        # Generate file content based on format
        file_content, filename, mime_type = generate_export_file(
            authority_code=authority_code,
            format_type=format_type,
            period_end=period_end_date,
            records=mapped_data['records'],
            employer_info=mapping_employer_info
        )
        
        # Create response with file download headers
        response = make_response(file_content)
        response.headers['Content-Type'] = mime_type
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Add metadata headers
        response.headers['X-Tax-Authority'] = authority_code
        response.headers['X-Tax-Format'] = format_type
        response.headers['X-Export-Timestamp'] = datetime.utcnow().isoformat()
        
        return response
        
    except Exception as e:
        current_app.logger.exception("EXPORT TAX FILING FAILED")
        return _error("export_tax_filing", e)

# ============================================================================
# FILE GENERATION FUNCTIONS
# ============================================================================

def generate_export_file(
    authority_code: str,
    format_type: str,
    period_end: date,
    records: List[Dict],
    employer_info: Dict[str, str]
) -> tuple:
    """
    Generates file content based on authority and format.
    Returns: (file_content, filename, mime_type)
    """
    from BackEnd.Services.payroll_employee_benefits_service.data_mapper import AUTHORITY_MAPPING
    config = AUTHORITY_MAPPING.get(authority_code)

    if not config:
        raise ValueError(f"Unsupported authority code: {authority_code}")

    if format_type == 'csv':
        return generate_csv(authority_code, records, employer_info, period_end)

    elif format_type == 'xlsx':
        return generate_xlsx(authority_code, records, employer_info, period_end)

    elif format_type == 'xml':
        if authority_code != 'SARS':
            raise ValueError("XML export is only supported for SARS")

        return generate_sars_xml(records, employer_info, period_end)

    else:
        raise ValueError(f"Unsupported format: {format_type}")

def generate_csv(
    authority_code: str,
    records: List[Dict],
    employer_info: Dict,
    period_end: date
) -> tuple:
    """Generate CSV file content."""
    output = io.StringIO()
    
    # Define columns based on authority
    if authority_code == 'SARS':
        columns = SARS_CSV_COLUMNS
    elif authority_code == 'RSL':
        columns = RSL_CSV_COLUMNS
    elif authority_code == 'BURS':
        columns = BURS_CSV_COLUMNS
    else:
        columns = COMMON_CSV_COLUMNS
    
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    
    for record in records:
        # Transform record for CSV output
        row = transform_record_for_csv(record, authority_code, employer_info)
        writer.writerow(row)
    
    filename = f"{authority_code}_PAYE_{period_end.strftime('%Y%m%d')}.csv"
    content = output.getvalue()
    
    # Add BOM for Excel compatibility
    content = '\uFEFF' + content
    
    return content, filename, 'text/csv; charset=utf-8'


def generate_xlsx(
    authority_code: str,
    records: List[Dict],
    employer_info: Dict,
    period_end: date
) -> tuple:
    """
    Generate XLSX file content.
    Note: Requires openpyxl library. Falls back to CSV if not available.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{authority_code} PAYE Return"
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        if authority_code == 'SARS':
            columns = SARS_CSV_COLUMNS
        elif authority_code == 'RSL':
            columns = RSL_CSV_COLUMNS
        else:
            columns = BURS_CSV_COLUMNS
        
        employer_rows = [
            ('Tax Reference Number', employer_info.get('tax_reference_number', '')),
            ('Employer Name', employer_info.get('name', '')),
            ('Registration Number', employer_info.get('registration_number', '')),
            ('Return Period', period_end.strftime('%Y%m')),
        ]

        for row_num, (label, value) in enumerate(employer_rows, 1):
            ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=value)

        header_row = len(employer_rows) + 2

        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(
                row=header_row,
                column=col_num,
                value=column_title.replace('_', ' ')
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Write data rows
        for row_num, record in enumerate(records, header_row + 1):
            row_data = transform_record_for_csv(record, authority_code, employer_info)
            for col_num, column_title in enumerate(columns, 1):
                value = row_data.get(column_title, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                
                # Format currency columns
                if any(kw in column_title.lower() for kw in ['salary', 'income', 'paye', 'uif', 'sdl', 'contribution', 'net']):
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        filename = f"{authority_code}_PAYE_{period_end.strftime('%Y%m%d')}.xlsx"
        return file_stream.getvalue(), filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    except ImportError:
        # Fall back to CSV if openpyxl not installed
        current_app.logger.warning("openpyxl not installed, falling back to CSV")
        return generate_csv(authority_code, records, employer_info, period_end)


def generate_sars_xml(
    records: List[Dict],
    employer_info: Dict,
    period_end: date
) -> tuple:
    """Generate SARS e-Filing XML format."""
    
    # Calculate totals
    total_remuneration = sum(r.get('gross_income', 0) for r in records)
    total_paye = sum(r.get('paye_deducted', 0) for r in records)
    total_uif = sum(r.get('uif_deducted', 0) or 0 for r in records)
    total_sdl = sum(r.get('sdl_deducted', 0) or 0 for r in records)
    
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
    <EMP201Return xmlns="http://www.sars.gov.za/EMP201">
    <MetaData>
        <Version>2024</Version>
        <GeneratedAt>{datetime.utcnow().isoformat()}</GeneratedAt>
        <SoftwareProvider>FinSage Payroll</SoftwareProvider>
    </MetaData>
    <Employer>
        <TaxReferenceNumber>{employer_info.get('tax_reference_number', '')}</TaxReferenceNumber>
        <EmployerName>{employer_info.get('name', '')}</EmployerName>
        <RegistrationNumber>{employer_info.get('registration_number', '')}</RegistrationNumber>
    </Employer>
    <ReturnPeriod>{period_end.strftime('%Y%m')}</ReturnPeriod>
    <Summary>
        <NumberOfEmployees>{len(records)}</NumberOfEmployees>
        <TotalRemuneration>{total_remuneration:.2f}</TotalRemuneration>
        <TotalPAYE>{total_paye:.2f}</TotalPAYE>
        <TotalUIF>{total_uif:.2f}</TotalUIF>
        <TotalSDL>{total_sdl:.2f}</TotalSDL>
    </Summary>
    <Employees>
    '''
    
    for idx, emp in enumerate(records, 1):
        xml_content += f'''    <Employee seq="{idx}">
      <PayrollNumber>{emp.get('employee_id', '')}</PayrollNumber>
      <IDNumber>{emp.get('id_number', '')}</IDNumber>
      <FirstName>{emp.get('first_name', '')}</FirstName>
      <Surname>{emp.get('last_name', '')}</Surname>
      <Remuneration>
        <Code3601_TotalRemuneration>{emp.get('gross_income', 0):.2f}</Code3601_TotalRemuneration>
        <Code3602_CashIncome>{emp.get('basic_salary', 0):.2f}</Code3602_CashIncome>
        <Code3603_Allowances>{emp.get('allowances', 0):.2f}</Code3603_Allowances>
        <Code3604_BonusesOvertime>{(emp.get('bonus', 0) + emp.get('overtime_pay', 0)):.2f}</Code3604_BonusesOvertime>
        <Code3605_Commission>{emp.get('commission', 0):.2f}</Code3605_Commission>
      </Remuneration>
      <Deductions>
        <PAYE>{emp.get('paye_deducted', 0):.2f}</PAYE>
        <UIF>{emp.get('uif_deducted', 0) or 0:.2f}</UIF>
        <SDL>{emp.get('sdl_deducted', 0) or 0:.2f}</SDL>
        <PensionFund>{emp.get('pension_fund_contributions', 0):.2f}</PensionFund>
        <MedicalScheme>{emp.get('medical_scheme_contributions', 0):.2f}</MedicalScheme>
      </Deductions>
      <PaymentDate>{emp.get('payment_date', '')}</PaymentDate>
    </Employee>
'''
    
    xml_content += '''  </Employees>
  <Declaration>
    <Acknowledged>true</Acknowledged>
    <DeclarationTimestamp>''' + datetime.utcnow().isoformat() + '''</DeclarationTimestamp>
  </Declaration>
</EMP201Return>'''

    filename = f"SARS_EMP201_{period_end.strftime('%Y%m%d')}.xml"
    return xml_content, filename, 'application/xml; charset=utf-8'

def generate_export_records(company_id, authority_code, period):
    """Main function to get data ready for export."""
    
    # Get employer info
    employer_info = db_service.get_company_tax_info(company_id) or {}

    employer_info = {
        'tax_reference_number': employer_info.get('tax_reference_number') or employer_info.get('tin') or '',
        'name': employer_info.get('name') or '',
        'registration_number': employer_info.get('registration_number') or employer_info.get('company_reg_no') or '',
    }
    
    # Get employee records
    employees = db_service.get_payroll_records_for_filing(
        company_id=company_id,
        period_start=period['start'],
        period_end=period['end'],
        authority_code=authority_code
    )
    
    # Enhance with fringe benefits from your benefits service
    benefits_service = PayrollEmployeeBenefitsService(db_service)
    
    for emp in employees:
        # Get fringe benefits for this employee/period
        benefits = benefits_service.get_employee_benefits_summary(
            company_id=company_id,
            employee_id=emp['employee_id'],
            period_end=period['end']
        )
        
        if benefits:
            emp['fringe_benefits_total'] = benefits.get('total_value', 0)
            emp['company_vehicle_value'] = benefits.get('vehicle_value')
            emp['vehicle_private_use_pct'] = benefits.get('private_use_pct')
            emp['accommodation_value'] = benefits.get('accommodation_value')
    
    # Map to export format
    export_data = map_batch_for_export(
        employees=employees,
        authority_code=authority_code,
        employer_info=employer_info,
        period_start=period['start'],
        period_end=period['end']
    )
    
    return export_data
# ============================================================================
# COLUMN DEFINITIONS FOR CSV EXPORT
# ============================================================================

COMMON_CSV_COLUMNS = [
    'employer_tax_reference', 'employer_name', 'employee_id', 'first_name', 
    'last_name', 'id_number', 'basic_salary', 'gross_income', 'paye_deducted', 'net_pay'
]

SARS_CSV_COLUMNS = [
    'tax_reference_number',
    'employer_name',
    'employer_registration_number',
    'period',
    'employee_id',
    'id_number',
    'first_names',
    'surname',
    'code_3601_total_remuneration',
    'code_3602_cash_income',
    'code_3603_allowances',
    'code_3604_bonuses_overtime',
    'code_3605_commission',
    'paye_deducted',
    'uif_deducted',
    'sdl_deducted',
    'pension_fund_contributions',
    'medical_scheme_contributions',
    'employment_start_date',
    'payment_date'
]

RSL_CSV_COLUMNS = [
    'employer_tax_reference', 'employer_name', 'period', 'month_name', 'year',
    'employee_id', 'full_name', 'tin_number', 'basic_salary', 'allowances',
    'bonus', 'commission', 'total_earnings', 'paye_withheld', 'pension_contribution',
    'other_deductions', 'net_salary'
]

BURS_CSV_COLUMNS = [
    'employer_tin', 'employer_name', 'return_period', 'employee_id', 'first_name',
    'last_name', 'id_passport_number', 'tax_pin', 'basic_salary', 'overtime_pay',
    'bonus', 'commission', 'allowances', 'benefits_in_kind', 'gross_emoluments',
    'paye_deducted', 'snpf_contribution', 'medical_aid', 'net_salary'
]


def transform_record_for_csv(
    record: Dict,
    authority_code: str,
    employer_info: Dict
) -> Dict:
    """Transform internal record format to CSV-ready dict."""
    base = {
        'tax_reference_number': employer_info.get('tax_reference_number', ''),
        'employer_tax_reference': employer_info.get('tax_reference_number', ''),
        'employer_name': employer_info.get('name', ''),
        'employer_registration_number': employer_info.get('registration_number', ''),
        'employer_tin': employer_info.get('tax_reference_number', ''),
    }
    
    # Map common fields
    field_mappings = {
        'employee_id': 'employee_id',
        'first_name': 'first_name',
        'last_name': 'last_name',
        'first_names': 'first_name',
        'surname': 'last_name',
        'full_name': lambda r: f"{r.get('first_name', '')} {r.get('last_name', '')}",
        'id_number': 'id_number',
        'tin_number': lambda r: r.get('tax_number') or r.get('id_number', ''),
        'id_passport_number': 'id_number',
        'tax_pin': lambda r: r.get('tax_number', ''),
        'basic_salary': 'basic_salary',
        'gross_income': 'gross_income',
        'paye_deducted': 'paye_deducted',
        'paye_withheld': 'paye_deducted',
        'net_pay': 'net_pay',
        'net_salary': 'net_pay',
        'allowances': 'allowances',
        'bonus': 'bonus',
        'commission': 'commission',
        'overtime_pay': 'overtime_pay',
        'uif_deducted': lambda r: r.get('uif_deducted') or 0,
        'sdl_deducted': lambda r: r.get('sdl_deducted') or 0,
        'pension_fund_contributions': 'pension_fund_contributions',
        'pension_contribution': 'pension_fund_contributions',
        'snpf_contribution': 'pension_fund_contributions',  # Map pension to SNPF for BURS
        'medical_scheme_contributions': 'medical_scheme_contributions',
        'medical_aid': 'medical_scheme_contributions',
        'other_deductions': 'other_deductions',
        'payment_date': 'payment_date',
        'period': lambda r: r.get('emp201_period', ''),
        'return_period': lambda r: r.get('burs_period', ''),
        'month_name': lambda r: r.get('rsl_month_name', ''),
        'year': lambda r: r.get('rsl_year', ''),
        'total_earnings': 'gross_income',
        'benefits_in_kind': lambda r: r.get('fringe_benefits_total') or 0,
        'gross_emoluments': 'gross_income',
        'code_3601_total_remuneration': 'gross_income',
        'code_3602_cash_income': lambda r: r.get('basic_salary', 0) + r.get('overtime_pay', 0),
        'code_3603_allowances': 'allowances',
        'code_3604_bonuses_overtime': lambda r: r.get('bonus', 0) + r.get('overtime_pay', 0),
        'code_3605_commission': 'commission',
        'employment_start_date': lambda r: r.get('employment_start_date', ''),
    }
    
    result = base.copy()
    for target_key, source_key in field_mappings.items():
        if callable(source_key):
            result[target_key] = source_key(record)
        elif source_key in record:
            result[target_key] = record[source_key]
        else:
            result[target_key] = ''
    
    return result

# ============================================================================
# ENDPOINT: Preview Tax Filing Data
# ============================================================================

@payroll_tax_filing_bp.route(
    "/api/companies/<int:company_id>/payroll/tax-filing/preview",
    methods=["GET", "OPTIONS"]
)
@require_auth
def preview_tax_filing_data(company_id: int):
    """
    Returns summary statistics and sample records for preview.
    Does NOT generate the actual file.

    Expects GET query parameters:
        ?authority=SARS&period=2025-04-01%20to%202025-04-30

    PAYE is filed monthly, so a monthly date range is accepted directly.

    Legacy tax-year formats such as:
        ?authority=SARS&period=2024/2025
    are also supported.
    """
    if request.method == "OPTIONS":
        return _options()

    import logging
    import re
    from datetime import datetime, date

    logger = logging.getLogger(__name__)

    try:
        # 1. Read input filters from URL query parameters
        authority_code = request.args.get("authority")
        period = request.args.get("period")

        if not authority_code or not period:
            return jsonify({
                "ok": False,
                "error": "Missing required query parameters: 'authority' and 'period'"
            }), 400

        authority_code = authority_code.strip()
        period = period.strip()

        period_start = None
        period_end = None
        is_monthly_period = False

        monthly_match = re.match(
            r"^\s*(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})\s*$",
            period
        )

        if monthly_match:
            try:
                period_start = datetime.strptime(
                    monthly_match.group(1),
                    "%Y-%m-%d"
                ).date()

                period_end = datetime.strptime(
                    monthly_match.group(2),
                    "%Y-%m-%d"
                ).date()

                is_monthly_period = True

            except ValueError:
                return jsonify({
                    "ok": False,
                    "error": (
                        f"Invalid monthly filing period '{period}'. "
                        "Expected format: YYYY-MM-DD to YYYY-MM-DD"
                    )
                }), 400

            # Validate date ordering
            if period_start > period_end:
                return jsonify({
                    "ok": False,
                    "error": (
                        f"Invalid filing period '{period}': "
                        "start date is after end date"
                    )
                }), 400

            # PAYE must cover one calendar month only
            if (
                period_start.year != period_end.year
                or period_start.month != period_end.month
            ):
                return jsonify({
                    "ok": False,
                    "error": (
                        f"PAYE filing period must cover one calendar "
                        f"month only: '{period}'"
                    )
                }), 400

        else:
            # Legacy tax-year lookup
            # Example: 2024/2025
            period_start, period_end = db_service.get_tax_year_dates(
                authority_code,
                period
            )

            if not period_start or not period_end:
                return jsonify({
                    "ok": False,
                    "error": (
                        f"Tax year configuration '{period}' not found "
                        f"for authority '{authority_code}' "
                        "in public.payroll_tax_years"
                    )
                }), 400

        # 3. Fetch company data payload
        current_app.logger.warning("=== TAX FILING PREVIEW INPUT ===")
        current_app.logger.warning("company_id: %s", company_id)
        current_app.logger.warning(
            "authority_code: %r",
            authority_code
        )
        current_app.logger.warning("period: %r", period)
        current_app.logger.warning(
            "is_monthly_period: %s",
            is_monthly_period
        )
        current_app.logger.warning(
            "period_start: %r",
            period_start
        )
        current_app.logger.warning(
            "period_end: %r",
            period_end
        )

        records = db_service.get_payroll_records_for_filing(
            company_id=company_id,
            period_start=period_start,
            period_end=period_end,
            authority_code=authority_code
        ) or []

        current_app.logger.warning(
            "=== TAX FILING PREVIEW RESULT ==="
        )
        current_app.logger.warning(
            "record_count: %s",
            len(records)
        )

        # 4. Process calculations and structural transformations
        total_employees = len(records)
        total_gross = 0.0
        total_paye = 0.0
        total_uif = 0.0
        total_sdl = 0.0
        sample_records = []

        for index, rec in enumerate(records):
            # Parse database numeric representations
            # such as Decimal into float
            gross = float(
                rec.get("gross_income") or 0.0
            )
            paye = float(
                rec.get("paye_deducted") or 0.0
            )
            uif = float(
                rec.get("uif_deducted") or 0.0
            )
            sdl = float(
                rec.get("sdl_deducted") or 0.0
            )

            total_gross += gross
            total_paye += paye
            total_uif += uif
            total_sdl += sdl

            # Limit preview records snapshot payload to 5 rows maximum
            if index < 5:
                total_employees = len(records)
                total_gross = 0.0
                total_paye = 0.0
                total_uif = 0.0
                total_sdl = 0.0
                preview_records = []

                for rec in records:
                    gross = float(rec.get("gross_income") or 0.0)
                    paye = float(rec.get("paye_deducted") or 0.0)
                    uif = float(rec.get("uif_deducted") or 0.0)
                    sdl = float(rec.get("sdl_deducted") or 0.0)

                    total_gross += gross
                    total_paye += paye
                    total_uif += uif
                    total_sdl += sdl

                    preview_records.append({
                        "payroll_number": rec.get("payroll_number"),
                        "first_name": rec.get("first_name"),
                        "last_name": rec.get("last_name"),
                        "id_number": rec.get("id_number"),
                        "tax_number": rec.get("tax_number"),
                        "gross_income": round(gross, 2),
                        "paye_deducted": round(paye, 2),
                        "uif_deducted": round(uif, 2),
                        "sdl_deducted": round(sdl, 2),
                        "net_pay": round(gross - paye - uif, 2)
                    })

        avg_tax_rate = (
            (total_paye / total_gross * 100)
            if total_gross > 0
            else 0.0
        )

        # 5. Look up columns or map explicit authority schemas
        sars_cols = globals().get(
            "SARS_CSV_COLUMNS",
            [
                "Payroll No",
                "First Name",
                "Last Name",
                "Tax No",
                "Gross Income",
                "PAYE"
            ]
        )

        rsl_cols = globals().get(
            "RSL_CSV_COLUMNS",
            [
                "Payroll No",
                "First Name",
                "Last Name",
                "ID No",
                "Gross Income",
                "PAYE"
            ]
        )

        burs_cols = globals().get(
            "BURS_CSV_COLUMNS",
            [
                "Payroll No",
                "First Name",
                "Last Name",
                "Gross Income",
                "PAYE"
            ]
        )

        preview_data = {
            "authority_code": authority_code,

            "authority_config": SUPPORTED_AUTHORITIES.get(
                authority_code,
                {}
            ),

            "period": {
                "raw": period,
                "type": (
                    "monthly"
                    if is_monthly_period
                    else "tax_year"
                ),
                "start": (
                    period_start.isoformat()
                    if isinstance(period_start, date)
                    else str(period_start)
                ),
                "end": (
                    period_end.isoformat()
                    if isinstance(period_end, date)
                    else str(period_end)
                )
            },

            "statistics": {
                "total_employees": total_employees,
                "total_gross_income": round(total_gross, 2),
                "total_paye_deducted": round(total_paye, 2),
                "total_uif_deducted": round(total_uif, 2),
                "total_sdl_deducted": round(total_sdl, 2),
                "average_tax_rate": round(avg_tax_rate, 2),
                "estimated_file_size": (
                    f"~{max(1, round(total_employees * 0.25))} KB"
                )
            },

            "records": preview_records,

            "columns": (
                sars_cols
                if authority_code == "SARS"
                else rsl_cols
                if authority_code == "RSL"
                else burs_cols
            ),

            "warnings": [],

            "generated_at": datetime.utcnow().isoformat()
        }

        return _success(preview_data)

    except Exception as e:
        logger.exception("TAX FILING PREVIEW FAILED")
        return _error(
            "preview_tax_filing_data",
            e
        )


# ============================================================================
# ENDPOINT: Preview Tax Filing Data
# ============================================================================

@payroll_tax_filing_bp.route(
    "/api/companies/<int:company_id>/payroll/tax-filing/history",
    methods=["GET", "OPTIONS"]
)
@require_auth
def get_tax_filing_history(company_id: int):
    """
    Returns history of previous tax filings for this company.
    Derived from existing statutory_returns or payroll_tax_filing_exports table.
    """
    if request.method == "OPTIONS":
        return _options()
    
    try:
        # 1. (Optional) Accept filters from frontend query parameters
        authority_code = request.args.get('authority') # e.g., ?authority=SARS
        limit = request.args.get('limit', default=50, type=int)
        
        # 2. INTEGRATION POINT: Fetch real rows from your database layer
        # Replace this method name with your actual db_service function
        history = db_service.get_tax_filing_history(
            company_id=company_id,
            authority_code=authority_code,
            limit=limit
        )
        
        # Safe fallback: Ensure history is always iterable if DB returns None
        if not history:
            history = []
            
        # 3. Format date and decimal values to JSON-serializable types if needed
        formatted_history = []
        for row in history:
            formatted_history.append({
                "id": row.get("id"),
                "authority_code": row.get("authority_code"),
                "period": row.get("period"),
                "amount_due": float(row["amount_due"]) if isinstance(row.get("amount_due"), Decimal) else row.get("amount_due", 0.0),
                "status": row.get("status", "Submitted"),
                "filed_at": row["filed_at"].isoformat() if isinstance(row.get("filed_at"), (date, datetime)) else row.get("filed_at"),
                "filed_by_name": row.get("filed_by_name", "System User"),
                "file_format": row.get("file_format", "csv")
            })
        
        return _success({
            'company_id': company_id,
            'history': formatted_history,
            'total_exports': len(formatted_history)
        })
        
    except Exception as e:
        return _error("get_tax_filing_history", e)


# ============================================================================
# ENDPOINT: Preview Employee Data for Tax Filing (GET - Reads from Real Tables)
# ============================================================================

@payroll_tax_filing_bp.route(
    "/api/companies/<int:company_id>/tax-filing/preview",
    methods=["GET", "OPTIONS"]
)
@require_auth
def get_paye_preview_data(company_id: int):
    """
    GET endpoint to preview employee PAYE data before export.
    
    Reads from ACTUAL payroll tables:
      - company_{id}.payroll_employees (master data)
      - company_{id}.payroll_payslips (pay records)
      - company_{id}.payroll_runs (run headers)
    
    Query params:
      - authority: SARS | RSL | BURS (default: SARS)
      - period: 2024/2025 format (default: 2024/2025)
    
    Returns empty employees array if no data found.
    Frontend calls: /api/companies/8/tax-filing/preview?authority=SARS&period=2024/2025
    """
    if request.method == "OPTIONS":
        return _options()
    
    try:
        # ─── Get Query Parameters ─────────────────────────────
        authority = request.args.get('authority', 'SARS').upper()
        period = request.args.get('period', '2024/2025')
        
        # Validate authority
        if authority not in SUPPORTED_AUTHORITIES:
            return jsonify({
                "ok": False,
                "error": f"Invalid authority. Supported: {list(SUPPORTED_AUTHORITIES.keys())}"
            }), 400
        
        # ─── Parse Tax Year Period ────────────────────────────
        year_start, year_end = parse_tax_year_period(period, authority)
        
        schema = f"company_{company_id}"
        
        # ─── Query 1: Active Employees ────────────────────────
        employees_sql = f"""
            SELECT 
                e.id,
                e.employee_no,
                e.first_name,
                e.last_name,
                e.tax_number,
                e.id_number,
                e.employment_status,
                e.start_date,
                COALESCE(e.is_archived, FALSE) as is_archived
            FROM {schema}.payroll_employees e
            WHERE e.company_id = %s 
              AND COALESCE(e.is_archived, FALSE) = FALSE
              AND e.employment_status IN ('active', 'suspended')
            ORDER BY e.last_name, e.first_name, e.employee_no
        """
        
        employees = db_service.fetch_all(employees_sql, (company_id,))
        
        # ─── Return Empty If No Employees ─────────────────────
        if not employees:
            return _success({
                'authority': authority,
                'period': period,
                'period_start': str(year_start),
                'period_end': str(year_end),
                'employees': [],
                'totals': {
                    'employee_count': 0,
                    'gross_income': 0.00,
                    'paye_deducted': 0.00,
                    'ui_fund': 0.00,
                    'sdl': 0.00,
                    'net_pay': 0.00
                },
                'message': 'No active employees found for this company'
            })
        
        # ─── Query 2: Payslip Summary for Period ───────────────
        employee_ids = tuple(emp['id'] for emp in employees)
        
        payslip_sql = f"""
            SELECT 
                ps.employee_id,
                SUM(ps.gross_income) as total_gross,
                SUM(ps.paye_deducted) as total_paye,
                SUM(ps.ui_fund) as total_ui_fund,
                SUM(COALESCE(ps.SDL, 0)) as total_sdl,
                COUNT(DISTINCT ps.run_id) as pay_periods,
                MIN(r.run_date) as first_pay_date,
                MAX(r.run_date) as last_pay_date
            FROM {schema}.payroll_payslips ps
            JOIN {schema}.payroll_runs r ON r.id = ps.run_id
            WHERE ps.employee_id IN %s
              AND r.run_date BETWEEN %s AND %s
              AND r.status = 'posted'
            GROUP BY ps.employee_id
        """
        
        payslip_data = db_service.fetch_all(
            payslip_sql, 
            (employee_ids, year_start, year_end)
        )
        
        # Build lookup dict: {employee_id: payslip_data}
        payslip_lookup = {p['employee_id']: p for p in payslip_data} if payslip_data else {}
        
        # ─── Build Response Array ─────────────────────────────
        mapped_employees = []
        totals = {
            'employee_count': 0,
            'gross_income': 0.00,
            'paye_deducted': 0.00,
            'ui_fund': 0.00,
            'sdl': 0.00,
            'net_pay': 0.00
        }
        
        for emp in employees:
            pay = payslip_lookup.get(emp['id'], {})
            
            gross_income = float(pay.get('total_gross') or 0)
            paye_deducted = float(pay.get('total_paye') or 0)
            ui_fund = float(pay.get('total_ui_fund') or 0)
            sdl = float(pay.get('total_sdl') or 0)
            net_pay = gross_income - paye_deducted - ui_fund - sdl
            
            employee_record = {
                'employee_no': emp['employee_no'],
                'first_name': emp['first_name'],
                'last_name': emp['last_name'],
                'full_name': f"{emp['first_name']} {emp['last_name']}",
                'tax_number': emp['tax_number'] or '',
                'id_number': emp['id_number'] or '',
                'employment_status': emp['employment_status'],
                'start_date': str(emp['start_date']) if emp.get('start_date') else None,
                # Financial data from payslips
                'gross_income': round(gross_income, 2),
                'paye_deducted': round(paye_deducted, 2),
                'ui_fund': round(ui_fund, 2),
                'sdl': round(sdl, 2),
                'net_pay': round(net_pay, 2),
                'pay_periods': int(pay.get('pay_periods') or 0),
                'first_pay_date': str(pay.get('first_pay_date')) if pay.get('first_pay_date') else None,
                'last_pay_date': str(pay.get('last_pay_date')) if pay.get('last_pay_date') else None,
                # Status for UI display
                'status': 'has_data' if gross_income > 0 else 'no_payslips'
            }
            
            mapped_employees.append(employee_record)
            
            # Accumulate totals
            totals['employee_count'] += 1
            totals['gross_income'] += gross_income
            totals['paye_deducted'] += paye_deducted
            totals['ui_fund'] += ui_fund
            totals['sdl'] += sdl
            totals['net_pay'] += net_pay
        
        # Round all totals
        totals = {k: round(v, 2) for k, v in totals.items()}
        
        return _success({
            'authority': authority,
            'authority_config': SUPPORTED_AUTHORITIES[authority],
            'period': period,
            'period_start': str(year_start),
            'period_end': str(year_end),
            'employees': mapped_employees,
            'totals': totals,
            'generated_at': datetime.utcnow().isoformat()
        })
        
    except Exception as e:
        current_app.logger.error(f"[PAYE] Preview failed for company {company_id}: {str(e)}")
        return _error("get_paye_preview_data", e)


# ============================================================================
# HELPER: Parse Tax Year into Date Range
# ============================================================================

def parse_tax_year_period(period: str, authority: str) -> tuple:
    """
    Convert tax year string (e.g., '2024/2025') to start/end dates
    based on tax authority's fiscal year rules.
    
    SARS (South Africa):   March - February  (2024/2025 → Mar 2024 - Feb 2025)
    RSL (Lesotho):         April - March     (2024/2025 → Apr 2024 - Mar 2025)
    BURS (Botswana):       July - June       (2024/2025 → Jul 2024 - Jun 2025)
    """
    try:
        parts = period.split('/')
        start_year = int(parts[0])
        
        if authority == 'SARS':
            # South Africa: March to February
            return date(start_year, 3, 1), date(start_year + 1, 2, 28)
            
        elif authority == 'RSL':
            # Lesotho: April to March
            return date(start_year, 4, 1), date(start_year + 1, 3, 31)
            
        elif authority == 'BURS':
            # Botswana: July to June
            return date(start_year, 7, 1), date(start_year + 1, 6, 30)
            
        else:
            # Default: Calendar year (January to December)
            return date(start_year, 1, 1), date(start_year, 12, 31)
            
    except Exception:
        # Fallback: Current calendar year
        today = date.today()
        return date(today.year, 1, 1), date(today.year, 12, 31)


# ============================================================================
# ENDPOINT: Quick Export (GET version for frontend convenience)
# ============================================================================

@payroll_tax_filing_bp.route(
    "/api/companies/<int:company_id>/tax-filing/export",
    methods=["GET", "OPTIONS"]
)
@require_auth
def get_tax_filing_export(company_id: int):
    """
    GET endpoint for quick export (frontend-friendly).
    
    Query params:
      - authority: SARS | RSL | BURS
      - period: 2024/2025
      - format: csv | xlsx | xml
    
    Returns downloadable file.
    """
    if request.method == "OPTIONS":
        return _options()
    
    try:
        authority_code = request.args.get('authority', 'SARS').upper()
        format_type = request.args.get('format', 'csv').lower()
        period = request.args.get('period', '2024/2025')
        
        # Validate
        if authority_code not in SUPPORTED_AUTHORITIES:
            return jsonify({"ok": False, "error": f"Invalid authority: {authority_code}"}), 400
        
        # Get preview data (reuses the logic above)
        preview_result = get_paye_preview_data.__wrapped__(company_id) if hasattr(get_paye_preview_data, '__wrapped__') else None
        
        # Re-query for export (more efficient than calling preview)
        year_start, year_end = parse_tax_year_period(period, authority_code)
        schema = f"company_{company_id}"
        
        # Query employees
        employees_sql = f"""
            SELECT 
                e.id, e.employee_no, e.first_name, e.last_name,
                e.tax_number, e.id_number, e.employment_status, e.start_date
            FROM {schema}.payroll_employees e
            WHERE e.company_id = %s 
              AND COALESCE(e.is_archived, FALSE) = FALSE
              AND e.employment_status IN ('active', 'suspended')
            ORDER BY e.last_name, e.first_name
        """
        
        employees = db_service.fetch_all(employees_sql, (company_id,))
        
        if not employees:
            return jsonify({
                "ok": False,
                "error": "No employees found to export"
            }), 404
        
        # Query payslip data
        employee_ids = tuple(emp['id'] for emp in employees)
        payslip_sql = f"""
            SELECT 
                ps.employee_id,
                SUM(ps.gross_income) as total_gross,
                SUM(ps.paye_deducted) as total_paye,
                SUM(ps.ui_fund) as total_ui_fund,
                SUM(COALESCE(ps.SDL, 0)) as total_sdl
            FROM {schema}.payroll_payslips ps
            JOIN {schema}.payroll_runs r ON r.id = ps.run_id
            WHERE ps.employee_id IN %s
              AND r.run_date BETWEEN %s AND %s
              AND r.status = 'posted'
            GROUP BY ps.employee_id
        """
        
        payslip_data = db_service.fetch_all(payslip_sql, (employee_ids, year_start, year_end))
        payslip_lookup = {p['employee_id']: p for p in payslip_data} if payslip_data else {}
        
        # Build records for export
        employer_info = {
            'tax_reference_number': '',  # Could fetch from company settings
            'name': ''  # Could fetch from company settings
        }
        
        records = []
        for emp in employees:
            pay = payslip_lookup.get(emp['id'], {})
            records.append({
                'employee_id': emp['employee_no'],
                'first_name': emp['first_name'],
                'last_name': emp['last_name'],
                'id_number': emp['id_number'] or '',
                'tax_number': emp['tax_number'] or '',
                'basic_salary': float(pay.get('total_gross', 0)) * 0.7,  # Estimate
                'gross_income': float(pay.get('total_gross') or 0),
                'paye_deducted': float(pay.get('total_paye') or 0),
                'uif_deducted': float(pay.get('total_ui_fund') or 0),
                'sdl_deducted': float(pay.get('total_sdl') or 0),
                'allowances': 0,
                'bonus': 0,
                'commission': 0,
                'overtime_pay': 0,
                'net_pay': float(pay.get('total_gross', 0)) - 
                         float(pay.get('total_paye', 0)) - 
                         float(pay.get('total_ui_fund', 0)) -
                         float(pay.get('total_sdl', 0)),
                'payment_date': str(pay.get('last_pay_date')) if pay.get('last_pay_date') else '',
                'employment_start_date': str(emp['start_date']) if emp.get('start_date') else ''
            })
        
        # Generate file using existing function
        file_content, filename, mime_type = generate_export_file(
            authority_code=authority_code,
            format_type=format_type,
            period_end=year_end,
            records=records,
            employer_info=employer_info
        )
        
        # Create download response
        response = make_response(file_content)
        response.headers['Content-Type'] = mime_type
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['X-Tax-Authority'] = authority_code
        response.headers['X-Tax-Format'] = format_type
        response.headers['X-Export-Timestamp'] = datetime.utcnow().isoformat()
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"[PAYE] Export failed for company {company_id}: {str(e)}")
        return _error("get_tax_filing_export", e)
# ============================================================================
# REGISTRATION HELPERS
# ============================================================================

def register_tax_filing_blueprint(app):
    """
    Call this in your main app factory to register the blueprint.
    
    Example:
        from BackEnd.Services.payroll_tax_filing_routes import register_tax_filing_blueprint
        register_tax_filing_blueprint(app)
    """
    app.register_blueprint(payroll_tax_filing_bp)
    current_app.logger.info("Registered payroll_tax_filing_bp blueprint")

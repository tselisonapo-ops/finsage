from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Tuple

from flask import Response, request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.platypus import KeepTogether
from xml.sax.saxutils import escape
from BackEnd.Services.vat_pack_pdf_builder import _add_brand_header

THIN = Side(style="thin", color="D9E2F3")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBTOTAL_FILL = PatternFill("solid", fgColor="EEF4FB")
TITLE_FILL = PatternFill("solid", fgColor="BFD7EA")


def _pdf_amount(v):
    try:
        if v is None or v == "":
            return ""
        n = float(v)
        if abs(n) < 0.000001:
            n = 0.0
        return f"({abs(n):,.2f})" if n < 0 else f"{n:,.2f}"
    except Exception:
        return "" if v is None else str(v)


def _note_para(text, style):
    text = escape(str(text or "")).replace("\n", "<br/>")
    return Paragraph(text, style)

def _company_from_meta(meta: Dict[str, Any]) -> Dict[str, Any]:
    company = dict(meta.get("company") or {})
    company.setdefault("company_name", meta.get("company_name") or meta.get("name"))
    company.setdefault("currency", meta.get("currency"))

    for k in (
        "logo_path",
        "logo_file",
        "logo_local_path",
        "logo",
        "company_logo",
        "attachment_path",
        "logo_attachment_path",
        "logo_url",
        "branding_logo_url",
        "company_reg_no",
        "reg_no",
        "vat_no",
        "vat_number",
        "company_email",
        "email",
        "company_phone",
        "phone",
        "address",
        "physical_address",
        "postal_address",
    ):
        if meta.get(k) and not company.get(k):
            company[k] = meta.get(k)

    return company

def _financial_table(rows, amount_keys=None, amount_labels=None, page_width_mm=260):
    """
    FS-style table with dynamic widths.
    Supports wide disclosure notes such as PPE and IFRS 16.
    """
    amount_keys = amount_keys or ["amount"]
    amount_labels = amount_labels or {k: k.replace("_", " ").title() for k in amount_keys}

    data = []
    row_types = []

    # Header row
    header_style = ParagraphStyle(
        "tbl_header",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        alignment=TA_RIGHT,
    )

    label_header_style = ParagraphStyle(
        "tbl_header_label",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        alignment=TA_LEFT,
    )

    data.append(
        [Paragraph("Description", label_header_style)]
        + [Paragraph(escape(str(amount_labels.get(k, k))), header_style) for k in amount_keys]
    )
    row_types.append("header")

    for r in rows or []:
        label = r.get("label") or r.get("name") or ""
        values = r.get("values") or {}
        rt = r.get("row_type") or "normal"

        row = [Paragraph(escape(label), ParagraphStyle(
            "tbl_label",
            fontName="Helvetica-Bold" if rt in ("header", "subtotal", "total") else "Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
        ))]

        for k in amount_keys:
            row.append(Paragraph(_pdf_amount(values.get(k)), ParagraphStyle(
                "tbl_amt",
                fontName="Helvetica-Bold" if rt in ("subtotal", "total") else "Helvetica",
                fontSize=8,
                leading=10,
                alignment=TA_RIGHT,
            )))

        data.append(row)
        row_types.append(rt)

    if len(data) <= 1:
        return None

    if len(amount_keys) == 1:
        label_width = 115 * mm
        amount_width = 42 * mm
    else:
        label_width = 55 * mm
        available = page_width_mm * mm - label_width
        amount_width = max(20 * mm, available / max(len(amount_keys), 1))
        
    table = Table(
        data,
        colWidths=[label_width, *([amount_width] * len(amount_keys))],
        hAlign="LEFT",
        repeatRows=1,
    )

    style = TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.black),
    ])

    for idx, rt in enumerate(row_types):
        if rt == "header":
            style.add("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold")

        if rt == "subtotal":
            style.add("LINEABOVE", (1, idx), (-1, idx), 0.4, colors.black)

        if rt == "total":
            style.add("LINEABOVE", (1, idx), (-1, idx), 0.7, colors.black)
            style.add("LINEBELOW", (1, idx), (-1, idx), 0.7, colors.black)

    table.setStyle(style)
    return table

def _clean_number(v: Any) -> Any:
    try:
        if v is None or v == "":
            return ""
        return float(v)
    except Exception:
        return v


def _statement_title(meta: Dict[str, Any]) -> str:
    stmt = str((meta or {}).get("statement") or "").strip().lower()

    mapping = {
        "bs": "Statement of Financial Position",
        "balance_sheet": "Statement of Financial Position",
        "pnl": "Statement of Profit or Loss",
        "income_statement": "Statement of Profit or Loss",
        "cf": "Statement of Cash Flows",
        "cashflow": "Statement of Cash Flows",
        "socie": "Statement of Changes in Equity",
    }
    return mapping.get(stmt, (meta or {}).get("report_name") or "Financial Statement")


def _payload_columns(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    cols = payload.get("columns") or []
    if not cols:
        return [{"key": "amount", "label": "Amount"}]

    # hide comparison / extra columns that have no data anywhere
    used = set()

    def scan_values(values):
        if isinstance(values, dict):
            for k, v in values.items():
                if _has_value(v):
                    used.add(k)

    for r in payload.get("rows") or []:
        scan_values(r.get("values"))

    for sec in payload.get("sections") or []:
        for ln in sec.get("lines") or []:
            scan_values(ln.get("values"))
        scan_values(sec.get("totals"))

    def scan_bs_side(side):
        for sec in (side or {}).values():
            if isinstance(sec, dict):
                for ln in sec.get("lines") or []:
                    scan_values(ln.get("values"))
                scan_values(sec.get("totals"))
                scan_values(sec.get("values"))

    scan_bs_side(payload.get("assets"))
    scan_bs_side(payload.get("equity_and_liabilities"))

    for key in ("net_result", "net_change", "opening_balance", "closing_balance"):
        block = payload.get(key)
        if isinstance(block, dict):
            scan_values(block.get("values"))

    cash_pos = payload.get("cash_position") or {}
    for block in cash_pos.values():
        if isinstance(block, dict):
            scan_values(block.get("values"))

    reconciliation = payload.get("reconciliation") or {}
    for block in reconciliation.values():
        if isinstance(block, dict):
            scan_values(block.get("values"))

    # ✅ Keep the template/requested structure.
    # Do NOT collapse export columns just because values are zero.
    return cols

def _row_type(row: Dict[str, Any]) -> str:
    meta = row.get("meta") or {}
    return str(meta.get("row_type") or row.get("row_type") or "normal").strip().lower()


def _append_row(
    out_rows: List[Dict[str, Any]],
    label: str,
    values: Dict[str, Any],
    row_type: str = "normal",
):
    out_rows.append({
        "label": label,
        "values": values or {},
        "row_type": row_type,
    })

def _has_value(v: Any) -> bool:
    if v is None or v == "":
        return False
    try:
        return abs(float(v)) > 0.000001
    except Exception:
        return True

def _pretty_date(v):
    if not v:
        return ""
    try:
        from datetime import datetime, date

        if isinstance(v, date):
            d = v
        else:
            d = datetime.strptime(str(v)[:10], "%Y-%m-%d").date()

        return d.strftime("%d %B %Y")
    except Exception:
        return str(v)

def _year_from_period(period):
    if not isinstance(period, dict):
        return None
    to = period.get("to")
    if not to:
        return None
    return str(to)[:4]


def _ias_export_columns(meta, cols):
    meta = meta or {}
    cols = cols or []

    cur_year = _year_from_period(meta.get("period"))

    comparison_periods = meta.get("comparison_periods") or []
    prior_period = meta.get("prior_period")

    year_by_key = {}

    if cur_year:
        year_by_key["cur"] = cur_year

    if isinstance(prior_period, dict):
        y = _year_from_period(prior_period)
        if y:
            year_by_key["pri"] = y

    for idx, p in enumerate(comparison_periods or [], start=1):
        key = p.get("key") or ("pri" if idx == 1 else f"p{idx}")
        y = _year_from_period({"to": p.get("to")})
        if y:
            year_by_key[key] = y

    out = []

    for c in cols:
        key = c.get("key")

        # ✅ Do not show variance column in IAS-style exports
        if key in ("delta", "variance", "movement"):
            continue

        label = year_by_key.get(key) or c.get("label") or key

        # fallback labels
        if str(label).lower() in ("current", "amount"):
            label = cur_year or label

        out.append({**c, "label": label})

    return out

def _ias_period_label(meta):
    stmt = str((meta or {}).get("statement") or "").lower()
    period = (meta or {}).get("period") or {}

    period_from = period.get("from")
    period_to = period.get("to")

    if stmt in ("bs", "balance_sheet"):
        return f"As at {_pretty_date(period_to)}" if period_to else ""

    if period_to:
        return f"For the year ended {_pretty_date(period_to)}"

    if period_from:
        return f"For the period from {_pretty_date(period_from)}"

    return ""

def _flatten_payload(payload: Dict[str, Any]) -> Tuple[List[str], List[Dict[str, Any]]]:
    cols = _payload_columns(payload)
    col_labels = [c.get("label") or c.get("key") for c in cols]

    out_rows: List[Dict[str, Any]] = []

    # 1) Balance Sheet shape
    if payload.get("assets") and payload.get("equity_and_liabilities"):

        def push_section(label, section):
            if not section:
                return

            _append_row(out_rows, label, {}, "header")

            for line in section.get("lines") or []:
                _append_row(
                    out_rows,
                    line.get("name") or line.get("label") or "",
                    line.get("values") or {},
                    _row_type(line),
                )

            totals = section.get("totals")
            if totals:
                if isinstance(totals, dict):
                    vals = totals.get("values") or {
                        k: v for k, v in totals.items()
                        if k not in ("label", "name", "row_type", "meta")
                    }
                else:
                    vals = totals
                _append_row(out_rows, f"Total {label}", vals or {}, "total")

        assets = payload.get("assets") or {}
        push_section("Current assets", assets.get("current_assets"))
        push_section("Non-current assets", assets.get("non_current_assets"))

        if assets.get("totals"):
            _append_row(
                out_rows,
                assets["totals"].get("label") or "Total assets",
                assets["totals"].get("values") or {},
                "total",
            )

        eq = payload.get("equity_and_liabilities") or {}
        push_section("Equity", eq.get("equity"))
        push_section("Non-current liabilities", eq.get("non_current_liabilities"))
        push_section("Current liabilities", eq.get("current_liabilities"))

        if eq.get("totals"):
            _append_row(
                out_rows,
                eq["totals"].get("label") or "Total equity and liabilities",
                eq["totals"].get("values") or {},
                "total",
            )

        if payload.get("balance_check"):
            bc = payload["balance_check"]
            _append_row(
                out_rows,
                bc.get("label") or "Balance check",
                bc.get("values") or {},
                "subtotal",
            )

        return ["Line Item", *col_labels], out_rows

    # 2) SOCIE / row-based shape
    if payload.get("rows"):
        for r in payload.get("rows") or []:
            label = r.get("label") or r.get("name") or r.get("key") or ""
            rt = "total" if str(r.get("key") or "").lower() in {"closing_balance", "total"} else _row_type(r)
            _append_row(out_rows, label, r.get("values") or {}, rt)

        return ["Line Item", *col_labels], out_rows

    # 3A) P&L expanded / management shape (payload["blocks"])
    if payload.get("blocks"):
        for block in payload.get("blocks") or []:
            block_label = block.get("label") or block.get("key") or ""

            # Header
            if block_label:
                _append_row(out_rows, block_label, {}, "header")

            # Lines
            for line in block.get("lines") or []:
                rt = _row_type(line)
                if line.get("is_subtotal"):
                    rt = "subtotal"

                _append_row(
                    out_rows,
                    line.get("name") or line.get("label") or line.get("code") or "",
                    line.get("values") or {},
                    rt,
                )

            # Totals
            if block.get("totals"):
                _append_row(
                    out_rows,
                    f"Total {block_label}",
                    block.get("totals") or {},
                    "subtotal",
                )

            # Direct value blocks (e.g. gross profit)
            if block.get("values") and not block.get("lines") and not block.get("totals"):
                _append_row(
                    out_rows,
                    block_label,
                    block.get("values") or {},
                    "subtotal",
                )

        # Final net result
        # Final net result (support all builder variants)
        def _extract_net_result(payload):
            for key in ("net_result", "net_income", "net_profit", "profit_for_period", "net"):
                block = payload.get(key)
                if isinstance(block, dict):
                    values = block.get("values") or {}
                    if values:
                        return {
                            "label": block.get("label") or "Net Profit",
                            "values": values,
                        }

                    amt = block.get("amount")
                    if amt is not None:
                        cols = _payload_columns(payload)
                        k = cols[0].get("key") if cols else "cur"
                        return {
                            "label": block.get("label") or "Net Profit",
                            "values": {k: amt},
                        }

                elif block is not None:
                    cols = _payload_columns(payload)
                    k = cols[0].get("key") if cols else "cur"
                    return {
                        "label": "Net Profit",
                        "values": {k: block},
                    }

            return None


        nr = _extract_net_result(payload)
        if nr:
            _append_row(out_rows, nr["label"], nr["values"], "total")

        return ["Line Item", *col_labels], out_rows

    # 3) P&L / Cash Flow sections shape
    for sec in payload.get("sections") or []:
        sec_label = sec.get("label") or sec.get("key") or ""

        if sec_label:
            _append_row(out_rows, sec_label, {}, "header")

        for line in sec.get("lines") or []:
            label = line.get("name") or line.get("label") or line.get("code") or ""
            rt = _row_type(line)
            if line.get("is_subtotal"):
                rt = "subtotal"
            _append_row(out_rows, label, line.get("values") or {}, rt)

            # Optional: include breakdown details in Excel/PDF
            detail = line.get("detail") or {}
            for col_key, detail_rows in detail.items():
                if not isinstance(detail_rows, list):
                    continue
                for d in detail_rows:
                    _append_row(
                        out_rows,
                        f"   - {d.get('account_name') or d.get('name') or 'Detail'}",
                        {col_key: d.get("amount")},
                        "normal",
                    )

        totals = sec.get("totals")
        if totals:
            _append_row(out_rows, f"Total {sec_label}", totals or {}, "subtotal")

        # Some P&L blocks use values directly, not lines/totals
        if sec.get("values") and not sec.get("lines") and not sec.get("totals"):
            _append_row(out_rows, sec_label, sec.get("values") or {}, "subtotal")

    # 4) Statement-level totals / extras
    for key in ("net_result", "net_change", "opening_balance", "closing_balance"):
        block = payload.get(key)
        if isinstance(block, dict):
            _append_row(
                out_rows,
                block.get("label") or key.replace("_", " ").title(),
                block.get("values") or {},
                "total" if key in {"net_result", "net_change"} else "subtotal",
            )

    cash_pos = payload.get("cash_position") or {}
    for k in ("opening", "closing", "delta_from_tb", "reconciliation_gap"):
        block = cash_pos.get(k)
        if isinstance(block, dict):
            _append_row(
                out_rows,
                block.get("label") or k.replace("_", " ").title(),
                block.get("values") or {},
                "subtotal",
            )

    reconciliation = payload.get("reconciliation") or {}
    for k in ("delta_from_tb", "gap"):
        block = reconciliation.get(k)
        if isinstance(block, dict):
            _append_row(
                out_rows,
                block.get("label") or k.replace("_", " ").title(),
                block.get("values") or {},
                "subtotal",
            )

    return ["Line Item", *col_labels], out_rows

def _xlsx_apply_row_style(ws, row_idx: int, row_type: str, max_col: int):
    if row_type == "header":
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
    elif row_type in ("subtotal",):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = Font(bold=True)
            cell.fill = SUBTOTAL_FILL
    elif row_type in ("total",):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = Font(bold=True)
            cell.fill = TITLE_FILL

    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

def _write_statement_sheet(wb, sheet_name, payload, *, company_name="", currency=""):
    meta = payload.get("meta") or {}
    title = _statement_title(meta)

    cols = _ias_export_columns(meta, _payload_columns(payload))
    if len(cols) == 1:
        cols[0]["label"] = "Amount"

    payload = {**payload, "columns": cols}
    headers, flat_rows = _flatten_payload(payload)
    col_keys = [c.get("key") for c in cols]

    ws = wb.create_sheet(title=sheet_name[:31])

    ws["A1"] = str(company_name or meta.get("company_name") or "").upper()
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = str(title or "").upper()
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A3"] = _ias_period_label(meta)
    ws["A3"].alignment = Alignment(horizontal="center")

    ws["A4"] = f"(All amounts presented in {currency or meta.get('currency')})" if (currency or meta.get("currency")) else ""
    ws["A4"].alignment = Alignment(horizontal="center")

    max_col = max(1, len(headers))
    for row in range(1, 5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)

    start_row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    current_row = start_row + 1
    for item in flat_rows:
        ws.cell(row=current_row, column=1, value=item["label"])
        vals = item.get("values") or {}

        for i, key in enumerate(col_keys, start=2):
            val = _clean_number(vals.get(key))
            cell = ws.cell(row=current_row, column=i, value=val)

            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

        _xlsx_apply_row_style(ws, current_row, item.get("row_type") or "normal", len(headers))
        current_row += 1

    ws.column_dimensions["A"].width = 42
    for idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    return ws

def build_pnl_export_summary_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Converts detailed P&L payload into IAS 1-style summary payload for exports.
    Keeps the same columns/comparisons, but removes account-level detail.
    """

    if not isinstance(payload, dict):
        return payload

    meta = dict(payload.get("meta") or {})
    if str(meta.get("statement") or "").lower() not in ("pnl", "income_statement", "profit_loss"):
        return payload

    sections = payload.get("sections") or []
    if not isinstance(sections, list):
        return payload

    by_key = {
        str(s.get("key") or "").lower(): s
        for s in sections
        if isinstance(s, dict)
    }

    def vals(*keys):
        for key in keys:
            sec = by_key.get(key)
            if sec and isinstance(sec.get("totals"), dict):
                return dict(sec.get("totals") or {})
        return {}

    def row(key, label, values, row_type="normal"):
        return {
            "key": key,
            "label": label,
            "values": values or {},
            "row_type": row_type,
        }

    rows = []

    rows.append(row("revenue", "Revenue", vals("revenue"), "normal"))

    cogs_vals = vals("cogs", "cost_of_sales", "cost_of_revenue")
    if cogs_vals:
        rows.append(row("cost_of_sales", "Cost of sales", cogs_vals, "normal"))

    gross_vals = vals("gross_profit")
    if gross_vals:
        rows.append(row("gross_profit", "Gross profit", gross_vals, "subtotal"))

    exp_vals = vals("operating_expenses", "expenses")
    if exp_vals:
        rows.append(row("operating_expenses", "Operating expenses", exp_vals, "normal"))

    op_vals = vals("operating_profit", "operating_income")
    if op_vals:
        rows.append(row("operating_profit", "Operating profit", op_vals, "subtotal"))

    other_vals = vals("other", "other_income", "other_income_expense")
    if other_vals:
        rows.append(row("other_income_expense", "Other income/(expense)", other_vals, "normal"))

    pbt_vals = vals("profit_before_tax")
    if pbt_vals:
        rows.append(row("profit_before_tax", "Profit before tax", pbt_vals, "subtotal"))

    tax_vals = vals("tax", "income_tax")
    if tax_vals:
        rows.append(row("income_tax", "Income tax expense", tax_vals, "normal"))

    net = payload.get("net_result") or {}
    net_vals = dict(net.get("values") or {})
    if net_vals:
        rows.append(row("profit_for_the_year", net.get("label") or "Profit for the year", net_vals, "total"))

    out = dict(payload)
    out["rows"] = rows
    out["sections"] = []
    out.setdefault("meta", {})
    out["meta"] = {
        **meta,
        "statement_title": "Statement of Profit or Loss",
        "export_layout": "ias1_summary_pnl",
    }

    return out

def export_statement_xlsx(payload: Dict[str, Any], filename: str = "statement.xlsx") -> Response:
    meta = payload.get("meta") or {}
    title = _statement_title(meta)
    company_name = meta.get("company_name") or ""
    currency = meta.get("currency") or ""
    period = meta.get("period") or {}
    period_from = period.get("from")
    period_to = period.get("to")

    cols = _payload_columns(payload)

    # ✅ Apply single-column rename BEFORE flatten
    cols = _ias_export_columns(meta, _payload_columns(payload))

    if len(cols) == 1:
        cols[0]["label"] = "Amount"

    payload = {**payload, "columns": cols}

    headers, flat_rows = _flatten_payload(payload)
    col_keys = [c.get("key") for c in cols]

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    is_pnl_export = str((meta.get("statement") or "")).lower() in (
        "pnl",
        "income_statement",
        "profit_loss",
    )

    if is_pnl_export:
        summary_payload = build_pnl_export_summary_payload(payload)
        _write_statement_sheet(
            wb,
            "P&L Summary",
            summary_payload,
            company_name=company_name,
            currency=currency,
        )

        detail_payload = dict(payload)
        detail_payload.setdefault("meta", {})
        detail_payload["meta"] = {
            **(detail_payload.get("meta") or {}),
            "statement_title": "Detailed Profit or Loss",
        }

        _write_statement_sheet(
            wb,
            "Detailed P&L",
            detail_payload,
            company_name=company_name,
            currency=currency,
        )
    else:
        _write_statement_sheet(
            wb,
            "Statement",
            payload,
            company_name=company_name,
            currency=currency,
        )

    # SOCIE comparison statements - separate sheets
    comparison_statements = payload.get("comparison_statements") or []

    for idx, cmp_stmt in enumerate(comparison_statements, start=1):
        sheet_name = "Comparative" if idx == 1 else f"Comparative {idx}"
        ws_cmp = wb.create_sheet(title=sheet_name[:31])

        cmp_meta = cmp_stmt.get("meta") or {}
        cmp_title = _statement_title(cmp_meta)
        cmp_company = cmp_meta.get("company_name") or company_name
        cmp_currency = cmp_meta.get("currency") or currency

        cmp_cols = _payload_columns(cmp_stmt)
        if len(cmp_cols) == 1:
            cmp_cols[0]["label"] = "Amount"

        cmp_stmt = {**cmp_stmt, "columns": cmp_cols}
        cmp_headers, cmp_flat_rows = _flatten_payload(cmp_stmt)
        cmp_col_keys = [c.get("key") for c in cmp_cols]

        cmp_max_col = max(1, len(cmp_headers))

        ws_cmp["A1"] = str(cmp_company or "").upper()
        ws_cmp["A1"].font = Font(bold=True, size=13)
        ws_cmp["A1"].alignment = Alignment(horizontal="center")

        ws_cmp["A2"] = str(cmp_title or "").upper()
        ws_cmp["A2"].font = Font(bold=True, size=14)
        ws_cmp["A2"].alignment = Alignment(horizontal="center")

        ws_cmp["A3"] = _ias_period_label(cmp_meta)
        ws_cmp["A3"].alignment = Alignment(horizontal="center")

        ws_cmp["A4"] = f"(All amounts presented in {cmp_currency})" if cmp_currency else ""
        ws_cmp["A4"].alignment = Alignment(horizontal="center")

        for row in range(1, 5):
            ws_cmp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cmp_max_col)

        start_row_cmp = 6

        for col_idx, header in enumerate(cmp_headers, start=1):
            cell = ws_cmp.cell(row=start_row_cmp, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

        current_row_cmp = start_row_cmp + 1

        for item in cmp_flat_rows:
            ws_cmp.cell(row=current_row_cmp, column=1, value=item["label"])
            vals = item.get("values") or {}

            for i, key in enumerate(cmp_col_keys, start=2):
                val = _clean_number(vals.get(key))
                cell = ws_cmp.cell(row=current_row_cmp, column=i, value=val)

                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

            _xlsx_apply_row_style(
                ws_cmp,
                current_row_cmp,
                item.get("row_type") or "normal",
                len(cmp_headers),
            )

            current_row_cmp += 1

        ws_cmp.column_dimensions["A"].width = 42
        for col_idx in range(2, len(cmp_headers) + 1):
            ws_cmp.column_dimensions[get_column_letter(col_idx)].width = 18

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def export_statement_pdf(payload: Dict[str, Any], filename: str = "statement.pdf") -> Response:
    original_payload = payload
    meta = payload.get("meta") or {}

    if str(meta.get("statement") or "").lower() in ("pnl", "income_statement", "profit_loss"):
        payload = build_pnl_export_summary_payload(payload)
        meta = payload.get("meta") or {}
    title = _statement_title(meta)
    company_name = meta.get("company_name") or ""
    currency = meta.get("currency") or ""

    cols = _ias_export_columns(meta, _payload_columns(payload))

    if len(cols) == 1:
        cols[0] = {**cols[0], "label": "Amount"}
        
    payload = {**payload, "columns": cols}
    _, flat_rows = _flatten_payload(payload)
    col_keys = [c.get("key") for c in cols]

    wide_table = len(cols) > 6
    page_size = landscape(A4) if wide_table else A4
    page_width_mm = 260 if wide_table else 174

    amount_labels = {
        c.get("key"): c.get("label") or c.get("key")
        for c in cols
    }

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=12 * mm if wide_table else 18 * mm,
        rightMargin=12 * mm if wide_table else 18 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "fs_title",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "fs_meta",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        alignment=1,
    )

    story = []

    company = _company_from_meta(meta)

    if company.get("logo_path") or company.get("logo_file") or company.get("logo_url") or company.get("company_logo"):
        _add_brand_header(story, doc, title, company)
    else:
        if company_name:
            company_style = ParagraphStyle(
                "company_name",
                parent=styles["BodyText"],
                fontName="Helvetica-Bold",
                fontSize=12,
                alignment=1,  # centre
                spaceAfter=4,
            )

            story.append(
                Paragraph(
                    escape(str(company_name).upper()),
                    company_style
                )
            )

        story.append(
            Paragraph(
                escape(title.upper()),
                title_style
            )
        )
    period_label = _ias_period_label(meta)

    if period_label:
        story.append(Paragraph(escape(period_label), meta_style))
        
    if currency:
        story.append(Paragraph(
            escape(f"(All amounts presented in {currency})"),
            meta_style
        ))

    story.append(Spacer(1, 10))

    tbl = _financial_table(
        flat_rows,
        col_keys,
        amount_labels=amount_labels,
        page_width_mm=page_width_mm,
    )

    if tbl:
        story.append(tbl)

    is_pnl_export = str((payload.get("meta") or {}).get("statement") or "").lower() in (
        "pnl",
        "income_statement",
        "profit_loss",
    )

    include_detail = is_pnl_export

    if include_detail and str((payload.get("meta") or {}).get("statement") or "").lower() in ("pnl", "income_statement", "profit_loss"):
        story.append(PageBreak())
        detail_payload = dict(original_payload)
        detail_payload.setdefault("meta", {})
        detail_payload["meta"] = {
            **(detail_payload.get("meta") or {}),
            "statement_title": "Detailed Profit or Loss",
        }

        detail_cols = _ias_export_columns(detail_payload.get("meta") or {}, _payload_columns(detail_payload))
        detail_payload = {**detail_payload, "columns": detail_cols}
        _, detail_rows = _flatten_payload(detail_payload)
        detail_col_keys = [c.get("key") for c in detail_cols]
        detail_labels = {c.get("key"): c.get("label") or c.get("key") for c in detail_cols}

        story.append(Paragraph("DETAILED PROFIT OR LOSS", title_style))
        detail_tbl = _financial_table(
            detail_rows,
            detail_col_keys,
            amount_labels=detail_labels,
            page_width_mm=page_width_mm,
        )
        if detail_tbl:
            story.append(detail_tbl)

    # ✅ SOCIE comparison statements: render each comparison as its own table
    comparison_statements = payload.get("comparison_statements") or []

    if comparison_statements:
        for idx, cmp_stmt in enumerate(comparison_statements, start=1):
            cmp_meta = cmp_stmt.get("meta") or {}
            cmp_period = cmp_meta.get("period") or {}

            heading = "Comparative period" if idx == 1 else f"Comparative period {idx}"

            story.append(Spacer(1, 14))
            story.append(Paragraph(f"<b>{escape(heading)}</b>", meta_style))

            if cmp_period.get("from") or cmp_period.get("to"):
                story.append(Paragraph(
                    escape(_ias_period_label(cmp_meta)),
                    meta_style,
                ))

            cmp_cols = _payload_columns(cmp_stmt)
            if len(cmp_cols) == 1:
                cmp_cols[0] = {**cmp_cols[0], "label": "Amount"}

            cmp_payload = {**cmp_stmt, "columns": cmp_cols}
            _, cmp_flat_rows = _flatten_payload(cmp_payload)
            cmp_col_keys = [c.get("key") for c in cmp_cols]

            cmp_amount_labels = {
                c.get("key"): c.get("label") or c.get("key")
                for c in cmp_cols
            }

            cmp_tbl = _financial_table(
                cmp_flat_rows,
                cmp_col_keys,
                amount_labels=cmp_amount_labels,
                page_width_mm=page_width_mm,
            )

            if cmp_tbl:
                story.append(cmp_tbl)

    doc.build(story)

    pdf_bytes = buffer.getvalue()
    buffer.close()

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def export_fs_notes_pdf(notes: List[Dict[str, Any]], filename: str = "financial_statement_notes.pdf") -> Response:
    """
    notes shape:
    [
      {
        "title": "Leases",
        "text": "...policy wording...",
        "sections": [
          {"title": "Right-of-use assets", "rows": [...]},
          {"title": "Lease liabilities", "rows": [...]},
        ]
      }
    ]
    """
    all_section_keys = []

    for note in notes or []:
        for sec in note.get("sections") or []:
            rows = sec.get("rows") or []
            keys = sec.get("amount_keys") or []

            if not keys and sec.get("columns"):
                keys = [c.get("key") for c in sec.get("columns") or [] if c.get("key")]

            if not keys:
                for r in rows:
                    for k in (r.get("values") or {}).keys():
                        if k not in keys:
                            keys.append(k)

            all_section_keys.extend(keys)

    wide_notes = len(set(all_section_keys)) > 4
    page_size = landscape(A4) if wide_notes else A4

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=12 * mm if wide_notes else 18 * mm,
        rightMargin=12 * mm if wide_notes else 18 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()

    note_title = ParagraphStyle(
        "note_title",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=6,
    )

    section_title = ParagraphStyle(
        "section_title",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        spaceBefore=8,
        spaceAfter=4,
    )

    body = ParagraphStyle(
        "note_body",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        spaceAfter=7,
    )

    story = []

    for note in notes or []:
        title = note.get("title") or "Note"
        text = note.get("text") or ""

        block = [
            Paragraph(escape(title), note_title),
            _note_para(text, body),
        ]

        for sec in note.get("sections") or []:
            rows = sec.get("rows") or []
            if not rows:
                continue

            block.append(Paragraph(escape(sec.get("title") or ""), section_title))

            amount_keys = sec.get("amount_keys")

            if not amount_keys:
                # Prefer explicit columns if section provides them
                if sec.get("columns"):
                    amount_keys = [c.get("key") for c in sec.get("columns") or [] if c.get("key")]
                else:
                    # Infer keys from row values
                    keys = []
                    for r in rows:
                        vals = r.get("values") or {}
                        for k in vals.keys():
                            if k not in keys:
                                keys.append(k)

                    amount_keys = keys or ["amount"]

            amount_labels = sec.get("amount_labels") or {}

            if sec.get("columns"):
                amount_labels.update({
                    c.get("key"): c.get("label") or c.get("key")
                    for c in sec.get("columns") or []
                    if c.get("key")
                })

            wide_table = len(amount_keys) > 4
            page_width_mm = 260 if wide_table else 174

            tbl = _financial_table(
                rows,
                amount_keys,
                amount_labels=amount_labels,
                page_width_mm=page_width_mm,
            )
            if tbl:
                block.append(tbl)
                block.append(Spacer(1, 6))

        story.extend(block)
        story.append(Spacer(1, 10))

    doc.build(story)

    pdf_bytes = buffer.getvalue()
    buffer.close()

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
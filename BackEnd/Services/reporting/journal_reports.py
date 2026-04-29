
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from flask import Response

def export_xlsx(payload_or_rows, columns=None, filename: str = "report.xlsx", sheet_name: str = "Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    if columns is None:
        payload = payload_or_rows or {}
        data = payload.get("data") or {}

        rows = (
            payload.get("rows")
            or data.get("rows")
            or payload.get("items")
            or data.get("items")
            or []
        )

        columns = (
            payload.get("columns")
            or data.get("columns")
            or []
        )
    else:
        rows = payload_or_rows or []
        columns = columns or []

    rows = [dict(r) for r in rows]

    # Keep these for sorting, but do not export them
    hidden_keys = {"journal_id", "account_code", "source_id"}

    columns = [
        c for c in columns
        if c.get("key") not in hidden_keys
    ]

    keys = [c.get("key") for c in columns]
    headers = [c.get("label") or c.get("key") for c in columns]

    def money_num(v):
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    # Sort journal lines so debits appear first, then credits,
    # while keeping journals grouped together.
    rows.sort(
        key=lambda r: (
            str(r.get("date") or ""),
            int(r.get("journal_id") or 0),
            1 if money_num(r.get("credit")) > 0 else 0,
            str(r.get("account_name") or ""),
        )
    )

    ws.append(headers)

    last_journal_id = None

    for row in rows:
        row = dict(row)
        journal_id = row.get("journal_id")

        # Add spacing + journal header when journal changes
        if journal_id and journal_id != last_journal_id:
            if last_journal_id is not None:
                ws.append([])

            header_text = f"Journal {journal_id}"
            ref = row.get("ref") or ""
            date = row.get("date") or ""
            desc = row.get("journal_description") or row.get("description") or ""

            ws.append([f"{header_text} | {date} | {ref} | {desc}"] + [""] * (len(keys) - 1))

            group_row = ws.max_row
            ws.merge_cells(
                start_row=group_row,
                start_column=1,
                end_row=group_row,
                end_column=max(len(keys), 1),
            )

            cell = ws.cell(row=group_row, column=1)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEAF6")
            cell.alignment = Alignment(vertical="center")

            last_journal_id = journal_id

        clean_row = []

        for k in keys:
            v = row.get(k, "")

            if k in {"debit", "credit", "debit_total", "credit_total"}:
                try:
                    if float(v or 0) == 0:
                        v = ""
                except Exception:
                    pass

            clean_row.append(v)

        ws.append(clean_row)

    header_fill = PatternFill("solid", fgColor="EAF2F8") 
    debit_fill = PatternFill("solid", fgColor="F8FBFF")
    credit_fill = PatternFill("solid", fgColor="FFF8E7")
    thin = Side(style="thin", color="D9E2EC")
    row_border = Border(bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = row_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    money_keys = {"debit", "credit", "balance", "debit_total", "credit_total"}

    account_name_col = keys.index("account_name") + 1 if "account_name" in keys else None
    debit_col = keys.index("debit") + 1 if "debit" in keys else None
    credit_col = keys.index("credit") + 1 if "credit" in keys else None

    for row_idx in range(2, ws.max_row + 1):

        # Skip group header rows and blank separator rows
        if ws.cell(row=row_idx, column=1).value and all(
            ws.cell(row=row_idx, column=c).value in (None, "")
            for c in range(2, ws.max_column + 1)
        ):
            continue

        if all(
            ws.cell(row=row_idx, column=c).value in (None, "")
            for c in range(1, ws.max_column + 1)
        ):
            continue

        debit_num = 0
        credit_num = 0

        if debit_col:
            debit_num = money_num(ws.cell(row=row_idx, column=debit_col).value)

        if credit_col:
            credit_num = money_num(ws.cell(row=row_idx, column=credit_col).value)

        row_fill = credit_fill if credit_num > 0 else debit_fill if debit_num > 0 else None

        for col_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = row_border
            wrap = key in {"memo", "description"}

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=wrap
            )

            if row_fill:
                cell.fill = row_fill

        if account_name_col and credit_num > 0:
            ws.cell(row=row_idx, column=account_name_col).alignment = Alignment(
                indent=2,
                vertical="top"
            )

    for col_idx, key in enumerate(keys, start=1):
        max_len = len(str(headers[col_idx - 1] or ""))

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            max_len = max(max_len, len(str(value or "")))

            if key in money_keys and row_idx > 1 and value not in ("", None):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="top")

        # smarter autofit per column type
        if key == "memo":
            width = min(max(max_len + 4, 25), 80)   # allow wide memo
        elif key in {"description", "account_name"}:
            width = min(max(max_len + 4, 20), 60)
        elif key in {"date", "ref", "source"}:
            width = min(max(max_len + 2, 12), 25)
        else:
            width = min(max(max_len + 3, 12), 40)

        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 🔥 AUTO ROW HEIGHT FOR WRAPPED TEXT
    for row_idx in range(2, ws.max_row + 1):
        max_lines = 1

        for col_idx, key in enumerate(keys, start=1):
            if key in {"memo", "description"}:
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    text = str(val)
                    lines = text.count("\n") + (len(text) // 60) + 1
                    max_lines = max(max_lines, lines)

        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # existing line
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )

def build_journal_register(db, company_id, date_from=None, date_to=None, q=None):
    schema = db.company_schema(company_id)

    where = ["j.company_id = %s"]
    params = [int(company_id)]

    if date_from:
        where.append("j.date >= %s")
        params.append(date_from)

    if date_to:
        where.append("j.date <= %s")
        params.append(date_to)

    if q:
        where.append("""
            (
                COALESCE(j.ref,'') ILIKE %s
                OR COALESCE(j.description,'') ILIKE %s
                OR COALESCE(l.memo,'') ILIKE %s
                OR COALESCE(l.account,'') ILIKE %s
                OR COALESCE(c.name,'') ILIKE %s
            )
        """)
        like = f"%{q}%"
        params.extend([like, like, like, like, like])

    sql = f"""
    SELECT
        j.id AS journal_id,
        j.date,
        COALESCE(j.ref, '') AS ref,
        COALESCE(j.description, '') AS description,
        COALESCE(j.source, '') AS source,
        l.id AS line_id,
        l.account AS account_code,
        COALESCE(c.name, l.account) AS account_name,
        COALESCE(l.memo, '') AS memo,
        COALESCE(l.debit, 0) AS debit,
        COALESCE(l.credit, 0) AS credit
    FROM {schema}.journal j
    JOIN {schema}.ledger l
      ON l.journal_id = j.id
     AND l.company_id = j.company_id
    LEFT JOIN {schema}.coa c
      ON c.company_id = l.company_id
     AND c.code = l.account
    WHERE {" AND ".join(where)}
    ORDER BY j.date DESC, j.id DESC, l.id ASC
    """

    raw_rows = db.fetch_all(sql, tuple(params)) or []

    rows = []
    for r in raw_rows:
        rows.append({
            "date": str(r.get("date")) if r.get("date") else "",
            "ref": r.get("ref") or "",
            "journal_id": r.get("journal_id"),
            "source": r.get("source") or "",
            "description": r.get("description") or "",
            "account_code": r.get("account_code") or "",
            "account_name": r.get("account_name") or "",
            "memo": r.get("memo") or "",
            "debit": float(r.get("debit") or 0),
            "credit": float(r.get("credit") or 0),
        })

    columns = [
        {"key": "date", "label": "Date"},
        {"key": "ref", "label": "Ref"},
        {"key": "journal_id", "label": "Journal ID"},
        {"key": "source", "label": "Source"},
        {"key": "description", "label": "Description"},
        {"key": "account_code", "label": "Account"},
        {"key": "account_name", "label": "Account Name"},
        {"key": "memo", "label": "Memo"},
        {"key": "debit", "label": "Debit"},
        {"key": "credit", "label": "Credit"},
    ]

    return rows, columns
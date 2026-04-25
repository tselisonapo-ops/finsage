# utils/vat_utils.py
from collections import defaultdict
import os
from BackEnd.Services.auth_middleware import _corsify, require_auth
from BackEnd.Services.db_service import db_service
from BackEnd.Services.company_context import get_company_context
from datetime import datetime, date, timedelta
from typing import Optional, Tuple, Set
from collections import defaultdict
from datetime import date as date_cls
from BackEnd.Services.emailer import send_mail
from BackEnd.Services.utils.view_token import create_report_export_token, verify_report_export_token
from urllib.parse import urlencode
import csv
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from flask import (
    Blueprint,
    jsonify,
    request,
    g,
    make_response,
    current_app 
)
from BackEnd.Services.vat_pack_pdf_builder import (
    generate_vat_return_pdf,
    generate_vat_supporting_pdf,
)
from flask import Blueprint

vat_utils_bp = Blueprint("companies_vat", __name__)

def _safe_cell(v, default="-"):
    if v is None or str(v).strip() == "":
        return default
    return v


def _money(v):
    try:
        return round(float(v or 0), 2)
    except Exception:
        return 0.0


def _style_xlsx_sheet(ws):
    header_fill = PatternFill("solid", fgColor="0F2A3D")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="D9F0F2")
    thin = Side(style="thin", color="D6DAD8")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.protection = Protection(locked=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in ws.iter_rows():
        first = str(row[0].value or "").upper()
        if first in {
            "FILLED VAT RETURN",
            "VAT RETURN VALUES",
            "DECLARATION",
            "VAT SUPPORTING SCHEDULE",
            "OFFICIAL FILING TOTALS",
            "DETAIL RECONCILIATION",
            "SETTLEMENT JOURNAL",
            "JOURNAL BALANCE CHECK",
        }:
            for cell in row:
                cell.fill = section_fill
                cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))

        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    ws.freeze_panes = "A2"
    ws.protection.sheet = True
    ws.protection.password = "FinSage"


def _build_vat_summary_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Filled VAT Return"

    for row in rows:
        ws.append([_excel_safe(v) for v in row])

    _style_xlsx_sheet(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _get_company_vat_pack_profile(company_id: int) -> dict:
    company = db_service.fetch_one(
        """
        SELECT
          id,
          name,
          company_reg_no,
          vat,
          tin,
          company_email,
          company_phone,
          physical_address,
          postal_address,
          currency,
          logo_url
        FROM public.companies
        WHERE id = %s
        LIMIT 1;
        """,
        (int(company_id),),
    ) or {}

    logo_url = (company.get("logo_url") or "").strip()

    if logo_url and logo_url.startswith("/uploads/company_logos/"):
        filename = os.path.basename(logo_url)
        company["logo_path"] = os.path.join(
            current_app.root_path,
            "uploads",
            "company_logos",
            filename,
        )

    company["company_name"] = company.get("name") or f"Company {company_id}"
    return company

def _build_vat_detail_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Supporting Schedule"

    for row in rows:
        ws.append([_excel_safe(v) for v in row])

    _style_xlsx_sheet(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _days_in_month(year, month):
    # month = 1..12
    if month == 12:
        return (date(year + 1, 1, 1) - date(year, 12, 1)).days
    return (date(year, month + 1, 1) - date(year, month, 1)).days

def _make_vat_periods_for_year(year: int, cfg: dict):
    freq = (cfg.get("frequency") or "bi_monthly").lower()
    anchor_month = cfg.get("anchor_month") or 1
    anchor_month = max(1, min(12, anchor_month))

    anchor_month = cfg.get("anchor_month") or 1
    try:
        anchor_month = int(anchor_month)
    except Exception:
        anchor_month = 1
    anchor_month = max(1, min(12, anchor_month))

    if freq == "monthly":
        step = 1
    elif freq == "quarterly":
        step = 3
    elif freq in ("semi_annual", "semi-annual", "half_year"):
        step = 6
    elif freq == "annual":
        step = 12
    else:  # bi-monthly default
        step = 2

    periods = []
    # zero-based offset for easier calc
    anchor0 = anchor_month - 1
    max_loops = (12 // step) + 2

    for k in range(max_loops):
        # start
        s_month_idx = anchor0 + k * step  # 0-based
        s_year = year + (s_month_idx // 12)
        s_month = (s_month_idx % 12) + 1
        start_date = date(s_year, s_month, 1)

        # end
        e_month_idx = s_month_idx + step - 1
        e_year = year + (e_month_idx // 12)
        e_month = (e_month_idx % 12) + 1
        end_date = date(e_year, e_month, _days_in_month(e_year, e_month))

        # Keep if it touches the given year
        if start_date.year == year or end_date.year == year:
            label_start = start_date.strftime("%b")
            label_end = end_date.strftime("%b")

            if start_date.year == end_date.year:
                if step == 1:
                    label = f"{label_start} {end_date.year}"
                else:
                    label = f"{label_start}–{label_end} {end_date.year}"
            else:
                label = f"{label_start} {start_date.year}–{label_end} {end_date.year}"

            periods.append({
                "label": label,
                "start_date": start_date,
                "end_date": end_date,
                "due_date": None,  # filled later
            })

    return periods

def _get_vat_lines(company_id: int, start_date, end_date):
    schema = db_service.company_schema(company_id)

    input_codes, output_codes = _get_vat_accounts(company_id)
    vat_codes = list(set(input_codes) | set(output_codes))

    if not vat_codes:
        return []

    placeholders = ",".join(["%s"] * len(vat_codes))

    sql = f"""
    WITH vat_lines AS (
        SELECT
            l.id,
            l.journal_id,
            l.date,
            l.ref,
            l.account AS vat_account_code,
            vat_acc.name AS vat_account_name,
            l.debit,
            l.credit
        FROM {schema}.ledger l
        LEFT JOIN {schema}.coa vat_acc
          ON vat_acc.company_id = l.company_id
         AND vat_acc.code = l.account
        WHERE l.company_id = %s
          AND l.date >= %s
          AND l.date <= %s
          AND l.account IN ({placeholders})
    ),
    source_lines AS (
        SELECT DISTINCT ON (vl.id)
            vl.id AS vat_line_id,
            src.account AS source_account_code,
            src_acc.name AS source_account_name,
            src.debit AS source_debit,
            src.credit AS source_credit
        FROM vat_lines vl
        LEFT JOIN {schema}.ledger src
          ON src.company_id = %s
         AND src.journal_id = vl.journal_id
         AND src.id <> vl.id
         AND src.account NOT IN ({placeholders})
        LEFT JOIN {schema}.coa src_acc
          ON src_acc.company_id = src.company_id
         AND src_acc.code = src.account
        WHERE src.id IS NOT NULL

          -- exclude control/settlement accounts
          AND COALESCE(src.account, '') NOT IN (
            'BS_CA_1000',
            'BS_CA_1010',
            'BS_CA_9002',
            'BS_CL_2200',
            'BS_CL_9001'
          )

          -- exclude by account names too
            AND COALESCE(src_acc.name, '') NOT ILIKE '%%cash%%'
            AND COALESCE(src_acc.name, '') NOT ILIKE '%%bank%%'
            AND COALESCE(src_acc.name, '') NOT ILIKE '%%receivable%%'
            AND COALESCE(src_acc.name, '') NOT ILIKE '%%debtor%%'
            AND COALESCE(src_acc.name, '') NOT ILIKE '%%payable%%'
            AND COALESCE(src_acc.name, '') NOT ILIKE '%%creditor%%'
            AND COALESCE(src_acc.name, '') NOT ILIKE '%%vat%%'

        ORDER BY
            vl.id,

            -- prefer P&L / real source accounts over BS controls
            CASE
            WHEN COALESCE(src_acc.category, '') ILIKE '%%revenue%%' THEN 1
            WHEN COALESCE(src_acc.category, '') ILIKE '%%income%%' THEN 1
            WHEN COALESCE(src_acc.category, '') ILIKE '%%expense%%' THEN 1
            WHEN COALESCE(src_acc.category, '') ILIKE '%%cost%%' THEN 1
            WHEN COALESCE(src_acc.category, '') ILIKE '%%asset%%' THEN 2
              ELSE 9
            END,

            ABS(COALESCE(src.debit, 0) - COALESCE(src.credit, 0)) DESC
    )
    SELECT
        vl.date,
        vl.ref,
        vl.vat_account_code,
        vl.vat_account_name,
        sl.source_account_code,
        sl.source_account_name,
        vl.debit,
        vl.credit
    FROM vat_lines vl
    LEFT JOIN source_lines sl
      ON sl.vat_line_id = vl.id
    ORDER BY vl.date DESC, vl.id DESC
    LIMIT 500;
    """

    params = (
        int(company_id),
        start_date,
        end_date,
        *vat_codes,
        int(company_id),
        *vat_codes,
    )

    with db_service._conn_cursor() as (_conn, cur):
        cur.execute(sql, params)
        rows = cur.fetchall() or []

    lines = []

    for r in rows:
        vat_code = str(r.get("vat_account_code") or "")
        side = "input" if vat_code in input_codes else (
            "output" if vat_code in output_codes else None
        )

        debit = float(r.get("debit") or 0)
        credit = float(r.get("credit") or 0)

        source_code = str(r.get("source_account_code") or "")
        source_name = r.get("source_account_name") or ""

        vat_name = r.get("vat_account_name") or ""

        lines.append({
            "date": r.get("date").isoformat() if r.get("date") else None,
            "ref": r.get("ref") or "",

            "source_account_code": source_code,
            "source_account_name": source_name,

            "vat_account_code": vat_code,
            "vat_account_name": vat_name,

            "account_code": source_code or vat_code,
            "account_name": source_name or vat_name,

            "debit": debit,
            "credit": credit,
            "vat_amount": abs(debit - credit),
            "vat_side": side,
        })

    return lines

def _excel_safe(value):
    """
    Convert values into Excel-safe format
    - remove timezone from datetime
    - stringify complex objects
    """
    from datetime import datetime

    if isinstance(value, datetime):
        # remove timezone if present
        return value.replace(tzinfo=None)

    # handle ISO string with timezone
    if isinstance(value, str) and "+" in value:
        try:
            from datetime import datetime
            dt = datetime.fromisoformat(value)
            return dt.replace(tzinfo=None)
        except Exception:
            return value

    return value

def compute_current_vat_period(today: date, cfg: dict):
    """Return the VAT period that contains today."""
    if not cfg:
        return None

    periods = (
        _make_vat_periods_for_year(today.year - 1, cfg)
        + _make_vat_periods_for_year(today.year, cfg)
        + _make_vat_periods_for_year(today.year + 1, cfg)
    )

    filing_lag_days = cfg.get("filing_lag_days")
    if not isinstance(filing_lag_days, int):
        filing_lag_days = 25

    for p in periods:
        if p["start_date"] <= today <= p["end_date"]:
            p["due_date"] = p["end_date"] + timedelta(days=filing_lag_days)
            return p

    return None

def compute_next_vat_period(today: date, cfg: dict):
    """Return the period with the soonest dueDate >= today, or the
    last one if all are overdue."""
    if not cfg:
        return None

    filing_lag_days = cfg.get("filing_lag_days")
    if not isinstance(filing_lag_days, int):
        filing_lag_days = 25

    year = today.year
    periods = (
        _make_vat_periods_for_year(year - 1, cfg)
        + _make_vat_periods_for_year(year, cfg)
        + _make_vat_periods_for_year(year + 1, cfg)
    )
    if not periods:
        return None

    for p in periods:
        p["due_date"] = p["end_date"] + timedelta(days=filing_lag_days)

    upcoming = [p for p in periods if p["due_date"] >= today]
    if upcoming:
        return sorted(upcoming, key=lambda p: p["due_date"])[0]

    # no upcoming, pick most recent overdue
    return sorted(periods, key=lambda p: p["due_date"], reverse=True)[0]

def assign_period_label(tx_date: date, cfg: dict):
    """Find the VAT period label for a given transaction date."""
    year = tx_date.year
    periods = (
        _make_vat_periods_for_year(year - 1, cfg)
        + _make_vat_periods_for_year(year, cfg)
        + _make_vat_periods_for_year(year + 1, cfg)
    )
    for p in periods:
        if p["start_date"] <= tx_date <= p["end_date"]:
          return p["label"]
    return tx_date.strftime("%b %Y")  # fallback monthly label


def _parse_date(value: str, default: Optional[date] = None) -> Optional[date]:
    if not value:
        return default
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except Exception:
        return default


def _get_vat_accounts(company_id: int) -> Tuple[Set[str], Set[str]]:
    """
    Resolve VAT Input/Output account codes from per-company COA table: {schema}.coa

    Your COA columns available:
      code, name, category, subcategory, template_code, code_numeric, etc.
    """
    schema = db_service.company_schema(company_id)

    sql = f"""
      SELECT
        code,
        name,
        category,
        subcategory,
        template_code,
        code_numeric
      FROM {schema}.coa
      WHERE company_id = %s
    """

    input_codes: Set[str] = set()
    output_codes: Set[str] = set()

    with db_service._conn_cursor() as (_conn, cur):
        cur.execute(sql, (int(company_id),))
        rows = cur.fetchall() or []

    for acc in rows:
        code = str(acc.get("code") or "").strip()
        if not code:
            continue

        name = (acc.get("name") or "").lower()
        cat  = (acc.get("category") or "").lower()
        sub  = (acc.get("subcategory") or "").lower()

        tcode = str(acc.get("template_code") or "").strip()
        num = acc.get("code_numeric")
        num_str = str(num) if num is not None else ""

        # ✅ Heuristics by label
        if ("vat input" in name) or ("vat input" in cat) or ("vat input" in sub):
            input_codes.add(code)

        if ("vat output" in name) or ("vat output" in cat) or ("vat output" in sub):
            output_codes.add(code)

        # ✅ Template/numeric fallbacks
        if tcode == "1410" or num_str == "1410" or code.endswith("_1410"):
            input_codes.add(code)

        if tcode == "2310" or num_str == "2310" or code.endswith("_2310"):
            output_codes.add(code)

    # ✅ Last-resort defaults (keeps system working even if COA naming differs)
    if not input_codes:
        input_codes.add("BS_CA_1410")
    if not output_codes:
        output_codes.add("BS_CL_2310")

    return input_codes, output_codes

def _serialise_vat_filing(row: dict):
    if not row:
        return None

    def iso(v):
        if v is None:
            return None
        try:
            return v.isoformat()
        except Exception:
            return str(v)

    return {
        "id": row.get("id"),
        "company_id": row.get("company_id"),
        "period_start": iso(row.get("period_start")),
        "period_end": iso(row.get("period_end")),
        "period_label": row.get("period_label"),
        "due_date": iso(row.get("due_date")),
        "input_total": float(row.get("input_total") or 0),
        "output_total": float(row.get("output_total") or 0),
        "net_vat": float(row.get("net_vat") or 0),
        "status": row.get("status"),
        "reference": row.get("reference"),
        "notes": row.get("notes"),
        "prepared_at": iso(row.get("prepared_at")),
        "prepared_by_user_id": row.get("prepared_by_user_id"),
        "submitted_at": iso(row.get("submitted_at")),
        "submitted_by_user_id": row.get("submitted_by_user_id"),
        "source": row.get("source"),
        "source_id": row.get("source_id"),
        "created_at": iso(row.get("created_at")),
        "updated_at": iso(row.get("updated_at")),
    }

def _build_vat_settlement_preview(input_total: float, output_total: float):
    input_total = round(float(input_total or 0), 2)
    output_total = round(float(output_total or 0), 2)
    net_vat = round(output_total - input_total, 2)

    lines = []

    if output_total > 0:
        lines.append({
            "account": "BS_CL_2310",
            "name": "VAT Output",
            "debit": output_total,
            "credit": 0,
        })

    if input_total > 0:
        lines.append({
            "account": "BS_CA_1410",
            "name": "VAT Input",
            "debit": 0,
            "credit": input_total,
        })

    if net_vat > 0:
        lines.append({
            "account": "BS_CL_2320",
            "name": "VAT Payable",
            "debit": 0,
            "credit": net_vat,
        })
        settlement_type = "payable"

    elif net_vat < 0:
        lines.append({
            "account": "BS_CA_1420",
            "name": "VAT Receivable / Refund Due",
            "debit": abs(net_vat),
            "credit": 0,
        })
        settlement_type = "refund"

    else:
        settlement_type = "nil"

    return {
        "input_total": input_total,
        "output_total": output_total,
        "net_vat": net_vat,
        "settlement_type": settlement_type,
        "journal_lines": lines,
    }

@vat_utils_bp.route("/api/companies/<int:company_id>/vat/periods", methods=["GET", "OPTIONS"])
@require_auth
def vat_periods(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", {}) or {}
    if int(user.get("company_id") or 0) != int(company_id):
        return jsonify({"error": "Not authorised"}), 403

    cfg = db_service.get_vat_settings(company_id) or {}
    today = date.today()

    filing_lag_days = cfg.get("filing_lag_days")
    if not isinstance(filing_lag_days, int):
        filing_lag_days = 25

    periods = (
        _make_vat_periods_for_year(today.year - 1, cfg)
        + _make_vat_periods_for_year(today.year, cfg)
        + _make_vat_periods_for_year(today.year + 1, cfg)
    )

    out = []
    for p in periods:
        due_date = p["end_date"] + timedelta(days=filing_lag_days)
        out.append({
            "label": p["label"],
            "start_date": p["start_date"].isoformat(),
            "end_date": p["end_date"].isoformat(),
            "due_date": due_date.isoformat(),
        })

    out.sort(key=lambda x: x["start_date"], reverse=True)

    current = compute_current_vat_period(today, cfg)

    return jsonify({
        "periods": out,
        "current": {
            "label": current["label"],
            "start_date": current["start_date"].isoformat(),
            "end_date": current["end_date"].isoformat(),
            "due_date": current["due_date"].isoformat() if current.get("due_date") else None,
        } if current else None,
    }), 200

@vat_utils_bp.route("/api/companies/<int:company_id>/vat_summary", methods=["GET", "OPTIONS"])
@require_auth
def vat_summary(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", {}) or {}
    if int(user.get("company_id") or 0) != int(company_id):
        return jsonify({"error": "Not authorised"}), 403

    from_str = (request.args.get("from") or "").strip()
    to_str   = (request.args.get("to") or "").strip()

    today = date_cls.today()
    start_date = _parse_date(from_str, None)
    end_date   = _parse_date(to_str, today)

    if not start_date or not end_date:
        return jsonify({"error": "from and to are required"}), 400

    schema = db_service.company_schema(company_id)
    cfg = db_service.get_vat_settings(company_id) or {}

    input_codes, output_codes = _get_vat_accounts(company_id)
    vat_codes = list(set(input_codes) | set(output_codes))
    if not vat_codes:
        return jsonify({
            "input_total": 0.0,
            "output_total": 0.0,
            "net_vat": 0.0,
            "periods": [],
            "filing": None,
        }), 200

    placeholders = ",".join(["%s"] * len(vat_codes))

    sql = f"""
      SELECT
        account,
        date,
        debit,
        credit
      FROM {schema}.ledger
      WHERE company_id = %s
        AND date >= %s
        AND date <= %s
        AND account IN ({placeholders})
    """

    with db_service._conn_cursor() as (_conn, cur):
        cur.execute(sql, (int(company_id), start_date, end_date, *vat_codes))
        rows = cur.fetchall() or []

    input_total = 0.0
    output_total = 0.0
    buckets = defaultdict(lambda: {"input": 0.0, "output": 0.0})

    for r in rows:
        code = str(r.get("account") or "")
        dr = float(r.get("debit") or 0)
        cr = float(r.get("credit") or 0)
        bal = dr - cr

        label = assign_period_label(r["date"], cfg) if cfg else "Period"

        if code in input_codes:
            input_total += bal
            buckets[label]["input"] += bal
        elif code in output_codes:
            amt = -bal
            output_total += amt
            buckets[label]["output"] += amt

    periods = [
        {"label": k, "input": v["input"], "output": v["output"]}
        for k, v in buckets.items()
    ]

    current_period = compute_current_vat_period(today, cfg) if cfg else None

    return jsonify({
        "input_total": input_total,
        "output_total": output_total,
        "net_vat": output_total - input_total,
        "periods": periods,
        "filing": {
            "label": current_period["label"] if current_period else None,
            "start_date": current_period["start_date"].isoformat() if current_period else None,
            "end_date": current_period["end_date"].isoformat() if current_period else None,
            "due_date": current_period["due_date"].isoformat() if current_period and current_period.get("due_date") else None,
            "status": "open",
            "frequency": cfg.get("frequency"),
            "anchor_month": cfg.get("anchor_month"),
        } if current_period else None,
    }), 200


@vat_utils_bp.route("/api/companies/<int:company_id>/vat/lines", methods=["GET", "OPTIONS"])
@require_auth
def vat_lines(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", {}) or {}
    if int(user.get("company_id") or 0) != int(company_id):
        return jsonify({"error": "Not authorised"}), 403

    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()

    today = date.today()
    start_date = _parse_date(from_str, None)
    end_date = _parse_date(to_str, today)

    if not start_date or not end_date:
        return jsonify({"error": "from and to are required"}), 400

    lines = _get_vat_lines(company_id, start_date, end_date)

    return jsonify({"lines": lines}), 200
@vat_utils_bp.route("/api/companies/<int:company_id>/vat/filings", methods=["GET", "OPTIONS"])
@require_auth
def vat_filings(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", {}) or {}
    if int(user.get("company_id") or 0) != int(company_id):
        return jsonify({"error": "Not authorised"}), 403

    try:
        db_service.ensure_company_schema(company_id)
        db_service.ensure_company_vat_filings(company_id)
    except Exception:
        pass

    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()

    start_date = _parse_date(from_str, None)
    end_date = _parse_date(to_str, None)

    if start_date and end_date:
        filing = db_service.get_vat_filing(company_id, start_date, end_date)
        return jsonify({
            "filing": _serialise_vat_filing(filing)
        }), 200

    rows = db_service.list_vat_filings(company_id, limit=100)
    return jsonify({
        "filings": [_serialise_vat_filing(r) for r in rows]
    }), 200

@vat_utils_bp.route("/api/companies/<int:company_id>/vat/filings/prepare", methods=["POST", "OPTIONS"])
@require_auth
def vat_prepare_filing(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    current_user = getattr(g, "current_user", {}) or {}
    if int(current_user.get("company_id") or 0) != int(company_id):
        return jsonify({"ok": False, "error": "Not authorised for this company"}), 403

    payload = request.get_json(force=True) or {}
    from_str = (payload.get("from") or "").strip()
    to_str = (payload.get("to") or "").strip()
    notes = (payload.get("notes") or "").strip() or None
    preview_only = bool(payload.get("preview_only"))

    start_date = _parse_date(from_str, None)
    end_date = _parse_date(to_str, None)

    if not start_date or not end_date:
        return jsonify({"ok": False, "error": "from and to are required"}), 400

    try:
        db_service.ensure_company_schema(company_id)
        db_service.ensure_company_vat_filings(company_id)
    except Exception:
        pass

    cfg = db_service.get_vat_settings(company_id) or {}
    filing_lag_days = cfg.get("filing_lag_days")
    if not isinstance(filing_lag_days, int):
        filing_lag_days = 25

    due_date = end_date + timedelta(days=filing_lag_days)

    period_label = None
    periods = (
        _make_vat_periods_for_year(start_date.year - 1, cfg)
        + _make_vat_periods_for_year(start_date.year, cfg)
        + _make_vat_periods_for_year(start_date.year + 1, cfg)
    )

    for p in periods:
        if p["start_date"] == start_date and p["end_date"] == end_date:
            period_label = p["label"]
            break

    if not period_label:
        period_label = f"{start_date.isoformat()} to {end_date.isoformat()}"

    schema = db_service.company_schema(company_id)
    input_codes, output_codes = _get_vat_accounts(company_id)
    vat_codes = list(set(input_codes) | set(output_codes))

    input_total = 0.0
    output_total = 0.0

    if vat_codes:
        placeholders = ",".join(["%s"] * len(vat_codes))
        sql = f"""
          SELECT account, debit, credit
          FROM {schema}.ledger
          WHERE company_id = %s
            AND date >= %s
            AND date <= %s
            AND account IN ({placeholders})
        """

        with db_service._conn_cursor() as (_conn, cur):
            cur.execute(sql, (int(company_id), start_date, end_date, *vat_codes))
            rows = cur.fetchall() or []

        for r in rows:
            code = str(r.get("account") or "")
            dr = float(r.get("debit") or 0)
            cr = float(r.get("credit") or 0)
            bal = dr - cr

            if code in input_codes:
                input_total += bal
            elif code in output_codes:
                output_total += -bal

    input_total = round(float(input_total or 0), 2)
    output_total = round(float(output_total or 0), 2)
    net_vat = round(output_total - input_total, 2)

    preview = _build_vat_settlement_preview(input_total, output_total)

    if preview_only:
        return jsonify({
            "ok": True,
            "period": {
                "label": period_label,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "due_date": due_date.isoformat(),
            },
            "preview": preview,
        }), 200

    existing = db_service.get_vat_filing(company_id, start_date, end_date)

    if existing and existing.get("settlement_journal_id"):
        return jsonify({
            "ok": True,
            "already_posted": True,
            "filing": _serialise_vat_filing(existing),
            "preview": preview,
        }), 200

    filing_id = db_service.upsert_vat_filing_prepared(
        company_id,
        period_start=start_date,
        period_end=end_date,
        period_label=period_label,
        due_date=due_date,
        input_total=input_total,
        output_total=output_total,
        net_vat=net_vat,
        notes=notes,
        prepared_by_user_id=int(current_user.get("id") or 0) or None,
        source="api",
    )

    journal_id = None

    if preview.get("settlement_type") != "nil":
        journal_id = db_service.post_simple_journal(
            company_id=company_id,
            journal_date=end_date,
            ref=f"VAT-{start_date.isoformat()}-{end_date.isoformat()}",
            description=f"VAT settlement for {period_label}",
            source="vat_filing",
            source_id=filing_id,
            lines=preview["journal_lines"],
        )

        db_service.mark_vat_filing_settlement_posted(
            company_id,
            filing_id,
            journal_id=journal_id,
        )

    saved = db_service.get_vat_filing(company_id, start_date, end_date)

    try:
        db_service.audit_log(
            company_id=company_id,
            actor_user_id=int(current_user.get("id") or 0),
            module="tax",
            action="prepare_vat_filing",
            severity="info",
            entity_type="vat_filing",
            entity_id=str(filing_id),
            entity_ref=(period_label or f"vat_filing_{filing_id}"),
            amount=float(net_vat or 0.0),
            currency=(get_company_context(db_service, company_id) or {}).get("currency"),
            before_json=payload if isinstance(payload, dict) else {},
            after_json=_serialise_vat_filing(saved) or {},
            message="VAT return prepared and settlement journal posted",
            source="api",
        )
    except Exception:
        pass

    return jsonify({
        "ok": True,
        "filing_id": filing_id,
        "journal_id": journal_id,
        "preview": preview,
        "filing": _serialise_vat_filing(saved),
    }), 200

@vat_utils_bp.route("/api/companies/<int:company_id>/vat/filings/export", methods=["GET", "OPTIONS"])
@require_auth
def vat_filing_export(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", {}) or {}
    if int(user.get("company_id") or 0) != int(company_id):
        return jsonify({"error": "Not authorised"}), 403

    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()

    start_date = _parse_date(from_str, None)
    end_date = _parse_date(to_str, None)

    if not start_date or not end_date:
        return jsonify({"error": "from and to are required"}), 400

    filing = db_service.get_vat_filing(company_id, start_date, end_date)
    if not filing:
        return jsonify({"error": "Prepare VAT return first"}), 400

    lines = _get_vat_lines(company_id, start_date, end_date)

    import csv, io

    output = io.StringIO()
    writer = csv.writer(output)

    ctx = get_company_context(db_service, company_id) or {}
    company_name = ctx.get("company_name") or f"Company {company_id}"

    # HEADER
    writer.writerow(["VAT FILING REPORT"])
    writer.writerow(["Company", company_name])
    writer.writerow(["Period", f"{start_date} to {end_date}"])
    writer.writerow(["Due date", filing.get("due_date")])
    writer.writerow(["Status", filing.get("status")])
    writer.writerow(["Prepared at", filing.get("prepared_at")])
    writer.writerow([])

    # SUMMARY
    writer.writerow(["SUMMARY"])
    writer.writerow(["Output VAT", filing.get("output_total")])
    writer.writerow(["Input VAT", filing.get("input_total")])
    writer.writerow(["Net VAT", filing.get("net_vat")])
    writer.writerow([])

    # DETAIL
    writer.writerow([
        "Date",
        "Reference",
        "Source Account",
        "VAT Side",
        "VAT Account",
        "VAT Amount",
    ])

    for l in lines:
        writer.writerow([
            l.get("date"),
            l.get("ref"),
            f"{l.get('source_account_code')} {l.get('source_account_name')}",
            l.get("vat_side"),
            f"{l.get('vat_account_code')} {l.get('vat_account_name')}",
            l.get("vat_amount"),
        ])

    csv_data = output.getvalue()
    output.close()

    filename = f"vat_filing_{company_id}_{start_date}_{end_date}.csv"

    resp = make_response(csv_data)
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return _corsify(resp)

@vat_utils_bp.route("/api/companies/<int:company_id>/vat/filings/export-pack", methods=["GET", "OPTIONS"])
def vat_filing_export_pack(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    token = (request.args.get("t") or "").strip()

    if token:
        data = verify_report_export_token(token)

        if not data:
            return jsonify({"error": "Invalid or expired token"}), 401

        if int(data.get("company_id") or 0) != int(company_id):
            return jsonify({"error": "Token company mismatch"}), 403

        if data.get("report_key") != "vat_pack":
            return jsonify({"error": "Invalid report key"}), 403
    else:
        user = getattr(g, "current_user", {}) or {}
        if int(user.get("company_id") or 0) != int(company_id):
            return jsonify({"error": "Not authorised"}), 403

    start_date = _parse_date((request.args.get("from") or "").strip(), None)
    end_date = _parse_date((request.args.get("to") or "").strip(), None)

    if not start_date or not end_date:
        return jsonify({"error": "from and to are required"}), 400

    filing = db_service.get_vat_filing(company_id, start_date, end_date)
    if not filing:
        return jsonify({"error": "Prepare VAT return first"}), 400

    lines = _get_vat_lines(company_id, start_date, end_date)

    ctx = _get_company_vat_pack_profile(company_id)

    company_name = ctx.get("company_name") or ctx.get("name") or f"Company {company_id}"
    currency = ctx.get("currency") or ""
    company_reg_no = ctx.get("company_reg_no") or "-"
    vat_number = ctx.get("vat") or "-"
    tin = ctx.get("tin") or "-"
    company_email = ctx.get("company_email") or "-"

    input_total = _money(filing.get("input_total"))
    output_total = _money(filing.get("output_total"))
    net_vat = _money(filing.get("net_vat"))

    if net_vat > 0:
        net_label = "VAT Payable"
    elif net_vat < 0:
        net_label = "VAT Refundable"
    else:
        net_label = "Nil VAT"

    summary_rows = [
        ["FILLED VAT RETURN"],
        [],
        ["Company", company_name],
        ["Company Registration No", company_reg_no],
        ["VAT Number", vat_number],
        ["TIN", tin],
        ["Company Email", company_email],
        ["Currency", currency],
        ["VAT Period", f"{start_date} to {end_date}"],
        ["Due Date", filing.get("due_date")],
        ["Status", filing.get("status")],
        ["Prepared At", filing.get("prepared_at")],
        ["Submitted At", filing.get("submitted_at")],
        ["Submission Reference", filing.get("reference") or ""],
        [],
        ["VAT RETURN VALUES"],
        ["Output VAT", output_total],
        ["Input VAT", input_total],
        [net_label, abs(net_vat)],
        [],
        ["DECLARATION"],
        ["Declaration", "This VAT return was prepared from ledger VAT records in FinSage."],
        ["Prepared By", "FinSage"],
        ["Notice", "This document is prepared by FinSage, no stamp required"],
    ]

    detail_rows = [
        ["VAT SUPPORTING SCHEDULE"],
        [],
        ["Company", company_name],
        ["VAT Period", f"{start_date} to {end_date}"],
        ["Status", filing.get("status")],
        ["Prepared At", filing.get("prepared_at")],
        [],
        [
            "Date",
            "Reference",
            "Source Account Code",
            "Source Account Name",
            "VAT Side",
            "VAT Account Code",
            "VAT Account Name",
            "Debit",
            "Credit",
            "VAT Amount",
        ],
    ]

    extracted_input = 0.0
    extracted_output = 0.0

    for l in lines:
        side = str(l.get("vat_side") or "").lower()
        vat_amount = _money(l.get("vat_amount"))
        debit = _money(l.get("debit"))
        credit = _money(l.get("credit"))

        if side == "input":
            extracted_input += vat_amount
        elif side == "output":
            extracted_output += vat_amount

        detail_rows.append([
            l.get("date"),
            l.get("ref"),
            l.get("source_account_code"),
            l.get("source_account_name"),
            side,
            l.get("vat_account_code"),
            l.get("vat_account_name"),
            debit,
            credit,
            vat_amount,
        ])

    extracted_net = round(extracted_output - extracted_input, 2)

    detail_rows += [
        [],
        ["OFFICIAL FILING TOTALS"],
        ["Total Output VAT", output_total],
        ["Total Input VAT", input_total],
        ["Net VAT", net_vat],
        [],
        ["DETAIL RECONCILIATION"],
        ["Extracted Detail Output VAT", extracted_output],
        ["Extracted Detail Input VAT", extracted_input],
        ["Extracted Detail Net VAT", extracted_net],
        ["Detail vs Filing Difference", round(extracted_net - net_vat, 2)],
        [],
    ]

    settlement_preview = _build_vat_settlement_preview(input_total, output_total)
    journal_lines = settlement_preview.get("journal_lines") or []

    detail_rows += [
        ["SETTLEMENT JOURNAL"],
        ["Account Code", "Account Name", "Debit", "Credit"],
    ]

    debit_total = 0.0
    credit_total = 0.0

    for jl in journal_lines:
        debit = _money(jl.get("debit"))
        credit = _money(jl.get("credit"))

        debit_total += debit
        credit_total += credit

        detail_rows.append([
            jl.get("account"),
            jl.get("name"),
            debit,
            credit,
        ])

    detail_rows += [
        [],
        ["JOURNAL BALANCE CHECK"],
        ["Debit Total", debit_total],
        ["Credit Total", credit_total],
        ["Difference", round(debit_total - credit_total, 2)],
        [],
        ["This document is prepared by FinSage, no stamp required"],
    ]

    # CSV bytes
    summary_csv_io = io.StringIO()
    csv.writer(summary_csv_io).writerows(summary_rows)
    summary_csv = summary_csv_io.getvalue()

    detail_csv_io = io.StringIO()
    csv.writer(detail_csv_io).writerows(detail_rows)
    detail_csv = detail_csv_io.getvalue()

    # XLSX bytes
    summary_xlsx = _build_vat_summary_xlsx(summary_rows)
    detail_xlsx = _build_vat_detail_xlsx(detail_rows)

    # PDF bytes
    summary_pdf = generate_vat_return_pdf(
        filing,
        ctx,
        start_date=start_date,
        end_date=end_date,
    )

    detail_pdf = generate_vat_supporting_pdf(
        filing,
        ctx,
        lines,
        start_date=start_date,
        end_date=end_date,
    )

    safe_start = start_date.isoformat()
    safe_end = end_date.isoformat()

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            f"vat_return_summary_{company_id}_{safe_start}_{safe_end}.csv",
            summary_csv,
        )
        zf.writestr(
            f"vat_supporting_schedule_{company_id}_{safe_start}_{safe_end}.csv",
            detail_csv,
        )
        zf.writestr(
            f"vat_return_summary_{company_id}_{safe_start}_{safe_end}.xlsx",
            summary_xlsx,
        )
        zf.writestr(
            f"vat_supporting_schedule_{company_id}_{safe_start}_{safe_end}.xlsx",
            detail_xlsx,
        )
        zf.writestr(
            f"vat_return_summary_{company_id}_{safe_start}_{safe_end}.pdf",
            summary_pdf,
        )
        zf.writestr(
            f"vat_supporting_schedule_{company_id}_{safe_start}_{safe_end}.pdf",
            detail_pdf,
        )

    zip_buffer.seek(0)

    filename = f"vat_pack_{company_id}_{safe_start}_{safe_end}.zip"

    resp = make_response(zip_buffer.getvalue())
    resp.headers["Content-Type"] = "application/zip"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'

    return _corsify(resp)

@vat_utils_bp.route("/api/companies/<int:company_id>/vat/filings/email-pack", methods=["POST", "OPTIONS"])
@require_auth
def vat_filing_email_pack(company_id: int):
    if request.method == "OPTIONS":
        return _corsify(make_response("", 204))

    user = getattr(g, "current_user", {}) or {}

    if int(user.get("company_id") or 0) != int(company_id):
        return jsonify({"ok": False, "error": "Not authorised"}), 403

    payload = request.get_json(force=True) or {}

    start_date = _parse_date((payload.get("from") or "").strip(), None)
    end_date = _parse_date((payload.get("to") or "").strip(), None)

    if not start_date or not end_date:
        return jsonify({"ok": False, "error": "from and to are required"}), 400

    filing = db_service.get_vat_filing(company_id, start_date, end_date)
    if not filing:
        return jsonify({"ok": False, "error": "Prepare VAT return first"}), 400

    user_email = (
        payload.get("email")
        or user.get("email")
        or user.get("user_email")
    )

    if not user_email:
        return jsonify({"ok": False, "error": "No user email address found"}), 400

    ctx = get_company_context(db_service, company_id) or {}
    company_name = ctx.get("company_name") or ctx.get("name") or f"Company {company_id}"

    # This should be the same signed/export URL pattern your frontend uses.

    public_base = (
        os.getenv("PUBLIC_APP_URL")
        or os.getenv("FRONTEND_URL")
        or "https://finspheresolutions.com"
    ).rstrip("/")

    token = create_report_export_token(
        company_id=company_id,
        report_key="vat_pack",
        user_id=int(user.get("id") or 0),
        ttl_seconds=3600,
    )

    qs = urlencode({
        "from": start_date.isoformat(),
        "to": end_date.isoformat(),
        "t": token,
    })

    pack_url = (
        f"{public_base}/api/api/companies/{company_id}/vat/filings/export-pack"
        f"?{qs}"
    )

    subject = f"VAT Pack - {company_name} - {start_date} to {end_date}"

    body = f"""
Hello,

Your VAT Pack is ready.

Company: {company_name}
Period: {start_date} to {end_date}
Status: {filing.get("status")}
Net VAT: {filing.get("net_vat")}

Download VAT Pack:
{pack_url}

This document is prepared by FinSage, no stamp required.

Regards,
FinSage
""".strip()

    try:
        send_mail(
            to_email=user_email,
            subject=subject,
            html_body=body.replace("\n", "<br>"),
            text_body=body,
            from_name="FinSage",
        )
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": f"Failed to send VAT Pack email: {str(e)}"
        }), 500

    return jsonify({
        "ok": True,
        "message": f"VAT Pack email sent to {user_email}",
    }), 200
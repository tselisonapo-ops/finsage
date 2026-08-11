from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER,TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle,getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,Paragraph,SimpleDocTemplate,Spacer,
    Table,TableStyle
)


DOC_META={
    "PURCHASE_REQUISITION":{
        "title":"PURCHASE REQUISITION",
        "accent":"#0B6B5E",
        "section":"GOODS / SERVICES REQUIRED",
        "total":"REQUISITION TOTAL",
    },
    "SERVICE_REQUISITION":{
        "title":"SERVICE REQUISITION",
        "accent":"#34558B",
        "section":"SCOPE & ESTIMATED COSTS",
        "total":"ESTIMATED SERVICE COST",
    },
    "CAPEX_REQUISITION":{
        "title":"CAPITAL EXPENDITURE REQUISITION",
        "accent":"#6D4D2F",
        "section":"ASSET / CAPITAL ITEMS",
        "total":"TOTAL CAPITAL REQUIREMENT",
    },
    "LEASE_REQUISITION":{
        "title":"LEASE REQUISITION",
        "accent":"#66508F",
        "section":"ESTIMATED LEASE COMMITMENT",
        "total":"ESTIMATED COMMITMENT",
    },
    "PAYMENT_REQUEST":{
        "title":"PAYMENT REQUEST",
        "accent":"#9A5A25",
        "section":"PAYMENT LINES",
        "total":"AMOUNT REQUESTED",
    },
    "TRAVEL_AUTHORISATION":{
        "title":"TRAVEL AUTHORISATION",
        "accent":"#286C82",
        "section":"ESTIMATED TRAVEL COSTS",
        "total":"ESTIMATED TRAVEL COST",
    },
    "GENERAL_REQUISITION":{
        "title":"GENERAL REQUISITION",
        "accent":"#4E625D",
        "section":"REQUEST DETAILS",
        "total":"REQUISITION TOTAL",
    },
}


def _meta(payload):
    doc_type=(
        payload.get("document",{}).get("type")
        or "GENERAL_REQUISITION"
    )

    return DOC_META.get(
        doc_type,
        DOC_META["GENERAL_REQUISITION"],
    )


def _money(value,currency=""):
    return f"{currency or ''} {float(value or 0):,.2f}".strip()


def _logo_path(payload,root_path):
    company=payload.get("company") or {}
    logo_url=company.get("logo_url")

    if not logo_url or not root_path:
        return None

    if logo_url.startswith("/static/"):
        path=Path(root_path)/logo_url.lstrip("/")

        if path.exists():
            return path

    return None


def build_requisition_pdf(payload,output_path,*,root_path=None):
    output_path=Path(output_path)
    output_path.parent.mkdir(parents=True,exist_ok=True)

    meta=_meta(payload)
    company=payload.get("company") or {}
    request=payload.get("request") or {}
    items=request.get("items") or payload.get("items") or []
    budget=payload.get("budget") or {}
    approvals=payload.get("approvals") or []
    finance=payload.get("finance_review") or {}

    currency=request.get("currency_code") or company.get("currency") or ""

    styles=getSampleStyleSheet()

    title_style=ParagraphStyle(
        "FinSage NexusTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        alignment=TA_CENTER,
        textColor=colors.HexColor(meta["accent"]),
    )

    small=ParagraphStyle(
        "FinSage NexusSmall",
        parent=styles["Normal"],
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor("#56645F"),
    )

    body=ParagraphStyle(
        "FinSage NexusBody",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
    )

    section=ParagraphStyle(
        "FinSage NexusSection",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor(meta["accent"]),
    )

    doc=SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=14*mm,
        leftMargin=14*mm,
        topMargin=13*mm,
        bottomMargin=13*mm,
        title=meta["title"],
        author="FinSage Nexus",
    )

    story=[]

    logo_path=_logo_path(payload,root_path)

    if logo_path:
        try:
            logo=Image(str(logo_path),width=28*mm,height=18*mm)
            logo._restrictSize(28*mm,18*mm)
        except Exception:
            logo=Paragraph(
                company.get("name") or "",
                small,
            )
    else:
        logo=Paragraph(
            f"<b>{company.get('name') or ''}</b>",
            small,
        )

    company_lines=[
        f"<b>{company.get('name') or ''}</b>",
    ]

    if company.get("company_reg_no"):
        company_lines.append(
            f"Reg: {company['company_reg_no']}"
        )

    if company.get("vat"):
        company_lines.append(
            f"VAT: {company['vat']}"
        )

    if company.get("company_email"):
        company_lines.append(
            company["company_email"]
        )

    if company.get("company_phone"):
        company_lines.append(
            company["company_phone"]
        )

    header=Table(
        [[
            logo,
            Paragraph(meta["title"],title_style),
            Paragraph(
                "<br/>".join(company_lines),
                ParagraphStyle(
                    "Company",
                    parent=small,
                    alignment=TA_RIGHT,
                ),
            ),
        ]],
        colWidths=[43*mm,82*mm,50*mm],
    )

    header.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LINEBELOW",(0,0),(-1,-1),2,colors.HexColor(meta["accent"])),
    ]))

    story.append(header)
    story.append(Spacer(1,5*mm))

    requester=request.get("requester_name") or "-"
    department=request.get("department_name") or "-"
    branch=request.get("branch_name") or "Head office"

    info=[
        ["Document No.",request.get("request_no") or "-","Revision",str(request.get("revision_no") or 1)],
        ["Requester",requester,"Department",department],
        ["Branch",branch,"Priority",request.get("priority") or "-"],
        ["Request Date",str(request.get("created_at") or "")[:10],"Required Date",str(request.get("required_date") or "-")[:10]],
    ]

    info_table=Table(
        info,
        colWidths=[25*mm,62*mm,25*mm,63*mm],
    )

    info_table.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,0),(-1,-1),7.5),
        ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
        ("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
        ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F3F6F5")),
        ("BACKGROUND",(2,0),(2,-1),colors.HexColor("#F3F6F5")),
        ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#D9E1DE")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))

    story.append(info_table)
    story.append(Spacer(1,5*mm))

    story.append(
        Paragraph(
            "BUSINESS PURPOSE",
            section,
        )
    )

    story.append(
        Paragraph(
            request.get("business_purpose")
            or request.get("description")
            or "—",
            body,
        )
    )

    story.append(Spacer(1,5*mm))
    story.append(Paragraph(meta["section"],section))

    item_rows=[
        ["Description","Qty","Unit","Unit Cost","Amount"]
    ]

    total=0.0

    for item in items:
        qty=float(item.get("quantity") or 0)
        unit_cost=float(
            item.get("estimated_unit_cost") or 0
        )
        amount=float(
            item.get("estimated_total")
            or qty*unit_cost
        )
        total+=amount

        item_rows.append([
            item.get("description") or "",
            f"{qty:g}",
            item.get("unit_of_measure") or "",
            _money(unit_cost,currency),
            _money(amount,currency),
        ])

    if len(item_rows)==1:
        item_rows.append([
            "No requisition lines",
            "",
            "",
            "",
            "",
        ])

    item_table=Table(
        item_rows,
        colWidths=[73*mm,15*mm,22*mm,31*mm,34*mm],
        repeatRows=1,
    )

    item_table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor(meta["accent"])),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,0),(-1,-1),7.5),
        ("ALIGN",(1,1),(1,-1),"RIGHT"),
        ("ALIGN",(3,1),(-1,-1),"RIGHT"),
        ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#DDE4E1")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))

    story.append(item_table)

    total_table=Table(
        [[meta["total"],_money(total,currency)]],
        colWidths=[130*mm,45*mm],
    )

    total_table.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ALIGN",(1,0),(1,0),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),7),
        ("LINEABOVE",(0,0),(-1,0),0.8,colors.HexColor(meta["accent"])),
    ]))

    story.append(total_table)
    story.append(Spacer(1,5*mm))

    story.append(Paragraph("FINANCIAL CONTROL",section))

    if finance:
        finance_rows=[
            ["Classification",str(finance.get("classification") or "-").replace("_"," ").title()],
            ["GL Account",f"{finance.get('account_code') or ''} {finance.get('account_name') or ''}".strip() or "-"],
            ["Cost Centre",finance.get("cost_centre_name") or "-"],
            ["Tax Treatment",str(finance.get("tax_treatment") or "-").replace("_"," ").title()],
        ]

        finance_table=Table(
            finance_rows,
            colWidths=[42*mm,133*mm],
        )

        finance_table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F3F6F5")),
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
            ("FONTNAME",(1,0),(1,-1),"Helvetica"),
            ("FONTSIZE",(0,0),(-1,-1),7.5),
            ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#DDE4E1")),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))

        story.append(finance_table)
        story.append(Spacer(1,4*mm))

    if budget:
        budget_rows=[
            ["Approved Budget",_money(budget.get("budget_amount"),currency)],
            ["Actual Expenditure",_money(budget.get("actual_amount"),currency)],
            ["Commitments",_money(budget.get("committed_amount"),currency)],
            ["This Requisition",_money(budget.get("requested_amount"),currency)],
            ["Remaining",_money(budget.get("available_after"),currency)],
            ["Status",str(budget.get("result") or "").upper()],
        ]

        budget_table=Table(
            budget_rows,
            colWidths=[80*mm,95*mm],
        )

        budget_table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F3F6F5")),
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
            ("FONTNAME",(1,0),(1,-1),"Helvetica"),
            ("FONTSIZE",(0,0),(-1,-1),7.5),
            ("ALIGN",(1,0),(1,-1),"RIGHT"),
            ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#DDE4E1")),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))

        story.append(budget_table)

    story.append(Spacer(1,5*mm))
    story.append(Paragraph("APPROVAL TRAIL",section))

    approval_rows=[
        ["Stage","Approver","Decision","Date"]
    ]

    for approval in approvals:
        approval_rows.append([
            approval.get("step_name") or "",
            approval.get("decided_by_name")
            or approval.get("assignee_name")
            or "",
            approval.get("decision")
            or approval.get("status")
            or "",
            str(
                approval.get("decided_at")
                or approval.get("acted_at")
                or ""
            )[:19],
        ])

    if len(approval_rows)==1:
        approval_rows.append([
            "Pending",
            "",
            "",
            "",
        ])

    approval_table=Table(
        approval_rows,
        colWidths=[47*mm,55*mm,32*mm,41*mm],
        repeatRows=1,
    )

    approval_table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F3F6F5")),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,0),(-1,-1),7),
        ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#DDE4E1")),
        ("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))

    story.append(approval_table)
    story.append(Spacer(1,7*mm))

    story.append(
        Paragraph(
            "This document was prepared and approved electronically through FinSage Nexus.",
            small,
        )
    )

    doc.build(story)

    return output_path


def build_requisition_xlsx(payload,output_path):
    output_path=Path(output_path)
    output_path.parent.mkdir(parents=True,exist_ok=True)

    meta=_meta(payload)
    company=payload.get("company") or {}
    request=payload.get("request") or {}
    items=request.get("items") or payload.get("items") or []
    budget=payload.get("budget") or {}
    approvals=payload.get("approvals") or []
    finance=payload.get("finance_review") or {}

    currency=request.get("currency_code") or company.get("currency") or ""

    wb=Workbook()
    ws=wb.active
    ws.title="Requisition"

    accent=meta["accent"].replace("#","")
    thin=Side(style="thin",color="D9E1DE")

    ws.merge_cells("A1:E1")
    ws["A1"]=company.get("name") or ""
    ws["A1"].font=Font(size=16,bold=True,color=accent)

    ws.merge_cells("A2:E2")
    ws["A2"]=meta["title"]
    ws["A2"].font=Font(size=13,bold=True,color=accent)

    ws["A4"]="Document No."
    ws["B4"]=request.get("request_no") or ""
    ws["D4"]="Revision"
    ws["E4"]=request.get("revision_no") or 1

    ws["A5"]="Requester"
    ws["B5"]=request.get("requester_name") or ""
    ws["D5"]="Department"
    ws["E5"]=request.get("department_name") or ""

    ws["A6"]="Branch"
    ws["B6"]=request.get("branch_name") or ""
    ws["D6"]="Priority"
    ws["E6"]=request.get("priority") or ""

    ws["A8"]="Business Purpose"
    ws["A8"].font=Font(bold=True,color=accent)

    ws.merge_cells("A9:E10")
    ws["A9"]=request.get("business_purpose") or request.get("description") or ""
    ws["A9"].alignment=Alignment(wrap_text=True,vertical="top")

    row=12

    headers=["Description","Quantity","Unit","Unit Cost","Amount"]

    for col,value in enumerate(headers,1):
        cell=ws.cell(row=row,column=col,value=value)
        cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor=accent)
        cell.border=Border(
            left=thin,right=thin,
            top=thin,bottom=thin,
        )

    row+=1
    total=0.0

    for item in items:
        qty=float(item.get("quantity") or 0)
        unit_cost=float(item.get("estimated_unit_cost") or 0)
        amount=float(item.get("estimated_total") or qty*unit_cost)
        total+=amount

        values=[
            item.get("description") or "",
            qty,
            item.get("unit_of_measure") or "",
            unit_cost,
            amount,
        ]

        for col,value in enumerate(values,1):
            cell=ws.cell(row=row,column=col,value=value)
            cell.border=Border(
                left=thin,right=thin,
                top=thin,bottom=thin,
            )

        ws.cell(row=row,column=4).number_format='#,##0.00'
        ws.cell(row=row,column=5).number_format='#,##0.00'
        row+=1

    ws.cell(row=row,column=4,value=meta["total"]).font=Font(bold=True)
    ws.cell(row=row,column=5,value=total).font=Font(bold=True)
    ws.cell(row=row,column=5).number_format='#,##0.00'

    row+=3

    ws.cell(row=row,column=1,value="Financial Control").font=Font(bold=True,color=accent)
    row+=1

    finance_rows=[
        ("Classification",finance.get("classification") or ""),
        ("GL Account",f"{finance.get('account_code') or ''} {finance.get('account_name') or ''}".strip()),
        ("Cost Centre",finance.get("cost_centre_name") or ""),
        ("Tax Treatment",finance.get("tax_treatment") or ""),
    ]

    for label,value in finance_rows:
        ws.cell(row=row,column=1,value=label).font=Font(bold=True)
        ws.cell(row=row,column=2,value=value)
        row+=1

    if budget:
        row+=1

        budget_rows=[
            ("Approved Budget",budget.get("budget_amount")),
            ("Actual Expenditure",budget.get("actual_amount")),
            ("Commitments",budget.get("committed_amount")),
            ("This Requisition",budget.get("requested_amount")),
            ("Remaining",budget.get("available_after")),
            ("Status",str(budget.get("result") or "").upper()),
        ]

        for label,value in budget_rows:
            ws.cell(row=row,column=1,value=label).font=Font(bold=True)
            ws.cell(row=row,column=2,value=value)

            if isinstance(value,(int,float)):
                ws.cell(row=row,column=2).number_format='#,##0.00'

            row+=1

    row+=2

    ws.cell(row=row,column=1,value="Approval Trail").font=Font(bold=True,color=accent)
    row+=1

    approval_headers=["Stage","Approver","Decision","Date"]

    for col,value in enumerate(approval_headers,1):
        cell=ws.cell(row=row,column=col,value=value)
        cell.font=Font(bold=True)
        cell.fill=PatternFill("solid",fgColor="F3F6F5")

    row+=1

    for approval in approvals:
        ws.cell(row=row,column=1,value=approval.get("step_name") or "")
        ws.cell(row=row,column=2,value=approval.get("decided_by_name") or approval.get("assignee_name") or "")
        ws.cell(row=row,column=3,value=approval.get("decision") or approval.get("status") or "")
        ws.cell(row=row,column=4,value=str(approval.get("decided_at") or approval.get("acted_at") or ""))
        row+=1

    widths=[42,14,18,20,20]

    for index,width in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(index)].width=width

    ws.freeze_panes="A12"
    ws.sheet_view.showGridLines=False

    wb.save(output_path)

    return output_path